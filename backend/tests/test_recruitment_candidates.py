"""Integration tests — Plan 13.3: Candidate Management."""
from __future__ import annotations

import io
from datetime import date

import pytest
from fastapi.testclient import TestClient

BASE_CAND = "/api/v1/recruitment/candidates"
BASE_JR   = "/api/v1/recruitment/job-requisitions"
BASE_DEPT = "/api/v1/departments"
BASE_POS  = "/api/v1/job-positions"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_TEST_YEAR      = 2097  # Năm không có dữ liệu thực (khác 13.1=2099, 13.2=2098)


# ── Helpers ───────────────────────────────────────────────────────────────────


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _get_dept_id(client: TestClient, h: dict) -> int:
    items = client.get(BASE_DEPT, headers=h, params={"page_size": 1}).json()
    rows = items if isinstance(items, list) else items.get("items", items)
    assert rows, "Cần ít nhất 1 phòng ban"
    return rows[0]["id"]


def _get_pos_id(client: TestClient, h: dict) -> int:
    items = client.get(BASE_POS, headers=h, params={"page_size": 1}).json()
    rows = items if isinstance(items, list) else items.get("items", items)
    assert rows, "Cần ít nhất 1 vị trí công việc"
    return rows[0]["id"]


def _create_approved_jr(client: TestClient, h: dict) -> int:
    dept_id = _get_dept_id(client, h)
    pos_id = _get_pos_id(client, h)
    jr = client.post(
        BASE_JR,
        json={"job_position_id": pos_id, "department_id": dept_id, "quantity": 1, "reason_type": "new"},
        headers=h,
    ).json()
    jr_id = jr["id"]
    client.post(f"{BASE_JR}/{jr_id}/submit", headers=h)
    client.post(f"{BASE_JR}/{jr_id}/approve", headers=h)
    return jr_id


def _create_candidate(client: TestClient, h: dict, suffix: str = "") -> dict:
    res = client.post(
        BASE_CAND,
        json={
            "full_name": f"Nguyễn Test {suffix}",
            "personal_email": f"test_{suffix}_{_TEST_YEAR}@example.com",
            "phone_number": f"090000{suffix[:4] if suffix else '0000'}",
        },
        headers=h,
    )
    assert res.status_code == 201, res.text
    return res.json()


def _get_education_level_id(client: TestClient, h: dict) -> int:
    rows = client.get("/api/v1/lookups/education-levels", headers=h, params={"limit": 20}).json()
    assert rows, "Cần ít nhất 1 trình độ học vấn"
    return rows[0]["id"]


def _get_institution_id(client: TestClient, h: dict) -> int:
    rows = client.get("/api/v1/lookups/educational-institutions", headers=h, params={"limit": 20}).json()
    assert rows, "Cần ít nhất 1 trường học"
    return rows[0]["id"]


def _get_major_id(client: TestClient, h: dict) -> int:
    rows = client.get("/api/v1/lookups/education-majors", headers=h, params={"limit": 20}).json()
    assert rows, "Cần ít nhất 1 chuyên ngành"
    return rows[0]["id"]


def _get_skill_id(client: TestClient, h: dict) -> int:
    rows = client.get("/api/v1/lookups/skills", headers=h, params={"limit": 20}).json()
    assert rows, "Cần ít nhất 1 kỹ năng"
    return rows[0]["id"]


# ── TestCandidateCRUD ─────────────────────────────────────────────────────────


class TestCandidateCRUD:
    def test_create_candidate(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "crud01")
        assert c["full_name"] == "Nguyễn Test crud01"
        assert c["is_active"] is True
        assert c["educations"] == []
        assert c["work_experiences"] == []
        assert c["skills"] == []
        assert c["attachments"] == []
        assert c["active_applications"] == 0
        assert c["identity_strength"] == "medium"
        assert c["conversion_ready"] is False

    def test_create_candidate_requires_identity_anchor(self, client: TestClient) -> None:
        h = _admin(client)
        res = client.post(
            BASE_CAND,
            json={"full_name": "Chỉ Có Họ Tên"},
            headers=h,
        )
        assert res.status_code == 422, res.text
        detail = res.json()["detail"]
        assert any("ít nhất một thông tin định danh" in item["msg"] for item in detail)

    def test_create_candidate_with_id_number_is_strong_identity(self, client: TestClient) -> None:
        h = _admin(client)
        res = client.post(
            BASE_CAND,
            json={
                "full_name": "Nguyễn Strong Identity",
                "id_number": "012345678901",
            },
            headers=h,
        )
        assert res.status_code == 201, res.text
        data = res.json()
        assert data["identity_strength"] == "strong"
        assert data["identity_strength_label"] == "Định danh mạnh"

    def test_create_candidate_accepts_legacy_contact_and_normalizes_nationality(self, client: TestClient) -> None:
        h = _admin(client)
        res = client.post(
            BASE_CAND,
            json={
                "full_name": "Nguyễn Văn Legacy",
                "email": f"legacy_{_TEST_YEAR}@example.com",
                "phone": "0912345678",
                "nationality": "VN",
            },
            headers=h,
        )
        assert res.status_code == 201, res.text
        data = res.json()
        assert data["last_name"] == "Nguyễn Văn"
        assert data["first_name"] == "Legacy"
        assert data["personal_email"] == f"legacy_{_TEST_YEAR}@example.com"
        assert data["phone_number"] == "0912345678"
        assert data["raw_nationality_text"] == "VN"
        assert data["nationality_id"] is not None
        assert data["nationality_name"] == "Việt Nam"

    def test_get_candidate(self, client: TestClient) -> None:
        h = _admin(client)
        created = _create_candidate(client, h, "crud02")
        res = client.get(f"{BASE_CAND}/{created['id']}", headers=h)
        assert res.status_code == 200
        assert res.json()["id"] == created["id"]

    def test_update_candidate(self, client: TestClient) -> None:
        h = _admin(client)
        created = _create_candidate(client, h, "crud03")
        res = client.put(
            f"{BASE_CAND}/{created['id']}",
            json={"full_name": "Nguyễn Đã Sửa", "current_company": "Công ty ABC"},
            headers=h,
        )
        assert res.status_code == 200
        data = res.json()
        assert data["full_name"] == "Nguyễn Đã Sửa"
        assert data["current_company"] == "Công ty ABC"

    def test_list_candidates(self, client: TestClient) -> None:
        h = _admin(client)
        _create_candidate(client, h, "list01")
        res = client.get(BASE_CAND, headers=h)
        assert res.status_code == 200
        body = res.json()
        assert "items" in body
        assert body["total"] >= 1

    def test_soft_delete_candidate(self, client: TestClient) -> None:
        h = _admin(client)
        created = _create_candidate(client, h, "del01")
        cid = created["id"]
        res = client.delete(f"{BASE_CAND}/{cid}", headers=h)
        assert res.status_code == 204
        # Deleted candidate returns 404
        res2 = client.get(f"{BASE_CAND}/{cid}", headers=h)
        assert res2.status_code == 404

    def test_search_candidates(self, client: TestClient) -> None:
        h = _admin(client)
        _create_candidate(client, h, "search01")
        res = client.get(BASE_CAND, headers=h, params={"search": "Nguyễn Test search01"})
        assert res.status_code == 200
        body = res.json()
        assert any("search01" in item["full_name"] for item in body["items"])


class TestCandidateDuplicateCheck:
    def test_check_duplicates_returns_exact_match(self, client: TestClient) -> None:
        import uuid
        h = _admin(client)
        token = uuid.uuid4().hex[:8]
        email = f"dup_exact_{token}_{_TEST_YEAR}@example.com"
        phone = f"09{token[:8]}"
        existing = client.post(
            BASE_CAND,
            json={
                "full_name": f"Nguyễn Văn Trùng {token}",
                "personal_email": email,
                "phone_number": phone,
                "id_number": "012345678999",
            },
            headers=h,
        ).json()

        res = client.post(
            f"{BASE_CAND}/check-duplicates",
            json={
                "full_name": f"Ứng viên mới {token}",
                "personal_email": email,
                "phone_number": phone,
            },
            headers=h,
        )
        assert res.status_code == 200, res.text
        body = res.json()
        assert body["exact_matches"]
        match = next(item for item in body["exact_matches"] if item["candidate_id"] == existing["id"])
        assert "same_personal_email" in match["reason_codes"]
        assert "same_phone_number" in match["reason_codes"]

    def test_check_duplicates_returns_possible_match_for_same_full_name(self, client: TestClient) -> None:
        import uuid
        h = _admin(client)
        token = uuid.uuid4().hex[:8]
        full_name = f"Trần Thị Trùng Tên {token}"
        existing = client.post(
            BASE_CAND,
            json={
                "full_name": full_name,
                "date_of_birth": "1995-05-20",
                "phone_number": f"08{token[:8]}",
            },
            headers=h,
        ).json()

        res = client.post(
            f"{BASE_CAND}/check-duplicates",
            json={
                "full_name": full_name,
                "date_of_birth": "1995-05-20",
            },
            headers=h,
        )
        assert res.status_code == 200, res.text
        body = res.json()
        assert body["exact_matches"] == []
        assert body["possible_matches"]
        match = next(item for item in body["possible_matches"] if item["candidate_id"] == existing["id"])
        assert match["reason_codes"] == ["same_full_name_and_date_of_birth"]

    def test_check_duplicates_excludes_current_candidate(self, client: TestClient) -> None:
        import uuid
        h = _admin(client)
        suffix = f"dup_exclude_{uuid.uuid4().hex[:6]}"
        existing = _create_candidate(client, h, suffix)

        res = client.post(
            f"{BASE_CAND}/check-duplicates",
            json={
                "full_name": existing["full_name"],
                "personal_email": existing["personal_email"],
                "exclude_candidate_id": existing["id"],
            },
            headers=h,
        )
        assert res.status_code == 200, res.text
        body = res.json()
        assert body["exact_matches"] == []
        assert body["possible_matches"] == []


# ── TestCandidateSubResources ─────────────────────────────────────────────────


class TestCandidateSubResources:
    def test_add_education(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "edu01")
        institution_id = _get_institution_id(client, h)
        major_id = _get_major_id(client, h)
        education_level_id = _get_education_level_id(client, h)
        res = client.post(
            f"{BASE_CAND}/{c['id']}/educations",
            json={
                "institution_id": institution_id,
                "major_id": major_id,
                "education_level_id": education_level_id,
                "graduation_year": 2018,
                "diploma_type": "Kỹ sư",
                "is_main_education": True,
            },
            headers=h,
        )
        assert res.status_code == 201
        data = res.json()
        assert data["institution_id"] == institution_id
        assert data["major_id"] == major_id
        assert data["education_level_id"] == education_level_id
        assert data["diploma_type"] == "Kỹ sư"
        assert data["is_main_education"] is True

    def test_update_education(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "edu02")
        institution_id = _get_institution_id(client, h)
        education_level_id = _get_education_level_id(client, h)
        edu = client.post(
            f"{BASE_CAND}/{c['id']}/educations",
            json={
                "institution_id": institution_id,
                "education_level_id": education_level_id,
            },
            headers=h,
        ).json()
        major_id = _get_major_id(client, h)
        res = client.put(
            f"{BASE_CAND}/{c['id']}/educations/{edu['id']}",
            json={
                "institution_id": institution_id,
                "major_id": major_id,
                "education_level_id": education_level_id,
                "graduation_year": 2020,
            },
            headers=h,
        )
        assert res.status_code == 200
        assert res.json()["major_id"] == major_id
        assert res.json()["graduation_year"] == 2020

    def test_delete_education(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "edu03")
        institution_id = _get_institution_id(client, h)
        education_level_id = _get_education_level_id(client, h)
        edu = client.post(
            f"{BASE_CAND}/{c['id']}/educations",
            json={
                "institution_id": institution_id,
                "education_level_id": education_level_id,
            },
            headers=h,
        ).json()
        res = client.delete(f"{BASE_CAND}/{c['id']}/educations/{edu['id']}", headers=h)
        assert res.status_code == 204

    def test_add_work_experience(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "exp01")
        res = client.post(
            f"{BASE_CAND}/{c['id']}/work-experiences",
            json={"company_name": "Công ty ABC", "position_name": "Developer"},
            headers=h,
        )
        assert res.status_code == 201
        assert res.json()["company_name"] == "Công ty ABC"

    def test_add_skill(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "skill01")
        skill_id = _get_skill_id(client, h)
        res = client.post(
            f"{BASE_CAND}/{c['id']}/skills",
            json={"skill_id": skill_id, "proficiency_level": "advanced", "note": "Từ hồ sơ năng lực"},
            headers=h,
        )
        assert res.status_code == 201
        data = res.json()
        assert data["skill_id"] == skill_id
        assert data["skill_name"] != ""
        assert data["note"] == "Từ hồ sơ năng lực"

    def test_add_duplicate_skill_returns_409(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "skill02")
        payload = {"skill_id": _get_skill_id(client, h), "proficiency_level": "beginner"}
        client.post(f"{BASE_CAND}/{c['id']}/skills", json=payload, headers=h)
        res = client.post(f"{BASE_CAND}/{c['id']}/skills", json=payload, headers=h)
        assert res.status_code == 409

    def test_delete_skill(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "skill03")
        skill_id = _get_skill_id(client, h)
        sk = client.post(
            f"{BASE_CAND}/{c['id']}/skills",
            json={"skill_id": skill_id, "proficiency_level": "advanced"},
            headers=h,
        ).json()
        res = client.delete(f"{BASE_CAND}/{c['id']}/skills/{sk['id']}", headers=h)
        assert res.status_code == 204


# ── TestCandidateApplication ──────────────────────────────────────────────────


class TestCandidateApplication:
    def test_apply_candidate_to_approved_jr(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply01")
        jr_id = _create_approved_jr(client, h)
        res = client.post(
            f"{BASE_CAND}/{c['id']}/apply",
            json={"job_requisition_id": jr_id, "applied_date": str(date.today())},
            headers=h,
        )
        assert res.status_code == 201
        data = res.json()
        assert data["candidate_id"] == c["id"]
        assert data["job_requisition_id"] == jr_id
        assert data["current_stage"] == "new"

    def test_apply_sets_jr_in_progress(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply02")
        jr_id = _create_approved_jr(client, h)
        client.post(
            f"{BASE_CAND}/{c['id']}/apply",
            json={"job_requisition_id": jr_id, "applied_date": str(date.today())},
            headers=h,
        )
        jr = client.get(f"{BASE_JR}/{jr_id}", headers=h).json()
        assert jr["status"] == "in_progress"

    def test_apply_duplicate_returns_409(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply03")
        jr_id = _create_approved_jr(client, h)
        payload = {"job_requisition_id": jr_id, "applied_date": str(date.today())}
        client.post(f"{BASE_CAND}/{c['id']}/apply", json=payload, headers=h)
        res = client.post(f"{BASE_CAND}/{c['id']}/apply", json=payload, headers=h)
        assert res.status_code == 409

    def test_apply_draft_jr_returns_422(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply04")
        dept_id = _get_dept_id(client, h)
        pos_id = _get_pos_id(client, h)
        jr = client.post(
            BASE_JR,
            json={"job_position_id": pos_id, "department_id": dept_id, "quantity": 1, "reason_type": "new"},
            headers=h,
        ).json()
        res = client.post(
            f"{BASE_CAND}/{c['id']}/apply",
            json={"job_requisition_id": jr["id"], "applied_date": str(date.today())},
            headers=h,
        )
        assert res.status_code == 422

    def test_list_applications_for_jr(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply05")
        jr_id = _create_approved_jr(client, h)
        client.post(
            f"{BASE_CAND}/{c['id']}/apply",
            json={"job_requisition_id": jr_id, "applied_date": str(date.today())},
            headers=h,
        )
        res = client.get(f"/api/v1/recruitment/job-requisitions/{jr_id}/applications", headers=h)
        assert res.status_code == 200
        body = res.json()
        assert body["total"] >= 1
        assert any(item["candidate_id"] == c["id"] for item in body["items"])

    def test_list_applications_for_candidate(self, client: TestClient) -> None:
        h = _admin(client)
        c = _create_candidate(client, h, "apply06")
        jr_id = _create_approved_jr(client, h)
        client.post(
            f"{BASE_CAND}/{c['id']}/apply",
            json={"job_requisition_id": jr_id, "applied_date": str(date.today())},
            headers=h,
        )
        res = client.get(f"{BASE_CAND}/{c['id']}/applications", headers=h)
        assert res.status_code == 200
        body = res.json()
        assert body["total"] >= 1
        assert any(item["job_requisition_id"] == jr_id for item in body["items"])


# ── TestImportTemplate ────────────────────────────────────────────────────────


class TestImportTemplate:
    def test_download_template_returns_xlsx(self, client: TestClient) -> None:
        h = _admin(client)
        res = client.get(f"{BASE_CAND}/import-template", headers=h)
        assert res.status_code == 200
        assert "spreadsheetml" in res.headers["content-type"]

    def test_import_excel_creates_candidate(self, client: TestClient) -> None:
        # Columns: full_name(0), dob(1), gender(2), id_number(3), phone(4), email(5), ...
        import uuid
        from openpyxl import Workbook
        unique_email = f"import_{uuid.uuid4().hex[:8]}@example.com"
        wb = Workbook()
        ws = wb.active
        ws.append(["Họ và tên (*)", "Ngày sinh", "Giới tính", "CCCD", "Điện thoại", "Email"])
        ws.append([f"Import Test {_TEST_YEAR}", None, None, None, None, unique_email])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        h = _admin(client)
        res = client.post(
            f"{BASE_CAND}/import",
            files={"file": ("candidates.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=h,
        )
        assert res.status_code == 200
        body = res.json()
        assert body["created"] >= 1
        assert body["errors"] == []

    def test_import_excel_updates_existing_by_email(self, client: TestClient) -> None:
        # Columns: full_name(0), dob(1), gender(2), id_number(3), phone(4), email(5), address(6), company(7)
        from openpyxl import Workbook
        email = f"upsert_test_{_TEST_YEAR}@example.com"
        h = _admin(client)

        def _make_row(name, company):
            return [name, None, None, None, None, email, None, company]

        # First import — create
        wb = Workbook()
        ws = wb.active
        ws.append(["full_name", "dob", "gender", "id", "phone", "email", "addr", "company"])
        ws.append(_make_row(f"Upsert Test {_TEST_YEAR}", "Công ty Cũ"))
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        client.post(
            f"{BASE_CAND}/import",
            files={"file": ("c.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=h,
        )

        # Second import — update
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.append(["full_name", "dob", "gender", "id", "phone", "email", "addr", "company"])
        ws2.append(_make_row(f"Upsert Test {_TEST_YEAR}", "Công ty Mới"))
        buf2 = io.BytesIO()
        wb2.save(buf2)
        buf2.seek(0)
        res = client.post(
            f"{BASE_CAND}/import",
            files={"file": ("c2.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
            headers=h,
        )
        assert res.status_code == 200
        body = res.json()
        assert body["updated"] >= 1
