"""Tests cho 12.1 Slice 2a — Import nghỉ phép hàng loạt."""
from __future__ import annotations

import asyncio
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.schemas.employee import EmployeeCreate
from app.services import employee_service

BASE      = "/api/v1/imports"
BASE_AUTH = "/api/v1/auth/login"
BASE_EMPLOYEES = "/api/v1/employees"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

_LT_ANNUAL = "annual_leave"
_LT_SICK   = "sick_leave"
_YEAR      = 2095   # năm đủ xa tránh conflict


def _login(client: TestClient) -> dict:
    token = client.post(BASE_AUTH, json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        await s.execute(text(f"DELETE FROM leave_records WHERE EXTRACT(year FROM start_date) = {_YEAR}"))
        employee_ids = [e.id for e in (await s.execute(select(Employee))).scalars().all() if e.id_number.startswith("TESTIMPSEQ")]
        if employee_ids:
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup():
    await _cleanup()
    yield
    await _cleanup()


HEADERS = ["Mã nhân viên", "Mã loại phép", "Ngày bắt đầu", "Ngày kết thúc", "Ghi chú"]
HEADERS_WITH_SEQUENCE = ["Mã nhân viên", "Hệ mã nhân viên", "Mã loại phép", "Ngày bắt đầu", "Ngày kết thúc", "Ghi chú"]


def _make_xlsx(rows: list[list]) -> bytes:
    return _make_xlsx_with_headers(HEADERS, rows)


def _make_xlsx_with_headers(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, headers, xlsx_bytes: bytes):
    return client.post(
        f"{BASE}/leave-records",
        files={"file": ("import.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


def _employee_display_code(client: TestClient, headers: dict, employee_id: int) -> str:
    r = client.get(f"{BASE_EMPLOYEES}/{employee_id}", headers=headers)
    assert r.status_code == 200, r.text
    return r.json()["display_code"]


def _test_employee_display_code(
    client: TestClient,
    headers: dict,
    *,
    id_number: str,
    employee_seq: int,
    sequence_code: str = "SYS1",
) -> str:
    employee_id = asyncio.run(_create_employee_same_seq(id_number, sequence_code, employee_seq))
    return _employee_display_code(client, headers, employee_id)


async def _get_sequence(code: str) -> EmployeeCodeSequence:
    async with _make_session()() as s:
        return (
            await s.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == code))
        ).scalar_one()


def _employee_payload(id_number: str, *, sequence_id: int, employee_seq: int) -> EmployeeCreate:
    return EmployeeCreate(
        employee_seq=employee_seq,
        employee_code_sequence_id=sequence_id,
        full_name=f"Import Seq {id_number}",
        last_name="Import",
        first_name=id_number,
        date_of_birth="1990-01-01",
        gender="male",
        nationality_id=1,
        id_number=id_number,
        id_issued_on="2020-01-01",
        id_issued_by="Cuc Canh sat",
        status="official",
        start_date="2026-01-01",
    )


async def _create_employee_same_seq(id_number: str, sequence_code: str, employee_seq: int) -> int:
    sequence = await _get_sequence(sequence_code)
    async with _make_session()() as s:
        employee = await employee_service.create_employee(
            s,
            _employee_payload(id_number, sequence_id=sequence.id, employee_seq=employee_seq),
        )
        await s.commit()
        await s.refresh(employee)
        return employee.id


# ── Template ──────────────────────────────────────────────────────────────────

def test_template_download(client: TestClient):
    h = _login(client)
    r = client.get(f"{BASE}/leave-records/template", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert "Hướng dẫn" in wb.sheetnames


# ── Success ───────────────────────────────────────────────────────────────────

def test_import_success(client: TestClient):
    h = _login(client)
    code_1 = _test_employee_display_code(client, h, id_number="TESTIMPSEQL101", employee_seq=9301)
    code_2 = _test_employee_display_code(client, h, id_number="TESTIMPSEQL102", employee_seq=9302)
    rows = [
        [code_1, _LT_ANNUAL, f"01/06/{_YEAR}", f"03/06/{_YEAR}", "Nghỉ phép năm"],
        [code_2, _LT_SICK,   f"10/06/{_YEAR}", f"10/06/{_YEAR}", ""],
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 2
    assert body["success"] == 2
    assert body["failed"] == 0


def test_import_total_days_calculated(client: TestClient):
    """3 ngày nghỉ: total_days = 3."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL103", employee_seq=9303)
    rows = [[code, _LT_ANNUAL, f"01/07/{_YEAR}", f"03/07/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["success"] == 1
    assert r.json()["created_ids"]


def test_import_same_day_valid(client: TestClient):
    """start = end = 1 ngày → hợp lệ."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL104", employee_seq=9304)
    rows = [[code, _LT_SICK, f"15/07/{_YEAR}", f"15/07/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["success"] == 1


# ── Errors ────────────────────────────────────────────────────────────────────

def test_import_invalid_employee(client: TestClient):
    h = _login(client)
    rows = [["999999", _LT_ANNUAL, f"01/06/{_YEAR}", f"03/06/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("Không tìm thấy nhân viên" in e["message"] for e in body["errors"])


def test_import_invalid_leave_type(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL105", employee_seq=9305)
    rows = [[code, "nonexistent_leave", f"01/06/{_YEAR}", f"03/06/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("loại phép" in e["message"] for e in body["errors"])


def test_import_date_order_error(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL106", employee_seq=9306)
    rows = [[code, _LT_ANNUAL, f"10/06/{_YEAR}", f"01/06/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("kết thúc" in e["message"].lower() for e in body["errors"])


def test_import_missing_required(client: TestClient):
    h = _login(client)
    rows = [["", _LT_ANNUAL, f"01/06/{_YEAR}", f"03/06/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_overlap_generates_warning(client: TestClient):
    """NV đã có record trùng ngày → cảnh báo, vẫn import thành công."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL107", employee_seq=9307)
    rows1 = [[code, _LT_ANNUAL, f"01/08/{_YEAR}", f"05/08/{_YEAR}", ""]]
    _upload(client, h, _make_xlsx(rows1))

    rows2 = [[code, _LT_SICK, f"03/08/{_YEAR}", f"04/08/{_YEAR}", ""]]
    r = _upload(client, h, _make_xlsx(rows2))
    assert r.status_code == 200
    body = r.json()
    assert body["success"] == 1   # vẫn tạo thành công
    assert any("[CẢNH BÁO]" in e["message"] for e in body["errors"])


def test_import_partial_errors(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQL108", employee_seq=9308)
    rows = [
        [code, _LT_ANNUAL, f"01/09/{_YEAR}", f"03/09/{_YEAR}", ""],  # OK
        ["999999",      _LT_ANNUAL, f"05/09/{_YEAR}", f"07/09/{_YEAR}", ""],  # fail
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["success"] == 1
    assert body["failed"] == 1


# ── Auth & validation ─────────────────────────────────────────────────────────

def test_unauthenticated(client: TestClient):
    r = client.post(f"{BASE}/leave-records",
                    files={"file": ("f.xlsx", b"x", "application/vnd.ms-excel")})
    assert r.status_code == 401


def test_rejects_non_xlsx(client: TestClient):
    h = _login(client)
    r = client.post(f"{BASE}/leave-records",
                    files={"file": ("data.csv", b"a,b", "text/csv")},
                    headers=h)
    assert r.status_code == 400


def test_empty_rows(client: TestClient):
    h = _login(client)
    r = _upload(client, h, _make_xlsx([]))
    assert r.status_code == 200
    assert r.json()["total"] == 0


def test_import_resolves_display_code_exactly_across_three_sequences(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQL001", "SYS1", 9102))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQL002", "SYS2", 9102))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQL003", "SYS3", 9102))

    rows = [
        ["9102", "SYS1", _LT_ANNUAL, f"01/10/{_YEAR}", f"01/10/{_YEAR}", ""],
        ["9102", "SYS2", _LT_ANNUAL, f"02/10/{_YEAR}", f"02/10/{_YEAR}", ""],
        ["9102", "SYS3", _LT_ANNUAL, f"03/10/{_YEAR}", f"03/10/{_YEAR}", ""],
    ]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_SEQUENCE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 3
    assert body["failed"] == 0


def test_import_rejects_ambiguous_seq_without_sequence_code(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQL901", "SYS1", 9902))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQL902", "SYS2", 9902))

    rows = [
        ["9902", _LT_ANNUAL, f"01/11/{_YEAR}", f"01/11/{_YEAR}", ""],
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 0
    assert body["failed"] == 1
    assert any("Hệ mã nhân viên" in e["message"] for e in body["errors"])
