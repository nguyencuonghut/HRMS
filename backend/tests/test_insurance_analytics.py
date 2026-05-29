"""Tests cho 11.4 — Insurance Analytics."""
from __future__ import annotations

import asyncio
import io
from datetime import date
import pytest
from fastapi.testclient import TestClient
import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings

BASE = "/api/v1/reports/insurance"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

_TEST_YEAR = 2098
_TEST_MONTH = 5


def _login(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        # Xóa các event dùng trong test
        await s.execute(
            text("DELETE FROM insurance_change_events WHERE period_year = :y"),
            {"y": _TEST_YEAR},
        )
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup_db():
    await _cleanup()
    yield
    await _cleanup()


async def _get_employee_id() -> int:
    async with _make_session()() as s:
        # Lấy employee đầu tiên sẵn có trong DB test
        r = await s.execute(text("SELECT id FROM employees LIMIT 1"))
        row = r.fetchone()
        assert row is not None, "Không tìm thấy employee trong DB"
        return row[0]


async def _create_test_event(
    employee_id: int,
    change_type: str,
    basis_amount: float,
    effective_date: date,
):
    async with _make_session()() as s:
        await s.execute(
            text(
                """
                INSERT INTO insurance_change_events (
                    employee_id, change_type, change_reason, ibhxh_reason_code,
                    effective_date, period_year, period_month,
                    employee_name_snapshot, date_of_birth_snapshot, gender_snapshot,
                    basis_amount, allowances_amount, suggested_declaration_year,
                    suggested_declaration_month, is_manual, new_status
                ) VALUES (
                    :eid, :ctype, 'new_hire', 'T-01', :eff_date, :year, :month,
                    'Test Employee', '1990-01-01', 'male', :basis, 0, :year, :month, TRUE, 'active'
                )
                """
            ),
            {
                "eid": employee_id,
                "ctype": change_type,
                "eff_date": effective_date,
                "year": _TEST_YEAR,
                "month": _TEST_MONTH,
                "basis": basis_amount,
            },
        )
        await s.commit()


def test_insurance_dashboard_kpi(client: TestClient):
    headers = _login(client)
    
    # 1. Gọi khi chưa có test event
    resp = client.get(
        f"{BASE}/dashboard",
        params={"year": _TEST_YEAR, "month": _TEST_MONTH},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["year"] == _TEST_YEAR
    assert data["month"] == _TEST_MONTH
    assert data["increased_count"] == 0
    assert data["decreased_count"] == 0
    assert data["net_change"] == 0

    # 2. Tạo test event
    emp_id = asyncio.run(_get_employee_id())
    asyncio.run(
        _create_test_event(emp_id, "increase", 15000000.0, date(_TEST_YEAR, _TEST_MONTH, 1))
    )

    # 3. Gọi lại và kiểm tra số tăng
    resp = client.get(
        f"{BASE}/dashboard",
        params={"year": _TEST_YEAR, "month": _TEST_MONTH},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["increased_count"] == 1
    assert data["net_change"] == 1


def test_insurance_monthly_changes(client: TestClient):
    headers = _login(client)
    
    resp = client.get(
        f"{BASE}/monthly-changes",
        params={"year": _TEST_YEAR},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["year"] == _TEST_YEAR
    assert len(data["data"]) == 12
    # Đảm bảo các tháng trống đều có giá trị 0
    for point in data["data"]:
        assert point["increased"] == 0
        assert point["decreased"] == 0
        assert point["net"] == 0


def test_insurance_payroll_fund(client: TestClient):
    headers = _login(client)

    resp = client.get(
        f"{BASE}/payroll-fund",
        params={"year": _TEST_YEAR},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["year"] == _TEST_YEAR
    assert len(data["data"]) == 12


def test_insurance_non_participants(client: TestClient):
    headers = _login(client)

    resp = client.get(
        f"{BASE}/non-participants",
        params={"page": 1, "page_size": 5},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert "items" in data
    assert "total" in data
    assert data["page"] == 1
    assert data["page_size"] == 5


def test_insurance_department_breakdown(client: TestClient):
    headers = _login(client)

    resp = client.get(
        f"{BASE}/by-department",
        params={"year": _TEST_YEAR, "month": _TEST_MONTH},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert "items" in data
    assert data["year"] == _TEST_YEAR
    assert data["month"] == _TEST_MONTH


def test_insurance_employee_history(client: TestClient):
    headers = _login(client)
    emp_id = asyncio.run(_get_employee_id())

    resp = client.get(
        f"{BASE}/employee-history",
        params={"employee_id": emp_id, "year": _TEST_YEAR},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["employee_id"] == emp_id
    assert "history" in data


def test_insurance_export_excel(client: TestClient):
    headers = _login(client)

    resp = client.get(
        f"{BASE}/export",
        params={"year": _TEST_YEAR, "month": _TEST_MONTH},
        headers=headers,
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers.get("content-type", "")

    # Đọc lại file Excel để kiểm thử openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) == 3
    assert wb.sheetnames[0] == "Tổng quan"
    assert wb.sheetnames[1] == "Cơ cấu phòng ban"
    assert wb.sheetnames[2] == "Không tham gia BHXH"
