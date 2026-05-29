"""Integration tests cho Plan 10.1 — KPI Tháng (Slice 1: CRUD + Slice 2: Import/Template)."""
from __future__ import annotations

import io
import uuid

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings

BASE_KPI       = "/api/v1/performance/kpi"
BASE_EMPLOYEES = "/api/v1/employees"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_RUN_ID = uuid.uuid4().hex[:8]

# Dùng RUN_ID để sinh year duy nhất mỗi lần chạy, tránh conflict dữ liệu cũ
_BASE       = 2000 + int(_RUN_ID[:2], 16) % 80   # 2000–2079
_YEAR       = _BASE        # năm dùng cho TestKpiCRUD
_MONTH      = 6
_LIST_YEAR  = _BASE + 1    # năm riêng cho TestKpiList
_IMPORT_YEAR = _BASE + 2   # năm riêng cho TestKpiImport


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup_kpi_test_rows():
    async with _make_session()() as session:
        await session.execute(
            text(
                """
                DELETE FROM employee_kpi_monthly
                WHERE year = ANY(:years)
                """
            ),
            {"years": [_YEAR, _LIST_YEAR, _IMPORT_YEAR]},
        )
        await session.commit()


@pytest.fixture(autouse=True)
async def cleanup_kpi_rows():
    await _cleanup_kpi_test_rows()
    yield
    await _cleanup_kpi_test_rows()


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _get_emp_id(client, h) -> int:
    r = client.get(BASE_EMPLOYEES, headers=h, params={"page_size": 1})
    items = r.json()["items"]
    assert items, "Cần ít nhất 1 nhân viên trong DB"
    return items[0]["id"]


def _create_kpi(client, h, emp_id: int, year: int = _YEAR, month: int = _MONTH, score: float = 85.0) -> dict:
    r = client.post(
        BASE_KPI,
        json={"employee_id": emp_id, "year": year, "month": month, "score": score},
        headers=h,
    )
    assert r.status_code == 201, r.text
    return r.json()


# ── TestKpiCRUD ───────────────────────────────────────────────────────────────


class TestKpiCRUD:
    def test_create_kpi_success(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        kpi = _create_kpi(client, h, emp_id)
        assert kpi["id"] > 0
        assert kpi["employee_id"] == emp_id
        assert kpi["year"] == _YEAR
        assert kpi["month"] == _MONTH
        assert float(kpi["score"]) == 85.0

    def test_create_kpi_stores_note(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        r = client.post(
            BASE_KPI,
            json={"employee_id": emp_id, "year": _YEAR, "month": 7, "score": 90.0, "note": f"note_{_RUN_ID}"},
            headers=h,
        )
        assert r.status_code == 201, r.text
        assert r.json()["note"] == f"note_{_RUN_ID}"

    def test_create_kpi_duplicate_returns_409(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_kpi(client, h, emp_id, month=8)
        r = client.post(
            BASE_KPI,
            json={"employee_id": emp_id, "year": _YEAR, "month": 8, "score": 70.0},
            headers=h,
        )
        assert r.status_code == 409, r.text

    def test_create_kpi_invalid_score_returns_422(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        r = client.post(
            BASE_KPI,
            json={"employee_id": emp_id, "year": _YEAR, "month": 9, "score": 150.0},
            headers=h,
        )
        assert r.status_code == 422

    def test_create_kpi_negative_score_returns_422(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        r = client.post(
            BASE_KPI,
            json={"employee_id": emp_id, "year": _YEAR, "month": 10, "score": -1.0},
            headers=h,
        )
        assert r.status_code == 422

    def test_create_kpi_invalid_month_returns_422(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        r = client.post(
            BASE_KPI,
            json={"employee_id": emp_id, "year": _YEAR, "month": 13, "score": 80.0},
            headers=h,
        )
        assert r.status_code == 422

    def test_create_kpi_nonexistent_employee_returns_404(self, client: TestClient):
        h = _admin(client)
        r = client.post(
            BASE_KPI,
            json={"employee_id": 999999, "year": _YEAR, "month": 11, "score": 80.0},
            headers=h,
        )
        assert r.status_code == 404

    def test_get_kpi_by_id(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        created = _create_kpi(client, h, emp_id, month=1)
        r = client.get(f"{BASE_KPI}/{created['id']}", headers=h)
        assert r.status_code == 200
        assert r.json()["id"] == created["id"]

    def test_get_nonexistent_returns_404(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_KPI}/999999", headers=h)
        assert r.status_code == 404

    def test_update_kpi_score(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        kpi = _create_kpi(client, h, emp_id, month=2)
        r = client.put(f"{BASE_KPI}/{kpi['id']}", json={"score": 95.5}, headers=h)
        assert r.status_code == 200
        assert float(r.json()["score"]) == 95.5

    def test_update_kpi_note(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        kpi = _create_kpi(client, h, emp_id, month=3)
        r = client.put(f"{BASE_KPI}/{kpi['id']}", json={"note": "updated note"}, headers=h)
        assert r.status_code == 200
        assert r.json()["note"] == "updated note"

    def test_delete_kpi(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        kpi = _create_kpi(client, h, emp_id, month=4)
        r = client.delete(f"{BASE_KPI}/{kpi['id']}", headers=h)
        assert r.status_code == 204
        r2 = client.get(f"{BASE_KPI}/{kpi['id']}", headers=h)
        assert r2.status_code == 404

    def test_delete_nonexistent_returns_404(self, client: TestClient):
        h = _admin(client)
        r = client.delete(f"{BASE_KPI}/999999", headers=h)
        assert r.status_code == 404


# ── TestKpiList ───────────────────────────────────────────────────────────────


class TestKpiList:
    def test_list_returns_items(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_kpi(client, h, emp_id, year=_LIST_YEAR, month=1)
        r = client.get(BASE_KPI, headers=h)
        assert r.status_code == 200
        data = r.json()
        assert "items" in data
        assert data["total"] >= 1

    def test_list_filter_by_year(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_kpi(client, h, emp_id, year=_LIST_YEAR, month=2)
        r = client.get(BASE_KPI, headers=h, params={"year": _LIST_YEAR})
        assert r.status_code == 200
        items = r.json()["items"]
        assert len(items) >= 1
        assert all(i["year"] == _LIST_YEAR for i in items)

    def test_list_filter_by_month(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_kpi(client, h, emp_id, year=_LIST_YEAR, month=3)
        r = client.get(BASE_KPI, headers=h, params={"year": _LIST_YEAR, "month": 3})
        assert r.status_code == 200
        items = r.json()["items"]
        assert len(items) >= 1
        assert all(i["month"] == 3 for i in items)

    def test_list_search_by_employee_name(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_kpi(client, h, emp_id, year=_LIST_YEAR, month=4)
        emp_name = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["full_name"]
        r = client.get(BASE_KPI, headers=h, params={"search": emp_name[:4]})
        assert r.status_code == 200
        items = r.json()["items"]
        assert len(items) >= 1

    def test_pagination(self, client: TestClient):
        h = _admin(client)
        r = client.get(BASE_KPI, headers=h, params={"page": 1, "page_size": 2})
        assert r.status_code == 200
        data = r.json()
        assert data["page"] == 1
        assert data["page_size"] == 2
        assert len(data["items"]) <= 2

    def test_list_response_includes_employee_info(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        kpi = _create_kpi(client, h, emp_id, year=_LIST_YEAR, month=5)
        r = client.get(BASE_KPI, headers=h, params={"year": _LIST_YEAR, "month": 5})
        assert r.status_code == 200
        items = r.json()["items"]
        found = next((i for i in items if i["id"] == kpi["id"]), None)
        assert found is not None
        assert found["employee_code"] != ""
        assert found["employee_name"] != ""


# ── Helpers cho Import ────────────────────────────────────────────────────────


def _make_xlsx(rows: list[tuple]) -> bytes:
    """Tạo file xlsx với header + data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Mã nhân viên (*)", "Họ và tên (tham khảo)", "Năm (*)", "Tháng (*)", "Điểm KPI (*)", "Ghi chú"])
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client: TestClient, h: dict, content: bytes, filename: str = "kpi.xlsx") -> dict:
    r = client.post(
        f"{BASE_KPI}/import",
        headers=h,
        files={"file": (filename, content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    return r


# ── TestKpiImport ─────────────────────────────────────────────────────────────


class TestKpiImport:
    def test_import_creates_new_records(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        emp_code = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["display_code"]
        content = _make_xlsx([(emp_code, "", _IMPORT_YEAR, 1, 88.0, "")])
        r = _upload(client, h, content)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["created"] == 1
        assert data["updated"] == 0
        assert data["skipped"] == 0

    def test_import_updates_existing_records(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        emp_code = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["display_code"]
        content = _make_xlsx([(emp_code, "", _IMPORT_YEAR, 2, 70.0, "")])
        _upload(client, h, content)  # first import: creates
        content2 = _make_xlsx([(emp_code, "", _IMPORT_YEAR, 2, 95.0, "updated")])
        r = _upload(client, h, content2)  # second import: updates
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["updated"] == 1
        assert data["created"] == 0

    def test_import_skips_invalid_employee_code(self, client: TestClient):
        h = _admin(client)
        content = _make_xlsx([("KHONGCO999", "", _IMPORT_YEAR, 3, 80.0, "")])
        r = _upload(client, h, content)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["skipped"] == 1
        assert data["created"] == 0
        assert len(data["errors"]) == 1

    def test_import_skips_invalid_score(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        emp_code = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["display_code"]
        content = _make_xlsx([(emp_code, "", _IMPORT_YEAR, 4, 150.0, "")])
        r = _upload(client, h, content)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["skipped"] == 1
        assert data["created"] == 0

    def test_import_skips_invalid_month(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        emp_code = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["display_code"]
        content = _make_xlsx([(emp_code, "", _IMPORT_YEAR, 13, 80.0, "")])
        r = _upload(client, h, content)
        assert r.status_code == 200, r.text
        assert r.json()["skipped"] == 1

    def test_import_result_counts_correct(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        emp_code = client.get(f"{BASE_EMPLOYEES}/{emp_id}", headers=h).json()["display_code"]
        # 1 valid new + 1 invalid code + 1 invalid score
        content = _make_xlsx([
            (emp_code, "", _IMPORT_YEAR, 5, 75.0, ""),
            ("BADCODE", "", _IMPORT_YEAR, 5, 75.0, ""),
            (emp_code, "", _IMPORT_YEAR, 6, 999.0, ""),
        ])
        r = _upload(client, h, content)
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["created"] == 1
        assert data["skipped"] == 2
        assert len(data["errors"]) == 2

    def test_import_non_xlsx_returns_422(self, client: TestClient):
        h = _admin(client)
        r = client.post(
            f"{BASE_KPI}/import",
            headers=h,
            files={"file": ("kpi.txt", b"not an xlsx file", "text/plain")},
        )
        assert r.status_code == 422


# ── TestKpiTemplate ───────────────────────────────────────────────────────────


class TestKpiTemplate:
    def test_template_returns_xlsx_content_type(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_KPI}/template", headers=h)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_template_is_valid_xlsx(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_KPI}/template", headers=h)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        assert ws.cell(1, 1).value is not None  # header exists

    def test_template_has_correct_headers(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_KPI}/template", headers=h)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, 7)]
        assert "Mã nhân viên (*)" in headers
        assert "Điểm KPI (*)" in headers
