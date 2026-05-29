"""Integration tests cho Plan 12.2 Slice 1 — Xuất biểu mẫu BHXH D02-LT và D03-TS."""
from __future__ import annotations

import asyncio
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings

BASE = "/api/v1/exports/bhxh"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ── Auth helper ───────────────────────────────────────────────────────────────

def _login(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


# ── DB helper — seed employee với insurance profile xác định ─────────────────

def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _set_basis_amount(employee_id: int, amount: int) -> int | None:
    """Đặt insurance_basis_amount cho employee, trả về giá trị cũ (None nếu chưa có)."""
    async with _make_session()() as s:
        row = (await s.execute(text(
            "SELECT insurance_basis_amount FROM employee_insurance_profiles WHERE employee_id = :eid"
        ), {"eid": employee_id})).fetchone()
        old_val = int(row[0]) if row and row[0] is not None else None

        await s.execute(text("""
            UPDATE employee_insurance_profiles
            SET insurance_basis_amount = :amt, insurance_basis_source = 'manual_fixed'
            WHERE employee_id = :eid
        """), {"amt": amount, "eid": employee_id})
        await s.commit()
    return old_val


async def _restore_basis_amount(employee_id: int, old_val: int | None) -> None:
    async with _make_session()() as s:
        await s.execute(text("""
            UPDATE employee_insurance_profiles
            SET insurance_basis_amount = :amt
            WHERE employee_id = :eid
        """), {"amt": old_val, "eid": employee_id})
        await s.commit()


async def _find_active_insured_employee_id() -> int | None:
    """Tìm employee đang tham gia BHXH active."""
    async with _make_session()() as s:
        row = (await s.execute(text("""
            SELECT e.id FROM employees e
            JOIN employee_insurance_profiles eip ON eip.employee_id = e.id
            WHERE eip.participation_status = 'active'
            ORDER BY e.id
            LIMIT 1
        """))).fetchone()
    return row[0] if row else None


# ── Tests D02-LT ──────────────────────────────────────────────────────────────

class TestD02LT:
    def test_d02_lt_returns_xlsx(self, client: TestClient):
        headers = _login(client)
        resp = client.get(f"{BASE}/d02-lt?year=2026&month=5", headers=headers)
        assert resp.status_code == 200
        assert _XLSX_MIME in resp.headers["content-type"]

    def test_d02_lt_has_correct_headers(self, client: TestClient):
        headers = _login(client)
        resp = client.get(f"{BASE}/d02-lt?year=2026&month=5", headers=headers)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Row 4 col A = "STT"
        assert ws.cell(row=4, column=1).value == "STT"
        # Verify một số header khác
        assert ws.cell(row=4, column=2).value == "Họ và tên"
        assert ws.cell(row=4, column=11).value == "Trạng thái"

    def test_d02_lt_requires_permission(self, client: TestClient):
        resp = client.get(f"{BASE}/d02-lt?year=2026&month=5")
        assert resp.status_code == 401

    def test_d02_lt_title_row(self, client: TestClient):
        headers = _login(client)
        resp = client.get(f"{BASE}/d02-lt?year=2026&month=5", headers=headers)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        assert ws["A1"].value == "DANH SÁCH LAO ĐỘNG THAM GIA BHXH, BHYT, BHTN"


# ── Tests D03-TS ──────────────────────────────────────────────────────────────

class TestD03TS:
    def test_d03_ts_returns_xlsx(self, client: TestClient):
        headers = _login(client)
        resp = client.get(f"{BASE}/d03-ts?year=2026&month=5", headers=headers)
        assert resp.status_code == 200
        assert _XLSX_MIME in resp.headers["content-type"]

    def test_d03_ts_has_total_row(self, client: TestClient):
        headers = _login(client)
        resp = client.get(f"{BASE}/d03-ts?year=2026&month=5", headers=headers)
        assert resp.status_code == 200
        wb = openpyxl.load_workbook(BytesIO(resp.content))
        ws = wb.active
        # Tìm dòng TỔNG CỘNG ở cột B trong toàn sheet
        found = any(
            ws.cell(row=r, column=2).value == "TỔNG CỘNG"
            for r in range(1, ws.max_row + 1)
        )
        assert found, "Không tìm thấy dòng TỔNG CỘNG trong workbook"

    def test_d03_ts_requires_permission(self, client: TestClient):
        resp = client.get(f"{BASE}/d03-ts?year=2026&month=5")
        assert resp.status_code == 401

    def test_d03_ts_calculation_correct(self, client: TestClient):
        """Với employee có insurance_basis_amount=5000000, BHXH NLĐ (8%) = 400000."""
        loop = asyncio.new_event_loop()
        emp_id = loop.run_until_complete(_find_active_insured_employee_id())
        assert emp_id is not None, "Không có employee BHXH active trong DB"

        old_val = loop.run_until_complete(_set_basis_amount(emp_id, 5_000_000))
        try:
            headers = _login(client)
            resp = client.get(f"{BASE}/d03-ts?year=2026&month=5", headers=headers)
            assert resp.status_code == 200
            wb = openpyxl.load_workbook(BytesIO(resp.content))
            ws = wb.active

            # Tìm dòng của employee theo emp_id trong workbook
            # (tất cả employee active đều có trong file, tìm dòng có basis = 5_000_000)
            found_row = None
            for r in range(5, ws.max_row + 1):
                val = ws.cell(row=r, column=3).value
                if val is not None and abs(float(val) - 5_000_000) < 1:
                    found_row = r
                    break

            assert found_row is not None, "Không tìm thấy dòng có basis=5000000 trong workbook"
            # Col 4 = BHXH NLĐ 8% = 400_000
            assert ws.cell(row=found_row, column=4).value == pytest.approx(400_000)
        finally:
            loop.run_until_complete(_restore_basis_amount(emp_id, old_val))
            loop.close()
