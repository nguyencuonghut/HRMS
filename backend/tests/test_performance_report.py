"""Integration tests cho Plan 10.4 — Báo cáo hiệu suất."""
from __future__ import annotations

import io
import uuid

import pytest
from fastapi.testclient import TestClient

BASE_KPI      = "/api/v1/performance/kpi"
BASE_REVIEWS  = "/api/v1/performance/yearly-reviews"
BASE_REPORT   = "/api/v1/performance/report"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_RUN_ID = uuid.uuid4().hex[:8]

# Dùng năm rất cao để tránh va chạm với dữ liệu thực
_YEAR = 2097


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _get_emp_id(client: TestClient, h: dict) -> int:
    items = client.get("/api/v1/employees", headers=h, params={"page_size": 1}).json()["items"]
    assert items, "Cần ít nhất 1 nhân viên trong DB"
    return items[0]["id"]


def _seed_kpi(client: TestClient, h: dict, emp_id: int, year: int, scores: list):
    r = client.get(BASE_KPI, headers=h, params={"year": year, "page_size": 50})
    for item in r.json().get("items", []):
        if item["employee_id"] == emp_id:
            client.delete(f"{BASE_KPI}/{item['id']}", headers=h)
    for month, score in scores:
        r = client.post(BASE_KPI, json={"employee_id": emp_id, "year": year, "month": month, "score": score}, headers=h)
        assert r.status_code == 201, r.text


def _create_review(client: TestClient, h: dict, emp_id: int, year: int, rating: str) -> dict:
    r = client.get(BASE_REVIEWS, headers=h, params={"year": year, "page_size": 50})
    for item in r.json().get("items", []):
        if item["employee_id"] == emp_id:
            client.delete(f"{BASE_REVIEWS}/{item['id']}", headers=h)
    r = client.post(BASE_REVIEWS, json={"employee_id": emp_id, "year": year, "rating": rating}, headers=h)
    assert r.status_code == 201, r.text
    return r.json()


# ── TestRatingDistribution ────────────────────────────────────────────────────


class TestRatingDistribution:
    def test_distribution_counts_correct(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_review(client, h, emp_id, _YEAR, "tot")

        r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": _YEAR})
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["total_reviewed"] >= 1
        total_in_dist = sum(d["count"] for d in data["distribution"])
        assert total_in_dist == data["total_reviewed"]

    def test_distribution_includes_all_four_ratings(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": _YEAR})
        assert r.status_code == 200
        ratings = {d["rating"] for d in r.json()["distribution"]}
        assert ratings == {"xuat_sac", "tot", "dat", "can_cai_thien"}

    def test_coverage_rate_calculation(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": _YEAR})
        assert r.status_code == 200
        data = r.json()
        if data["total_employees"] > 0:
            expected = round(data["total_reviewed"] / data["total_employees"] * 100, 1)
            assert abs(data["coverage_rate"] - expected) < 0.2

    def test_percentage_sums_to_100(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _create_review(client, h, emp_id, _YEAR, "dat")

        r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": _YEAR})
        assert r.status_code == 200
        data = r.json()
        if data["total_reviewed"] > 0:
            total_pct = sum(d["percentage"] for d in data["distribution"])
            assert abs(total_pct - 100.0) < 1.0  # rounding tolerance

    def test_empty_year_returns_zero_counts(self, client: TestClient):
        h = _admin(client)
        # Dùng năm không có dữ liệu
        r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": 2096})
        assert r.status_code == 200
        data = r.json()
        assert data["total_reviewed"] == 0
        assert all(d["count"] == 0 for d in data["distribution"])


# ── TestDepartmentKpi ─────────────────────────────────────────────────────────


class TestDepartmentKpi:
    def test_avg_score_calculated_per_department(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _seed_kpi(client, h, emp_id, _YEAR, [(1, 80.0), (2, 90.0)])

        r = client.get(f"{BASE_REPORT}/department-kpi", headers=h, params={"year": _YEAR})
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        assert len(items) >= 1
        for item in items:
            assert "avg_score" in item
            assert "employee_count" in item

    def test_filter_by_month(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        _seed_kpi(client, h, emp_id, _YEAR, [(3, 75.0)])

        r = client.get(f"{BASE_REPORT}/department-kpi", headers=h, params={"year": _YEAR, "month": 3})
        assert r.status_code == 200, r.text
        items = r.json()
        assert isinstance(items, list)
        # Ít nhất 1 phòng ban có dữ liệu tháng 3
        assert any(i["months_data_count"] >= 1 for i in items)

    def test_filter_by_department_id(self, client: TestClient):
        h = _admin(client)
        # Lấy danh sách phòng ban
        depts = client.get("/api/v1/departments", headers=h, params={"page_size": 5}).json()
        if not depts:
            pytest.skip("Không có phòng ban trong DB")
        dept_id = depts[0]["id"]

        r = client.get(f"{BASE_REPORT}/department-kpi", headers=h, params={"year": _YEAR, "department_id": dept_id})
        assert r.status_code == 200, r.text
        items = r.json()
        # Chỉ trả phòng ban được lọc hoặc danh sách rỗng
        for item in items:
            assert item["department_id"] == dept_id

    def test_no_data_returns_empty_list(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/department-kpi", headers=h, params={"year": 2100})
        assert r.status_code == 200
        assert r.json() == []


# ── TestMonthlyTrend ──────────────────────────────────────────────────────────


class TestMonthlyTrend:
    def test_trend_has_12_months(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/monthly-trend", headers=h, params={"year": _YEAR})
        assert r.status_code == 200, r.text
        data = r.json()
        assert len(data["points"]) == 12
        assert [p["month"] for p in data["points"]] == list(range(1, 13))

    def test_months_without_data_return_null_avg(self, client: TestClient):
        h = _admin(client)
        # Năm 2100 không có dữ liệu → tất cả tháng avg_score=null
        r = client.get(f"{BASE_REPORT}/monthly-trend", headers=h, params={"year": 2100})
        assert r.status_code == 200
        data = r.json()
        assert all(p["avg_score"] is None for p in data["points"])
        assert all(p["employee_count"] == 0 for p in data["points"])

    def test_filter_by_department(self, client: TestClient):
        h = _admin(client)
        depts = client.get("/api/v1/departments", headers=h, params={"page_size": 1}).json()
        if not depts:
            pytest.skip("Không có phòng ban trong DB")
        dept_id = depts[0]["id"]

        r = client.get(f"{BASE_REPORT}/monthly-trend", headers=h, params={"year": _YEAR, "department_id": dept_id})
        assert r.status_code == 200
        data = r.json()
        assert data["department_id"] == dept_id
        assert len(data["points"]) == 12


# ── TestPerformanceExport ─────────────────────────────────────────────────────


class TestPerformanceExport:
    def test_export_returns_xlsx_content_type(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={"year": _YEAR})
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_export_valid_xlsx_file(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={"year": _YEAR})
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert wb is not None

    def test_export_has_two_sheets(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={"year": _YEAR})
        assert r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "Phân phối xếp loại"
        assert wb.sheetnames[1] == "KPI theo phòng ban"

    def test_sheet1_rating_counts_match_distribution(self, client: TestClient):
        h = _admin(client)
        dist_r = client.get(f"{BASE_REPORT}/rating-distribution", headers=h, params={"year": _YEAR})
        total_reviewed = dist_r.json()["total_reviewed"]

        exp_r = client.get(f"{BASE_REPORT}/export", headers=h, params={"year": _YEAR})
        assert exp_r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(exp_r.content))
        ws1 = wb["Phân phối xếp loại"]
        # Row 9 is the total row (row 5 + 4 ratings = row 9), column B = total_reviewed
        total_in_excel = ws1.cell(row=9, column=2).value
        assert total_in_excel == total_reviewed

    def test_sheet2_row_count_matches_departments(self, client: TestClient):
        h = _admin(client)
        dept_r = client.get(f"{BASE_REPORT}/department-kpi", headers=h, params={"year": _YEAR})
        n_depts = len(dept_r.json())

        exp_r = client.get(f"{BASE_REPORT}/export", headers=h, params={"year": _YEAR})
        assert exp_r.status_code == 200
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(exp_r.content))
        ws2 = wb["KPI theo phòng ban"]
        # Data rows start at row 5; total row at 5+n_depts
        # Count non-empty rows between 5 and the total row
        data_rows = 0
        row = 5
        while ws2.cell(row=row, column=1).value not in (None, "TỔNG CỘNG"):
            data_rows += 1
            row += 1
        assert data_rows == n_depts
