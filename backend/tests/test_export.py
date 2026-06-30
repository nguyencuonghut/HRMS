from __future__ import annotations

import asyncio
import io
import uuid
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pytest
from pypdf import PdfReader
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import encrypt, hash_sensitive
from app.models.employee import Employee
from app.utils.excel_style import ExcelStyler
from app.workers.export_tasks import run_export_task
from app.services.export_service import cleanup_expired_exports_async

BASE = "/api/v1/reports/export"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_PREFIX = f"T11EXP_{uuid.uuid4().hex[:8]}"


def _engine():
    return create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})


def _admin_headers(client: TestClient) -> dict[str, str]:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


async def _cleanup() -> None:
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)
    async with Session() as s:
        await s.execute(text("""
            DELETE FROM export_jobs WHERE filename LIKE :prefix
        """), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("""
            DELETE FROM report_templates WHERE name LIKE :prefix
        """), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("DELETE FROM employee_contracts WHERE contract_number LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        hashes = [hash_sensitive(f"{_PREFIX}{index:03d}") for index in range(1, 4)]
        employee_ids = list((await s.execute(select(Employee.id).where(Employee.id_number_hash.in_(hashes)))).scalars().all())
        if employee_ids:
            await s.execute(text("DELETE FROM employee_job_records WHERE employee_id = ANY(:employee_ids)"), {"employee_ids": employee_ids})
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        await s.execute(text("DELETE FROM contract_categories WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("DELETE FROM job_titles WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("DELETE FROM departments WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.commit()
    await engine.dispose()


async def _seed() -> None:
    await _cleanup()
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)
    today = date.today()
    async with Session() as s:
        nationality_id = (await s.execute(text("SELECT id FROM nationalities ORDER BY id LIMIT 1"))).scalar_one()
        sequence_id = (await s.execute(text("SELECT id FROM employee_code_sequences ORDER BY id LIMIT 1"))).scalar_one()
        max_seq = (await s.execute(text("SELECT COALESCE(MAX(employee_seq), 0) FROM employees"))).scalar_one()

        dept_id = (await s.execute(text("""
            INSERT INTO departments (code, name, short_name, dept_type, order_no, is_active, created_at)
            VALUES (:code, :name, 'EXP', 'PHONG', 0, TRUE, NOW()) RETURNING id
        """), {"code": f"{_PREFIX}_DEPT", "name": f"{_PREFIX} Dept"})).scalar_one()
        job_title_id = (await s.execute(text("""
            INSERT INTO job_titles (code, name, level, is_active, created_at)
            VALUES (:code, :name, 4, TRUE, NOW()) RETURNING id
        """), {"code": f"{_PREFIX}_JT", "name": f"{_PREFIX} Staff"})).scalar_one()
        category_id = (await s.execute(text("""
            INSERT INTO contract_categories (code, name, normalized_name, document_kind, is_active, created_at)
            VALUES (:code, :name, :normalized_name, 'labor_contract', TRUE, NOW()) RETURNING id
        """), {
            "code": f"{_PREFIX}_CAT",
            "name": f"{_PREFIX} Labor",
            "normalized_name": f"{_PREFIX.lower()} labor",
        })).scalar_one()

        for index in range(1, 4):
            employee_id = (await s.execute(text("""
                INSERT INTO employees (
                    employee_seq, employee_code_sequence_id, full_name, normalized_name,
                    last_name, first_name, date_of_birth, gender, nationality_id,
                    ethnicity_id, religion_id, id_number, id_number_hash, id_issued_on, id_issued_by,
                    id_expires_on, passport_number, passport_issued_on, passport_expires_on,
                    work_permit_number, work_permit_issued_on, work_permit_expires_on,
                    phone_number, personal_email, personal_tax_code, bhxh_code, avatar_path,
                    status, start_date, resigned_date, user_id, is_active, created_at, updated_at
                ) VALUES (
                    :employee_seq, :sequence_id, :full_name, :normalized_name,
                    :last_name, :first_name, :date_of_birth, 'male', :nationality_id,
                    NULL, NULL, :id_number, :id_number_hash, :id_issued_on, 'CA Test',
                    NULL, NULL, NULL, NULL,
                    NULL, NULL, NULL,
                    NULL, :personal_email, NULL, NULL, NULL,
                    'official', :start_date, NULL, NULL, TRUE, NOW(), NOW()
                ) RETURNING id
            """), {
                "employee_seq": max_seq + index,
                "sequence_id": sequence_id,
                "full_name": f"{_PREFIX} Employee {index}",
                "normalized_name": f"{_PREFIX.lower()} employee {index}",
                "last_name": f"Employee{index}",
                "first_name": _PREFIX,
                "date_of_birth": date(1990, 1, min(index, 28)),
                "nationality_id": nationality_id,
                "id_number": encrypt(f"{_PREFIX}{index:03d}"),
                "id_number_hash": hash_sensitive(f"{_PREFIX}{index:03d}"),
                "id_issued_on": date(2020, 1, 1),
                "personal_email": f"{_PREFIX.lower()}_{index}@example.com",
                "start_date": today - timedelta(days=120 * index),
            })).scalar_one()

            await s.execute(text("""
                INSERT INTO employee_job_records (
                    employee_id, department_id, job_title_id, job_position_id,
                    probation_start_date, probation_end_date, official_date,
                    effective_from, effective_to, is_current, notes, changed_by, created_at, updated_at
                ) VALUES (
                    :employee_id, :department_id, :job_title_id, NULL,
                    NULL, NULL, NULL,
                    :effective_from, NULL, TRUE, NULL, NULL, NOW(), NOW()
                )
            """), {
                "employee_id": employee_id,
                "department_id": dept_id,
                "job_title_id": job_title_id,
                "effective_from": today - timedelta(days=120 * index),
            })
            await s.execute(text("""
                INSERT INTO employee_contracts (
                    employee_id, parent_contract_id, contract_category_id, document_kind,
                    contract_number, signed_date, effective_from, effective_to,
                    insurance_salary, insurance_salary_mode, insurance_salary_fixed_amount,
                    status, created_at, updated_at
                ) VALUES (
                    :employee_id, NULL, :category_id, 'labor_contract',
                    :contract_number, :signed_date, :effective_from, :effective_to,
                    NULL, 'fixed_manual', NULL, 'active', NOW(), NOW()
                )
            """), {
                "employee_id": employee_id,
                "category_id": category_id,
                "contract_number": f"{_PREFIX}-C-{index:03d}",
                "signed_date": today - timedelta(days=120 * index),
                "effective_from": today - timedelta(days=120 * index),
                "effective_to": today + timedelta(days=30 * index),
            })
        await s.commit()
    await engine.dispose()


@pytest.fixture(scope="module", autouse=True)
def _seed_data():
    asyncio.run(_seed())
    yield
    asyncio.run(_cleanup())


def test_sync_export_xlsx(client: TestClient):
    headers = _admin_headers(client)
    response = client.post(
        BASE,
        headers=headers,
        json={
            "report_type": "hr-employee-list",
            "format": "xlsx",
            "filename": f"{_PREFIX}_sync",
            "filters": {"page_size": 20},
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["status"] == "done"
    assert payload["download_url"]

    download = client.get(payload["download_url"], headers=headers)
    assert download.status_code == 200
    assert download.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    workbook = openpyxl.load_workbook(io.BytesIO(download.content))
    assert workbook.active.title == "Danh sách nhân sự"
    assert workbook.active["A1"].value == "Mã NV"


def test_export_meta_contract(client: TestClient):
    headers = _admin_headers(client)
    response = client.get(f"{BASE}/meta", headers=headers)
    assert response.status_code == 200
    payload = response.json()
    report_types = payload["report_types"]
    assert any(item["code"] == "hr-movement" and item["label"] == "Nhân sự: Biến động nhân sự" for item in report_types)
    assert [item["code"] for item in payload["formats"]] == ["xlsx", "pdf"]
    assert [item["code"] for item in payload["statuses"]] == ["pending", "processing", "done", "failed"]


def test_sync_export_dashboard_xlsx(client: TestClient):
    headers = _admin_headers(client)
    response = client.post(
        BASE,
        headers=headers,
        json={
            "report_type": "dashboard",
            "format": "xlsx",
            "filename": f"{_PREFIX}_dashboard",
            "filters": {},
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["status"] == "done"
    assert payload["download_url"]

    download = client.get(payload["download_url"], headers=headers)
    assert download.status_code == 200
    workbook = openpyxl.load_workbook(io.BytesIO(download.content))
    assert workbook.active.title == "Dashboard"
    assert workbook.active["A1"].value == "Dashboard nhân sự"
    values = {
        row[0]
        for row in workbook.active.iter_rows(min_row=4, max_col=1, values_only=True)
        if row[0] is not None
    }
    assert "Phòng ban" in values
    assert f"{_PREFIX} Dept" in values


def test_sync_export_pdf_contains_report_content(client: TestClient):
    headers = _admin_headers(client)
    response = client.post(
        BASE,
        headers=headers,
        json={
            "report_type": "hr-employee-list",
            "format": "pdf",
            "filename": f"{_PREFIX}_sync_pdf",
            "filters": {"page_size": 20},
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["status"] == "done"
    assert payload["download_url"]

    download = client.get(payload["download_url"], headers=headers)
    assert download.status_code == 200
    assert download.headers["content-type"].startswith("application/pdf")
    assert download.content.startswith(b"%PDF-1.4")
    extracted = "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(download.content)).pages)
    assert "Danh sách nhân sự" in extracted
    assert "Mã NV" in extracted
    assert "Họ tên" in extracted
    assert "Phòng ban" in extracted
    assert _PREFIX in extracted


def test_export_normalizes_string_dates_for_date_based_reports(client: TestClient, monkeypatch: pytest.MonkeyPatch):
    headers = _admin_headers(client)
    captured: dict[str, tuple[object, object]] = {}

    async def fake_movement_getter(session, *, start_date, end_date, group_by="month", department_id=None):
        captured["hr-movement-estimate"] = (start_date, end_date)
        return type("MovementEstimate", (), {"rows": []})()

    async def fake_movement_exporter(session, *, start_date, end_date, group_by="month", department_id=None):
        captured["hr-movement-export"] = (start_date, end_date)
        return ExcelStyler.build_table_workbook("Movement", ["A"], [["ok"]], "Movement")

    async def fake_funnel(session, start_date, end_date, department_id=None, job_requisition_id=None):
        captured["recruitment-estimate"] = (start_date, end_date)
        return type("FunnelReport", (), {"stages": []})()

    async def fake_recruitment_export(session, start_date, end_date, department_id=None):
        captured["recruitment-export"] = (start_date, end_date)
        return ExcelStyler.build_table_workbook("Recruitment", ["A"], [["ok"]], "Recruitment")

    async def fake_probation_history(
        session,
        *,
        start_date=None,
        end_date=None,
        department_id=None,
        keyword=None,
        page=1,
        page_size=1,
    ):
        captured["probation-estimate"] = (start_date, end_date)
        return type("ProbationHistory", (), {"total": 0})()

    async def fake_probation_export(session, *, start_date, end_date, department_id=None):
        captured["probation-export"] = (start_date, end_date)
        return ExcelStyler.build_table_workbook("Probation", ["A"], [["ok"]], "Probation")

    monkeypatch.setattr("app.services.export_service.hr_report_service.get_movement_report", fake_movement_getter)
    monkeypatch.setattr("app.services.export_service.hr_report_service.export_movement_excel", fake_movement_exporter)
    monkeypatch.setattr("app.services.export_service.recruitment_report_service.get_funnel", fake_funnel)
    monkeypatch.setattr("app.services.export_service.recruitment_report_service.export_excel", fake_recruitment_export)
    monkeypatch.setattr("app.services.export_service.probation_report_service.get_probation_history", fake_probation_history)
    monkeypatch.setattr("app.services.export_service.probation_report_service.export_excel", fake_probation_export)

    for report_type in ("hr-movement", "recruitment", "probation"):
        response = client.post(
            BASE,
            headers=headers,
            json={
                "report_type": report_type,
                "format": "xlsx",
                "filename": f"{_PREFIX}_{report_type}",
                "filters": {
                    "start_date": "2026-01-01",
                    "end_date": "2026-03-31",
                },
            },
        )
        assert response.status_code == 200, response.text
        assert response.json()["status"] == "done"

    for key, value in captured.items():
        start_date, end_date = value
        assert isinstance(start_date, date), key
        assert isinstance(end_date, date), key
        assert start_date.isoformat() == "2026-01-01", key
        assert end_date.isoformat() == "2026-03-31", key


def test_async_export_history_status_and_delete(client: TestClient, monkeypatch: pytest.MonkeyPatch):
    headers = _admin_headers(client)

    class _TaskResult:
        id = "fake-export-task"

    monkeypatch.setattr("app.services.export_service.ASYNC_THRESHOLD_ROWS", 1)
    monkeypatch.setattr("app.workers.export_tasks.run_export_task.delay", lambda *args, **kwargs: _TaskResult())

    response = client.post(
        BASE,
        headers=headers,
        json={
            "report_type": "hr-employee-list",
            "format": "xlsx",
            "filename": f"{_PREFIX}_async",
            "filters": {"page_size": 20},
        },
    )
    assert response.status_code == 200, response.text
    payload = response.json()
    assert payload["status"] == "pending"
    job_id = payload["id"]

    status_pending = client.get(f"{BASE}/{job_id}/status", headers=headers)
    assert status_pending.status_code == 200
    assert status_pending.json()["status"] == "pending"

    run_export_task.run(job_id, {
        "report_type": "hr-employee-list",
        "format": "xlsx",
        "filename": f"{_PREFIX}_async.xlsx",
        "filters": {"page_size": 20},
    })

    status_done = client.get(f"{BASE}/{job_id}/status", headers=headers)
    assert status_done.status_code == 200
    assert status_done.json()["status"] == "done"
    assert status_done.json()["download_url"]

    history = client.get(f"{BASE}/history?page=1&page_size=10", headers=headers)
    assert history.status_code == 200
    assert history.json()["total"] >= 1
    assert any(item["id"] == job_id for item in history.json()["items"])

    delete_response = client.delete(f"{BASE}/{job_id}", headers=headers)
    assert delete_response.status_code == 204

    status_missing = client.get(f"{BASE}/{job_id}/status", headers=headers)
    assert status_missing.status_code == 404


def test_report_template_crud_and_history_paging(client: TestClient):
    headers = _admin_headers(client)

    create = client.post(
        f"{BASE}/templates",
        headers=headers,
        json={
            "name": f"{_PREFIX}_template",
            "description": "Mau bao cao nhan su",
            "report_type": "hr-employee-list",
            "format": "xlsx",
            "filters": {"page_size": 50},
            "is_default": True,
        },
    )
    assert create.status_code == 200, create.text
    template = create.json()
    assert template["name"] == f"{_PREFIX}_template"
    assert template["is_default"] is True
    template_id = template["id"]

    listing = client.get(f"{BASE}/templates", headers=headers)
    assert listing.status_code == 200
    assert any(item["id"] == template_id for item in listing.json())

    update = client.put(
        f"{BASE}/templates/{template_id}",
        headers=headers,
        json={"description": "Cap nhat", "filters": {"page_size": 100}, "is_default": False},
    )
    assert update.status_code == 200
    assert update.json()["description"] == "Cap nhat"
    assert update.json()["filters"]["page_size"] == 100
    assert update.json()["is_default"] is False

    export_response = client.post(
        BASE,
        headers=headers,
        json={
            "report_type": "hr-employee-list",
            "format": "xlsx",
            "filename": f"{_PREFIX}_history",
            "filters": {"page_size": 20},
        },
    )
    assert export_response.status_code == 200

    history = client.get(f"{BASE}/history?page=1&page_size=1", headers=headers)
    assert history.status_code == 200
    assert history.json()["page"] == 1
    assert history.json()["page_size"] == 1
    assert history.json()["total"] >= 1
    assert len(history.json()["items"]) == 1

    delete = client.delete(f"{BASE}/templates/{template_id}", headers=headers)
    assert delete.status_code == 204

    listing_after = client.get(f"{BASE}/templates", headers=headers)
    assert all(item["id"] != template_id for item in listing_after.json())


def test_cleanup_expired_exports_removes_jobs_and_files():
    async def _seed_expired_job() -> tuple[str, str]:
        engine = _engine()
        Session = async_sessionmaker(engine, expire_on_commit=False)
        async with Session() as s:
            job_id = str(uuid.uuid4())
            file_path = f"/tmp/hrms_exports/{job_id}.xlsx"
            Path(file_path).write_bytes(b"expired-export")
            await s.execute(text("""
                INSERT INTO export_jobs (
                    id, user_id, report_type, format, filters, status, filename,
                    file_path, file_size_bytes, row_count, created_at, expires_at
                ) VALUES (
                    :id, 1, 'hr-employee-list', 'xlsx', '{}'::jsonb, 'done', :filename,
                    :file_path, 14, 1, NOW() - INTERVAL '8 days', NOW() - INTERVAL '1 day'
                )
            """), {"id": job_id, "filename": f"{_PREFIX}_expired.xlsx", "file_path": file_path})
            await s.commit()
        await engine.dispose()
        return job_id, file_path

    job_id, file_path = asyncio.run(_seed_expired_job())
    result = asyncio.run(cleanup_expired_exports_async())
    assert result["deleted_jobs"] >= 1
    assert not Path(file_path).exists()

    async def _assert_missing() -> None:
        engine = _engine()
        Session = async_sessionmaker(engine, expire_on_commit=False)
        async with Session() as s:
            remaining = (await s.execute(text("SELECT COUNT(*) FROM export_jobs WHERE id = :id"), {"id": job_id})).scalar_one()
            assert remaining == 0
        await engine.dispose()

    asyncio.run(_assert_missing())
