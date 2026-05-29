"""Integration tests cho service báo cáo nhân sự (11.2 Slice 1)."""
from __future__ import annotations

import asyncio
from datetime import date, timedelta
import io
import uuid

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import encrypt, hash_sensitive
from app.models.employee import Employee
from app.services import hr_report_service

BASE = "/api/v1/reports/hr"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_PREFIX = f"T11HR_{uuid.uuid4().hex[:8]}"


def _engine():
    return create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})


def _admin_headers(client: TestClient) -> dict[str, str]:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


async def _cleanup() -> None:
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)
    async with Session() as s:
        await s.execute(
            text("DELETE FROM employee_contracts WHERE contract_number LIKE :prefix"),
            {"prefix": f"{_PREFIX}%"},
        )
        dept_ids = list(
            (
                await s.execute(
                    text("SELECT id FROM departments WHERE code LIKE :prefix"),
                    {"prefix": f"{_PREFIX}%"},
                )
            )
            .scalars()
            .all()
        )
        employee_ids = list(
            (
                await s.execute(
                    select(Employee.id).where(
                        Employee.personal_email.like(f"{_PREFIX.lower()}_%@example.com")
                    )
                )
            )
            .scalars()
            .all()
        )
        if employee_ids:
            await s.execute(
                text("DELETE FROM employee_contracts WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_job_records WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        if dept_ids:
            await s.execute(
                text("DELETE FROM employee_job_records WHERE department_id = ANY(:department_ids)"),
                {"department_ids": dept_ids},
            )
        await s.execute(text("DELETE FROM contract_categories WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("DELETE FROM job_titles WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.execute(text("DELETE FROM departments WHERE code LIKE :prefix"), {"prefix": f"{_PREFIX}%"})
        await s.commit()
    await engine.dispose()


async def _seed_hr_report_data() -> dict[str, int]:
    await _cleanup()
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)
    today = date.today()

    async with Session() as s:
        nationality_id = (await s.execute(text("SELECT id FROM nationalities ORDER BY id LIMIT 1"))).scalar_one()
        sequence_id = (await s.execute(text("SELECT id FROM employee_code_sequences ORDER BY id LIMIT 1"))).scalar_one()
        max_seq = (await s.execute(text("SELECT COALESCE(MAX(employee_seq), 0) FROM employees"))).scalar_one()

        dept_root = (await s.execute(text("""
            INSERT INTO departments (code, name, short_name, dept_type, order_no, is_active, created_at)
            VALUES (:code, :name, 'ROOT', 'PHONG', 0, TRUE, NOW())
            RETURNING id
        """), {"code": f"{_PREFIX}_ROOT", "name": f"{_PREFIX} Root"})).scalar_one()
        dept_sales = (await s.execute(text("""
            INSERT INTO departments (code, name, short_name, parent_id, dept_type, order_no, is_active, created_at)
            VALUES (:code, :name, 'SALE', :parent_id, 'BO_PHAN', 0, TRUE, NOW())
            RETURNING id
        """), {"code": f"{_PREFIX}_SALE", "name": f"{_PREFIX} Sales", "parent_id": dept_root})).scalar_one()
        dept_ops = (await s.execute(text("""
            INSERT INTO departments (code, name, short_name, parent_id, dept_type, order_no, is_active, created_at)
            VALUES (:code, :name, 'OPS', :parent_id, 'BO_PHAN', 1, TRUE, NOW())
            RETURNING id
        """), {"code": f"{_PREFIX}_OPS", "name": f"{_PREFIX} Ops", "parent_id": dept_root})).scalar_one()

        manager_title = (await s.execute(text("""
            INSERT INTO job_titles (code, name, level, is_active, created_at)
            VALUES (:code, :name, 2, TRUE, NOW())
            RETURNING id
        """), {"code": f"{_PREFIX}_MGR", "name": f"{_PREFIX} Manager"})).scalar_one()
        staff_title = (await s.execute(text("""
            INSERT INTO job_titles (code, name, level, is_active, created_at)
            VALUES (:code, :name, 5, TRUE, NOW())
            RETURNING id
        """), {"code": f"{_PREFIX}_STF", "name": f"{_PREFIX} Staff"})).scalar_one()

        labor_category = (await s.execute(text("""
            INSERT INTO contract_categories (
                code, name, normalized_name, document_kind, is_active, created_at
            ) VALUES (:code, :name, :normalized_name, 'labor', TRUE, NOW())
            RETURNING id
        """), {
            "code": f"{_PREFIX}_LABOR",
            "name": f"{_PREFIX} Labor",
            "normalized_name": f"{_PREFIX.lower()} labor",
        })).scalar_one()
        probation_category = (await s.execute(text("""
            INSERT INTO contract_categories (
                code, name, normalized_name, document_kind, is_active, created_at
            ) VALUES (:code, :name, :normalized_name, 'probation', TRUE, NOW())
            RETURNING id
        """), {
            "code": f"{_PREFIX}_PROB",
            "name": f"{_PREFIX} Probation",
            "normalized_name": f"{_PREFIX.lower()} probation",
        })).scalar_one()

        employees = [
            ("Alpha", "male", "official", True, today - timedelta(days=120), None, dept_sales, manager_title, labor_category),
            ("Bravo", "female", "probation", True, today - timedelta(days=500), None, dept_sales, staff_title, probation_category),
            ("Charlie", "male", "long_leave", True, today - timedelta(days=900), None, dept_sales, staff_title, labor_category),
            ("Delta", "female", "official", True, today - timedelta(days=1500), None, dept_ops, staff_title, labor_category),
            ("Echo", "other", "official", True, today - timedelta(days=2400), None, dept_ops, manager_title, labor_category),
            ("Foxtrot", "male", "official", True, today - timedelta(days=3650), None, dept_sales, staff_title, labor_category),
            ("Golf", "female", "official", True, today - timedelta(days=4200), None, dept_ops, staff_title, labor_category),
            ("Hotel", "male", "official", True, today - timedelta(days=30), None, dept_ops, staff_title, probation_category),
            ("India", "female", "resigned", False, today - timedelta(days=240), today - timedelta(days=10), dept_sales, staff_title, labor_category),
            ("Juliet", "female", "official", True, today - timedelta(days=70), None, dept_sales, staff_title, labor_category),
        ]

        ids: dict[str, int] = {}
        for index, (name, gender, status, is_active, start_date, resigned_date, department_id, title_id, category_id) in enumerate(employees, start=1):
            employee_id = (await s.execute(text("""
                INSERT INTO employees (
                    employee_seq, employee_code_sequence_id, full_name, normalized_name,
                    last_name, first_name, date_of_birth, gender, nationality_id,
                    ethnicity_id, religion_id, id_number, id_number_hash, id_issued_on, id_issued_by,
                    id_expires_on, passport_number, passport_issued_on, passport_expires_on,
                    work_permit_number, work_permit_issued_on, work_permit_expires_on,
                    phone_number, personal_email, personal_tax_code, bhxh_code, avatar_path,
                    status, start_date, resigned_date, user_id, is_active, created_at, updated_at
                ) VALUES (
                    :employee_seq, :sequence_id, :full_name, :normalized_name,
                    :last_name, :first_name, :date_of_birth, :gender, :nationality_id,
                    NULL, NULL, :id_number, :id_number_hash, :id_issued_on, :id_issued_by,
                    NULL, NULL, NULL, NULL,
                    NULL, NULL, NULL,
                    NULL, :personal_email, NULL, NULL, NULL,
                    :status, :start_date, :resigned_date, NULL, :is_active, NOW(), NOW()
                )
                RETURNING id
            """), {
                "employee_seq": max_seq + index,
                "sequence_id": sequence_id,
                "full_name": f"{_PREFIX} {name}",
                "normalized_name": f"{_PREFIX.lower()} {name.lower()}",
                "last_name": name,
                "first_name": _PREFIX,
                "date_of_birth": date(1990, 1, min(index, 28)),
                "gender": gender,
                "nationality_id": nationality_id,
                "id_number": encrypt(f"{_PREFIX}{index:03d}"),
                "id_number_hash": hash_sensitive(f"{_PREFIX}{index:03d}"),
                "id_issued_on": date(2020, 1, 1),
                "id_issued_by": "CA Test",
                "personal_email": f"{_PREFIX.lower()}_{index}@example.com",
                "status": status,
                "start_date": start_date,
                "resigned_date": resigned_date,
                "is_active": is_active,
            })).scalar_one()
            ids[name.lower()] = employee_id

            await s.execute(text("""
                INSERT INTO employee_job_records (
                    employee_id, department_id, job_title_id, job_position_id,
                    probation_start_date, probation_end_date, official_date,
                    effective_from, effective_to, is_current, notes, changed_by,
                    created_at, updated_at
                ) VALUES (
                    :employee_id, :department_id, :job_title_id, NULL,
                    NULL, NULL, NULL,
                    :effective_from, NULL, TRUE, NULL, NULL, NOW(), NOW()
                )
            """), {
                "employee_id": employee_id,
                "department_id": department_id,
                "job_title_id": title_id,
                "effective_from": start_date,
            })

            await s.execute(text("""
                INSERT INTO employee_contracts (
                    employee_id, parent_contract_id, contract_category_id, document_kind,
                    contract_number, signed_date, effective_from, effective_to,
                    insurance_salary, status, created_at, updated_at
                ) VALUES (
                    :employee_id, NULL, :contract_category_id, :document_kind,
                    :contract_number, :signed_date, :effective_from, NULL,
                    NULL, 'active', NOW(), NOW()
                )
            """), {
                "employee_id": employee_id,
                "contract_category_id": category_id,
                "document_kind": "probation" if category_id == probation_category else "labor",
                "contract_number": f"{_PREFIX}-C-{index:03d}",
                "signed_date": start_date,
                "effective_from": start_date,
            })

        await s.execute(text("""
            UPDATE employee_job_records
            SET is_current = FALSE, effective_to = :effective_to
            WHERE employee_id = :employee_id AND effective_from = :effective_from
        """), {
            "employee_id": ids["juliet"],
            "effective_from": today - timedelta(days=70),
            "effective_to": today - timedelta(days=20),
        })
        await s.execute(text("""
            INSERT INTO employee_job_records (
                employee_id, department_id, job_title_id, job_position_id,
                probation_start_date, probation_end_date, official_date,
                effective_from, effective_to, is_current, notes, changed_by,
                created_at, updated_at
            ) VALUES (
                :employee_id, :department_id, :job_title_id, NULL,
                NULL, NULL, NULL,
                :effective_from, NULL, TRUE, 'Transfer', NULL, NOW(), NOW()
            )
        """), {
            "employee_id": ids["juliet"],
            "department_id": dept_ops,
            "job_title_id": staff_title,
            "effective_from": today - timedelta(days=20),
        })

        await s.commit()
    await engine.dispose()
    return {
        "dept_root": dept_root,
        "dept_sales": dept_sales,
        "dept_ops": dept_ops,
    }


@pytest.fixture(autouse=True)
def _hr_report_fixture_cleanup():
    asyncio.run(_cleanup())
    yield
    asyncio.run(_cleanup())


def test_get_employee_list_filters():
    refs = asyncio.run(_seed_hr_report_data())
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)

    async def run():
        async with Session() as s:
            return await hr_report_service.get_employee_list(
                s,
                department_id=refs["dept_sales"],
                gender="female",
                page=1,
                page_size=20,
            )

    report = asyncio.run(run())
    asyncio.run(engine.dispose())

    assert report.total == 2
    assert [item.full_name for item in report.items] == [f"{_PREFIX} Bravo", f"{_PREFIX} India"]
    assert all(item.department_id == refs["dept_sales"] for item in report.items)
    assert all(item.gender == "female" for item in report.items)


def test_get_movement_report_groups_monthly_changes():
    refs = asyncio.run(_seed_hr_report_data())
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)
    today = date.today()
    start_date = today - timedelta(days=120)
    end_date = today

    async def run():
        async with Session() as s:
            return await hr_report_service.get_movement_report(
                s,
                start_date=start_date,
                end_date=end_date,
                group_by="month",
                department_id=refs["dept_root"],
            )

    report = asyncio.run(run())
    asyncio.run(engine.dispose())

    assert report.total_hired >= 3
    assert report.total_resigned == 1
    assert report.total_transfers == 1
    assert any(row.transfer_count == 1 for row in report.rows)


def test_get_tenure_report_builds_all_buckets():
    refs = asyncio.run(_seed_hr_report_data())
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)

    async def run():
        async with Session() as s:
            return await hr_report_service.get_tenure_report(
                s,
                department_id=refs["dept_root"],
            )

    report = asyncio.run(run())
    asyncio.run(engine.dispose())

    groups = {group.group_key: group for group in report.groups}
    assert report.total_active == 9
    assert set(groups) == {"lt_1", "1_3", "3_5", "5_10", "gt_10"}
    assert groups["lt_1"].headcount == 3
    assert groups["1_3"].headcount == 2
    assert groups["3_5"].headcount == 1
    assert groups["5_10"].headcount == 1
    assert groups["gt_10"].headcount == 2


def test_get_org_structure_returns_tree_with_rollups():
    refs = asyncio.run(_seed_hr_report_data())
    engine = _engine()
    Session = async_sessionmaker(engine, expire_on_commit=False)

    async def run():
        async with Session() as s:
            return await hr_report_service.get_org_structure(
                s,
                department_id=refs["dept_root"],
            )

    report = asyncio.run(run())
    asyncio.run(engine.dispose())

    assert report.department_count == 3
    assert len(report.tree) == 1
    root = report.tree[0]
    assert root.department_id == refs["dept_root"]
    assert root.total_headcount == 9
    assert root.direct_headcount == 0
    assert [child.department_id for child in root.children] == [refs["dept_sales"], refs["dept_ops"]]
    assert sum(child.total_headcount for child in root.children) == root.total_headcount


def test_employee_list_endpoint_returns_filtered_page(client: TestClient):
    refs = asyncio.run(_seed_hr_report_data())
    headers = _admin_headers(client)

    resp = client.get(
        f"{BASE}/employee-list",
        headers=headers,
        params={"department_id": refs["dept_sales"], "gender": "female", "page": 1, "page_size": 20},
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total"] == 2
    assert [item["full_name"] for item in data["items"]] == [f"{_PREFIX} Bravo", f"{_PREFIX} India"]


def test_movement_endpoint_returns_monthly_rows(client: TestClient):
    refs = asyncio.run(_seed_hr_report_data())
    headers = _admin_headers(client)
    today = date.today()

    resp = client.get(
        f"{BASE}/movement",
        headers=headers,
        params={
            "start_date": (today - timedelta(days=120)).isoformat(),
            "end_date": today.isoformat(),
            "group_by": "month",
            "department_id": refs["dept_root"],
        },
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total_resigned"] == 1
    assert data["total_transfers"] == 1
    assert any(row["transfer_count"] == 1 for row in data["rows"])


def test_export_employee_list_returns_valid_xlsx(client: TestClient):
    refs = asyncio.run(_seed_hr_report_data())
    headers = _admin_headers(client)

    resp = client.get(
        f"{BASE}/export",
        headers=headers,
        params={"type": "employee-list", "department_id": refs["dept_sales"]},
    )
    assert resp.status_code == 200, resp.text
    assert "spreadsheetml" in resp.headers.get("content-type", "")
    assert "bao_cao_nhan_su_employee-list" in resp.headers.get("content-disposition", "")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Danh sách nhân sự"]
    assert ws.max_row >= 2
    assert ws["B2"].value is not None


def test_export_requires_auth(client: TestClient):
    resp = client.get(f"{BASE}/export", params={"type": "tenure"})
    assert resp.status_code in (401, 403)
