from __future__ import annotations

from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient

IMPORT_BASE = "/api/v1/imports"
DEPT_BASE = "/api/v1/departments"
JT_BASE = "/api/v1/job-titles"
POS_BASE = "/api/v1/job-positions"
AUTH_BASE = "/api/v1/auth/login"

ADMIN_EMAIL = "admin@hrms.local"
ADMIN_PASSWORD = "Hrms@2026"

DEPT_ROOT = "IMP_DEPT_ROOT"
DEPT_CHILD = "IMP_DEPT_CHILD"
JT_CODE = "IMP_JT_001"
POS_CODE = "IMP_POS_001"


def _login(client: TestClient) -> dict[str, str]:
    token = client.post(AUTH_BASE, json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_xlsx(headers: list[str], rows: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client: TestClient, headers: dict[str, str], endpoint: str, content: bytes):
    return client.post(
        endpoint,
        files={"file": ("import.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


def _delete_position_by_code(client: TestClient, code: str) -> None:
    for item in client.get(POS_BASE).json():
        if item["code"] == code:
            client.delete(f"{POS_BASE}/{item['id']}")


def _delete_title_by_code(client: TestClient, code: str) -> None:
    for item in client.get(JT_BASE).json():
        if item["code"] == code:
            client.delete(f"{JT_BASE}/{item['id']}")


def _delete_dept_by_code(client: TestClient, code: str) -> None:
    for item in client.get(DEPT_BASE).json():
        if item["code"] == code:
            client.delete(f"{DEPT_BASE}/{item['id']}")


@pytest.fixture(autouse=True)
def cleanup(client: TestClient):
    _delete_position_by_code(client, POS_CODE)
    _delete_title_by_code(client, JT_CODE)
    _delete_dept_by_code(client, DEPT_CHILD)
    _delete_dept_by_code(client, DEPT_ROOT)
    yield
    _delete_position_by_code(client, POS_CODE)
    _delete_title_by_code(client, JT_CODE)
    _delete_dept_by_code(client, DEPT_CHILD)
    _delete_dept_by_code(client, DEPT_ROOT)


def test_department_import_supports_parent_child_upsert(client: TestClient):
    headers = _login(client)
    xlsx = _make_xlsx(
        ["Mã phòng ban", "Tên phòng ban", "Tên viết tắt", "Mã phòng ban cha", "Tiền tố hiển thị", "Loại đơn vị", "Thứ tự", "Kích hoạt"],
        [
            [DEPT_ROOT, "Khối nhập liệu", "KNL", "", "KN", "BAN", 1, 1],
            [DEPT_CHILD, "Tổ kiểm nhập", "TKN", DEPT_ROOT, "TK", "TO", 2, 0],
        ],
    )

    resp = _upload(client, headers, f"{IMPORT_BASE}/departments", xlsx)
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["success"] == 2
    assert body["failed"] == 0

    departments = client.get(DEPT_BASE).json()
    root = next(item for item in departments if item["code"] == DEPT_ROOT)
    child = next(item for item in departments if item["code"] == DEPT_CHILD)
    assert child["parent_id"] == root["id"]
    assert child["is_active"] is False
    assert root["display_prefix"] == "KN"

    xlsx_update = _make_xlsx(
        ["Mã phòng ban", "Tên phòng ban", "Tên viết tắt", "Mã phòng ban cha", "Tiền tố hiển thị", "Loại đơn vị", "Thứ tự", "Kích hoạt"],
        [
            [DEPT_ROOT, "Khối nhập liệu cập nhật", "KNL", "", "KN", "BAN", 5, 1],
        ],
    )
    updated = _upload(client, headers, f"{IMPORT_BASE}/departments", xlsx_update)
    assert updated.status_code == 200, updated.text
    root_after = next(item for item in client.get(DEPT_BASE).json() if item["code"] == DEPT_ROOT)
    assert root_after["name"] == "Khối nhập liệu cập nhật"
    assert root_after["order_no"] == 5


def test_job_title_import_supports_upsert(client: TestClient):
    headers = _login(client)
    xlsx = _make_xlsx(
        ["Mã chức danh", "Tên chức danh", "Cấp bậc", "Kích hoạt"],
        [[JT_CODE, "Chức danh import", 4, 1]],
    )
    resp = _upload(client, headers, f"{IMPORT_BASE}/job-titles", xlsx)
    assert resp.status_code == 200, resp.text
    assert resp.json()["success"] == 1

    titles = client.get(JT_BASE).json()
    title = next(item for item in titles if item["code"] == JT_CODE)
    assert title["name"] == "Chức danh import"
    assert title["level"] == 4

    xlsx_update = _make_xlsx(
        ["Mã chức danh", "Tên chức danh", "Cấp bậc", "Kích hoạt"],
        [[JT_CODE, "Chức danh import cập nhật", 2, 0]],
    )
    updated = _upload(client, headers, f"{IMPORT_BASE}/job-titles", xlsx_update)
    assert updated.status_code == 200, updated.text
    title_after = next(item for item in client.get(JT_BASE).json() if item["code"] == JT_CODE)
    assert title_after["name"] == "Chức danh import cập nhật"
    assert title_after["level"] == 2
    assert title_after["is_active"] is False


def test_job_title_import_ignores_trailing_formatted_blank_rows(client: TestClient):
    headers = _login(client)
    created_codes: list[str] = []
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Mã chức danh", "Tên chức danh", "Cấp bậc", "Kích hoạt"])
        for idx in range(1, 21):
            code = f"{JT_CODE}_{idx}"
            created_codes.append(code)
            ws.append([code, f"Chức danh import {idx}", 4, 1])
        ws["A1505"].fill = openpyxl.styles.PatternFill("solid", fgColor="FFFF00")

        buf = BytesIO()
        wb.save(buf)
        xlsx = buf.getvalue()

        resp = _upload(client, headers, f"{IMPORT_BASE}/job-titles", xlsx)
        assert resp.status_code == 200, resp.text
        body = resp.json()
        assert body["total"] == 20
        assert body["success"] == 20
        assert body["failed"] == 0
    finally:
        for code in created_codes:
            _delete_title_by_code(client, code)


def test_job_position_import_requires_existing_dependencies_and_upserts(client: TestClient):
    headers = _login(client)

    missing_dep = _make_xlsx(
        ["Mã vị trí", "Tên vị trí", "Mã phòng ban", "Mã chức danh", "Bậc mặc định", "Phụ cấp BHXH", "Phụ cấp ngoài BHXH", "Kích hoạt", "Mô tả", "Yêu cầu"],
        [[POS_CODE, "Vị trí lỗi", DEPT_ROOT, JT_CODE, 1, 100000, 50000, 1, "", ""]],
    )
    failed = _upload(client, headers, f"{IMPORT_BASE}/job-positions", missing_dep)
    assert failed.status_code == 200, failed.text
    assert failed.json()["failed"] == 1
    assert any("Không tìm thấy phòng ban" in err["message"] for err in failed.json()["errors"])

    dept_xlsx = _make_xlsx(
        ["Mã phòng ban", "Tên phòng ban", "Tên viết tắt", "Mã phòng ban cha", "Tiền tố hiển thị", "Loại đơn vị", "Thứ tự", "Kích hoạt"],
        [[DEPT_ROOT, "Phòng kinh doanh import", "KD", "", "KD", "PHONG", 1, 1]],
    )
    title_xlsx = _make_xlsx(
        ["Mã chức danh", "Tên chức danh", "Cấp bậc", "Kích hoạt"],
        [[JT_CODE, "Nhân viên kinh doanh import", 7, 1]],
    )
    assert _upload(client, headers, f"{IMPORT_BASE}/departments", dept_xlsx).status_code == 200
    assert _upload(client, headers, f"{IMPORT_BASE}/job-titles", title_xlsx).status_code == 200

    pos_xlsx = _make_xlsx(
        ["Mã vị trí", "Tên vị trí", "Mã phòng ban", "Mã chức danh", "Bậc mặc định", "Phụ cấp BHXH", "Phụ cấp ngoài BHXH", "Nhóm thử việc pháp lý", "Kích hoạt", "Mô tả", "Yêu cầu"],
        [[POS_CODE, "Nhân viên kinh doanh import", DEPT_ROOT, JT_CODE, 2, 150000, 80000, "college_plus", 1, "Chăm sóc khách hàng", "Giao tiếp tốt"]],
    )
    created = _upload(client, headers, f"{IMPORT_BASE}/job-positions", pos_xlsx)
    assert created.status_code == 200, created.text
    assert created.json()["success"] == 1

    position = next(item for item in client.get(POS_BASE).json() if item["code"] == POS_CODE)
    assert position["department_name"] == "Phòng kinh doanh import"
    assert position["job_title_name"] == "Nhân viên kinh doanh import"
    assert position["bhxh_allowance"] == 150000
    assert position["non_bhxh_allowance"] == 80000
    assert position["probation_legal_group"] == "college_plus"
    assert position["probation_days_limit"] == 60

    pos_update_xlsx = _make_xlsx(
        ["Mã vị trí", "Tên vị trí", "Mã phòng ban", "Mã chức danh", "Bậc mặc định", "Phụ cấp BHXH", "Phụ cấp ngoài BHXH", "Nhóm thử việc pháp lý", "Kích hoạt", "Mô tả", "Yêu cầu"],
        [[POS_CODE, "Nhân viên kinh doanh import cập nhật", DEPT_ROOT, JT_CODE, 3, 200000, 90000, "intermediate_technical_clerical", 0, "Mô tả mới", "Yêu cầu mới"]],
    )
    updated = _upload(client, headers, f"{IMPORT_BASE}/job-positions", pos_update_xlsx)
    assert updated.status_code == 200, updated.text
    position_after = next(item for item in client.get(POS_BASE).json() if item["code"] == POS_CODE)
    assert position_after["name"] == "Nhân viên kinh doanh import cập nhật"
    assert position_after["bhxh_allowance"] == 200000
    assert position_after["non_bhxh_allowance"] == 90000
    assert position_after["probation_legal_group"] == "intermediate_technical_clerical"
    assert position_after["probation_days_limit"] == 30
    assert position_after["is_active"] is False
