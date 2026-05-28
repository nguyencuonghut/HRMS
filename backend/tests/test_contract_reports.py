"""Tests cho 11.5 — Contract Reports."""
from __future__ import annotations

import io
import asyncio
from datetime import date, timedelta
import pytest
from fastapi.testclient import TestClient
import openpyxl
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings

BASE = "/api/v1/reports/contracts"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"


def _login(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        await s.execute(text("DELETE FROM employee_contracts WHERE contract_number LIKE 'TEST-CON-%'"))
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup_db():
    await _cleanup()
    yield
    await _cleanup()


async def _get_employee_and_category_ids() -> tuple[int, int]:
    async with _make_session()() as s:
        # Lấy employee đầu tiên sẵn có trong DB test
        r_emp = await s.execute(text("SELECT id FROM employees WHERE is_active = TRUE LIMIT 1"))
        row_emp = r_emp.fetchone()
        assert row_emp is not None, "Không tìm thấy employee active trong DB"

        # Lấy contract category đầu tiên
        r_cat = await s.execute(text("SELECT id FROM contract_categories LIMIT 1"))
        row_cat = r_cat.fetchone()
        assert row_cat is not None, "Không tìm thấy contract category trong DB"

        return row_emp[0], row_cat[0]


async def _create_test_contract(
    employee_id: int,
    category_id: int,
    contract_number: str,
    signed_date: date,
    effective_from: date,
    effective_to: date | None,
    status: str = "active",
):
    async with _make_session()() as s:
        # Đảm bảo nhân viên có job record is_current = True
        job_exists = (await s.execute(
            text("SELECT 1 FROM employee_job_records WHERE employee_id = :eid AND is_current = TRUE"),
            {"eid": employee_id}
        )).scalar()

        if not job_exists:
            # Lấy phòng ban đầu tiên
            dept_id = (await s.execute(text("SELECT id FROM departments LIMIT 1"))).scalar()
            if dept_id:
                await s.execute(
                    text(
                        """
                        INSERT INTO employee_job_records (
                            employee_id, department_id, effective_from, is_current
                        ) VALUES (:eid, :did, :eff_from, TRUE)
                        """
                    ),
                    {"eid": employee_id, "did": dept_id, "eff_from": effective_from}
                )

        await s.execute(
            text(
                """
                INSERT INTO employee_contracts (
                    employee_id, contract_category_id, document_kind,
                    contract_number, signed_date, effective_from, effective_to,
                    insurance_salary, status
                ) VALUES (
                    :eid, :cid, 'labor_contract',
                    :cnum, :sdate, :eff_from, :eff_to,
                    12000000.00, :status
                )
                """
            ),
            {
                "eid": employee_id,
                "cid": category_id,
                "cnum": contract_number,
                "sdate": signed_date,
                "eff_from": effective_from,
                "eff_to": effective_to,
                "status": status,
            },
        )
        await s.commit()


def test_contract_summary(client: TestClient):
    headers = _login(client)

    # 1. Gọi summary ban đầu
    resp = client.get(f"{BASE}/summary", headers=headers)
    assert resp.status_code == 200, resp.text
    data_before = resp.json()
    assert "total_active" in data_before

    # 2. Tạo contract sắp hết hạn trong 15 ngày (nằm trong bucket 0-30)
    emp_id, cat_id = asyncio.get_event_loop().run_until_complete(_get_employee_and_category_ids())
    today = date.today()
    asyncio.get_event_loop().run_until_complete(
        _create_test_contract(
            employee_id=emp_id,
            category_id=cat_id,
            contract_number="TEST-CON-001",
            signed_date=today - timedelta(days=300),
            effective_from=today - timedelta(days=300),
            effective_to=today + timedelta(days=15),
            status="active"
        )
    )

    # 3. Tạo contract đã hết hạn
    asyncio.get_event_loop().run_until_complete(
        _create_test_contract(
            employee_id=emp_id,
            category_id=cat_id,
            contract_number="TEST-CON-002",
            signed_date=today - timedelta(days=400),
            effective_from=today - timedelta(days=400),
            effective_to=today - timedelta(days=10),
            status="active"
        )
    )

    # 4. Gọi lại summary và đối chiếu
    resp2 = client.get(f"{BASE}/summary", headers=headers)
    assert resp2.status_code == 200
    data_after = resp2.json()

    assert data_after["total_active"] == data_before["total_active"] + 2
    assert data_after["expiring_0_30"] == data_before["expiring_0_30"] + 1
    assert data_after["already_expired"] == data_before["already_expired"] + 1


def test_contract_expiring_list(client: TestClient):
    headers = _login(client)
    emp_id, cat_id = asyncio.get_event_loop().run_until_complete(_get_employee_and_category_ids())
    today = date.today()

    # Tạo HĐ hết hạn trong 5 ngày (CRITICAL) và 45 ngày (NOTICE)
    asyncio.get_event_loop().run_until_complete(
        _create_test_contract(
            employee_id=emp_id,
            category_id=cat_id,
            contract_number="TEST-CON-003",
            signed_date=today - timedelta(days=100),
            effective_from=today - timedelta(days=100),
            effective_to=today + timedelta(days=5),
            status="active"
        )
    )
    asyncio.get_event_loop().run_until_complete(
        _create_test_contract(
            employee_id=emp_id,
            category_id=cat_id,
            contract_number="TEST-CON-004",
            signed_date=today - timedelta(days=100),
            effective_from=today - timedelta(days=100),
            effective_to=today + timedelta(days=45),
            status="active"
        )
    )

    # Gọi API
    resp = client.get(f"{BASE}/expiring", params={"days_ahead": 90}, headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"] >= 2

    # Tìm các record test
    test_records = [item for item in data["items"] if item["contract_number"] in ["TEST-CON-003", "TEST-CON-004"]]
    assert len(test_records) == 2

    critical_item = next(item for item in test_records if item["contract_number"] == "TEST-CON-003")
    assert critical_item["urgency"] == "CRITICAL"
    assert critical_item["days_remaining"] == 5

    notice_item = next(item for item in test_records if item["contract_number"] == "TEST-CON-004")
    # Đúng rồi, 45 ngày > 30 nên là "NOTICE"
    assert notice_item["urgency"] == "NOTICE"

    # Test keyword search
    resp_search = client.get(f"{BASE}/expiring", params={"keyword": "TEST-CON-003"}, headers=headers)
    assert resp_search.status_code == 200
    data_search = resp_search.json()
    assert len(data_search["items"]) == 1
    assert data_search["items"][0]["contract_number"] == "TEST-CON-003"


def test_contract_by_type(client: TestClient):
    headers = _login(client)
    resp = client.get(f"{BASE}/by-type", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "total_contracts" in data


def test_contract_expiry_forecast(client: TestClient):
    headers = _login(client)
    resp = client.get(f"{BASE}/expiry-forecast", params={"months_ahead": 12}, headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert len(data["months"]) == 12
    assert data["months_ahead"] == 12


def test_contract_history(client: TestClient):
    headers = _login(client)
    emp_id, cat_id = asyncio.get_event_loop().run_until_complete(_get_employee_and_category_ids())
    today = date.today()

    asyncio.get_event_loop().run_until_complete(
        _create_test_contract(
            employee_id=emp_id,
            category_id=cat_id,
            contract_number="TEST-CON-005",
            signed_date=today - timedelta(days=200),
            effective_from=today - timedelta(days=200),
            effective_to=today + timedelta(days=200),
            status="active"
        )
    )

    resp = client.get(f"{BASE}/history", params={"employee_id": emp_id}, headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["employee_id"] == emp_id
    assert len(data["items"]) >= 1
    assert any(item["contract_number"] == "TEST-CON-005" for item in data["items"])


def test_contract_export(client: TestClient):
    headers = _login(client)
    resp = client.get(f"{BASE}/export", params={"status": "all"}, headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    assert len(wb.sheetnames) == 1
    assert wb.sheetnames[0] == "Hợp đồng"
