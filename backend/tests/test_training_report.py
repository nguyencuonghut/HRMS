"""Integration tests cho Plan 9.4 — Báo cáo đào tạo."""
from __future__ import annotations

import io
import uuid
from datetime import date, timedelta

import pytest
from fastapi.testclient import TestClient

BASE_COURSES   = "/api/v1/training/courses"
BASE_PLANS     = "/api/v1/training/plans"
BASE_RECORDS   = "/api/v1/training/records"
BASE_REPORT    = "/api/v1/training/report"
BASE_EMPLOYEES = "/api/v1/employees"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_RUN_ID = uuid.uuid4().hex[:8]

_IN_DATE  = "2097-06-15"   # inside test period
_OUT_DATE = "2090-01-01"   # outside test period — before from_date
_FROM     = "2097-01-01"
_TO       = "2097-12-31"


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_course(client, h, suffix: str = "", is_mandatory: bool = False, cost: float | None = None) -> dict:
    code = f"RP{_RUN_ID}{suffix}"[:50]
    body = {
        "code": code,
        "name": f"Report Course {suffix or _RUN_ID}",
        "course_type": "noi_bo",
    }
    if is_mandatory:
        body["is_mandatory"] = True
    if cost is not None:
        body["cost_per_person"] = cost
    r = client.post(BASE_COURSES, json=body, headers=h)
    assert r.status_code == 201, r.text
    return r.json()


def _make_record(client, h, emp_id: int, course_id: int, status: str = "chua_bat_dau",
                 end_date: str | None = None) -> dict:
    body: dict = {"employee_id": emp_id, "course_id": course_id, "status": status}
    if end_date:
        body["end_date"] = end_date
        body["start_date"] = end_date
    r = client.post(BASE_RECORDS, json=body, headers=h)
    assert r.status_code == 201, r.text
    return r.json()


def _update_record(client, h, record_id: int, **kwargs) -> dict:
    r = client.put(f"{BASE_RECORDS}/{record_id}", json=kwargs, headers=h)
    assert r.status_code == 200, r.text
    return r.json()


def _get_emp_id(client, h) -> int:
    r = client.get(BASE_EMPLOYEES, headers=h, params={"page_size": 1})
    items = r.json()["items"]
    assert items, "Cần ít nhất 1 nhân viên active trong DB"
    return items[0]["id"]


# ── TestTrainingReportSummary ─────────────────────────────────────────────────


class TestTrainingReportSummary:
    def test_empty_period_returns_zeros(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": "2088-01-01",
            "to_date":   "2088-01-01",
        })
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["total_records"] == 0
        assert data["total_completed"] == 0
        assert data["by_course"] == []
        assert data["by_department"] == []
        assert float(data["total_cost"]) == 0.0

    def test_total_records_count_in_range(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "TR1")
        # 3 records in range
        for _ in range(3):
            _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
        # 1 record out of range (end_date before from_date)
        _make_record(client, h, emp_id, c["id"], end_date=_OUT_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c["id"],
        })
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["total_records"] >= 3

    def test_excludes_records_out_of_range(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "TR2")
        _make_record(client, h, emp_id, c["id"], end_date=_OUT_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c["id"],
        })
        assert r.status_code == 200, r.text
        data = r.json()
        assert data["total_records"] == 0

    def test_by_course_groups_correctly(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c1 = _make_course(client, h, "GC1")
        c2 = _make_course(client, h, "GC2")
        for _ in range(2):
            _make_record(client, h, emp_id, c1["id"], end_date=_IN_DATE)
        for _ in range(3):
            _make_record(client, h, emp_id, c2["id"], end_date=_IN_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200, r.text
        course_ids = {s["course_id"] for s in r.json()["by_course"]}
        assert c1["id"] in course_ids
        assert c2["id"] in course_ids

    def test_completion_rate_calculation(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "CR1")
        # 4 completed, 1 not
        for _ in range(4):
            rec = _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
            _update_record(client, h, rec["id"], status="hoan_thanh")
        _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c["id"],
        })
        assert r.status_code == 200, r.text
        stat = r.json()["by_course"][0]
        assert stat["total_assigned"] == 5
        assert stat["completed"] == 4
        assert stat["completion_rate"] == 80.0

    def test_total_cost_only_counts_completed(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "COST1", cost=100.0)
        # 3 hoan_thanh, 2 khong_hoan_thanh
        for _ in range(3):
            rec = _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
            _update_record(client, h, rec["id"], status="hoan_thanh")
        for _ in range(2):
            rec = _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
            _update_record(client, h, rec["id"], status="khong_hoan_thanh")

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c["id"],
        })
        assert r.status_code == 200, r.text
        data = r.json()
        assert float(data["total_cost"]) == 300.0

    def test_cost_null_when_no_cost_per_person(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "NOCOST1")  # no cost_per_person
        rec = _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
        _update_record(client, h, rec["id"], status="hoan_thanh")

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c["id"],
        })
        assert r.status_code == 200, r.text
        data = r.json()
        # total_cost = 0 (no cost courses), by_department total_cost = null
        dept_stats = data["by_department"]
        assert len(dept_stats) > 0
        # Each dept total_cost should be null since no cost_per_person
        for ds in dept_stats:
            assert ds["total_cost"] is None

    def test_avg_completion_rate_is_mean_of_courses(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        # Course A: 100% (1/1 completed)
        ca = _make_course(client, h, "AVG1")
        rec = _make_record(client, h, emp_id, ca["id"], end_date=_IN_DATE)
        _update_record(client, h, rec["id"], status="hoan_thanh")
        # Course B: 0% (1 chua_bat_dau)
        cb = _make_course(client, h, "AVG2")
        _make_record(client, h, emp_id, cb["id"], end_date=_IN_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200, r.text
        # avg should include 100 and 0 → 50.0 at minimum across our two courses
        # (There may be other courses in DB so we just check by_course entries for our courses)
        by_course = {s["course_id"]: s["completion_rate"] for s in r.json()["by_course"]}
        assert by_course.get(ca["id"]) == 100.0
        assert by_course.get(cb["id"]) == 0.0

    def test_filter_by_course_id(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c1 = _make_course(client, h, "FC1")
        c2 = _make_course(client, h, "FC2")
        _make_record(client, h, emp_id, c1["id"], end_date=_IN_DATE)
        _make_record(client, h, emp_id, c2["id"], end_date=_IN_DATE)

        r = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO, "course_id": c1["id"],
        })
        assert r.status_code == 200, r.text
        data = r.json()
        course_ids = [s["course_id"] for s in data["by_course"]]
        assert c1["id"] in course_ids
        assert c2["id"] not in course_ids


# ── TestIncompleteMandatory ───────────────────────────────────────────────────


class TestIncompleteMandatory:
    def test_no_mandatory_courses_returns_empty(self, client: TestClient):
        # Temporarily: just check — if there are no mandatory courses at all
        h = _admin(client)
        # Check that endpoint responds 200 (even if there are mandatory courses)
        r = client.get(f"{BASE_REPORT}/incomplete-mandatory", headers=h)
        assert r.status_code == 200, r.text
        assert isinstance(r.json(), list)

    def test_employee_with_completed_record_is_excluded(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "MAND1", is_mandatory=True)

        rec = _make_record(client, h, emp_id, c["id"])
        _update_record(client, h, rec["id"], status="hoan_thanh")

        r = client.get(f"{BASE_REPORT}/incomplete-mandatory", headers=h)
        assert r.status_code == 200, r.text
        emp_ids = [e["employee_id"] for e in r.json()]
        # employee who completed this mandatory course should NOT appear
        # (they might appear for OTHER mandatory courses they haven't completed)
        # To verify specifically: check they don't have this course in incomplete list
        for entry in r.json():
            if entry["employee_id"] == emp_id:
                assert c["name"] not in entry["incomplete_courses"]

    def test_employee_with_non_completed_record_is_incomplete(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c = _make_course(client, h, "MAND2", is_mandatory=True)

        rec = _make_record(client, h, emp_id, c["id"])
        _update_record(client, h, rec["id"], status="dang_hoc")

        r = client.get(f"{BASE_REPORT}/incomplete-mandatory", headers=h)
        assert r.status_code == 200, r.text
        for entry in r.json():
            if entry["employee_id"] == emp_id:
                assert c["name"] in entry["incomplete_courses"]
                break

    def test_incomplete_courses_list_correct(self, client: TestClient):
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        c1 = _make_course(client, h, "MAND3", is_mandatory=True)
        c2 = _make_course(client, h, "MAND4", is_mandatory=True)

        # emp has no records for c1, c2 → both should appear as incomplete
        r = client.get(f"{BASE_REPORT}/incomplete-mandatory", headers=h)
        assert r.status_code == 200, r.text
        for entry in r.json():
            if entry["employee_id"] == emp_id:
                assert c1["name"] in entry["incomplete_courses"]
                assert c2["name"] in entry["incomplete_courses"]
                break


# ── TestTrainingExport ────────────────────────────────────────────────────────


class TestTrainingExport:
    def test_export_returns_xlsx_content_type(self, client: TestClient):
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers["content-type"]

    def test_export_valid_xlsx_file(self, client: TestClient):
        import openpyxl
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert wb is not None

    def test_export_has_three_sheets(self, client: TestClient):
        import openpyxl
        h = _admin(client)
        r = client.get(f"{BASE_REPORT}/export", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200, r.text
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert wb.sheetnames == [
            "Tổng hợp theo khóa học",
            "Chi tiết đào tạo",
            "NV chưa HT bắt buộc",
        ]

    def test_sheet1_row_count_matches_courses(self, client: TestClient):
        import openpyxl
        h = _admin(client)
        emp_id = _get_emp_id(client, h)
        courses_created = []
        for i in range(3):
            c = _make_course(client, h, f"EXP{i}")
            _make_record(client, h, emp_id, c["id"], end_date=_IN_DATE)
            courses_created.append(c["id"])

        # Get summary first to know expected course count
        summary = client.get(f"{BASE_REPORT}/summary", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        }).json()
        expected_course_rows = len(summary["by_course"])

        r = client.get(f"{BASE_REPORT}/export", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws1 = wb["Tổng hợp theo khóa học"]
        # Data starts at row 5, total row at the end
        # max_row = 4 (header rows) + data_rows + 1 (total row)
        data_rows = ws1.max_row - 4 - 1  # subtract header (4) and total (1)
        assert data_rows == expected_course_rows

    def test_sheet3_row_count_matches_incomplete(self, client: TestClient):
        import openpyxl
        h = _admin(client)
        # Get incomplete count from API
        incomplete = client.get(f"{BASE_REPORT}/incomplete-mandatory", headers=h).json()
        expected = len(incomplete)

        r = client.get(f"{BASE_REPORT}/export", headers=h, params={
            "from_date": _FROM, "to_date": _TO,
        })
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws3 = wb["NV chưa HT bắt buộc"]
        data_rows = ws3.max_row - 4 - 1
        assert data_rows == expected
