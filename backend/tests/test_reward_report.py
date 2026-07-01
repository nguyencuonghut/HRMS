"""Integration tests cho Plan 8.3 Slice 4 — Báo cáo khen thưởng / kỷ luật.

Covers:
  - TestRewardDisciplineReport: summary report (counts, grouping, date range filter)
  - TestRewardDisciplineExport: Excel export (content-type, sheets, row counts)
"""
from __future__ import annotations

import io
import json
import uuid

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.security import hash_password

BASE = "/api/v1/rewards/report"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_RUN_ID = uuid.uuid4().hex[:8]

# Ngày test — tương lai xa để tránh collision với dữ liệu thật
_YEAR = 2097
_IN_RANGE = f"{_YEAR}-06-15"
_BEFORE = f"{_YEAR - 1}-12-31"
_AFTER = f"{_YEAR + 1}-01-01"
_FROM = f"{_YEAR}-01-01"
_TO = f"{_YEAR}-12-31"

# Năm hoàn toàn trống (không có test nào tạo dữ liệu ở đây)
_EMPTY_YEAR = 2095
_EMPTY_FROM = f"{_EMPTY_YEAR}-01-01"
_EMPTY_TO = f"{_EMPTY_YEAR}-12-31"


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _login(client: TestClient, email: str, password: str) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": password},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


async def _get_active_employee() -> dict:
    async with _make_session()() as s:
        r = await s.execute(text(
            "SELECT id, full_name FROM employees WHERE status != 'resigned' ORDER BY id LIMIT 1"
        ))
        row = r.fetchone()
        assert row is not None, "Không tìm thấy nhân viên active"
        return {"id": row[0], "name": row[1]}


async def _get_active_reward_type() -> dict:
    """Trả về loại khen thưởng có tiền (is_monetary=true)."""
    async with _make_session()() as s:
        r = await s.execute(text(
            "SELECT id, name, is_monetary FROM reward_types "
            "WHERE is_active = true AND is_monetary = true LIMIT 1"
        ))
        row = r.fetchone()
        assert row is not None, "Không tìm thấy loại khen thưởng có tiền"
        return {"id": row[0], "name": row[1], "is_monetary": row[2]}


async def _get_non_monetary_reward_type() -> dict:
    """Trả về loại khen thưởng không có tiền (is_monetary=false)."""
    async with _make_session()() as s:
        r = await s.execute(text(
            "SELECT id, name, is_monetary FROM reward_types "
            "WHERE is_active = true AND is_monetary = false LIMIT 1"
        ))
        row = r.fetchone()
        assert row is not None, "Không tìm thấy loại khen thưởng không tiền"
        return {"id": row[0], "name": row[1], "is_monetary": row[2]}


def _create_reward(client: TestClient, h: dict, emp_id: int, type_id: int,
                   reward_date: str, value: str | None = None) -> dict:
    data: dict = {
        "employee_id": emp_id,
        "reward_type_id": type_id,
        "reward_date": reward_date,
        "title": f"Test reward {_RUN_ID}",
    }
    if value is not None:
        data["value"] = value
    payload = json.dumps(data)
    return client.post("/api/v1/rewards", data={"body": payload}, headers=h).json()


def _create_discipline(client: TestClient, h: dict, emp_id: int,
                       effective_date: str) -> dict:
    payload = json.dumps({
        "employee_id": emp_id,
        "discipline_form": "khien_trach",
        "violation_date": effective_date,
        "effective_date": effective_date,
        "title": f"Test disc {_RUN_ID}",
    })
    return client.post("/api/v1/disciplines", data={"body": payload}, headers=h).json()


def _get_summary(client: TestClient, h: dict, from_date: str = _FROM,
                 to_date: str = _TO, department_id: int | None = None) -> dict:
    params: dict = {"from_date": from_date, "to_date": to_date}
    if department_id is not None:
        params["department_id"] = department_id
    r = client.get(f"{BASE}/summary", params=params, headers=h)
    assert r.status_code == 200, f"Expected 200, got {r.status_code}: {r.text}"
    return r.json()


async def _create_disciplines_report_viewer() -> tuple[str, str]:
    email = f"disciplines_report_{uuid.uuid4().hex[:8]}@hrms.local"
    password = "Hrms@2026"
    role_code = f"disciplines_report_{uuid.uuid4().hex[:8]}"

    async with _make_session()() as s:
        await s.execute(
            text("""
                INSERT INTO roles (code, name, description, is_system, created_at)
                VALUES (:code, :name, :description, false, now())
            """),
            {
                "code": role_code,
                "name": f"Disciplines report {role_code}",
                "description": "Test role for reward/discipline report",
            },
        )
        await s.execute(
            text("""
                INSERT INTO users (email, full_name, hashed_password, is_active, is_superuser, created_at)
                VALUES (:email, :full_name, :hashed_password, true, false, now())
            """),
            {
                "email": email,
                "full_name": "Disciplines Report Viewer",
                "hashed_password": hash_password(password),
            },
        )
        await s.execute(
            text("""
                INSERT INTO user_roles (user_id, role_id)
                SELECT u.id, r.id
                FROM users u, roles r
                WHERE u.email = :email AND r.code = :role_code
            """),
            {"email": email, "role_code": role_code},
        )
        await s.execute(
            text("""
                INSERT INTO role_permissions (role_id, permission_id)
                SELECT r.id, p.id
                FROM roles r
                JOIN permissions p ON p.code IN ('reports:view', 'disciplines:view')
                WHERE r.code = :role_code
                ON CONFLICT DO NOTHING
            """),
            {"role_code": role_code},
        )
        await s.commit()

    return email, password


# ── TestRewardDisciplineReport ─────────────────────────────────────────────────


class TestRewardDisciplineReport:
    def test_empty_period_returns_zero_counts(self, client: TestClient):
        """Kỳ không có dữ liệu → total_rewards=0, total_disciplines=0."""
        h = _admin(client)
        data = _get_summary(client, h, from_date=_EMPTY_FROM, to_date=_EMPTY_TO)
        assert data["summary"]["total_rewards"] == 0
        assert data["summary"]["total_disciplines"] == 0
        assert len(data["summary"]["by_discipline_form"]) == 4

    @pytest.mark.asyncio
    async def test_counts_rewards_in_date_range(self, client: TestClient):
        """Tạo 2 khen thưởng trong kỳ → total_rewards >= 2."""
        emp = await _get_active_employee()
        rt = await _get_non_monetary_reward_type()
        h = _admin(client)

        r1 = _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)
        r2 = _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)
        assert r1.get("id") is not None, f"Reward 1 creation failed: {r1}"
        assert r2.get("id") is not None, f"Reward 2 creation failed: {r2}"

        data = _get_summary(client, h)
        assert data["summary"]["total_rewards"] >= 2
        assert data["total_rewards"] >= 2

    @pytest.mark.asyncio
    async def test_excludes_rewards_outside_date_range(self, client: TestClient):
        """Khen thưởng ngoài kỳ không được tính vào báo cáo."""
        emp = await _get_active_employee()
        rt = await _get_non_monetary_reward_type()
        h = _admin(client)

        # Lấy count trước khi tạo
        before = _get_summary(client, h)
        count_before = before["summary"]["total_rewards"]

        # Tạo 1 reward ngoài kỳ (_BEFORE) và 1 ngoài kỳ (_AFTER)
        _create_reward(client, h, emp["id"], rt["id"], _BEFORE)
        _create_reward(client, h, emp["id"], rt["id"], _AFTER)

        after = _get_summary(client, h)
        count_after = after["summary"]["total_rewards"]

        # Count không được tăng thêm
        assert count_after == count_before

    @pytest.mark.asyncio
    async def test_counts_disciplines_in_date_range(self, client: TestClient):
        """Tạo 1 kỷ luật trong kỳ → total_disciplines >= 1."""
        emp = await _get_active_employee()
        h = _admin(client)

        _create_discipline(client, h, emp["id"], _IN_RANGE)

        data = _get_summary(client, h)
        assert data["summary"]["total_disciplines"] >= 1

    def test_by_discipline_form_always_has_4_entries(self, client: TestClient):
        """by_discipline_form luôn có đúng 4 hình thức, theo thứ tự chuẩn."""
        h = _admin(client)
        data = _get_summary(client, h, from_date=_EMPTY_FROM, to_date=_EMPTY_TO)
        forms = [item["discipline_form"] for item in data["summary"]["by_discipline_form"]]
        assert len(forms) == 4
        assert forms == ["khien_trach", "keo_dai_nang_luong", "cach_chuc", "sa_thai"]

    @pytest.mark.asyncio
    async def test_by_reward_type_groups_correctly(self, client: TestClient):
        """2 khen thưởng cùng loại → by_reward_type có count >= 2 cho loại đó."""
        emp = await _get_active_employee()
        rt = await _get_non_monetary_reward_type()
        h = _admin(client)

        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)
        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)

        data = _get_summary(client, h)
        by_type = data["summary"]["by_reward_type"]
        matching = [x for x in by_type if x["reward_type_id"] == rt["id"]]
        assert len(matching) == 1
        assert matching[0]["count"] >= 2

    @pytest.mark.asyncio
    async def test_monetary_reward_value_summed_correctly(self, client: TestClient):
        """Khen thưởng có giá trị tiền → total_reward_value > 0."""
        emp = await _get_active_employee()
        rt = await _get_active_reward_type()
        h = _admin(client)

        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE, value="500000")

        data = _get_summary(client, h)
        total_val = float(data["summary"]["total_reward_value"])
        assert total_val > 0

        by_type = data["summary"]["by_reward_type"]
        matching = [x for x in by_type if x["reward_type_id"] == rt["id"]]
        assert len(matching) == 1
        assert matching[0]["total_value"] is not None
        assert float(matching[0]["total_value"]) > 0

    @pytest.mark.asyncio
    async def test_by_discipline_form_counts_correctly(self, client: TestClient):
        """Tạo 1 kỷ luật khiển trách → by_discipline_form[khien_trach].count >= 1."""
        emp = await _get_active_employee()
        h = _admin(client)

        _create_discipline(client, h, emp["id"], _IN_RANGE)

        data = _get_summary(client, h)
        form_stats = data["summary"]["by_discipline_form"]
        kt = next((x for x in form_stats if x["discipline_form"] == "khien_trach"), None)
        assert kt is not None
        assert kt["count"] >= 1

    @pytest.mark.asyncio
    async def test_by_department_merges_reward_and_discipline_stats(self, client: TestClient):
        """1 khen thưởng + 1 kỷ luật cùng nhân viên → by_department không rỗng."""
        emp = await _get_active_employee()
        rt = await _get_non_monetary_reward_type()
        h = _admin(client)

        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)
        _create_discipline(client, h, emp["id"], _IN_RANGE)

        data = _get_summary(client, h)
        by_dept = data["summary"]["by_department"]
        assert len(by_dept) > 0
        # Phải có ít nhất 1 dept với cả reward và discipline
        has_both = any(
            d["reward_count"] >= 1 and d["discipline_count"] >= 1
            for d in by_dept
        )
        assert has_both

    def test_unauthenticated_returns_401(self, client: TestClient):
        """Không có auth header → 401."""
        r = client.get(f"{BASE}/summary", params={"from_date": _FROM, "to_date": _TO})
        assert r.status_code == 401

    @pytest.mark.asyncio
    async def test_summary_accepts_disciplines_view_with_reports_view(self, client: TestClient):
        email, password = await _create_disciplines_report_viewer()
        headers = _login(client, email, password)

        response = client.get(
            f"{BASE}/summary",
            params={"from_date": _FROM, "to_date": _TO},
            headers=headers,
        )
        assert response.status_code == 200, response.text


# ── TestRewardDisciplineExport ─────────────────────────────────────────────────


class TestRewardDisciplineExport:
    def _get_export(self, client: TestClient, h: dict,
                    from_date: str = _FROM, to_date: str = _TO):
        r = client.get(f"{BASE}/export", params={"from_date": from_date, "to_date": to_date},
                       headers=h)
        return r

    def test_returns_xlsx_content_type(self, client: TestClient):
        """Export trả về content-type spreadsheetml."""
        h = _admin(client)
        r = self._get_export(client, h)
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_file_is_valid_xlsx(self, client: TestClient):
        """Bytes trả về là file xlsx hợp lệ."""
        h = _admin(client)
        r = self._get_export(client, h)
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert wb is not None

    def test_sheet_khen_thuong_exists(self, client: TestClient):
        """Workbook có sheet 'Khen thưởng'."""
        h = _admin(client)
        r = self._get_export(client, h)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Khen thưởng" in wb.sheetnames

    def test_sheet_ky_luat_exists(self, client: TestClient):
        """Workbook có sheet 'Kỷ luật'."""
        h = _admin(client)
        r = self._get_export(client, h)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        assert "Kỷ luật" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_reward_sheet_correct_row_count(self, client: TestClient):
        """Sheet 'Khen thưởng': số dòng dữ liệu khớp với total_rewards trong báo cáo."""
        emp = await _get_active_employee()
        rt = await _get_non_monetary_reward_type()
        h = _admin(client)

        # Tạo thêm 2 phần tử và lấy tổng từ báo cáo
        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)
        _create_reward(client, h, emp["id"], rt["id"], _IN_RANGE)

        # Lấy tổng số reward trong kỳ từ report (page_size lớn)
        report_r = client.get(
            f"{BASE}/summary",
            params={"from_date": _FROM, "to_date": _TO,
                    "reward_page_size": 200, "discipline_page_size": 200},
            headers=h,
        )
        assert report_r.status_code == 200
        total_rewards = report_r.json()["total_rewards"]

        # Export và đếm dòng
        r = self._get_export(client, h)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Khen thưởng"]
        # rows 1-4: header (công ty, tiêu đề, trống, cột); dòng cuối: tổng cộng
        data_rows = ws.max_row - 5
        assert data_rows >= total_rewards

    @pytest.mark.asyncio
    async def test_discipline_sheet_correct_row_count(self, client: TestClient):
        """Sheet 'Kỷ luật': số dòng dữ liệu >= total_disciplines trong báo cáo."""
        emp = await _get_active_employee()
        h = _admin(client)

        _create_discipline(client, h, emp["id"], _IN_RANGE)

        report_r = client.get(
            f"{BASE}/summary",
            params={"from_date": _FROM, "to_date": _TO,
                    "reward_page_size": 200, "discipline_page_size": 200},
            headers=h,
        )
        assert report_r.status_code == 200
        total_disciplines = report_r.json()["total_disciplines"]

        r = self._get_export(client, h)
        wb = openpyxl.load_workbook(io.BytesIO(r.content))
        ws = wb["Kỷ luật"]
        data_rows = ws.max_row - 5
        assert data_rows >= total_disciplines

    def test_filename_contains_date_range(self, client: TestClient):
        """Content-Disposition header chứa from_date và to_date."""
        h = _admin(client)
        r = self._get_export(client, h, from_date=_FROM, to_date=_TO)
        assert r.status_code == 200
        disposition = r.headers.get("content-disposition", "")
        assert _FROM in disposition
        assert _TO in disposition
