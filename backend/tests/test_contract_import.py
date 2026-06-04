"""Tests cho 12.1 Slice 1 — Import hợp đồng hàng loạt."""
from __future__ import annotations

import asyncio
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.models.salary import BhxhPositionGroup
from app.schemas.employee import EmployeeCreate
from app.seeds import bootstrap
from app.services import employee_service

BASE      = "/api/v1/imports"
BASE_AUTH = "/api/v1/auth/login"
BASE_EMPLOYEES = "/api/v1/employees"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

# Contract category codes từ seed
_CAT_INDEFINITE = "labor_indefinite"    # document_kind = labor_contract
_CAT_DEFINITE   = "labor_definite"      # document_kind = labor_contract
_CAT_APPENDIX   = "appendix_salary_change"  # document_kind = contract_appendix

_PREFIX = "TESTIMP"   # prefix số HĐ test → dễ cleanup


def _login(client: TestClient) -> dict:
    token = client.post(BASE_AUTH, json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        await s.execute(text(f"DELETE FROM employee_contracts WHERE contract_number LIKE '{_PREFIX}%'"))
        employee_ids = [e.id for e in (await s.execute(select(Employee))).scalars().all() if e.id_number.startswith("TESTIMPSEQ")]
        if employee_ids:
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup():
    await _cleanup()
    yield
    await _cleanup()


HEADERS = [
    "Mã nhân viên", "Số hợp đồng", "Loại văn bản",
    "Mã loại hợp đồng", "Ngày ký", "Ngày hiệu lực",
    "Ngày hết hạn", "Mức lương BHXH",
]
HEADERS_WITH_SEQUENCE = [
    "Mã nhân viên", "Hệ mã nhân viên", "Số hợp đồng", "Loại văn bản",
    "Mã loại hợp đồng", "Ngày ký", "Ngày hiệu lực",
    "Ngày hết hạn", "Mức lương BHXH",
]
HEADERS_WITH_BHXH_MODE = [
    "Mã nhân viên", "Hệ mã nhân viên", "Số hợp đồng", "Loại văn bản",
    "Mã loại hợp đồng", "Ngày ký", "Ngày hiệu lực",
    "Ngày hết hạn", "Mode lương BHXH", "Mã nhóm vị trí BHXH",
    "Bậc hệ số BHXH", "Ngày bắt đầu tính thâm niên BHXH", "Mức lương BHXH",
]


def _make_xlsx(rows: list[list]) -> bytes:
    return _make_xlsx_with_headers(HEADERS, rows)


def _make_xlsx_with_headers(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, headers, xlsx_bytes: bytes):
    return client.post(
        f"{BASE}/contracts",
        files={"file": ("import.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


def _employee_display_code(client: TestClient, headers: dict, employee_id: int) -> str:
    r = client.get(f"{BASE_EMPLOYEES}/{employee_id}", headers=headers)
    assert r.status_code == 200, r.text
    return r.json()["display_code"]


def _test_employee_display_code(
    client: TestClient,
    headers: dict,
    *,
    id_number: str,
    employee_seq: int,
    sequence_code: str = "SYS1",
) -> str:
    employee_id = asyncio.run(_create_employee_same_seq(id_number, sequence_code, employee_seq))
    return _employee_display_code(client, headers, employee_id)


async def _get_sequence(code: str) -> EmployeeCodeSequence:
    async with _make_session()() as s:
        return (
            await s.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == code))
        ).scalar_one()


def _employee_payload(id_number: str, *, sequence_id: int, employee_seq: int) -> EmployeeCreate:
    return EmployeeCreate(
        employee_seq=employee_seq,
        employee_code_sequence_id=sequence_id,
        full_name=f"Import Seq {id_number}",
        last_name="Import",
        first_name=id_number,
        date_of_birth="1990-01-01",
        gender="male",
        nationality_id=1,
        id_number=id_number,
        id_issued_on="2020-01-01",
        id_issued_by="Cuc Canh sat",
        status="official",
        start_date="2026-01-01",
    )


async def _create_employee_same_seq(id_number: str, sequence_code: str, employee_seq: int) -> int:
    sequence = await _get_sequence(sequence_code)
    async with _make_session()() as s:
        employee = await employee_service.create_employee(
            s,
            _employee_payload(id_number, sequence_id=sequence.id, employee_seq=employee_seq),
        )
        await s.commit()
        await s.refresh(employee)
        return employee.id


async def _ensure_bhxh_foundation() -> None:
    async with _make_session()() as s:
        await bootstrap.seed_salary_scale(s)
        await s.commit()


async def _get_bhxh_group(code: str) -> BhxhPositionGroup:
    async with _make_session()() as s:
        row = await s.execute(select(BhxhPositionGroup).where(BhxhPositionGroup.code == code))
        group = row.scalar_one_or_none()
        assert group is not None
        return group


async def _get_contract_state(contract_number: str) -> dict | None:
    async with _make_session()() as s:
        row = await s.execute(
            text(
                """
                SELECT insurance_salary_mode,
                       bhxh_position_group_id,
                       insurance_salary_grade_no,
                       bhxh_seniority_start_date,
                       insurance_salary,
                       insurance_salary_fixed_amount
                FROM employee_contracts
                WHERE contract_number = :num
                """
            ),
            {"num": contract_number},
        )
        data = row.fetchone()
        if data is None:
            return None
        return {
            "mode": data[0],
            "group_id": data[1],
            "grade_no": data[2],
            "seniority_start_date": data[3],
            "insurance_salary": data[4],
            "fixed_amount": data[5],
        }


async def _get_profile_basis_state(employee_id: int) -> dict | None:
    async with _make_session()() as s:
        row = await s.execute(
            text(
                """
                SELECT insurance_basis_source, insurance_basis_amount, company_bhxh_joined_date
                FROM employee_insurance_profiles
                WHERE employee_id = :eid
                """
            ),
            {"eid": employee_id},
        )
        data = row.fetchone()
        if data is None:
            return None
        return {
            "source": data[0],
            "amount": data[1],
            "joined_date": data[2],
        }


# ── Template ──────────────────────────────────────────────────────────────────

def test_template_download(client: TestClient):
    h = _login(client)
    r = client.get(f"{BASE}/contracts/template", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "mau_import_hop_dong" in r.headers.get("content-disposition", "")
    # Verify xlsx is readable và có sheet Hướng dẫn
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert "Hướng dẫn" in wb.sheetnames


# ── Import success ────────────────────────────────────────────────────────────

def test_import_success(client: TestClient):
    h = _login(client)
    code_1 = _test_employee_display_code(client, h, id_number="TESTIMPSEQC101", employee_seq=9201)
    code_2 = _test_employee_display_code(client, h, id_number="TESTIMPSEQC102", employee_seq=9202)
    rows = [
        [code_1, f"{_PREFIX}-001", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", "5000000"],
        [code_2, f"{_PREFIX}-002", "labor_contract", _CAT_DEFINITE,   "01/01/2026", "01/01/2026", "31/12/2026", "4500000"],
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 2
    assert body["success"] == 2
    assert body["failed"] == 0
    assert body["errors"] == []


def test_import_no_expiry_is_valid(client: TestClient):
    """Ngày hết hạn bỏ trống = vô thời hạn → hợp lệ."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC103", employee_seq=9203)
    rows = [[code, f"{_PREFIX}-003", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["success"] == 1


# ── Validation errors ─────────────────────────────────────────────────────────

def test_import_invalid_employee(client: TestClient):
    h = _login(client)
    rows = [["999999", f"{_PREFIX}-004", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("Không tìm thấy nhân viên" in e["message"] for e in body["errors"])


def test_import_invalid_document_kind(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC104", employee_seq=9204)
    rows = [[code, f"{_PREFIX}-005", "invalid_kind", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("labor_contract" in e["message"] for e in body["errors"])


def test_import_invalid_category_code(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC105", employee_seq=9205)
    rows = [[code, f"{_PREFIX}-006", "labor_contract", "nonexistent_cat", "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("loại hợp đồng" in e["message"].lower() for e in body["errors"])


def test_import_date_order_error(client: TestClient):
    """effective_to < effective_from → lỗi."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC106", employee_seq=9206)
    rows = [[code, f"{_PREFIX}-007", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/06/2026", "01/01/2026", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("hết hạn" in e["message"].lower() for e in body["errors"])


def test_import_duplicate_contract_number(client: TestClient):
    h = _login(client)
    code_1 = _test_employee_display_code(client, h, id_number="TESTIMPSEQC107", employee_seq=9207)
    code_2 = _test_employee_display_code(client, h, id_number="TESTIMPSEQC108", employee_seq=9208)
    rows_first = [[code_1, f"{_PREFIX}-008", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""]]
    _upload(client, h, _make_xlsx(rows_first))  # lần 1 thành công

    rows_dup = [[code_2, f"{_PREFIX}-008", "labor_contract", _CAT_DEFINITE, "01/02/2026", "01/02/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows_dup))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("đã tồn tại" in e["message"] for e in body["errors"])


def test_import_missing_required_field(client: TestClient):
    h = _login(client)
    # Bỏ trống "Số hợp đồng"
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC109", employee_seq=9209)
    rows = [[code, "", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_document_kind_mismatch_category(client: TestClient):
    """document_kind='labor_contract' nhưng category là appendix → lỗi mismatch."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC110", employee_seq=9210)
    rows = [[code, f"{_PREFIX}-009", "labor_contract", _CAT_APPENDIX, "01/01/2026", "01/01/2026", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("không khớp" in e["message"] for e in body["errors"])


def test_import_partial_errors(client: TestClient):
    """Dòng valid và invalid xen kẽ — dòng valid vẫn được tạo."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQC111", employee_seq=9211)
    rows = [
        [code, f"{_PREFIX}-010", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],  # OK
        ["999999", f"{_PREFIX}-011", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],       # NV không tồn tại
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 2
    assert body["success"] == 1
    assert body["failed"] == 1


# ── Permission ────────────────────────────────────────────────────────────────

def test_import_unauthenticated(client: TestClient):
    r = client.post(f"{BASE}/contracts",
                    files={"file": ("f.xlsx", b"x", "application/vnd.ms-excel")})
    assert r.status_code == 401


def test_template_unauthenticated(client: TestClient):
    r = client.get(f"{BASE}/contracts/template")
    assert r.status_code == 401


# ── File validation ───────────────────────────────────────────────────────────

def test_import_rejects_non_xlsx(client: TestClient):
    h = _login(client)
    r = client.post(f"{BASE}/contracts",
                    files={"file": ("data.csv", b"a,b,c", "text/csv")},
                    headers=h)
    assert r.status_code == 400
    assert "xlsx" in r.json()["detail"].lower()


def test_import_empty_file_rows(client: TestClient):
    """File chỉ có header, không có data → success=0."""
    h = _login(client)
    r = _upload(client, h, _make_xlsx([]))
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 0
    assert body["success"] == 0


def test_import_resolves_display_code_exactly_across_three_sequences(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC001", "SYS1", 9101))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC002", "SYS2", 9101))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC003", "SYS3", 9101))

    rows = [
        ["9101", "SYS1", f"{_PREFIX}-S1", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],
        ["9101", "SYS2", f"{_PREFIX}-S2", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],
        ["9101", "SYS3", f"{_PREFIX}-S3", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],
    ]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_SEQUENCE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 3
    assert body["failed"] == 0


def test_import_rejects_ambiguous_seq_without_sequence_code(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC901", "SYS1", 9901))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC902", "SYS2", 9901))

    rows = [
        ["9901", f"{_PREFIX}-AMB", "labor_contract", _CAT_INDEFINITE, "01/01/2026", "01/01/2026", "", ""],
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 0
    assert body["failed"] == 1
    assert any("Hệ mã nhân viên" in e["message"] for e in body["errors"])


def test_import_computed_contract_uses_bhxh_engine_and_syncs_profile(client: TestClient):
    h = _login(client)
    asyncio.run(_ensure_bhxh_foundation())
    employee_id = asyncio.run(_create_employee_same_seq("TESTIMPSEQC777", "SYS1", 9777))
    group = asyncio.run(_get_bhxh_group("OFFICE_STAFF"))
    contract_number = f"{_PREFIX}-CMP"

    rows = [[
        "9777", "SYS1", contract_number, "labor_contract", _CAT_INDEFINITE,
        "01/01/2026", "01/01/2026", "", "computed_by_position_group", "OFFICE_STAFF",
        "3", "01/01/2023", "",
    ]]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_BHXH_MODE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 1
    assert body["failed"] == 0

    contract = asyncio.run(_get_contract_state(contract_number))
    assert contract is not None
    assert contract["mode"] == "computed_by_position_group"
    assert contract["group_id"] == group.id
    assert contract["grade_no"] == 3
    assert str(contract["seniority_start_date"]) == "2023-01-01"
    assert str(contract["insurance_salary"]) == "5423400.00"
    assert contract["fixed_amount"] is None

    profile = asyncio.run(_get_profile_basis_state(employee_id))
    assert profile is not None
    assert profile["source"] == "computed"
    assert str(profile["amount"]) == "5423400.00"
    assert str(profile["joined_date"]) == "2023-01-01"


def test_import_rejects_manual_salary_on_computed_mode(client: TestClient):
    h = _login(client)
    asyncio.run(_ensure_bhxh_foundation())
    asyncio.run(_create_employee_same_seq("TESTIMPSEQC778", "SYS1", 9778))

    rows = [[
        "9778", "SYS1", f"{_PREFIX}-BAD-CMP", "labor_contract", _CAT_INDEFINITE,
        "01/01/2026", "01/01/2026", "", "computed_by_position_group", "OFFICE_STAFF",
        "3", "", "5000000",
    ]]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_BHXH_MODE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 0
    assert body["failed"] == 1
    assert any("không nhận nhập tay" in e["message"] for e in body["errors"])
