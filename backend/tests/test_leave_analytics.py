"""Tests cho 11.3 — Leave Analytics."""
from __future__ import annotations

from datetime import date, timedelta

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings

BASE     = "/api/v1/reports/leaves"
BASE_ENT = "/api/v1/leave-entitlements"
BASE_REC = "/api/v1/leave-records"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

_EMP_ID  = 1
_EMP_ID2 = 2
_LT_ID   = 1   # annual_leave — carryover_allowed=True
_YEAR    = 2093   # năm đủ xa, không conflict với test khác


def _login(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        await s.execute(text(f"DELETE FROM leave_records WHERE EXTRACT(year FROM start_date) = {_YEAR}"))
        await s.execute(text(f"DELETE FROM leave_entitlements WHERE year = {_YEAR}"))
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup():
    await _cleanup()
    yield
    await _cleanup()


def _make_entitlement(client, headers, *, emp_id=_EMP_ID, allocated=12.0,
                      carryover=0.0, carryover_expires=None):
    body: dict = {
        "employee_id":   emp_id,
        "leave_type_id": _LT_ID,
        "year":          _YEAR,
        "allocated_days": allocated,
        "carryover_days": carryover,
    }
    if carryover_expires:
        body["carryover_expires"] = str(carryover_expires)
    r = client.post(BASE_ENT, json=body, headers=headers)
    assert r.status_code == 201, r.text
    return r.json()


def _make_record(client, headers, *, emp_id=_EMP_ID, start=None, end=None, lt_id=_LT_ID):
    start = start or str(date(_YEAR, 6, 1))
    end   = end   or str(date(_YEAR, 6, 3))
    r = client.post(BASE_REC, json={
        "employee_id":   emp_id,
        "leave_type_id": lt_id,
        "start_date":    start,
        "end_date":      end,
    }, headers=headers)
    assert r.status_code == 201, r.text
    return r.json()


# ── Analytics Summary ─────────────────────────────────────────────────────────

def test_analytics_summary_kpis_correct(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    # 3 ngày T6, 2 ngày T8
    _make_record(client, headers, start=str(date(_YEAR, 6, 1)), end=str(date(_YEAR, 6, 3)))
    _make_record(client, headers, start=str(date(_YEAR, 8, 5)), end=str(date(_YEAR, 8, 6)))

    r = client.get(f"{BASE}/analytics-summary", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["year"] == _YEAR
    assert body["total_days_ytd"] == 5.0
    assert isinstance(body["avg_usage_rate"], float)
    assert isinstance(body["employees_not_taken"], int)
    assert isinstance(body["days_expiring_30d"], float)


def test_trend_fills_12_months(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    # Chỉ có dữ liệu tháng 3
    _make_record(client, headers, start=str(date(_YEAR, 3, 5)), end=str(date(_YEAR, 3, 6)))

    r = client.get(f"{BASE}/analytics-summary", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    trend = r.json()["monthly_trend"]
    assert len(trend) == 12
    months = [t["month"] for t in trend]
    assert months == list(range(1, 13))
    march = next(t for t in trend if t["month"] == 3)
    assert march["total_days"] == 2.0
    other = [t for t in trend if t["month"] != 3]
    assert all(t["total_days"] == 0 for t in other)


def test_employees_not_taken_counts_zero_users(client: TestClient):
    headers = _login(client)
    # emp1 có entitlement nhưng chưa nghỉ
    _make_entitlement(client, headers, emp_id=_EMP_ID, allocated=12.0)
    # emp2 có entitlement và có nghỉ
    _make_entitlement(client, headers, emp_id=_EMP_ID2, allocated=10.0)
    _make_record(client, headers, emp_id=_EMP_ID2,
                 start=str(date(_YEAR, 5, 1)), end=str(date(_YEAR, 5, 1)))

    r = client.get(f"{BASE}/analytics-summary", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    # emp1 chưa nghỉ → đếm vào employees_not_taken
    assert body["employees_not_taken"] >= 1


def test_unauthenticated_rejected(client: TestClient):
    endpoints = [
        f"{BASE}/analytics-summary?year={_YEAR}",
        f"{BASE}/by-type?year={_YEAR}",
        f"{BASE}/monthly-heatmap?year={_YEAR}",
        f"{BASE}/top-employees?year={_YEAR}",
        f"{BASE}/expiring-balance?year={_YEAR}",
        f"{BASE}/department-comparison?year={_YEAR}",
    ]
    for url in endpoints:
        r = client.get(url)
        assert r.status_code == 401, f"{url} expected 401, got {r.status_code}"


# ── By Type ───────────────────────────────────────────────────────────────────

def test_by_type_percentage_sums_100(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    _make_record(client, headers, start=str(date(_YEAR, 4, 1)), end=str(date(_YEAR, 4, 3)))
    _make_record(client, headers, start=str(date(_YEAR, 5, 1)), end=str(date(_YEAR, 5, 2)))

    r = client.get(f"{BASE}/by-type", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["grand_total_days"] > 0
    total_pct = sum(i["percentage"] for i in body["items"])
    assert abs(total_pct - 100.0) < 0.1


def test_by_type_returns_correct_days(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    # 3 ngày
    _make_record(client, headers, start=str(date(_YEAR, 7, 1)), end=str(date(_YEAR, 7, 3)))

    r = client.get(f"{BASE}/by-type", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    row = next((i for i in body["items"] if i["leave_type_id"] == _LT_ID), None)
    assert row is not None
    assert row["total_days"] == 3.0
    assert row["record_count"] >= 1


# ── Monthly Heatmap ───────────────────────────────────────────────────────────

def test_heatmap_pivot_correct(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    # 2 ngày tháng 4, 3 ngày tháng 9
    _make_record(client, headers, start=str(date(_YEAR, 4, 10)), end=str(date(_YEAR, 4, 11)))
    _make_record(client, headers, start=str(date(_YEAR, 9, 1)), end=str(date(_YEAR, 9, 3)))

    r = client.get(f"{BASE}/monthly-heatmap", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["departments"]) >= 1
    # company_monthly có đủ 12 tháng
    assert set(int(k) for k in body["company_monthly"].keys()) == set(range(1, 13))
    assert body["company_monthly"]["4"] == 2.0
    assert body["company_monthly"]["9"] == 3.0


# ── Top Employees ─────────────────────────────────────────────────────────────

def test_top_employees_ordering(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, emp_id=_EMP_ID, allocated=12.0)
    _make_entitlement(client, headers, emp_id=_EMP_ID2, allocated=10.0)
    # emp1 nghỉ 1 ngày, emp2 nghỉ 5 ngày
    _make_record(client, headers, emp_id=_EMP_ID,
                 start=str(date(_YEAR, 6, 1)), end=str(date(_YEAR, 6, 1)))
    _make_record(client, headers, emp_id=_EMP_ID2,
                 start=str(date(_YEAR, 6, 5)), end=str(date(_YEAR, 6, 9)))

    r = client.get(f"{BASE}/top-employees", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    items = body["items"]
    assert len(items) >= 2
    assert items[0]["rank"] == 1
    # NV nghỉ nhiều nhất ở vị trí 1
    assert items[0]["total_days_taken"] >= items[1]["total_days_taken"]
    assert items[0]["employee_id"] == _EMP_ID2


def test_top_employees_limit(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    _make_record(client, headers, start=str(date(_YEAR, 7, 1)), end=str(date(_YEAR, 7, 3)))

    r = client.get(f"{BASE}/top-employees", params={"year": _YEAR, "limit": 1}, headers=headers)
    assert r.status_code == 200, r.text
    assert len(r.json()["items"]) <= 1


# ── Expiring Balance ──────────────────────────────────────────────────────────

def test_expiring_balance_threshold_30d(client: TestClient):
    headers = _login(client)
    today = date.today()
    expires_soon = today + timedelta(days=15)
    _make_entitlement(client, headers, allocated=10.0, carryover=5.0,
                      carryover_expires=expires_soon)

    r = client.get(f"{BASE}/expiring-balance", params={"year": _YEAR, "expire_days": 30},
                   headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["items"]) >= 1
    row = next((i for i in body["items"] if i["employee_id"] == _EMP_ID), None)
    assert row is not None
    assert row["days_until_expire"] == (expires_soon - today).days
    assert body["total_expiring_days"] > 0


def test_expiring_balance_threshold_miss(client: TestClient):
    headers = _login(client)
    today = date.today()
    expires_far = today + timedelta(days=45)
    _make_entitlement(client, headers, allocated=10.0, carryover=5.0,
                      carryover_expires=expires_far)

    r = client.get(f"{BASE}/expiring-balance", params={"year": _YEAR, "expire_days": 30},
                   headers=headers)
    assert r.status_code == 200, r.text
    # expires_far > 30 ngày → không xuất hiện
    ids = [i["employee_id"] for i in r.json()["items"]]
    assert _EMP_ID not in ids


# ── Department Comparison ─────────────────────────────────────────────────────

def test_dept_comparison_usage_rate(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=10.0)
    # 5 ngày nghỉ → usage_rate = 5/10 * 100 = 50%
    _make_record(client, headers, start=str(date(_YEAR, 5, 1)), end=str(date(_YEAR, 5, 5)))

    r = client.get(f"{BASE}/department-comparison", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["items"]) >= 1
    assert all(0 <= i["usage_rate"] <= 100 for i in body["items"])


# ── Export Analytics ──────────────────────────────────────────────────────────

def test_export_analytics_returns_xlsx(client: TestClient):
    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    _make_record(client, headers, start=str(date(_YEAR, 6, 1)), end=str(date(_YEAR, 6, 2)))

    r = client.get(f"{BASE}/export-analytics", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    assert "spreadsheetml" in r.headers.get("content-type", "")
    assert "content-disposition" in r.headers


def test_export_analytics_sheet_names(client: TestClient):
    from io import BytesIO
    from openpyxl import load_workbook

    headers = _login(client)
    _make_entitlement(client, headers, allocated=12.0)
    _make_record(client, headers, start=str(date(_YEAR, 6, 1)), end=str(date(_YEAR, 6, 2)))

    r = client.get(f"{BASE}/export-analytics", params={"year": _YEAR}, headers=headers)
    assert r.status_code == 200, r.text
    wb = load_workbook(BytesIO(r.content))
    assert "Tổng quan" in wb.sheetnames
    assert "Theo loại phép" in wb.sheetnames
    assert "Heatmap tháng" in wb.sheetnames
    assert "Top nhân viên" in wb.sheetnames
