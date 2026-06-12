"""Tests cho 12.1 Slice 2b — Import bảo hiểm hàng loạt."""
from __future__ import annotations

import asyncio
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.models.catalog import ContractCategory
from app.core.config import settings
from app.models.catalog import Nationality
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.models.salary import BhxhPositionGroup
from app.schemas.employee import EmployeeCreate
from app.seeds import bootstrap
from app.services import employee_service

BASE      = "/api/v1/imports"
BASE_AUTH = "/api/v1/auth/login"
BASE_EMPLOYEES = "/api/v1/employees"

_ADMIN_EMAIL    = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

_BHYT_CODE = "01001"   # code từ seed bhyt_clinics


def _login(client: TestClient) -> dict:
    token = client.post(BASE_AUTH, json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        # Xoá insurance profile của emp 1, 2, 3 (nếu có) sau test
        # Dùng bhxh_code test để tránh xóa nhầm dữ liệu thật
        await s.execute(text("DELETE FROM employee_insurance_profiles WHERE bhxh_code LIKE 'TESTIMP_%'"))
        employee_ids = [e.id for e in (await s.execute(select(Employee))).scalars().all() if e.id_number.startswith("TESTIMPSEQ")]
        if employee_ids:
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        # Reset về null profile nếu cần
        await s.commit()


@pytest.fixture(autouse=True)
async def cleanup():
    await _cleanup()
    yield
    await _cleanup()


HEADERS = ["Mã nhân viên", "Mã BHXH", "Ngày tham gia", "Mức lương đóng", "Mã bệnh viện KCB", "Trạng thái"]
HEADERS_WITH_SEQUENCE = ["Mã nhân viên", "Hệ mã nhân viên", "Mã BHXH", "Ngày tham gia", "Mức lương đóng", "Mã bệnh viện KCB", "Trạng thái"]
HEADERS_WITH_SOURCE = ["Mã nhân viên", "Hệ mã nhân viên", "Mã BHXH", "Ngày tham gia", "Mức lương đóng", "Nguồn mức lương đóng", "Mã bệnh viện KCB", "Trạng thái"]


def _make_xlsx(rows: list[list]) -> bytes:
    return _make_xlsx_with_headers(HEADERS, rows)


def _make_xlsx_with_headers(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(client, headers, xlsx_bytes: bytes):
    return client.post(
        f"{BASE}/insurance",
        files={"file": ("import.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )


def _employee_display_code(client: TestClient, headers: dict, employee_id: int) -> str:
    r = client.get(f"{BASE_EMPLOYEES}/{employee_id}", headers=headers)
    assert r.status_code == 200, r.text
    return r.json()["display_code"]


def _test_employee_display_code(
    client: TestClient,
    headers: dict,
    *,
    id_number: str,
    employee_seq: int,
    sequence_code: str = "SYS1",
) -> str:
    employee_id = asyncio.run(_create_employee_same_seq(id_number, sequence_code, employee_seq))
    return _employee_display_code(client, headers, employee_id)


async def _get_sequence(code: str) -> EmployeeCodeSequence:
    async with _make_session()() as s:
        return (
            await s.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == code))
        ).scalar_one()


async def _get_nationality_id(code: str = "VN") -> int:
    async with _make_session()() as s:
        nationality = (
            await s.execute(select(Nationality).where(Nationality.code == code))
        ).scalar_one()
        return nationality.id


def _employee_payload(id_number: str, *, sequence_id: int, employee_seq: int, nationality_id: int) -> EmployeeCreate:
    return EmployeeCreate(
        employee_seq=employee_seq,
        employee_code_sequence_id=sequence_id,
        full_name=f"Import Seq {id_number}",
        last_name="Import",
        first_name=id_number,
        date_of_birth="1990-01-01",
        gender="male",
        nationality_id=nationality_id,
        id_number=id_number,
        id_issued_on="2020-01-01",
        id_issued_by="Cuc Canh sat",
        status="official",
        start_date="2026-01-01",
    )


async def _create_employee_same_seq(id_number: str, sequence_code: str, employee_seq: int) -> int:
    sequence = await _get_sequence(sequence_code)
    nationality_id = await _get_nationality_id()
    async with _make_session()() as s:
        employee = await employee_service.create_employee(
            s,
            _employee_payload(
                id_number,
                sequence_id=sequence.id,
                employee_seq=employee_seq,
                nationality_id=nationality_id,
            ),
        )
        await s.commit()
        await s.refresh(employee)
        return employee.id


async def _ensure_bhxh_foundation() -> None:
    async with _make_session()() as s:
        await bootstrap.seed_salary_scale(s)
        await s.commit()


async def _get_bhxh_group(code: str) -> BhxhPositionGroup:
    async with _make_session()() as s:
        row = await s.execute(select(BhxhPositionGroup).where(BhxhPositionGroup.code == code))
        group = row.scalar_one_or_none()
        assert group is not None
        return group


async def _get_contract_category_id(code: str) -> int:
    async with _make_session()() as s:
        row = await s.execute(select(ContractCategory).where(ContractCategory.code == code))
        category = row.scalar_one_or_none()
        assert category is not None
        return category.id


async def _get_profile_basis_state(employee_id: int) -> dict | None:
    async with _make_session()() as s:
        row = await s.execute(
            text(
                """
                SELECT insurance_basis_source, insurance_basis_amount, company_bhxh_joined_date
                FROM employee_insurance_profiles
                WHERE employee_id = :eid
                """
            ),
            {"eid": employee_id},
        )
        data = row.fetchone()
        if data is None:
            return None
        return {
            "source": data[0],
            "amount": data[1],
            "joined_date": data[2],
        }


def _create_computed_contract(
    client: TestClient,
    headers: dict,
    employee_id: int,
    *,
    group_code: str = "OFFICE_STAFF",
    grade_no: int = 3,
) -> None:
    asyncio.run(_ensure_bhxh_foundation())
    group = asyncio.run(_get_bhxh_group(group_code))
    category_id = asyncio.run(_get_contract_category_id("labor_indefinite"))
    resp = client.post(
        f"/api/v1/employees/{employee_id}/contracts",
        json={
            "contract_category_id": category_id,
            "contract_number": f"INS-{employee_id}-{group_code}",
            "signed_date": "2026-01-01",
            "effective_from": "2026-01-01",
            "effective_to": None,
            "insurance_salary_mode": "computed_by_position_group",
            "bhxh_position_group_id": group.id,
            "insurance_salary_grade_no": grade_no,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text


# ── Template ──────────────────────────────────────────────────────────────────

def test_template_download(client: TestClient):
    h = _login(client)
    r = client.get(f"{BASE}/insurance/template", headers=h)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = openpyxl.load_workbook(BytesIO(r.content))
    assert "Hướng dẫn" in wb.sheetnames


# ── Success ───────────────────────────────────────────────────────────────────

def test_import_success_create(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI101", employee_seq=9401)
    rows = [[code, "TESTIMP_001", "01/01/2026", "5000000", _BHYT_CODE, "active"]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 1
    assert body["success"] == 1
    assert body["failed"] == 0


def test_import_without_optional_fields(client: TestClient):
    """Không điền mã BHXH, bệnh viện, trạng thái → vẫn hợp lệ."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI102", employee_seq=9402)
    rows = [[code, "", "01/02/2026", "4500000", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["success"] == 1


def test_import_upsert_existing_profile(client: TestClient):
    """NV đã có hồ sơ BHXH → update (không báo lỗi, không tạo trùng)."""
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI103", employee_seq=9403)
    # Tạo lần 1
    rows1 = [[code, "TESTIMP_003", "01/01/2026", "5000000", "", "active"]]
    r1 = _upload(client, h, _make_xlsx(rows1))
    assert r1.json()["success"] == 1

    # Update lần 2 cùng employee → upsert thành công
    rows2 = [[code, "TESTIMP_003b", "01/02/2026", "6000000", _BHYT_CODE, "active"]]
    r2 = _upload(client, h, _make_xlsx(rows2))
    assert r2.status_code == 200
    body = r2.json()
    assert body["success"] == 1   # không fail
    assert body["failed"] == 0


# ── Errors ────────────────────────────────────────────────────────────────────

def test_import_invalid_employee(client: TestClient):
    h = _login(client)
    rows = [["999999", "TESTIMP_X", "01/01/2026", "5000000", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("Không tìm thấy nhân viên" in e["message"] for e in body["errors"])


def test_import_invalid_salary_zero(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI104", employee_seq=9404)
    rows = [[code, "TESTIMP_Z", "01/01/2026", "0", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("lương" in e["message"].lower() for e in body["errors"])


def test_import_invalid_salary_text(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI105", employee_seq=9405)
    rows = [[code, "TESTIMP_T", "01/01/2026", "abc", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_invalid_clinic_code(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI106", employee_seq=9406)
    rows = [[code, "TESTIMP_C", "01/01/2026", "5000000", "INVALID_CODE", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["failed"] == 1
    assert any("bệnh viện" in e["message"].lower() for e in body["errors"])


def test_import_invalid_status(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI107", employee_seq=9407)
    rows = [[code, "TESTIMP_S", "01/01/2026", "5000000", "", "unknown_status"]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_missing_required_salary(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI108", employee_seq=9408)
    rows = [[code, "TESTIMP_M", "01/01/2026", "", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_invalid_date(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI109", employee_seq=9409)
    rows = [[code, "TESTIMP_D", "not-a-date", "5000000", "", ""]]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    assert r.json()["failed"] == 1


def test_import_partial_errors(client: TestClient):
    h = _login(client)
    code = _test_employee_display_code(client, h, id_number="TESTIMPSEQI110", employee_seq=9410)
    rows = [
        [code, "TESTIMP_OK", "01/03/2026", "5000000", "", ""],   # OK
        ["999999",      "TESTIMP_BAD", "01/03/2026", "5000000", "", ""],  # fail
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200
    body = r.json()
    assert body["success"] == 1
    assert body["failed"] == 1


# ── Auth & validation ─────────────────────────────────────────────────────────

def test_unauthenticated(client: TestClient):
    r = client.post(f"{BASE}/insurance",
                    files={"file": ("f.xlsx", b"x", "application/vnd.ms-excel")})
    assert r.status_code == 401


def test_rejects_non_xlsx(client: TestClient):
    h = _login(client)
    r = client.post(f"{BASE}/insurance",
                    files={"file": ("data.csv", b"a,b", "text/csv")},
                    headers=h)
    assert r.status_code == 400


def test_empty_rows(client: TestClient):
    h = _login(client)
    r = _upload(client, h, _make_xlsx([]))
    assert r.status_code == 200
    assert r.json()["total"] == 0


def test_import_resolves_display_code_exactly_across_three_sequences(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQI001", "SYS1", 9103))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQI002", "SYS2", 9103))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQI003", "SYS3", 9103))

    rows = [
        ["9103", "SYS1", "TESTIMP_S1", "01/04/2026", "5100000", "", "active"],
        ["9103", "SYS2", "TESTIMP_S2", "01/04/2026", "5200000", "", "active"],
        ["9103", "SYS3", "TESTIMP_S3", "01/04/2026", "5300000", "", "active"],
    ]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_SEQUENCE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 3
    assert body["failed"] == 0


def test_import_rejects_ambiguous_seq_without_sequence_code(client: TestClient):
    h = _login(client)
    asyncio.run(_create_employee_same_seq("TESTIMPSEQI901", "SYS1", 9903))
    asyncio.run(_create_employee_same_seq("TESTIMPSEQI902", "SYS2", 9903))

    rows = [
        ["9903", "TESTIMP_AMB", "01/05/2026", "5500000", "", "active"],
    ]
    r = _upload(client, h, _make_xlsx(rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 0
    assert body["failed"] == 1
    assert any("Hệ mã nhân viên" in e["message"] for e in body["errors"])


def test_import_preserves_computed_profile_source_from_active_contract(client: TestClient):
    h = _login(client)
    employee_id = asyncio.run(_create_employee_same_seq("TESTIMPSEQI777", "SYS1", 9773))
    _create_computed_contract(client, h, employee_id)

    rows = [[
        "9773", "SYS1", "TESTIMP_COMPUTED", "01/02/2026", "", "", _BHYT_CODE, "active",
    ]]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_SOURCE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 1
    assert body["failed"] == 0

    profile = asyncio.run(_get_profile_basis_state(employee_id))
    assert profile is not None
    assert profile["source"] == "computed"
    assert str(profile["amount"]) == "5050800.00"


def test_import_rejects_computed_salary_mismatch_with_active_contract(client: TestClient):
    h = _login(client)
    employee_id = asyncio.run(_create_employee_same_seq("TESTIMPSEQI778", "SYS1", 9774))
    _create_computed_contract(client, h, employee_id)

    rows = [[
        "9774", "SYS1", "TESTIMP_COMPUTED_BAD", "01/02/2026", "5000000", "computed", _BHYT_CODE, "active",
    ]]
    r = _upload(client, h, _make_xlsx_with_headers(HEADERS_WITH_SOURCE, rows))
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["success"] == 0
    assert body["failed"] == 1
    assert any("không khớp với hợp đồng active" in e["message"] for e in body["errors"])
