from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import hash_sensitive
from app.models.employee import Employee, EmployeeAddress
from app.models.employee_code import EmployeeCodeSequence
from app.services.employee_import_service import IMPORT_COLUMNS, generate_template, process_import

TEST_ID_NUMBER_ADDR_NEW_ONLY = "IMPORTADDR001"
TEST_ID_NUMBER_ADDR_FULL = "IMPORTADDR002"
TEST_ID_NUMBER_ADDR_DETAIL_COMMA = "IMPORTADDR003"
TEST_ID_NUMBER_ADDR_MISSING_NEW = "IMPORTADDR004"
TEST_ID_NUMBER_ADDR_OLD_TOO_SHORT = "IMPORTADDR005"
TEST_ID_NUMBER_ADDR_UNKNOWN_NAME = "IMPORTADDR006"

VALID_NEW_ADDRESS = "thôn Vạc, Phường An Bình, Thành phố Cần Thơ"
VALID_OLD_ADDRESS = "thôn Vạc, Xã Thạnh Phú, Huyện Cờ Đỏ, Thành phố Cần Thơ"


def _make_engine_and_sessionmaker():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return engine, async_sessionmaker(engine, expire_on_commit=False)


def _make_xlsx(rows: list[list[str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_row(
    *,
    id_number: str,
    sequence_code: str,
    permanent_new: str = VALID_NEW_ADDRESS,
    permanent_old: str = "",
    contact_new: str = VALID_NEW_ADDRESS,
    contact_old: str = "",
) -> list[str]:
    values = {column: "" for column in IMPORT_COLUMNS}
    values.update(
        {
            "Họ và tên": "Import Address Test",
            "Họ": "Import",
            "Tên": "Address Test",
            "Ngày sinh": "01/01/1990",
            "Giới tính": "nam",
            "Số CCCD/CMND": id_number,
            "Ngày cấp CCCD": "01/01/2020",
            "Nơi cấp CCCD": "Cục Cảnh sát ĐKQLCƯ",
            "Trạng thái": "probation",
            "Ngày vào làm": date.today().strftime("%d/%m/%Y"),
            "Hệ mã nhân viên": sequence_code,
            "Địa chỉ thường trú (Hệ mới 2 cấp)": permanent_new,
            "Địa chỉ thường trú (Hệ cũ 3 cấp)": permanent_old,
            "Địa chỉ liên lạc (Hệ mới 2 cấp)": contact_new,
            "Địa chỉ liên lạc (Hệ cũ 3 cấp)": contact_old,
        }
    )
    return [values[column] for column in IMPORT_COLUMNS]


async def _cleanup(session: AsyncSession) -> None:
    for test_id_number in (
        TEST_ID_NUMBER_ADDR_NEW_ONLY,
        TEST_ID_NUMBER_ADDR_FULL,
        TEST_ID_NUMBER_ADDR_DETAIL_COMMA,
        TEST_ID_NUMBER_ADDR_MISSING_NEW,
        TEST_ID_NUMBER_ADDR_OLD_TOO_SHORT,
        TEST_ID_NUMBER_ADDR_UNKNOWN_NAME,
    ):
        employee = (
            await session.execute(
                select(Employee).where(Employee.id_number_hash == hash_sensitive(test_id_number))
            )
        ).scalar_one_or_none()
        if employee:
            await session.execute(delete(Employee).where(Employee.id == employee.id))
    await session.commit()


@pytest.mark.asyncio
async def test_import_address_new_system_only():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(id_number=TEST_ID_NUMBER_ADDR_NEW_ONLY, sequence_code=sequence.code),
            ])

            result = await process_import(session, workbook)
            assert result.success == 1, result.errors

            addresses = (
                await session.execute(
                    select(EmployeeAddress).where(EmployeeAddress.employee_id == result.created_ids[0])
                )
            ).scalars().all()
            by_type = {a.address_type: a for a in addresses}
            assert set(by_type) == {"permanent", "contact"}
            for addr in by_type.values():
                assert addr.new_province_unit_id is not None
                assert addr.new_ward_unit_id is not None
                assert addr.old_province_unit_id is None
                assert addr.old_district_unit_id is None
                assert addr.old_ward_unit_id is None

            await _cleanup(session)
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_import_address_new_and_old_system():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(
                    id_number=TEST_ID_NUMBER_ADDR_FULL,
                    sequence_code=sequence.code,
                    permanent_old=VALID_OLD_ADDRESS,
                ),
            ])

            result = await process_import(session, workbook)
            assert result.success == 1, result.errors

            permanent = (
                await session.execute(
                    select(EmployeeAddress).where(
                        EmployeeAddress.employee_id == result.created_ids[0],
                        EmployeeAddress.address_type == "permanent",
                    )
                )
            ).scalar_one()
            assert permanent.new_province_unit_id is not None
            assert permanent.new_ward_unit_id is not None
            assert permanent.old_province_unit_id is not None
            assert permanent.old_district_unit_id is not None
            assert permanent.old_ward_unit_id is not None
            assert permanent.old_address_line == "thôn Vạc"

            await _cleanup(session)
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_import_address_detail_with_comma_is_preserved():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            detail_new = "Số 12, ngõ 34, Phường An Bình, Thành phố Cần Thơ"
            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(
                    id_number=TEST_ID_NUMBER_ADDR_DETAIL_COMMA,
                    sequence_code=sequence.code,
                    permanent_new=detail_new,
                    contact_new=detail_new,
                ),
            ])

            result = await process_import(session, workbook)
            assert result.success == 1, result.errors

            permanent = (
                await session.execute(
                    select(EmployeeAddress).where(
                        EmployeeAddress.employee_id == result.created_ids[0],
                        EmployeeAddress.address_type == "permanent",
                    )
                )
            ).scalar_one()
            assert permanent.new_address_line == "Số 12, ngõ 34"

            await _cleanup(session)
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_import_address_missing_required_new_system_fails():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(
                    id_number=TEST_ID_NUMBER_ADDR_MISSING_NEW,
                    sequence_code=sequence.code,
                    permanent_new="",
                ),
            ])

            result = await process_import(session, workbook)
            assert result.success == 0
            assert result.failed == 1
            assert any(
                error.column == "Địa chỉ thường trú (Hệ mới 2 cấp)" and "bắt buộc" in error.message
                for error in result.errors
            )
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_import_address_old_system_too_few_parts_fails():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(
                    id_number=TEST_ID_NUMBER_ADDR_OLD_TOO_SHORT,
                    sequence_code=sequence.code,
                    permanent_old="Xã Thạnh Phú, Thành phố Cần Thơ",
                ),
            ])

            result = await process_import(session, workbook)
            assert result.success == 0
            assert result.failed == 1
            assert any(
                error.column == "Địa chỉ thường trú (Hệ cũ 3 cấp)" and "tối thiểu 3 phần" in error.message
                for error in result.errors
            )
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_import_address_unknown_name_fails_with_clear_message():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup(session)
            sequence = (
                await session.execute(select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1"))
            ).scalar_one()

            bogus = "thôn Vạc, Phường Không Tồn Tại Xyz, Thành phố Cần Thơ"
            workbook = _make_xlsx([
                IMPORT_COLUMNS,
                _make_row(
                    id_number=TEST_ID_NUMBER_ADDR_UNKNOWN_NAME,
                    sequence_code=sequence.code,
                    permanent_new=bogus,
                ),
            ])

            result = await process_import(session, workbook)
            assert result.success == 0
            assert result.failed == 1
            assert any(
                error.column == "Địa chỉ thường trú (Hệ mới 2 cấp)"
                and "Không tìm thấy xã/phường" in error.message
                and bogus in error.message
                for error in result.errors
            )
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_generate_template_includes_address_reference_sheets():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            content = await generate_template(session)
    finally:
        await engine.dispose()

    workbook = openpyxl.load_workbook(io.BytesIO(content))
    assert "DM Hệ mới" in workbook.sheetnames
    assert "DM Hệ cũ" in workbook.sheetnames

    new_sheet = workbook["DM Hệ mới"]
    old_sheet = workbook["DM Hệ cũ"]
    assert new_sheet.max_row > 1
    assert old_sheet.max_row > 1
    assert [c.value for c in new_sheet[1]] == ["Tỉnh/Thành phố", "Xã/Phường"]
    assert [c.value for c in old_sheet[1]] == ["Tỉnh/Thành phố", "Huyện/Quận", "Xã/Phường"]
