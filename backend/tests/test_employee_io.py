"""Tests cho Import/Export nhân viên (3.7)."""

import io
from datetime import date, timedelta

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import hash_sensitive
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.models.employee_insurance import EmployeeInsuranceProfile
from app.models.employee_job import EmployeeJobRecord
from app.models.org import JobPosition
from app.seeds import employees as employees_seed
from app.services.employee_import_service import IMPORT_COLUMNS, generate_template

BASE = "/api/v1/employees"

_ADMIN_EMAIL        = "admin@hrms.local"
_ADMIN_PASSWORD     = "Hrms@2026"
_LINE_MANAGER_EMAIL = "linemanager@hrms.local"
_PREFIX             = "TESTIO"


def _login(client: TestClient, email: str = _ADMIN_EMAIL, password: str = _ADMIN_PASSWORD) -> dict:
    token = client.post("/api/v1/auth/login", json={"email": email, "password": password}).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _admin(client: TestClient) -> dict:  return _login(client)
def _viewer(client: TestClient) -> dict: return _login(client, _LINE_MANAGER_EMAIL)


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


async def _cleanup():
    async with _make_session()() as s:
        hashes = [hash_sensitive(f"{_PREFIX}{suffix}") for suffix in (
            "0001", "0002", "0003", "0004", "0005", "0006", "0007", "0008", "0009",
            "0010", "0011", "0012", "0013", "0014", "0014BH", "0015", "EXP001", "EXPBH001",
        )]
        employee_ids = list((await s.execute(select(Employee.id).where(Employee.id_number_hash.in_(hashes)))).scalars().all())
        if employee_ids:
            await s.execute(text("DELETE FROM employee_job_records WHERE employee_id = ANY(:employee_ids)"), {"employee_ids": employee_ids})
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))
        await s.commit()


@pytest.fixture(scope="session", autouse=True)
async def seed_data():
    async with _make_session()() as session:
        await employees_seed.seed_sample_employees(session)
        await session.commit()
    yield


@pytest.fixture(autouse=True)
async def cleanup():
    yield
    await _cleanup()


# ── Helpers ────────────────────────────────────────────────────────────────────

def _get_dept_code(client: TestClient, headers: dict) -> str:
    depts = client.get("/api/v1/departments", headers=headers).json()
    items = depts["items"] if isinstance(depts, dict) else depts
    return items[0]["code"]


def _make_xlsx(rows: list[list]) -> bytes:
    """Tạo file xlsx từ danh sách rows (row 0 = header)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _valid_row(suffix: str, dept_code: str = "", sequence_code: str = "SYS1") -> list:
    today = date.today()
    return [
        f"Test IO {suffix}", "Test", f"IO {suffix}",
        "01/01/1990", "nam", f"{_PREFIX}{suffix}",
        "01/01/2020", "Cục Cảnh sát ĐKQLCƯ",
        "probation", today.strftime("%d/%m/%Y"),
        "0901234567", f"testio{suffix}@email.com",
        "", "",
        "", "", "",
        dept_code, "", "",
        sequence_code,
        "", "",
    ]


def _upload(client: TestClient, headers: dict, xlsx_bytes: bytes) -> dict:
    resp = client.post(
        f"{BASE}/import",
        files={"file": ("import.xlsx", io.BytesIO(xlsx_bytes), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    return resp


async def _get_employee_state(id_number: str) -> tuple[Employee, EmployeeJobRecord | None]:
    async with _make_session()() as s:
        employee = (
            await s.execute(select(Employee).where(Employee.id_number_hash == hash_sensitive(id_number)))
        ).scalar_one()
        job_record = (
            await s.execute(
                select(EmployeeJobRecord).where(
                    EmployeeJobRecord.employee_id == employee.id,
                    EmployeeJobRecord.is_current == True,
                )
            )
        ).scalar_one_or_none()
        return employee, job_record


async def _get_sequence_code(sequence_id: int | None) -> str | None:
    if sequence_id is None:
        return None
    async with _make_session()() as s:
        sequence = await s.get(EmployeeCodeSequence, sequence_id)
        return sequence.code if sequence else None


async def _get_employee_profile_by_id_number(id_number: str) -> EmployeeInsuranceProfile | None:
    async with _make_session()() as s:
        employee = (
            await s.execute(select(Employee).where(Employee.id_number_hash == hash_sensitive(id_number)))
        ).scalar_one_or_none()
        if not employee:
            return None
        return (
            await s.execute(
                select(EmployeeInsuranceProfile).where(
                    EmployeeInsuranceProfile.employee_id == employee.id
                )
            )
        ).scalar_one_or_none()


# ── Template ───────────────────────────────────────────────────────────────────

def test_download_template_200(client: TestClient):
    headers = _admin(client)
    resp = client.get(f"{BASE}/import/template", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    # Kiểm tra file parse được và có đúng header
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.worksheets[0]
    headers_in_file = [ws.cell(1, c).value for c in range(1, len(IMPORT_COLUMNS) + 1)]
    assert headers_in_file == IMPORT_COLUMNS


def test_download_template_viewer_403(client: TestClient):
    assert client.get(f"{BASE}/import/template", headers=_viewer(client)).status_code == 403


def test_download_template_unauth_401(client: TestClient):
    assert client.get(f"{BASE}/import/template").status_code == 401


# ── Import ─────────────────────────────────────────────────────────────────────

def test_import_success(client: TestClient):
    headers = _admin(client)
    dept = _get_dept_code(client, headers)
    xlsx = _make_xlsx([IMPORT_COLUMNS, _valid_row("0001", dept)])
    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"]   == 1
    assert data["success"] == 1
    assert data["failed"]  == 0
    assert len(data["errors"]) == 0
    assert len(data["created_ids"]) == 1


def test_import_success_without_dept_with_sequence_override(client: TestClient):
    """Không có phòng ban nhưng có hệ mã explicit thì vẫn import được."""
    headers = _admin(client)
    row = _valid_row("0002", "", "SYS2")
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["success"] == 1


def test_import_partial_errors(client: TestClient):
    """Dòng hợp lệ được tạo, dòng lỗi bị skip — không cancel cả batch."""
    headers = _admin(client)
    bad_row = _valid_row("0003")
    bad_row[4] = "xyz"          # giới tính sai
    good_row = _valid_row("0004")
    xlsx = _make_xlsx([IMPORT_COLUMNS, bad_row, good_row])
    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["total"]   == 2
    assert data["success"] == 1
    assert data["failed"]  == 1
    assert any(e["column"] == "Giới tính" for e in data["errors"])


def test_import_duplicate_id_number(client: TestClient):
    """id_number đã tồn tại → dòng đó bị failed."""
    headers = _admin(client)
    xlsx = _make_xlsx([IMPORT_COLUMNS, _valid_row("0005")])
    _upload(client, headers, xlsx)          # tạo lần đầu
    resp = _upload(client, headers, xlsx)   # upload lại cùng id_number
    data = resp.json()
    assert data["failed"] == 1
    assert any("CCCD" in e["column"] for e in data["errors"])


def test_import_missing_required_field(client: TestClient):
    headers = _admin(client)
    row = _valid_row("0006")
    row[0] = ""   # Họ và tên rỗng
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Họ và tên" for e in data["errors"])


def test_import_invalid_date_format(client: TestClient):
    headers = _admin(client)
    row = _valid_row("0007")
    row[3] = "1990-99-99"   # ngày sinh sai
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Ngày sinh" for e in data["errors"])


def test_import_invalid_gender(client: TestClient):
    headers = _admin(client)
    row = _valid_row("0008")
    row[4] = "unknown"
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Giới tính" for e in data["errors"])


def test_import_invalid_status(client: TestClient):
    headers = _admin(client)
    row = _valid_row("0009")
    row[8] = "invalid_status"
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Trạng thái" for e in data["errors"])


def test_import_unknown_department_warns(client: TestClient):
    """Phòng ban không tồn tại và không có fallback → dòng bị fail."""
    headers = _admin(client)
    row = _valid_row("0010", "DEPT_DOES_NOT_EXIST", "")
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Phòng ban" for e in data["errors"])


@pytest.mark.asyncio
async def test_import_department_creates_current_job_record(client: TestClient):
    headers = _admin(client)
    dept = _get_dept_code(client, headers)
    employee_id_number = f"{_PREFIX}0012"
    xlsx = _make_xlsx([IMPORT_COLUMNS, _valid_row("0012", dept)])

    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["success"] == 1

    employee, job_record = await _get_employee_state(employee_id_number)
    assert employee is not None
    assert job_record is not None
    assert job_record.effective_from == employee.start_date


def test_import_with_position_and_sequence_override(client: TestClient):
    headers = _admin(client)
    dept_code = _get_dept_code(client, headers)
    positions = client.get("/api/v1/job-positions", params={"is_active": True}, headers=headers).json()
    position = next((item for item in positions if item["department_name"] and item["department_id"]), None)
    assert position is not None
    dept_items = client.get("/api/v1/departments", headers=headers).json()
    depts = dept_items["items"] if isinstance(dept_items, dict) else dept_items
    position_dept = next(d for d in depts if d["id"] == position["department_id"])

    row = _valid_row("0013", position_dept["code"])
    row[16] = position["name"]
    row[17] = "SYS2"
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])

    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["success"] == 1


@pytest.mark.asyncio
async def test_import_with_position_and_sequence_override_persists_context(client: TestClient):
    headers = _admin(client)
    positions = client.get("/api/v1/job-positions", params={"is_active": True}, headers=headers).json()
    position = positions[0]
    dept_items = client.get("/api/v1/departments", headers=headers).json()
    depts = dept_items["items"] if isinstance(dept_items, dict) else dept_items
    position_dept = next(d for d in depts if d["id"] == position["department_id"])
    employee_id_number = f"{_PREFIX}0014"

    row = _valid_row("0014", position_dept["code"])
    row[16] = position["name"]
    row[17] = "SYS2"
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])
    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200

    employee, job_record = await _get_employee_state(employee_id_number)
    assert job_record is not None
    assert job_record.job_position_id == position["id"]
    assert await _get_sequence_code(employee.employee_code_sequence_id) == "SYS2"


@pytest.mark.asyncio
async def test_import_bhxh_code_syncs_employee_insurance_profile(client: TestClient):
    headers = _admin(client)
    dept = _get_dept_code(client, headers)
    row = _valid_row("0014BH", dept)
    row[13] = "  BHXH-IMPORT-001  "
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])

    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200, resp.text
    assert resp.json()["success"] == 1

    profile = await _get_employee_profile_by_id_number(f"{_PREFIX}0014BH")
    assert profile is not None
    assert profile.bhxh_code == "BHXH-IMPORT-001"


def test_import_invalid_sequence_override_fails_row(client: TestClient):
    headers = _admin(client)
    dept = _get_dept_code(client, headers)
    row = _valid_row("0015", dept)
    row[17] = "UNKNOWN_SYS"
    xlsx = _make_xlsx([IMPORT_COLUMNS, row])

    resp = _upload(client, headers, xlsx)
    assert resp.status_code == 200
    data = resp.json()
    assert data["failed"] == 1
    assert any(e["column"] == "Hệ mã nhân viên" for e in data["errors"])


def test_import_blank_rows_ignored(client: TestClient):
    """Dòng trống trong file bị bỏ qua."""
    headers = _admin(client)
    xlsx = _make_xlsx([IMPORT_COLUMNS, [""] * len(IMPORT_COLUMNS), _valid_row("0011")])
    resp = _upload(client, headers, xlsx)
    data = resp.json()
    assert data["total"]   == 1
    assert data["success"] == 1


def test_import_non_xlsx_400(client: TestClient):
    headers = _admin(client)
    resp = client.post(
        f"{BASE}/import",
        files={"file": ("data.csv", io.BytesIO(b"a,b,c"), "text/csv")},
        headers=headers,
    )
    assert resp.status_code == 400


def test_import_viewer_403(client: TestClient):
    xlsx = _make_xlsx([IMPORT_COLUMNS])
    resp = _upload(client, _viewer(client), xlsx)
    assert resp.status_code == 403


def test_import_audit_log(client: TestClient):
    headers = _admin(client)
    xlsx = _make_xlsx([IMPORT_COLUMNS, _valid_row("0012")])
    _upload(client, headers, xlsx)
    resp = client.get("/api/v1/audit-logs", params={"entity_type": "employee"}, headers=headers).json()
    logs = resp["items"] if isinstance(resp, dict) else resp
    actions = [l["action"] for l in logs]
    assert "IMPORT_EMPLOYEES" in actions


# ── Export list ────────────────────────────────────────────────────────────────

def test_export_list_200(client: TestClient):
    headers = _admin(client)
    resp = client.get(f"{BASE}/export", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws.max_row >= 1   # ít nhất có header


def test_export_list_has_data_rows(client: TestClient):
    headers = _admin(client)
    resp = client.get(f"{BASE}/export", headers=headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    # Seed data có nhân viên → phải có > 1 dòng
    assert ws.max_row > 1


def test_export_list_filter_status(client: TestClient):
    headers = _admin(client)
    # Tạo 1 nhân viên official
    client.post(BASE, json={
        "full_name": "Export Official", "last_name": "Export", "first_name": "Official",
        "date_of_birth": "1990-01-01", "gender": "male", "nationality_id": 1,
        "id_number": f"{_PREFIX}EXP001", "id_issued_on": "2020-01-01",
        "id_issued_by": "Cục", "status": "official", "start_date": "2020-01-01",
    }, headers=headers)

    resp_all      = client.get(f"{BASE}/export", headers=headers)
    resp_official = client.get(f"{BASE}/export", params={"status": "official"}, headers=headers)

    wb_all      = openpyxl.load_workbook(io.BytesIO(resp_all.content))
    wb_official = openpyxl.load_workbook(io.BytesIO(resp_official.content))
    assert wb_official.active.max_row <= wb_all.active.max_row


def test_export_list_viewer_200(client: TestClient):
    assert client.get(f"{BASE}/export", headers=_viewer(client)).status_code == 200


def test_export_list_unauth_401(client: TestClient):
    assert client.get(f"{BASE}/export").status_code == 401


# ── Export profile ─────────────────────────────────────────────────────────────

def test_export_profile_200(client: TestClient):
    headers = _admin(client)
    # Lấy employee đầu tiên từ seed
    emp_id = client.get(BASE, headers=headers).json()["items"][0]["id"]
    resp = client.get(f"{BASE}/{emp_id}/export", headers=headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]


def test_export_profile_has_sheets(client: TestClient):
    headers = _admin(client)
    emp_id = client.get(BASE, headers=headers).json()["items"][0]["id"]
    resp = client.get(f"{BASE}/{emp_id}/export", headers=headers)
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    sheet_names = wb.sheetnames
    assert "Thông tin cá nhân"   in sheet_names
    assert "Công việc"           in sheet_names
    assert "Học vấn"             in sheet_names
    assert "Người thân"          in sheet_names
    assert "Tài khoản ngân hàng" in sheet_names


def test_export_profile_404(client: TestClient):
    assert client.get(f"{BASE}/99999999/export", headers=_admin(client)).status_code == 404


def test_export_profile_viewer_200(client: TestClient):
    headers_admin  = _admin(client)
    headers_viewer = _viewer(client)
    emp_id = client.get(BASE, headers=headers_admin).json()["items"][0]["id"]
    assert client.get(f"{BASE}/{emp_id}/export", headers=headers_viewer).status_code == 200


@pytest.mark.asyncio
async def test_export_profile_reads_bhxh_code_from_insurance_profile(client: TestClient):
    headers = _admin(client)
    create_resp = client.post(
        BASE,
        json={
            "employee_code_sequence_id": 1,
            "full_name": "Export BHXH Profile",
            "last_name": "Export",
            "first_name": "BHXH Profile",
            "date_of_birth": "1990-01-01",
            "gender": "male",
            "nationality_id": 1,
            "id_number": f"{_PREFIX}EXPBH001",
            "id_issued_on": "2020-01-01",
            "id_issued_by": "Cục",
            "status": "official",
            "start_date": "2020-01-01",
            "bhxh_code": "BHXH-OLD",
        },
        headers=headers,
    )
    assert create_resp.status_code == 201, create_resp.text
    employee_id = create_resp.json()["id"]

    async with _make_session()() as s:
        employee = (
            await s.execute(select(Employee).where(Employee.id == employee_id))
        ).scalar_one()
        profile = (
            await s.execute(
                select(EmployeeInsuranceProfile).where(
                    EmployeeInsuranceProfile.employee_id == employee_id
                )
            )
        ).scalar_one()
        employee.bhxh_code = None
        profile.bhxh_code = "BHXH-PROFILE-SOURCE"
        await s.commit()

    resp = client.get(f"{BASE}/{employee_id}/export", headers=headers)
    assert resp.status_code == 200, resp.text
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb["Thông tin cá nhân"]
    rows = {
        ws.cell(row=row_idx, column=1).value: ws.cell(row=row_idx, column=2).value
        for row_idx in range(1, ws.max_row + 1)
    }
    assert rows["Số BHXH"] == "BHXH-PROFILE-SOURCE"
