from __future__ import annotations

from io import BytesIO

import openpyxl

from app.services import (
    contract_import_service,
    department_import_service,
    employee_import_service,
    insurance_import_service,
    job_position_import_service,
    job_title_import_service,
    leave_record_import_service,
)


def _font_rgb(cell) -> str | None:
    color = cell.font.color
    if color is None:
        return None
    if color.type == "rgb":
        return color.rgb
    return None


def _fill_rgb(cell) -> str | None:
    fill = cell.fill.fgColor
    if fill is None:
        return None
    if fill.type == "rgb":
        return fill.rgb
    return None


def _assert_header_colors(template_bytes: bytes, *, required_col: int, optional_col: int) -> None:
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb.worksheets[0]
    required_cell = ws.cell(row=1, column=required_col)
    optional_cell = ws.cell(row=1, column=optional_col)

    assert _fill_rgb(required_cell) == "001F4E79"
    assert _font_rgb(required_cell) == "00FFFFFF"
    assert _fill_rgb(optional_cell) == "00D6E4F0"
    assert _font_rgb(optional_cell) == "001F2937"


def test_department_template_optional_headers_use_dark_text():
    _assert_header_colors(
        department_import_service.generate_template(),
        required_col=1,
        optional_col=3,
    )


def test_job_title_template_optional_headers_use_dark_text():
    _assert_header_colors(
        job_title_import_service.generate_template(),
        required_col=1,
        optional_col=3,
    )


def test_job_position_template_optional_headers_use_dark_text():
    _assert_header_colors(
        job_position_import_service.generate_template(),
        required_col=1,
        optional_col=4,
    )


def test_contract_template_optional_headers_use_dark_text():
    _assert_header_colors(
        contract_import_service.generate_template(),
        required_col=1,
        optional_col=8,
    )


def test_leave_record_template_optional_headers_use_dark_text():
    _assert_header_colors(
        leave_record_import_service.generate_template(),
        required_col=1,
        optional_col=2,
    )


def test_insurance_template_optional_headers_use_dark_text():
    _assert_header_colors(
        insurance_import_service.generate_template(),
        required_col=1,
        optional_col=2,
    )


def test_employee_template_optional_headers_use_dark_text():
    _assert_header_colors(
        employee_import_service.generate_template(),
        required_col=1,
        optional_col=11,
    )
