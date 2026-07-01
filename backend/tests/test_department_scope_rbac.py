from __future__ import annotations

import json
from datetime import date
import io
import uuid

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.models.auth import User, UserRole
from app.models.employee import Employee
from app.models.org import Department

BASE_EMP = "/api/v1/employees"
BASE_DEPT = "/api/v1/departments"
BASE_USERS = "/api/v1/users"
BASE_ROLES = "/api/v1/roles"
BASE_LEAVE_ENT = "/api/v1/leave-entitlements"
BASE_LEAVE_REC = "/api/v1/leave-records"
BASE_KPI = "/api/v1/performance/kpi"
BASE_REVIEWS = "/api/v1/performance/yearly-reviews"
BASE_DASHBOARD = "/api/v1/reports/dashboard"
BASE_HR_REPORTS = "/api/v1/reports/hr"
BASE_JOB_TITLES = "/api/v1/job-titles"
BASE_JOB_POSITIONS = "/api/v1/job-positions"
BASE_ORG_HISTORY = "/api/v1/org-history"
BASE_CONTRACTS = "/api/v1/contracts"
BASE_CONTRACT_REPORTS = "/api/v1/reports/contracts"
BASE_REWARDS = "/api/v1/rewards"
BASE_DISCIPLINES = "/api/v1/disciplines"

_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"
_PREFIX = f"SCRB{uuid.uuid4().hex[:6].upper()}"


def _login(client: TestClient, email: str = _ADMIN_EMAIL, password: str = _ADMIN_PASSWORD) -> dict[str, str]:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": email, "password": password},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


@pytest.fixture(autouse=True)
async def cleanup():
    yield
    async with _make_session()() as s:
        employee_ids = list(
            (
                await s.execute(
                    select(Employee.id).where(
                        Employee.personal_email.like(f"{_PREFIX.lower()}%@example.com")
                    )
                )
            )
            .scalars()
            .all()
        )
        if employee_ids:
            await s.execute(
                text("DELETE FROM employee_rewards WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_disciplines WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_contracts WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM leave_records WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM leave_entitlements WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_kpi_monthly WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_yearly_reviews WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(
                text("DELETE FROM employee_job_records WHERE employee_id = ANY(:employee_ids)"),
                {"employee_ids": employee_ids},
            )
            await s.execute(delete(Employee).where(Employee.id.in_(employee_ids)))

        user_ids = list(
            (
                await s.execute(
                    select(User.id).where(User.email.like(f"{_PREFIX.lower()}%@example.com"))
                )
            )
            .scalars()
            .all()
        )
        if user_ids:
            await s.execute(
                delete(UserRole).where(UserRole.user_id.in_(user_ids))
            )
            await s.execute(delete(User).where(User.id.in_(user_ids)))

        dept_ids = list(
            (
                await s.execute(
                    select(Department.id).where(Department.code.like(f"{_PREFIX}%"))
                )
            )
            .scalars()
            .all()
        )
        if dept_ids:
            await s.execute(
                text("DELETE FROM department_job_positions WHERE department_id = ANY(:dept_ids)"),
                {"dept_ids": dept_ids},
            )

        position_ids = list(
            (
                await s.execute(
                    text("SELECT id FROM job_positions WHERE code LIKE :prefix"),
                    {"prefix": f"{_PREFIX}%"},
                )
            )
            .scalars()
            .all()
        )
        if position_ids:
            await s.execute(
                text("DELETE FROM department_job_positions WHERE job_position_id = ANY(:position_ids)"),
                {"position_ids": position_ids},
            )
            await s.execute(
                text("DELETE FROM job_position_attachments WHERE job_position_id = ANY(:position_ids)"),
                {"position_ids": position_ids},
            )
            await s.execute(
                text("DELETE FROM job_positions WHERE id = ANY(:position_ids)"),
                {"position_ids": position_ids},
            )

        if dept_ids:
            await s.execute(
                text("DELETE FROM employee_job_records WHERE department_id = ANY(:dept_ids)"),
                {"dept_ids": dept_ids},
            )
            await s.execute(delete(Department).where(Department.id.in_(dept_ids)))

        await s.commit()


def _create_department(client: TestClient, headers: dict[str, str], *, code: str, name: str, parent_id: int | None = None) -> dict:
    payload = {"code": code, "name": name, "dept_type": "PHONG"}
    if parent_id is not None:
        payload["parent_id"] = parent_id
        payload["dept_type"] = "BO_PHAN"
    resp = client.post(BASE_DEPT, json=payload, headers=headers)
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_employee(
    client: TestClient,
    headers: dict[str, str],
    *,
    suffix: str,
    full_name: str,
    department_id: int,
) -> dict:
    resp = client.post(
        BASE_EMP,
        json={
            "employee_code_sequence_id": 1,
            "full_name": full_name,
            "last_name": full_name.split()[0],
            "first_name": full_name.split()[-1],
            "date_of_birth": "1990-01-01",
            "gender": "male",
            "nationality_id": 1,
            "id_number": f"{_PREFIX}{suffix}",
            "id_issued_on": "2020-01-01",
            "id_issued_by": "CA Test",
            "status": "official",
            "start_date": "2026-01-01",
            "personal_email": f"{_PREFIX.lower()}{suffix}@example.com",
            "initial_department_id": department_id,
            "initial_job_effective_from": "2026-01-01",
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_scope_fixture(
    client: TestClient,
    headers: dict[str, str],
    *,
    code_suffix: str,
    visible_employee_suffix: str,
    hidden_employee_suffix: str,
    visible_name: str,
    hidden_name: str,
) -> dict:
    root = _create_department(client, headers, code=f"{_PREFIX}_{code_suffix}R", name=f"{_PREFIX} {code_suffix} Root")
    child = _create_department(
        client,
        headers,
        code=f"{_PREFIX}_{code_suffix}C",
        name=f"{_PREFIX} {code_suffix} Child",
        parent_id=root["id"],
    )
    outside = _create_department(
        client,
        headers,
        code=f"{_PREFIX}_{code_suffix}O",
        name=f"{_PREFIX} {code_suffix} Outside",
    )
    visible_emp = _create_employee(
        client,
        headers,
        suffix=visible_employee_suffix,
        full_name=visible_name,
        department_id=child["id"],
    )
    hidden_emp = _create_employee(
        client,
        headers,
        suffix=hidden_employee_suffix,
        full_name=hidden_name,
        department_id=outside["id"],
    )
    return {
        "root": root,
        "child": child,
        "outside": outside,
        "visible_emp": visible_emp,
        "hidden_emp": hidden_emp,
    }


def _create_scoped_line_manager(
    client: TestClient,
    admin_headers: dict[str, str],
    *,
    department_ids: list[int],
) -> dict[str, str]:
    return _create_scoped_role_user(
        client,
        admin_headers,
        role_code="line_manager",
        department_ids=department_ids,
        email=f"{_PREFIX.lower()}lm@example.com",
        full_name=f"{_PREFIX} Line Manager",
        password="LineMgr@1234",
    )


def _create_scoped_role_user(
    client: TestClient,
    admin_headers: dict[str, str],
    *,
    role_code: str,
    department_ids: list[int],
    email: str,
    full_name: str,
    password: str,
) -> dict[str, str]:
    user_resp = client.post(
        BASE_USERS,
        json={
            "email": email,
            "full_name": full_name,
            "password": password,
        },
        headers=admin_headers,
    )
    assert user_resp.status_code == 201, user_resp.text
    user = user_resp.json()

    roles_resp = client.get(BASE_ROLES, headers=admin_headers)
    assert roles_resp.status_code == 200, roles_resp.text
    target_role = next(role for role in roles_resp.json() if role["code"] == role_code)

    assign_resp = client.post(
        f"{BASE_USERS}/{user['id']}/roles",
        json={
            "role_id": target_role["id"],
            "scope_type": "department",
            "department_ids": department_ids,
        },
        headers=admin_headers,
    )
    assert assign_resp.status_code == 201, assign_resp.text
    return _login(client, email, password)


def _get_first_contract_category_id() -> int:
    import asyncio

    async def _fetch() -> int:
        async with _make_session()() as s:
            result = await s.execute(text("SELECT id FROM contract_categories ORDER BY id ASC LIMIT 1"))
            category_id = result.scalar_one_or_none()
            assert category_id is not None, "Không tìm thấy contract category"
            return int(category_id)

    return asyncio.run(_fetch())


def _get_first_reward_type_id() -> int:
    import asyncio

    async def _fetch() -> int:
        async with _make_session()() as s:
            result = await s.execute(
                text(
                    "SELECT id FROM reward_types "
                    "WHERE is_active = TRUE AND is_monetary = FALSE "
                    "ORDER BY id ASC LIMIT 1"
                )
            )
            reward_type_id = result.scalar_one_or_none()
            assert reward_type_id is not None, "Không tìm thấy reward type không tiền"
            return int(reward_type_id)

    return asyncio.run(_fetch())


def _create_contract(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    contract_number: str,
    effective_to: str | None = None,
) -> dict:
    resp = client.post(
        f"{BASE_EMP}/{employee_id}/contracts",
        json={
            "contract_category_id": _get_first_contract_category_id(),
            "contract_number": contract_number,
            "signed_date": "2026-01-01",
            "effective_from": "2026-01-01",
            "effective_to": effective_to,
            "insurance_salary_mode": "fixed_manual",
            "insurance_salary_fixed_amount": "5000000",
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_leave_entitlement(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    year: int,
    allocated_days: float = 12.0,
) -> dict:
    resp = client.post(
        BASE_LEAVE_ENT,
        json={
            "employee_id": employee_id,
            "leave_type_id": 1,
            "year": year,
            "allocated_days": allocated_days,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_leave_record(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    year: int,
) -> dict:
    resp = client.post(
        BASE_LEAVE_REC,
        json={
            "employee_id": employee_id,
            "leave_type_id": 1,
            "start_date": str(date(year, 6, 1)),
            "end_date": str(date(year, 6, 1)),
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_job_position(
    client: TestClient,
    headers: dict[str, str],
    *,
    code: str,
    name: str,
    department_id: int,
    assigned_department_ids: list[int] | None = None,
) -> dict:
    resp = client.post(
        BASE_JOB_POSITIONS,
        json={
            "code": code,
            "name": name,
            "department_id": department_id,
            "assigned_department_ids": assigned_department_ids or [department_id],
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_kpi(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    year: int,
    month: int,
    score: float,
) -> dict:
    resp = client.post(
        BASE_KPI,
        json={
            "employee_id": employee_id,
            "year": year,
            "month": month,
            "score": score,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_yearly_review(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    year: int,
    rating: str = "tot",
) -> dict:
    resp = client.post(
        BASE_REVIEWS,
        json={
            "employee_id": employee_id,
            "year": year,
            "rating": rating,
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_reward(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    title: str,
    reward_date: str,
) -> dict:
    import json

    resp = client.post(
        BASE_REWARDS,
        data={
            "body": json.dumps(
                {
                    "employee_id": employee_id,
                    "reward_type_id": _get_first_reward_type_id(),
                    "reward_date": reward_date,
                    "title": title,
                }
            )
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def _create_discipline(
    client: TestClient,
    headers: dict[str, str],
    *,
    employee_id: int,
    title: str,
    effective_date: str,
) -> dict:
    import json

    resp = client.post(
        BASE_DISCIPLINES,
        data={
            "body": json.dumps(
                {
                    "employee_id": employee_id,
                    "discipline_form": "khien_trach",
                    "violation_date": effective_date,
                    "effective_date": effective_date,
                    "title": title,
                }
            )
        },
        headers=headers,
    )
    assert resp.status_code == 201, resp.text
    return resp.json()


def test_scoped_line_manager_only_sees_assigned_departments_and_employees(client: TestClient):
    admin_headers = _login(client)
    fixture = _create_scope_fixture(
        client,
        admin_headers,
        code_suffix="BASE",
        visible_employee_suffix="001",
        hidden_employee_suffix="002",
        visible_name=f"{_PREFIX} Visible",
        hidden_name=f"{_PREFIX} Hidden",
    )
    root = fixture["root"]
    child = fixture["child"]
    outside = fixture["outside"]
    visible_emp = fixture["visible_emp"]
    hidden_emp = fixture["hidden_emp"]

    manager_headers = _create_scoped_line_manager(
        client,
        admin_headers,
        department_ids=[root["id"]],
    )

    dept_resp = client.get(BASE_DEPT, headers=manager_headers)
    assert dept_resp.status_code == 200, dept_resp.text
    dept_codes = {item["code"] for item in dept_resp.json()}
    assert dept_codes == {root["code"], child["code"]}

    tree_resp = client.get(f"{BASE_DEPT}/tree", headers=manager_headers)
    assert tree_resp.status_code == 200, tree_resp.text
    tree = tree_resp.json()
    assert len(tree) == 1
    assert tree[0]["id"] == root["id"]
    assert [node["id"] for node in tree[0]["children"]] == [child["id"]]

    employees_resp = client.get(
        BASE_EMP,
        params={"keyword": _PREFIX},
        headers=manager_headers,
    )
    assert employees_resp.status_code == 200, employees_resp.text
    employee_names = {item["full_name"] for item in employees_resp.json()["items"]}
    assert employee_names == {visible_emp["full_name"]}

    assert client.get(f"{BASE_EMP}/{visible_emp['id']}", headers=manager_headers).status_code == 200
    assert client.get(f"{BASE_EMP}/{hidden_emp['id']}", headers=manager_headers).status_code == 403
    assert client.get(f"{BASE_DEPT}/{outside['id']}", headers=manager_headers).status_code == 403

    me_resp = client.get("/api/v1/auth/me", headers=manager_headers)
    assert me_resp.status_code == 200, me_resp.text
    me_payload = me_resp.json()
    assert me_payload["department_scopes"]["org"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["employees"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["leaves"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["performance"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["rewards"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["disciplines"] == sorted([root["id"], child["id"]])
    assert me_payload["department_scopes"]["reports"] == sorted([root["id"], child["id"]])


def test_scoped_line_manager_nested_employee_view_is_blocked_outside_scope(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_R2", name=f"{_PREFIX} Root 2")
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_O2", name=f"{_PREFIX} Outside 2")
    hidden_emp = _create_employee(
        client,
        admin_headers,
        suffix="003",
        full_name=f"{_PREFIX} Hidden Nested",
        department_id=outside["id"],
    )

    manager_headers = _create_scoped_line_manager(
        client,
        admin_headers,
        department_ids=[root["id"]],
    )

    resp = client.get(f"{BASE_EMP}/{hidden_emp['id']}/attachments", headers=manager_headers)
    assert resp.status_code == 403


def test_scoped_line_manager_leave_actions_are_limited_to_assigned_departments(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_LR", name=f"{_PREFIX} Leave Root")
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_LO", name=f"{_PREFIX} Leave Outside")
    visible_emp = _create_employee(
        client,
        admin_headers,
        suffix="004",
        full_name=f"{_PREFIX} Leave Visible",
        department_id=root["id"],
    )
    hidden_emp = _create_employee(
        client,
        admin_headers,
        suffix="005",
        full_name=f"{_PREFIX} Leave Hidden",
        department_id=outside["id"],
    )
    year = 2096
    _create_leave_entitlement(client, admin_headers, employee_id=visible_emp["id"], year=year)
    _create_leave_entitlement(client, admin_headers, employee_id=hidden_emp["id"], year=year)
    hidden_record = _create_leave_record(client, admin_headers, employee_id=hidden_emp["id"], year=year)

    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    visible_create = client.post(
        BASE_LEAVE_REC,
        json={
            "employee_id": visible_emp["id"],
            "leave_type_id": 1,
            "start_date": str(date(year, 6, 2)),
            "end_date": str(date(year, 6, 2)),
        },
        headers=manager_headers,
    )
    assert visible_create.status_code == 201, visible_create.text

    hidden_create = client.post(
        BASE_LEAVE_REC,
        json={
            "employee_id": hidden_emp["id"],
            "leave_type_id": 1,
            "start_date": str(date(year, 6, 3)),
            "end_date": str(date(year, 6, 3)),
        },
        headers=manager_headers,
    )
    assert hidden_create.status_code == 403, hidden_create.text

    hidden_cancel = client.post(
        f"{BASE_LEAVE_REC}/{hidden_record['id']}/cancel",
        json={"cancel_reason": "scope test"},
        headers=manager_headers,
    )
    assert hidden_cancel.status_code == 403, hidden_cancel.text

    visible_list = client.get(
        BASE_LEAVE_REC,
        params={"employee_id": visible_emp["id"], "year": year},
        headers=manager_headers,
    )
    assert visible_list.status_code == 200, visible_list.text


def test_scoped_line_manager_performance_views_and_reports_are_limited_to_assigned_departments(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_PR", name=f"{_PREFIX} Perf Root")
    child = _create_department(
        client,
        admin_headers,
        code=f"{_PREFIX}_PC",
        name=f"{_PREFIX} Perf Child",
        parent_id=root["id"],
    )
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_PO", name=f"{_PREFIX} Perf Outside")
    visible_emp = _create_employee(
        client,
        admin_headers,
        suffix="006",
        full_name=f"{_PREFIX} Perf Visible",
        department_id=child["id"],
    )
    hidden_emp = _create_employee(
        client,
        admin_headers,
        suffix="007",
        full_name=f"{_PREFIX} Perf Hidden",
        department_id=outside["id"],
    )
    year = 2097
    visible_kpi = _create_kpi(
        client,
        admin_headers,
        employee_id=visible_emp["id"],
        year=year,
        month=6,
        score=88.0,
    )
    hidden_kpi = _create_kpi(
        client,
        admin_headers,
        employee_id=hidden_emp["id"],
        year=year,
        month=6,
        score=91.0,
    )
    _create_yearly_review(client, admin_headers, employee_id=visible_emp["id"], year=year)
    _create_yearly_review(client, admin_headers, employee_id=hidden_emp["id"], year=year)

    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    list_resp = client.get(BASE_KPI, params={"year": year, "search": _PREFIX}, headers=manager_headers)
    assert list_resp.status_code == 200, list_resp.text
    listed_ids = {item["id"] for item in list_resp.json()["items"]}
    assert visible_kpi["id"] in listed_ids
    assert hidden_kpi["id"] not in listed_ids

    hidden_detail = client.get(f"{BASE_KPI}/{hidden_kpi['id']}", headers=manager_headers)
    assert hidden_detail.status_code == 403, hidden_detail.text

    hidden_history = client.get(
        f"{BASE_EMP}/{hidden_emp['id']}/performance/kpi",
        headers=manager_headers,
    )
    assert hidden_history.status_code == 403, hidden_history.text

    report_resp = client.get(
        "/api/v1/performance/report/rating-distribution",
        params={"year": year},
        headers=manager_headers,
    )
    assert report_resp.status_code == 200, report_resp.text
    assert report_resp.json()["total_reviewed"] == 1


def test_scoped_line_manager_can_create_kpi_and_review_within_scope(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_PK", name=f"{_PREFIX} Perf Create Root")
    child = _create_department(
        client,
        admin_headers,
        code=f"{_PREFIX}_PKC",
        name=f"{_PREFIX} Perf Create Child",
        parent_id=root["id"],
    )
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_PKO", name=f"{_PREFIX} Perf Create Outside")
    visible_emp = _create_employee(
        client,
        admin_headers,
        suffix="008",
        full_name=f"{_PREFIX} Perf Create Visible",
        department_id=child["id"],
    )
    hidden_emp = _create_employee(
        client,
        admin_headers,
        suffix="009",
        full_name=f"{_PREFIX} Perf Create Hidden",
        department_id=outside["id"],
    )
    year = 2098
    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    visible_kpi = client.post(
        BASE_KPI,
        json={"employee_id": visible_emp["id"], "year": year, "month": 6, "score": 85.0},
        headers=manager_headers,
    )
    assert visible_kpi.status_code == 201, visible_kpi.text

    hidden_kpi = client.post(
        BASE_KPI,
        json={"employee_id": hidden_emp["id"], "year": year, "month": 6, "score": 91.0},
        headers=manager_headers,
    )
    assert hidden_kpi.status_code == 403, hidden_kpi.text

    visible_review = client.post(
        BASE_REVIEWS,
        json={"employee_id": visible_emp["id"], "year": year, "rating": "tot"},
        headers=manager_headers,
    )
    assert visible_review.status_code == 201, visible_review.text

    hidden_review = client.post(
        BASE_REVIEWS,
        json={"employee_id": hidden_emp["id"], "year": year, "rating": "tot"},
        headers=manager_headers,
    )
    assert hidden_review.status_code == 403, hidden_review.text


def test_scoped_line_manager_dashboard_and_hr_reports_are_limited_to_assigned_departments(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_DR", name=f"{_PREFIX} Dash Root")
    child = _create_department(
        client,
        admin_headers,
        code=f"{_PREFIX}_DC",
        name=f"{_PREFIX} Dash Child",
        parent_id=root["id"],
    )
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_DO", name=f"{_PREFIX} Dash Outside")

    visible_emp = _create_employee(
        client,
        admin_headers,
        suffix="010",
        full_name=f"{_PREFIX} Dash Visible",
        department_id=child["id"],
    )
    _create_employee(
        client,
        admin_headers,
        suffix="011",
        full_name=f"{_PREFIX} Dash Hidden",
        department_id=outside["id"],
    )

    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    dashboard_summary = client.get(f"{BASE_DASHBOARD}/summary", headers=manager_headers)
    assert dashboard_summary.status_code == 200, dashboard_summary.text
    assert dashboard_summary.json()["total_headcount"] == 1

    headcount_by_dept = client.get(f"{BASE_DASHBOARD}/headcount-by-dept", headers=manager_headers)
    assert headcount_by_dept.status_code == 200, headcount_by_dept.text
    root_items = headcount_by_dept.json()
    assert len(root_items) == 1
    assert root_items[0]["department_id"] == root["id"]
    assert root_items[0]["headcount"] == 1

    headcount_by_root = client.get(
        f"{BASE_DASHBOARD}/headcount-by-dept",
        params={"department_id": root["id"]},
        headers=manager_headers,
    )
    assert headcount_by_root.status_code == 200, headcount_by_root.text
    listed_department_ids = {item["department_id"] for item in headcount_by_root.json()}
    assert listed_department_ids == {child["id"]}

    forbidden_department_filter = client.get(
        f"{BASE_DASHBOARD}/summary",
        params={"department_id": outside["id"]},
        headers=manager_headers,
    )
    assert forbidden_department_filter.status_code == 403, forbidden_department_filter.text

    employee_list = client.get(
        f"{BASE_HR_REPORTS}/employee-list",
        params={"department_id": root["id"]},
        headers=manager_headers,
    )
    assert employee_list.status_code == 200, employee_list.text
    employee_items = employee_list.json()["items"]
    assert len(employee_items) == 1
    assert employee_items[0]["id"] == visible_emp["id"]

    hr_forbidden_department_filter = client.get(
        f"{BASE_HR_REPORTS}/employee-list",
        params={"department_id": outside["id"]},
        headers=manager_headers,
    )
    assert hr_forbidden_department_filter.status_code == 403, hr_forbidden_department_filter.text


def test_scoped_line_manager_org_module_is_limited_to_department_related_views(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_OR", name=f"{_PREFIX} Org Root")
    child = _create_department(
        client,
        admin_headers,
        code=f"{_PREFIX}_OC",
        name=f"{_PREFIX} Org Child",
        parent_id=root["id"],
    )
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_OX", name=f"{_PREFIX} Org Outside")

    visible_position = _create_job_position(
        client,
        admin_headers,
        code=f"{_PREFIX}_P1",
        name=f"{_PREFIX} Position Visible",
        department_id=root["id"],
        assigned_department_ids=[root["id"], child["id"]],
    )
    hidden_position = _create_job_position(
        client,
        admin_headers,
        code=f"{_PREFIX}_P2",
        name=f"{_PREFIX} Position Hidden",
        department_id=outside["id"],
        assigned_department_ids=[outside["id"]],
    )

    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    job_titles = client.get(BASE_JOB_TITLES, headers=manager_headers)
    assert job_titles.status_code == 403, job_titles.text

    org_history = client.get(BASE_ORG_HISTORY, headers=manager_headers)
    assert org_history.status_code == 403, org_history.text

    positions = client.get(BASE_JOB_POSITIONS, headers=manager_headers)
    assert positions.status_code == 200, positions.text
    position_ids = {item["id"] for item in positions.json()}
    assert visible_position["id"] in position_ids
    assert hidden_position["id"] not in position_ids

    visible_position_detail = client.get(f"{BASE_JOB_POSITIONS}/{visible_position['id']}", headers=manager_headers)
    assert visible_position_detail.status_code == 200, visible_position_detail.text

    hidden_position_detail = client.get(f"{BASE_JOB_POSITIONS}/{hidden_position['id']}", headers=manager_headers)
    assert hidden_position_detail.status_code == 403, hidden_position_detail.text


def test_scoped_line_manager_rewards_disciplines_and_reports_are_limited_to_assigned_departments(client: TestClient):
    admin_headers = _login(client)
    fixture = _create_scope_fixture(
        client,
        admin_headers,
        code_suffix="RD",
        visible_employee_suffix="014",
        hidden_employee_suffix="015",
        visible_name=f"{_PREFIX} Reward Visible",
        hidden_name=f"{_PREFIX} Reward Hidden",
    )
    root = fixture["root"]
    visible_emp = fixture["visible_emp"]
    hidden_emp = fixture["hidden_emp"]
    outside = fixture["outside"]
    report_date = "2099-06-15"

    hidden_reward = _create_reward(
        client,
        admin_headers,
        employee_id=hidden_emp["id"],
        title=f"{_PREFIX} Hidden Reward",
        reward_date=report_date,
    )
    hidden_discipline = _create_discipline(
        client,
        admin_headers,
        employee_id=hidden_emp["id"],
        title=f"{_PREFIX} Hidden Discipline",
        effective_date=report_date,
    )

    manager_headers = _create_scoped_line_manager(client, admin_headers, department_ids=[root["id"]])

    visible_reward_create = client.post(
        BASE_REWARDS,
        data={
            "body": json.dumps(
                {
                    "employee_id": visible_emp["id"],
                    "reward_type_id": _get_first_reward_type_id(),
                    "reward_date": report_date,
                    "title": f"{_PREFIX} Visible Reward",
                }
            )
        },
        headers=manager_headers,
    )
    assert visible_reward_create.status_code == 201, visible_reward_create.text
    visible_reward = visible_reward_create.json()

    hidden_reward_create = client.post(
        BASE_REWARDS,
        data={
            "body": json.dumps(
                {
                    "employee_id": hidden_emp["id"],
                    "reward_type_id": _get_first_reward_type_id(),
                    "reward_date": report_date,
                    "title": f"{_PREFIX} Blocked Reward",
                }
            )
        },
        headers=manager_headers,
    )
    assert hidden_reward_create.status_code == 403, hidden_reward_create.text

    visible_discipline_create = client.post(
        BASE_DISCIPLINES,
        data={
            "body": json.dumps(
                {
                    "employee_id": visible_emp["id"],
                    "discipline_form": "khien_trach",
                    "violation_date": report_date,
                    "effective_date": report_date,
                    "title": f"{_PREFIX} Visible Discipline",
                }
            )
        },
        headers=manager_headers,
    )
    assert visible_discipline_create.status_code == 201, visible_discipline_create.text
    visible_discipline = visible_discipline_create.json()

    hidden_discipline_create = client.post(
        BASE_DISCIPLINES,
        data={
            "body": json.dumps(
                {
                    "employee_id": hidden_emp["id"],
                    "discipline_form": "khien_trach",
                    "violation_date": report_date,
                    "effective_date": report_date,
                    "title": f"{_PREFIX} Blocked Discipline",
                }
            )
        },
        headers=manager_headers,
    )
    assert hidden_discipline_create.status_code == 403, hidden_discipline_create.text

    rewards_resp = client.get(
        BASE_REWARDS,
        params={"search": _PREFIX, "page_size": 100},
        headers=manager_headers,
    )
    assert rewards_resp.status_code == 200, rewards_resp.text
    reward_ids = {item["id"] for item in rewards_resp.json()["items"]}
    assert visible_reward["id"] in reward_ids
    assert hidden_reward["id"] not in reward_ids

    disciplines_resp = client.get(
        BASE_DISCIPLINES,
        params={"search": _PREFIX, "page_size": 100},
        headers=manager_headers,
    )
    assert disciplines_resp.status_code == 200, disciplines_resp.text
    discipline_ids = {item["id"] for item in disciplines_resp.json()["items"]}
    assert visible_discipline["id"] in discipline_ids
    assert hidden_discipline["id"] not in discipline_ids

    assert client.get(f"{BASE_REWARDS}/{visible_reward['id']}", headers=manager_headers).status_code == 200
    assert client.get(f"{BASE_REWARDS}/{hidden_reward['id']}", headers=manager_headers).status_code == 403
    assert client.get(f"{BASE_DISCIPLINES}/{visible_discipline['id']}", headers=manager_headers).status_code == 200
    assert client.get(f"{BASE_DISCIPLINES}/{hidden_discipline['id']}", headers=manager_headers).status_code == 403
    assert client.get(f"{BASE_EMP}/{hidden_emp['id']}/rewards", headers=manager_headers).status_code == 403
    assert client.get(f"{BASE_EMP}/{hidden_emp['id']}/disciplines", headers=manager_headers).status_code == 403

    summary_resp = client.get(
        f"{BASE_REWARDS}/report/summary",
        params={"from_date": "2099-01-01", "to_date": "2099-12-31"},
        headers=manager_headers,
    )
    assert summary_resp.status_code == 200, summary_resp.text
    summary_body = summary_resp.json()
    assert summary_body["total_rewards"] == 1
    assert summary_body["total_disciplines"] == 1
    assert {item["employee_id"] for item in summary_body["reward_items"]} == {visible_emp["id"]}
    assert {item["employee_id"] for item in summary_body["discipline_items"]} == {visible_emp["id"]}

    forbidden_summary = client.get(
        f"{BASE_REWARDS}/report/summary",
        params={"from_date": "2099-01-01", "to_date": "2099-12-31", "department_id": outside["id"]},
        headers=manager_headers,
    )
    assert forbidden_summary.status_code == 403, forbidden_summary.text

    export_resp = client.get(
        f"{BASE_REWARDS}/report/export",
        params={"from_date": "2099-01-01", "to_date": "2099-12-31"},
        headers=manager_headers,
    )
    assert export_resp.status_code == 200, export_resp.text
    assert export_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = openpyxl.load_workbook(io.BytesIO(export_resp.content))
    reward_rows = [
        row[2]
        for row in workbook["Khen thưởng"].iter_rows(min_row=5, values_only=True)
        if row[0] not in (None, "TỔNG CỘNG")
    ]
    discipline_rows = [
        row[2]
        for row in workbook["Kỷ luật"].iter_rows(min_row=5, values_only=True)
        if row[0] not in (None, "TỔNG CỘNG")
    ]
    assert visible_emp["full_name"] in reward_rows
    assert hidden_emp["full_name"] not in reward_rows
    assert visible_emp["full_name"] in discipline_rows
    assert hidden_emp["full_name"] not in discipline_rows


def test_scoped_contract_user_only_sees_contracts_within_assigned_departments(client: TestClient):
    admin_headers = _login(client)
    root = _create_department(client, admin_headers, code=f"{_PREFIX}_CR", name=f"{_PREFIX} Contract Root")
    child = _create_department(
        client,
        admin_headers,
        code=f"{_PREFIX}_CC",
        name=f"{_PREFIX} Contract Child",
        parent_id=root["id"],
    )
    outside = _create_department(client, admin_headers, code=f"{_PREFIX}_CO", name=f"{_PREFIX} Contract Outside")

    visible_emp = _create_employee(
        client,
        admin_headers,
        suffix="012",
        full_name=f"{_PREFIX} Contract Visible",
        department_id=child["id"],
    )
    hidden_emp = _create_employee(
        client,
        admin_headers,
        suffix="013",
        full_name=f"{_PREFIX} Contract Hidden",
        department_id=outside["id"],
    )

    visible_contract = _create_contract(
        client,
        admin_headers,
        employee_id=visible_emp["id"],
        contract_number=f"{_PREFIX}-CON-VISIBLE",
        effective_to="2026-07-15",
    )
    _create_contract(
        client,
        admin_headers,
        employee_id=hidden_emp["id"],
        contract_number=f"{_PREFIX}-CON-HIDDEN",
        effective_to="2026-07-20",
    )

    scoped_headers = _create_scoped_role_user(
        client,
        admin_headers,
        role_code="hr_officer",
        department_ids=[root["id"]],
        email=f"{_PREFIX.lower()}contract@example.com",
        full_name=f"{_PREFIX} Contract Scoped",
        password="Contract@1234",
    )

    me_resp = client.get("/api/v1/auth/me", headers=scoped_headers)
    assert me_resp.status_code == 200, me_resp.text
    assert me_resp.json()["department_scopes"]["contracts"] == sorted([root["id"], child["id"]])

    global_list = client.get(BASE_CONTRACTS, headers=scoped_headers)
    assert global_list.status_code == 200, global_list.text
    contract_numbers = {item["contract_number"] for item in global_list.json()["items"]}
    assert f"{_PREFIX}-CON-VISIBLE" in contract_numbers
    assert f"{_PREFIX}-CON-HIDDEN" not in contract_numbers

    visible_list = client.get(
        f"{BASE_EMP}/{visible_emp['id']}/contracts",
        headers=scoped_headers,
    )
    assert visible_list.status_code == 200, visible_list.text

    hidden_list = client.get(
        f"{BASE_EMP}/{hidden_emp['id']}/contracts",
        headers=scoped_headers,
    )
    assert hidden_list.status_code == 403, hidden_list.text

    expiring_resp = client.get(
        f"{BASE_CONTRACT_REPORTS}/expiring",
        params={"days_ahead": 90},
        headers=scoped_headers,
    )
    assert expiring_resp.status_code == 200, expiring_resp.text
    expiring_numbers = {item["contract_number"] for item in expiring_resp.json()["items"]}
    assert f"{_PREFIX}-CON-VISIBLE" in expiring_numbers
    assert f"{_PREFIX}-CON-HIDDEN" not in expiring_numbers

    history_visible = client.get(
        f"{BASE_CONTRACT_REPORTS}/history",
        params={"employee_id": visible_emp["id"]},
        headers=scoped_headers,
    )
    assert history_visible.status_code == 200, history_visible.text
    assert any(
        item["contract_number"] == visible_contract["contract_number"]
        for item in history_visible.json()["items"]
    )

    history_hidden = client.get(
        f"{BASE_CONTRACT_REPORTS}/history",
        params={"employee_id": hidden_emp["id"]},
        headers=scoped_headers,
    )
    assert history_hidden.status_code == 403, history_hidden.text
