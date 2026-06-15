from __future__ import annotations

import io
from datetime import date

import openpyxl
import pytest
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import hash_sensitive
from app.models.catalog import Ethnicity, Nationality, Religion
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.services.employee_import_service import IMPORT_COLUMNS, generate_template, process_import


TEST_ID_NUMBER = "IMPORTNATVN001"
TEST_ID_NUMBER_WITH_CATALOGS = "IMPORTNATVN002"
TEST_ID_NUMBER_INVALID_CATALOG = "IMPORTNATVN003"
TEST_ID_NUMBER_LEGACY_CODE = "IMPORTNATVN004"
TEST_ID_NUMBER_LEGACY_CODE_MISMATCH = "IMPORTNATVN005"


def _make_engine_and_sessionmaker():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return engine, async_sessionmaker(engine, expire_on_commit=False)


def _make_xlsx(rows: list[list[str]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_import_row(
    *,
    id_number: str,
    start_date_value: str,
    sequence_code: str,
    nationality: str = "",
    ethnicity: str = "",
    religion: str = "",
    employee_seq: str = "",
    display_code: str = "",
) -> list[str]:
    values = {column: "" for column in IMPORT_COLUMNS}
    values.update(
        {
            "Họ và tên": "Import Test",
            "Họ": "Import",
            "Tên": "Test",
            "Ngày sinh": "01/01/1990",
            "Giới tính": "nam",
            "Số CCCD/CMND": id_number,
            "Ngày cấp CCCD": "01/01/2020",
            "Nơi cấp CCCD": "Cục Cảnh sát ĐKQLCƯ",
            "Trạng thái": "probation",
            "Ngày vào làm": start_date_value,
            "Hệ mã nhân viên": sequence_code,
            "Quốc tịch": nationality,
            "Dân tộc": ethnicity,
            "Tôn giáo": religion,
            "Số thứ tự mã NV": employee_seq,
            "Mã NV hiện hữu": display_code,
        }
    )
    return [values[column] for column in IMPORT_COLUMNS]


async def _cleanup_employee(session: AsyncSession) -> None:
    for test_id_number in (
        TEST_ID_NUMBER,
        TEST_ID_NUMBER_WITH_CATALOGS,
        TEST_ID_NUMBER_INVALID_CATALOG,
        TEST_ID_NUMBER_LEGACY_CODE,
        TEST_ID_NUMBER_LEGACY_CODE_MISMATCH,
    ):
        employee = (
            await session.execute(
                select(Employee).where(Employee.id_number_hash == hash_sensitive(test_id_number))
            )
        ).scalar_one_or_none()
        if employee:
            await session.execute(delete(Employee).where(Employee.id == employee.id))
    await session.commit()


@pytest.mark.asyncio
async def test_employee_import_defaults_nationality_to_vn():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup_employee(session)

            vn = (
                await session.execute(select(Nationality).where(Nationality.code == "VN"))
            ).scalar_one()
            sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1")
                )
            ).scalar_one()

            workbook = _make_xlsx(
                [
                    IMPORT_COLUMNS,
                    _make_import_row(
                        id_number=TEST_ID_NUMBER,
                        start_date_value=date.today().strftime("%d/%m/%Y"),
                        sequence_code=sequence.code,
                    ),
                ]
            )

            result = await process_import(session, workbook)
            assert result.success == 1
            assert len(result.created_ids) == 1

            employee = (
                await session.execute(select(Employee).where(Employee.id == result.created_ids[0]))
            ).scalar_one()
            assert employee.nationality_id == vn.id

            await _cleanup_employee(session)
    finally:
        await engine.dispose()


def test_employee_import_template_explains_employee_code_sequences():
    workbook = openpyxl.load_workbook(io.BytesIO(generate_template()))
    guide = workbook["Hướng dẫn"]

    sequence_row = None
    for row_no in range(2, guide.max_row + 1):
        if guide.cell(row_no, 1).value == "Hệ mã nhân viên":
            sequence_row = row_no
            break

    assert sequence_row is not None
    assert guide.cell(sequence_row, 4).value == (
        "SYS1 = Hệ 1 (mặc định toàn công ty); "
        "SYS2 = Hệ 2 (công nhân bốc xếp / ra cám / tạp vụ); "
        "SYS3 = Hệ 3 (công nhân / bảo vệ thuộc Phòng trại)"
    )


@pytest.mark.asyncio
async def test_employee_import_preserves_legacy_employee_code_and_bumps_sequence():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup_employee(session)

            sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1")
                )
            ).scalar_one()
            original_next_value = sequence.next_value
            legacy_seq = max(original_next_value + 25, 1200)

            workbook = _make_xlsx(
                [
                    IMPORT_COLUMNS,
                    _make_import_row(
                        id_number=TEST_ID_NUMBER_LEGACY_CODE,
                        start_date_value=date.today().strftime("%d/%m/%Y"),
                        sequence_code=sequence.code,
                        employee_seq=str(legacy_seq),
                        display_code=f"{legacy_seq:04d}",
                    ),
                ]
            )

            result = await process_import(session, workbook)
            assert result.success == 1
            assert len(result.created_ids) == 1

            employee = (
                await session.execute(select(Employee).where(Employee.id == result.created_ids[0]))
            ).scalar_one()
            assert employee.employee_seq == legacy_seq
            assert employee.employee_code_sequence_id == sequence.id

            updated_sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.id == sequence.id)
                )
            ).scalar_one()
            assert updated_sequence.next_value == legacy_seq + 1

            await _cleanup_employee(session)
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_employee_import_fails_when_legacy_display_code_does_not_match():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup_employee(session)

            sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1")
                )
            ).scalar_one()

            workbook = _make_xlsx(
                [
                    IMPORT_COLUMNS,
                    _make_import_row(
                        id_number=TEST_ID_NUMBER_LEGACY_CODE_MISMATCH,
                        start_date_value=date.today().strftime("%d/%m/%Y"),
                        sequence_code=sequence.code,
                        employee_seq="1500",
                        display_code="SAI1500",
                    ),
                ]
            )

            result = await process_import(session, workbook)
            assert result.success == 0
            assert result.failed == 1
            assert any(
                error.column == "Mã NV hiện hữu" and "không khớp" in error.message
                for error in result.errors
            )

            employee = (
                await session.execute(
                    select(Employee).where(Employee.id_number_hash == hash_sensitive(TEST_ID_NUMBER_LEGACY_CODE_MISMATCH))
                )
            ).scalar_one_or_none()
            assert employee is None
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_employee_import_maps_nationality_ethnicity_religion_from_catalog():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup_employee(session)

            vn = (
                await session.execute(select(Nationality).where(Nationality.code == "VN"))
            ).scalar_one()
            kinh = (
                await session.execute(select(Ethnicity).where(Ethnicity.code == "KINH"))
            ).scalar_one()
            khong = (
                await session.execute(select(Religion).where(Religion.code == "NONE"))
            ).scalar_one()
            sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1")
                )
            ).scalar_one()

            workbook = _make_xlsx(
                [
                    IMPORT_COLUMNS,
                    _make_import_row(
                        id_number=TEST_ID_NUMBER_WITH_CATALOGS,
                        start_date_value=date.today().strftime("%d/%m/%Y"),
                        sequence_code=sequence.code,
                        nationality="VN",
                        ethnicity="Kinh",
                        religion="Không",
                    ),
                ]
            )

            result = await process_import(session, workbook)
            assert result.success == 1
            assert len(result.created_ids) == 1

            employee = (
                await session.execute(select(Employee).where(Employee.id == result.created_ids[0]))
            ).scalar_one()
            assert employee.nationality_id == vn.id
            assert employee.ethnicity_id == kinh.id
            assert employee.religion_id == khong.id

            await _cleanup_employee(session)
    finally:
        await engine.dispose()


@pytest.mark.asyncio
async def test_employee_import_fails_when_catalog_value_not_found():
    engine, session_factory = _make_engine_and_sessionmaker()
    try:
        async with session_factory() as session:
            await _cleanup_employee(session)

            sequence = (
                await session.execute(
                    select(EmployeeCodeSequence).where(EmployeeCodeSequence.code == "SYS1")
                )
            ).scalar_one()

            workbook = _make_xlsx(
                [
                    IMPORT_COLUMNS,
                    _make_import_row(
                        id_number=TEST_ID_NUMBER_INVALID_CATALOG,
                        start_date_value=date.today().strftime("%d/%m/%Y"),
                        sequence_code=sequence.code,
                        nationality="QUOC-TICH-KHONG-TON-TAI",
                    ),
                ]
            )

            result = await process_import(session, workbook)
            assert result.success == 0
            assert result.failed == 1
            assert any(error.column == "Quốc tịch" for error in result.errors)

            employee = (
                await session.execute(
                    select(Employee).where(Employee.id_number_hash == hash_sensitive(TEST_ID_NUMBER_INVALID_CATALOG))
                )
            ).scalar_one_or_none()
            assert employee is None
    finally:
        await engine.dispose()
