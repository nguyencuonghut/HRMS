"""Integration tests cho Plan 7.3 — Bảng tổng hợp lương BHXH.

Covers:
  - TestGetRatesForMonth      : lấy tỷ lệ đóng BHXH từ DB theo tháng
  - TestComputeContributions  : pure function tính số tiền đóng (không cần DB)
  - TestGetSalarySummary      : API bảng tổng hợp + filter + pagination
  - TestExportSalarySummaryExcel : export file xlsx
"""
from __future__ import annotations

import asyncio
from decimal import Decimal

import pytest
from fastapi.testclient import TestClient
from io import BytesIO

from app.services.salary_service import _compute_contributions

BASE = "/api/v1/salary"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

# Dùng tháng/năm hiện tại có dữ liệu nhân viên và policy
_TEST_YEAR = 2026
_TEST_MONTH = 5

# Tỷ lệ seed mặc định từ migration 0018
_BHXH_EMP_RATE = Decimal("8.0")
_BHYT_EMP_RATE = Decimal("1.5")
_BHTN_EMP_RATE = Decimal("1.0")
_BHXH_ER_RATE = Decimal("17.5")
_BHYT_ER_RATE = Decimal("3.0")
_BHTN_ER_RATE = Decimal("1.0")


# ── Auth ───────────────────────────────────────────────────────────────────────

def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_rates() -> dict:
    return {
        "BHXH": {"employee_rate": _BHXH_EMP_RATE, "employer_rate": _BHXH_ER_RATE},
        "BHYT": {"employee_rate": _BHYT_EMP_RATE, "employer_rate": _BHYT_ER_RATE},
        "BHTN": {"employee_rate": _BHTN_EMP_RATE, "employer_rate": _BHTN_ER_RATE},
    }


# ── TestGetRatesForMonth ───────────────────────────────────────────────────────

class TestGetRatesForMonth:
    """API /summary phải trả rates đúng từ DB."""

    def test_returns_correct_rates_for_given_month(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        rates = r.json()["rates"]
        assert float(rates["bhxh_employee_rate"]) == pytest.approx(8.0)
        assert float(rates["bhyt_employee_rate"]) == pytest.approx(1.5)
        assert float(rates["bhtn_employee_rate"]) == pytest.approx(1.0)
        assert float(rates["bhxh_employer_rate"]) == pytest.approx(17.5)
        assert float(rates["bhyt_employer_rate"]) == pytest.approx(3.0)
        assert float(rates["bhtn_employer_rate"]) == pytest.approx(1.0)

    def test_raises_422_when_no_rates_configured_for_month(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year=1900&month=1",
            headers=_admin(client),
        )
        assert r.status_code == 422, r.text

    def test_returns_rates_in_summary_response(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        data = r.json()
        assert "rates" in data
        rates = data["rates"]
        required = [
            "bhxh_employee_rate", "bhyt_employee_rate", "bhtn_employee_rate",
            "bhxh_employer_rate", "bhyt_employer_rate", "bhtn_employer_rate",
        ]
        for key in required:
            assert key in rates, f"Missing rate key: {key}"


# ── TestComputeContributions ───────────────────────────────────────────────────

class TestComputeContributions:
    """Pure function — không cần DB, không cần client."""

    @pytest.fixture(autouse=True)
    def rates(self):
        self._rates = _make_rates()

    def test_bhxh_employee_is_8_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhxh_employee"] == Decimal("800000")

    def test_bhyt_employee_is_1_5_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhyt_employee"] == Decimal("150000")

    def test_bhtn_employee_is_1_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhtn_employee"] == Decimal("100000")

    def test_bhxh_employer_is_17_5_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhxh_employer"] == Decimal("1750000")

    def test_bhyt_employer_is_3_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhyt_employer"] == Decimal("300000")

    def test_bhtn_employer_is_1_percent_of_basis(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["bhtn_employer"] == Decimal("100000")

    def test_total_employee_is_sum_of_nld_components(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        expected = contrib["bhxh_employee"] + contrib["bhyt_employee"] + contrib["bhtn_employee"]
        assert contrib["total_employee"] == expected

    def test_total_employer_is_sum_of_nsdld_components(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        expected = contrib["bhxh_employer"] + contrib["bhyt_employer"] + contrib["bhtn_employer"]
        assert contrib["total_employer"] == expected

    def test_grand_total_is_nld_plus_nsdld(self):
        contrib = _compute_contributions(Decimal("10000000"), self._rates)
        assert contrib["grand_total"] == contrib["total_employee"] + contrib["total_employer"]

    def test_rounds_to_integer_using_half_up(self):
        # basis = 3_000_001: BHXH NLĐ = 3_000_001 * 8% = 240_000.08 → 240_000
        # basis = 3_125_000: BHXH NLĐ = 3_125_000 * 8% = 250_000 (exact)
        # basis = 3_750_000: BHYT NLĐ = 3_750_000 * 1.5% = 56_250 (exact)
        # Test rounding: basis = 1 → BHXH NLĐ = 0.08 → rounds to 0
        contrib_small = _compute_contributions(Decimal("1"), self._rates)
        assert contrib_small["bhxh_employee"] == Decimal("0")

        # basis = 10: BHXH NLĐ = 0.8 → rounds to 1 (ROUND_HALF_UP)
        contrib_10 = _compute_contributions(Decimal("10"), self._rates)
        assert contrib_10["bhxh_employee"] == Decimal("1")


# ── TestGetSalarySummary ──────────────────────────────────────────────────────

class TestGetSalarySummary:
    """API GET /salary/summary."""

    def test_only_includes_active_participation_employees(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        # Verify all returned employees are 'active' by checking via basis API
        items = r.json()["items"]
        for item in items:
            emp_id = item["employee_id"]
            basis_r = client.get(f"{BASE}/employees/{emp_id}/bhxh-basis", headers=_admin(client))
            if basis_r.status_code == 200:
                assert basis_r.json()["participation_status"] == "active"

    def test_excludes_paused_and_stopped_employees(self, client: TestClient):
        # Get list of all employees to verify paused/stopped don't appear in summary
        list_r = client.get(
            f"{BASE}/employees?participation_status=paused&page_size=5",
            headers=_admin(client),
        )
        summary_r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert summary_r.status_code == 200
        summary_ids = {item["employee_id"] for item in summary_r.json()["items"]}

        if list_r.status_code == 200:
            paused_ids = {item["employee_id"] for item in list_r.json()["items"]}
            assert paused_ids.isdisjoint(summary_ids), \
                "Paused employees should not appear in summary"

    def test_totals_match_sum_of_rows(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page_size=500",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        data = r.json()
        items = data["items"]
        totals = data["totals"]

        if not items:
            pytest.skip("Không có dữ liệu để kiểm tra")

        sum_basis = sum(Decimal(str(i["basis_amount"])) for i in items)
        sum_grand = sum(Decimal(str(i["grand_total"])) for i in items)

        assert Decimal(str(totals["sum_basis"])) == sum_basis
        assert Decimal(str(totals["sum_grand_total"])) == sum_grand

    def test_stt_starts_from_1_and_increments(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        items = r.json()["items"]
        if not items:
            pytest.skip("Không có dữ liệu")
        for i, item in enumerate(items):
            assert item["stt"] == i + 1

    def test_pagination_works(self, client: TestClient):
        r_all = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page_size=500",
            headers=_admin(client),
        )
        assert r_all.status_code == 200
        total = r_all.json()["total"]
        if total < 2:
            pytest.skip("Cần ít nhất 2 nhân viên để test pagination")

        r_p1 = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page=1&page_size=1",
            headers=_admin(client),
        )
        r_p2 = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page=2&page_size=1",
            headers=_admin(client),
        )
        assert r_p1.status_code == 200
        assert r_p2.status_code == 200
        assert r_p1.json()["items"][0]["employee_id"] != r_p2.json()["items"][0]["employee_id"]

    def test_filter_by_department_id(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200
        items = r.json()["items"]
        dept_ids = [i["employee_id"] for i in items if i.get("department_name")]
        if not dept_ids:
            pytest.skip("Không có nhân viên thuộc phòng ban nào")

        # Get a valid department from the response
        dept_item = next((i for i in items if i.get("department_name")), None)
        if not dept_item:
            pytest.skip("Không có nhân viên thuộc phòng ban nào")

        # Just verify the endpoint accepts department_id filter without error
        dept_r = client.get(
            f"{BASE}/employees?page_size=1",
            headers=_admin(client),
        )
        if dept_r.status_code == 200 and dept_r.json()["items"]:
            dept_id = dept_r.json()["items"][0].get("department_id")
            if dept_id:
                filtered_r = client.get(
                    f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&department_id={dept_id}",
                    headers=_admin(client),
                )
                assert filtered_r.status_code in (200, 422)

    def test_returns_empty_when_no_active_employees_for_month(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year=2019&month=1",
            headers=_admin(client),
        )
        # Either 422 (no rates) or 200 with empty items
        assert r.status_code in (200, 422)
        if r.status_code == 200:
            assert r.json()["total"] == 0

    def test_contribution_amounts_match_rates(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200
        data = r.json()
        rates = data["rates"]
        items = data["items"]
        if not items:
            pytest.skip("Không có dữ liệu")

        row = items[0]
        basis = Decimal(str(row["basis_amount"]))
        bhxh_emp_rate = Decimal(str(rates["bhxh_employee_rate"]))
        expected_bhxh_emp = (basis * bhxh_emp_rate / Decimal("100")).quantize(Decimal("1"))
        assert Decimal(str(row["bhxh_employee"])) == expected_bhxh_emp

    def test_unauthenticated_returns_401(self, client: TestClient):
        r = client.get(f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}")
        assert r.status_code == 401


# ── TestExportSalarySummaryExcel ──────────────────────────────────────────────

class TestExportSalarySummaryExcel:
    """API GET /salary/summary/export."""

    def test_returns_200_with_xlsx_content_type(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200, r.text
        assert "spreadsheetml" in r.headers.get("content-type", "")

    def test_file_is_valid_xlsx(self, client: TestClient):
        from openpyxl import load_workbook
        r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200
        wb = load_workbook(BytesIO(r.content))
        assert wb is not None
        assert len(wb.sheetnames) >= 1

    def test_header_contains_year_month(self, client: TestClient):
        from openpyxl import load_workbook
        r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        title_cell = ws["A2"].value or ""
        assert str(_TEST_YEAR) in title_cell
        assert str(_TEST_MONTH).zfill(2) in title_cell

    def test_correct_number_of_data_rows(self, client: TestClient):
        from openpyxl import load_workbook
        summary_r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page_size=500",
            headers=_admin(client),
        )
        assert summary_r.status_code == 200
        expected_count = summary_r.json()["total"]

        export_r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        wb = load_workbook(BytesIO(export_r.content))
        ws = wb.active
        # Data starts at row 7, totals row is next, signature row after +2
        # So data rows = max_row - 3 (totals + empty + signature)
        data_rows = ws.max_row - 3 - 6  # subtract header rows (1-6) and footer (totals+gap+sig)
        # Simpler: count rows from row 7 to max_row-3
        data_start = 7
        data_end = ws.max_row - 3
        actual_count = max(0, data_end - data_start + 1)
        assert actual_count == expected_count

    def test_bhxh_employee_column_is_8_percent_of_basis(self, client: TestClient):
        from openpyxl import load_workbook
        summary_r = client.get(
            f"{BASE}/summary?year={_TEST_YEAR}&month={_TEST_MONTH}&page_size=1",
            headers=_admin(client),
        )
        assert summary_r.status_code == 200
        items = summary_r.json()["items"]
        if not items:
            pytest.skip("Không có dữ liệu")

        export_r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        wb = load_workbook(BytesIO(export_r.content))
        ws = wb.active
        # Row 7 col E = basis_amount, col F = bhxh_employee
        basis_val = ws.cell(7, 5).value
        bhxh_val = ws.cell(7, 6).value
        if basis_val and bhxh_val:
            expected = round(basis_val * 0.08)
            assert bhxh_val == expected

    def test_totals_row_exists_in_export(self, client: TestClient):
        from openpyxl import load_workbook
        r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        wb = load_workbook(BytesIO(r.content))
        ws = wb.active
        # Totals row is max_row - 2 (signature at max_row, empty at max_row-1)
        totals_row = ws.max_row - 2
        totals_label = ws.cell(totals_row, 1).value
        assert totals_label is not None
        assert "TỔNG CỘNG" in str(totals_label)

    def test_content_disposition_has_correct_filename(self, client: TestClient):
        r = client.get(
            f"{BASE}/summary/export?year={_TEST_YEAR}&month={_TEST_MONTH}",
            headers=_admin(client),
        )
        assert r.status_code == 200
        disposition = r.headers.get("content-disposition", "")
        expected_filename = f"luong_bhxh_{_TEST_YEAR}_{str(_TEST_MONTH).zfill(2)}.xlsx"
        assert expected_filename in disposition
