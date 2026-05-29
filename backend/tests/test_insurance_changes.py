"""Integration tests cho Plan 6.3 — Biến động BHXH.

Covers:
  - TestAutoRecordOnStatusChange: auto-event khi PUT đổi status
  - TestManualChangeEvents: CRUD manual event + guard auto-event
  - TestMonthlyReport: monthly summary + filter
  - TestRefactoring: seed data integrity (Slice 4a)
  - TestExport: VNPT D02-TS Excel output — 3 sheets đầu (Dữ Liệu, Phụ lục, Bảng kê)
"""
from __future__ import annotations

import asyncio
import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

from app.core.config import settings
from app.core.encryption import hash_sensitive
from app.models.employee import Employee

BASE = "/api/v1/insurance"
_ADMIN_EMAIL = "admin@hrms.local"
_ADMIN_PASSWORD = "Hrms@2026"

# Test period unlikely to collide with real events
_TEST_YEAR = 2099
_TEST_MONTH = 12
_TEST_EMPLOYEE_ID_NUMBER = "TESTINSCHANGE0001"
_TEST_EMPLOYEE_NAME = "Test Insurance Change"
_TEST_CONTRACT_NUMBER = "TEST-INS-CHANGE-001"


def _make_session():
    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    return async_sessionmaker(engine, expire_on_commit=False)


def _admin(client: TestClient) -> dict:
    token = client.post(
        "/api/v1/auth/login",
        json={"email": _ADMIN_EMAIL, "password": _ADMIN_PASSWORD},
    ).json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


async def _ensure_employee_with_basis() -> dict:
    """Đảm bảo có 1 employee test với profile active + active contract + insurance_salary > 0."""
    async with _make_session()() as s:
        id_hash = hash_sensitive(_TEST_EMPLOYEE_ID_NUMBER)
        employee = (
            await s.execute(
                text(
                    """
                    SELECT id, full_name
                    FROM employees
                    WHERE id_number_hash = :id_number_hash
                    LIMIT 1
                    """
                ),
                {"id_number_hash": id_hash},
            )
        ).fetchone()

        if employee is None:
            max_seq = (
                await s.execute(text("SELECT COALESCE(MAX(employee_seq), 0) FROM employees"))
            ).scalar_one()
            nationality_id = (
                await s.execute(
                    text("SELECT id FROM nationalities WHERE iso2_code IS NOT NULL ORDER BY id LIMIT 1")
                )
            ).scalar_one()
            ethnicity_id = (
                await s.execute(
                    text("SELECT id FROM ethnicities WHERE bhxh_code IS NOT NULL ORDER BY id LIMIT 1")
                )
            ).scalar_one()
            employee_row = Employee(
                employee_seq=max_seq + 1,
                employee_code_sequence_id=1,
                full_name=_TEST_EMPLOYEE_NAME,
                normalized_name="test insurance change",
                last_name="Test Insurance",
                first_name="Change",
                date_of_birth=date(1990, 1, 1),
                gender="male",
                nationality_id=nationality_id,
                ethnicity_id=ethnicity_id,
                religion_id=None,
                id_number=_TEST_EMPLOYEE_ID_NUMBER,
                id_number_hash=id_hash,
                id_issued_on=date(2020, 1, 1),
                id_issued_by="CA Test",
                id_expires_on=None,
                passport_number=None,
                passport_issued_on=None,
                passport_expires_on=None,
                work_permit_number=None,
                work_permit_issued_on=None,
                work_permit_expires_on=None,
                phone_number=None,
                personal_email="test-insurance-change@example.com",
                personal_tax_code=None,
                bhxh_code=None,
                avatar_path=None,
                status="official",
                start_date=date(2020, 1, 1),
                resigned_date=None,
                user_id=None,
                is_active=True,
            )
            s.add(employee_row)
            await s.flush()
            employee_id = employee_row.id
        else:
            employee_id = employee[0]

        department_id = (
            await s.execute(text("SELECT id FROM departments ORDER BY id LIMIT 1"))
        ).scalar_one()
        job_title_id = (
            await s.execute(text("SELECT id FROM job_titles ORDER BY id LIMIT 1"))
        ).scalar()
        category = (
            await s.execute(
                text(
                    """
                    SELECT id, document_kind
                    FROM contract_categories
                    ORDER BY id
                    LIMIT 1
                    """
                )
            )
        ).fetchone()
        assert category is not None, "Không tìm thấy contract category trong DB test"

        await s.execute(
            text(
                """
                UPDATE employees
                SET status = 'official',
                    is_active = TRUE,
                    resigned_date = NULL,
                    ethnicity_id = COALESCE(ethnicity_id, :ethnicity_id)
                WHERE id = :employee_id
                """
            ),
            {"employee_id": employee_id, "ethnicity_id": (
                await s.execute(
                    text("SELECT id FROM ethnicities WHERE bhxh_code IS NOT NULL ORDER BY id LIMIT 1")
                )
            ).scalar_one()},
        )
        await s.execute(
            text("DELETE FROM employee_job_records WHERE employee_id = :employee_id AND is_current = TRUE"),
            {"employee_id": employee_id},
        )
        await s.execute(
            text(
                """
                INSERT INTO employee_job_records (
                    employee_id, department_id, job_title_id, job_position_id,
                    probation_start_date, probation_end_date, official_date,
                    effective_from, effective_to, is_current, notes, changed_by, created_at, updated_at
                ) VALUES (
                    :employee_id, :department_id, :job_title_id, NULL,
                    NULL, NULL, :official_date,
                    :effective_from, NULL, TRUE, NULL, NULL, NOW(), NOW()
                )
                """
            ),
            {
                "employee_id": employee_id,
                "department_id": department_id,
                "job_title_id": job_title_id,
                "official_date": date(2020, 1, 1),
                "effective_from": date(2020, 1, 1),
            },
        )

        await s.execute(
            text(
                """
                INSERT INTO employee_insurance_profiles (
                    employee_id, bhxh_code, bhyt_initial_clinic_name, bhyt_initial_clinic_code,
                    company_bhxh_joined_date, participation_status, status_effective_from, status_note,
                    insurance_basis_source, insurance_basis_amount, insurance_policy_version_id,
                    created_at, updated_at
                ) VALUES (
                    :employee_id, NULL, NULL, NULL,
                    NULL, 'active', NULL, NULL,
                    'contract', NULL, NULL,
                    NOW(), NOW()
                )
                ON CONFLICT (employee_id) DO UPDATE
                SET participation_status = 'active',
                    status_effective_from = NULL,
                    status_note = NULL,
                    insurance_basis_source = 'contract',
                    insurance_basis_amount = NULL,
                    updated_at = NOW()
                """
            ),
            {"employee_id": employee_id},
        )

        await s.execute(
            text("DELETE FROM employee_contracts WHERE contract_number = :contract_number"),
            {"contract_number": _TEST_CONTRACT_NUMBER},
        )
        await s.execute(
            text(
                """
                INSERT INTO employee_contracts (
                    employee_id, contract_category_id, document_kind,
                    contract_number, signed_date, effective_from, effective_to,
                    insurance_salary, status, created_at, updated_at
                ) VALUES (
                    :employee_id, :category_id, :document_kind,
                    :contract_number, :signed_date, :effective_from, :effective_to,
                    :insurance_salary, 'active', NOW(), NOW()
                )
                """
            ),
            {
                "employee_id": employee_id,
                "category_id": category[0],
                "document_kind": category[1],
                "contract_number": _TEST_CONTRACT_NUMBER,
                "signed_date": date(2025, 1, 1),
                "effective_from": date(2025, 1, 1),
                "effective_to": date(2027, 12, 31),
                "insurance_salary": Decimal("12000000"),
            },
        )
        await s.commit()
        return {"id": employee_id, "name": _TEST_EMPLOYEE_NAME, "status": "active"}


# ── Shared helpers ─────────────────────────────────────────────────────────────

async def _get_employee_with_basis() -> dict:
    """Trả về employee test có profile active và contract active với insurance_salary."""
    return await _ensure_employee_with_basis()


async def _reset_profile_to_active(employee_id: int) -> None:
    async with _make_session()() as s:
        await s.execute(
            text("UPDATE employee_insurance_profiles SET participation_status='active', status_effective_from=NULL, status_note=NULL WHERE employee_id=:eid"),
            {"eid": employee_id},
        )
        await s.commit()


async def _delete_change_events_for(employee_id: int) -> None:
    async with _make_session()() as s:
        await s.execute(
            text("DELETE FROM insurance_change_events WHERE employee_id=:eid"),
            {"eid": employee_id},
        )
        await s.commit()


async def _delete_change_events_for_period(year: int, month: int) -> None:
    async with _make_session()() as s:
        await s.execute(
            text("DELETE FROM insurance_change_events WHERE period_year=:y AND period_month=:m"),
            {"y": year, "m": month},
        )
        await s.commit()


@pytest.fixture(autouse=True)
async def ensure_test_employee_basis():
    await _ensure_employee_with_basis()
    yield


def _upsert_payload(
    participation_status: str = "active",
    status_effective_from: str | None = None,
) -> dict:
    return {
        "bhxh_code": None,
        "bhyt_initial_clinic_name": None,
        "bhyt_initial_clinic_code": None,
        "company_bhxh_joined_date": None,
        "participation_status": participation_status,
        "status_effective_from": status_effective_from,
        "status_note": None,
        "insurance_basis_source": "contract",
        "insurance_basis_amount": None,
        "component_overrides": [],
    }


# ═══════════════════════════════════════════════════════════════════════════════
# TestAutoRecordOnStatusChange
# ═══════════════════════════════════════════════════════════════════════════════

class TestAutoRecordOnStatusChange:

    @pytest.mark.asyncio
    async def test_active_to_stopped_creates_decrease_event(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _reset_profile_to_active(eid)
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            resp = client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("stopped", "2026-06-01"),
                headers=headers,
            )
            assert resp.status_code == 200, resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}", headers=headers).json()["items"]
            decrease_events = [e for e in events if e["change_type"] == "decrease"]
            assert len(decrease_events) >= 1
            ev = decrease_events[0]
            assert ev["change_reason"] == "resignation"
            assert ev["ibhxh_reason_code"] == "G-01"
            assert ev["old_status"] == "active"
            assert ev["new_status"] == "stopped"
            assert ev["is_manual"] is False
        finally:
            await _reset_profile_to_active(eid)
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_active_to_paused_creates_decrease_with_g03_code(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _reset_profile_to_active(eid)
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            resp = client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("paused"),
                headers=headers,
            )
            assert resp.status_code == 200, resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}&change_type=decrease", headers=headers).json()["items"]
            assert len(events) >= 1
            ev = events[0]
            assert ev["ibhxh_reason_code"] == "G-03"
            assert ev["change_reason"] == "unpaid_leave"
        finally:
            await _reset_profile_to_active(eid)
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_paused_to_active_creates_increase_with_t02_code(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            # Set to paused first (bypass status guard via direct SQL to avoid creating extra events)
            async with _make_session()() as s:
                await s.execute(
                    text("UPDATE employee_insurance_profiles SET participation_status='paused' WHERE employee_id=:eid"),
                    {"eid": eid},
                )
                await s.commit()

            resp = client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("active"),
                headers=headers,
            )
            assert resp.status_code == 200, resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}&change_type=increase", headers=headers).json()["items"]
            assert len(events) >= 1
            ev = events[0]
            assert ev["ibhxh_reason_code"] == "T-02"
            assert ev["change_reason"] == "return_from_leave"
        finally:
            await _reset_profile_to_active(eid)
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_no_event_when_status_unchanged(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _reset_profile_to_active(eid)
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            # PUT with same status (active → active) should NOT create event
            resp = client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("active"),
                headers=headers,
            )
            assert resp.status_code == 200, resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}", headers=headers).json()["items"]
            assert len(events) == 0
        finally:
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_increase_event_has_full_snapshot_fields(self, client: TestClient):
        """Mọi auto event tăng phải có đủ snapshot: tên, ngày sinh, basis_amount."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            # stopped → active triggers increase (T-02 return_from_leave)
            async with _make_session()() as s:
                await s.execute(
                    text("UPDATE employee_insurance_profiles SET participation_status='stopped' WHERE employee_id=:eid"),
                    {"eid": eid},
                )
                await s.commit()

            resp = client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("active"),
                headers=headers,
            )
            assert resp.status_code == 200, resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}&change_type=increase", headers=headers).json()["items"]
            assert len(events) >= 1
            ev = events[0]
            assert ev["change_type"] == "increase"
            assert ev["is_manual"] is False
            assert ev["employee_name_snapshot"] is not None and ev["employee_name_snapshot"] != ""
            assert ev["date_of_birth_snapshot"] is not None
            assert ev["basis_amount"] is not None
            assert float(ev["basis_amount"]) > 0
            assert ev["nationality_code_snapshot"] is not None
        finally:
            await _reset_profile_to_active(eid)
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_ibhxh_reason_code_matches_mapping_table(self, client: TestClient):
        """Tất cả auto events phải có ibhxh_reason_code thuộc tập T-0x / G-0x."""
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT DISTINCT ibhxh_reason_code FROM insurance_change_events
                WHERE is_manual = FALSE
            """))
            codes = {row[0] for row in r.fetchall()}

        valid_t = {f"T-0{i}" for i in range(1, 6)}
        valid_g = {f"G-0{i}" for i in range(1, 8)}
        for code in codes:
            assert code in (valid_t | valid_g), f"ibhxh_reason_code không hợp lệ: {code}"


# ═══════════════════════════════════════════════════════════════════════════════
# TestManualChangeEvents
# ═══════════════════════════════════════════════════════════════════════════════

class TestManualChangeEvents:

    @pytest.mark.asyncio
    async def test_create_manual_event_with_valid_payload(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            resp = client.post(
                f"{BASE}/change-events",
                json={
                    "employee_id": eid,
                    "change_type": "decrease",
                    "change_reason": "resignation",
                    "effective_date": "2026-06-15",
                    "note": "Test manual event",
                },
                headers=headers,
            )
            assert resp.status_code == 201, resp.text
            body = resp.json()
            assert body["employee_id"] == eid
            assert body["change_type"] == "decrease"
            assert body["change_reason"] == "resignation"
            assert body["ibhxh_reason_code"] == "G-01"
            assert body["is_manual"] is True
            assert body["note"] == "Test manual event"
            assert body["employee_name_snapshot"] is not None
            assert float(body["basis_amount"]) > 0
        finally:
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_delete_manual_event(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            create_resp = client.post(
                f"{BASE}/change-events",
                json={
                    "employee_id": eid,
                    "change_type": "increase",
                    "change_reason": "new_hire",
                    "effective_date": "2026-06-15",
                },
                headers=headers,
            )
            assert create_resp.status_code == 201, create_resp.text
            event_id = create_resp.json()["id"]

            del_resp = client.delete(f"{BASE}/change-events/{event_id}", headers=headers)
            assert del_resp.status_code == 204, del_resp.text

            events = client.get(f"{BASE}/change-events?employee_id={eid}", headers=headers).json()["items"]
            assert all(e["id"] != event_id for e in events)
        finally:
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_cannot_delete_auto_event_returns_409(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _reset_profile_to_active(eid)
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            # Trigger an auto event
            client.put(
                f"{BASE}/employees/{eid}",
                json=_upsert_payload("stopped", "2026-06-01"),
                headers=headers,
            )

            events = client.get(f"{BASE}/change-events?employee_id={eid}", headers=headers).json()["items"]
            auto_events = [e for e in events if not e["is_manual"]]
            assert len(auto_events) >= 1, "Không có auto event để test"
            auto_id = auto_events[0]["id"]

            del_resp = client.delete(f"{BASE}/change-events/{auto_id}", headers=headers)
            assert del_resp.status_code == 409, del_resp.text
        finally:
            await _reset_profile_to_active(eid)
            await _delete_change_events_for(eid)

    @pytest.mark.asyncio
    async def test_create_event_invalid_reason_type_mismatch(self, client: TestClient):
        """change_type=increase không hợp với change_reason=resignation → 422."""
        emp = await _get_employee_with_basis()
        headers = _admin(client)
        resp = client.post(
            f"{BASE}/change-events",
            json={
                "employee_id": emp["id"],
                "change_type": "increase",
                "change_reason": "resignation",
                "effective_date": "2026-06-15",
            },
            headers=headers,
        )
        assert resp.status_code == 422, resp.text


# ═══════════════════════════════════════════════════════════════════════════════
# TestMonthlyReport
# ═══════════════════════════════════════════════════════════════════════════════

class TestMonthlyReport:

    @pytest.mark.asyncio
    async def test_monthly_summary_counts_correctly(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)
        try:
            headers = _admin(client)
            # Tạo 2 increase + 1 decrease manual events
            for _ in range(2):
                client.post(f"{BASE}/change-events", json={
                    "employee_id": eid,
                    "change_type": "increase",
                    "change_reason": "new_hire",
                    "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
                }, headers=headers)
            client.post(f"{BASE}/change-events", json={
                "employee_id": eid,
                "change_type": "decrease",
                "change_reason": "resignation",
                "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
            }, headers=headers)

            summary_resp = client.get(
                f"{BASE}/change-events/monthly-summary?year={_TEST_YEAR}&month={_TEST_MONTH}",
                headers=headers,
            )
            assert summary_resp.status_code == 200, summary_resp.text
            summary = summary_resp.json()
            assert summary["increase_count"] == 2
            assert summary["decrease_count"] == 1
            assert float(summary["total_basis_increase"]) > 0
            assert float(summary["total_basis_decrease"]) > 0
        finally:
            await _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)

    @pytest.mark.asyncio
    async def test_filter_by_period_returns_correct_events(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        await _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)
        try:
            headers = _admin(client)
            client.post(f"{BASE}/change-events", json={
                "employee_id": eid,
                "change_type": "increase",
                "change_reason": "new_hire",
                "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
            }, headers=headers)

            resp = client.get(
                f"{BASE}/change-events?period_year={_TEST_YEAR}&period_month={_TEST_MONTH}",
                headers=headers,
            )
            assert resp.status_code == 200
            items = resp.json()["items"]
            assert len(items) >= 1
            for item in items:
                assert item["period_year"] == _TEST_YEAR
                assert item["period_month"] == _TEST_MONTH
        finally:
            await _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)

    def test_monthly_summary_empty_period_returns_zero_counts(self, client: TestClient):
        headers = _admin(client)
        # Year 2050 has no events
        resp = client.get(f"{BASE}/change-events/monthly-summary?year=2050&month=1", headers=headers)
        assert resp.status_code == 200
        summary = resp.json()
        assert summary["increase_count"] == 0
        assert summary["decrease_count"] == 0


# ═══════════════════════════════════════════════════════════════════════════════
# TestRefactoring (Slice 4a seed data integrity)
# ═══════════════════════════════════════════════════════════════════════════════

class TestRefactoring:

    @pytest.mark.asyncio
    async def test_ethnicity_bhxh_code_seeded_for_all_56_entries(self):
        async with _make_session()() as s:
            r = await s.execute(text("SELECT COUNT(*) FROM ethnicities WHERE bhxh_code IS NOT NULL"))
            count = r.scalar_one()
        assert count == 56, f"Kỳ vọng 56 dân tộc có bhxh_code, nhận {count}"

    @pytest.mark.asyncio
    async def test_ethnicity_kinh_has_code_01(self):
        async with _make_session()() as s:
            r = await s.execute(text("SELECT bhxh_code FROM ethnicities WHERE name = 'Kinh'"))
            row = r.fetchone()
        assert row is not None, "Không tìm thấy dân tộc 'Kinh'"
        assert row[0] == "01", f"Kinh phải có bhxh_code='01', nhận '{row[0]}'"

    @pytest.mark.asyncio
    async def test_nationality_seed_has_iso2_codes(self):
        async with _make_session()() as s:
            r = await s.execute(text("SELECT COUNT(*) FROM nationalities WHERE iso2_code IS NOT NULL"))
            count = r.scalar_one()
        assert count >= 200, f"Kỳ vọng ≥200 nationalities có iso2_code, nhận {count}"

    @pytest.mark.asyncio
    async def test_bhyt_clinics_table_has_expected_row_count(self):
        async with _make_session()() as s:
            r = await s.execute(text("SELECT COUNT(*) FROM bhyt_clinics"))
            count = r.scalar_one()
        assert count >= 10_000, f"Kỳ vọng ≥10,000 bhyt_clinics, nhận {count}"

    @pytest.mark.asyncio
    async def test_bhyt_clinics_code_is_unique(self):
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT COUNT(*) FROM (
                    SELECT code FROM bhyt_clinics GROUP BY code HAVING COUNT(*) > 1
                ) dups
            """))
            dup_count = r.scalar_one()
        assert dup_count == 0, f"bhyt_clinics.code không unique: {dup_count} trùng lặp"

    @pytest.mark.asyncio
    async def test_bhyt_clinics_province_name_populated(self):
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT COUNT(*) FROM bhyt_clinics
                WHERE province_code IS NOT NULL AND province_name IS NOT NULL
            """))
            count = r.scalar_one()
        assert count >= 5_000, f"Kỳ vọng ≥5,000 bhyt_clinics có province_name, nhận {count}"

    @pytest.mark.asyncio
    async def test_create_event_snapshots_ethnicity_bhxh_code(self, client: TestClient):
        """Khi tạo event thủ công, ethnicity_bhxh_code_snapshot phải được lấy từ DB."""
        # Tìm employee có ethnicity với bhxh_code
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT e.id, eth.bhxh_code
                FROM employees e
                JOIN ethnicities eth ON eth.id = e.ethnicity_id
                JOIN employee_insurance_profiles eip ON eip.employee_id = e.id
                JOIN employee_contracts ec ON ec.employee_id = e.id AND ec.status = 'active'
                WHERE eth.bhxh_code IS NOT NULL
                  AND eip.participation_status = 'active'
                  AND ec.insurance_salary IS NOT NULL AND ec.insurance_salary > 0
                ORDER BY e.id LIMIT 1
            """))
            row = r.fetchone()
        if row is None:
            pytest.skip("Không có employee với ethnicity bhxh_code trong DB test")

        eid, expected_code = row[0], row[1]
        await _delete_change_events_for(eid)
        try:
            headers = _admin(client)
            resp = client.post(f"{BASE}/change-events", json={
                "employee_id": eid,
                "change_type": "increase",
                "change_reason": "new_hire",
                "effective_date": "2026-07-01",
            }, headers=headers)
            assert resp.status_code == 201, resp.text
            assert resp.json()["ethnicity_bhxh_code_snapshot"] == expected_code
        finally:
            await _delete_change_events_for(eid)


# ═══════════════════════════════════════════════════════════════════════════════
# TestExport — VNPT D02-TS Excel — 3 sheets đầu tiên
# ═══════════════════════════════════════════════════════════════════════════════

def _parse_export_response(resp) -> openpyxl.Workbook:
    assert resp.status_code == 200, resp.text
    content_type = resp.headers.get("content-type", "")
    assert "spreadsheetml" in content_type, f"Unexpected content-type: {content_type}"
    return openpyxl.load_workbook(io.BytesIO(resp.content))


class TestExport:

    @pytest.fixture(autouse=True)
    def cleanup_test_period(self):
        """Xóa events test period trước và sau mỗi test."""
        asyncio.run(
            _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)
        )
        yield
        asyncio.run(
            _delete_change_events_for_period(_TEST_YEAR, _TEST_MONTH)
        )

    def _create_increase_event(self, client, headers, eid: int) -> dict:
        resp = client.post(f"{BASE}/change-events", json={
            "employee_id": eid,
            "change_type": "increase",
            "change_reason": "new_hire",
            "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
        }, headers=headers)
        assert resp.status_code == 201, resp.text
        return resp.json()

    def _create_decrease_event(self, client, headers, eid: int) -> dict:
        resp = client.post(f"{BASE}/change-events", json={
            "employee_id": eid,
            "change_type": "decrease",
            "change_reason": "resignation",
            "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
        }, headers=headers)
        assert resp.status_code == 201, resp.text
        return resp.json()

    def _export(self, client, headers) -> openpyxl.Workbook:
        resp = client.get(
            f"{BASE}/change-events/export/vnpt-d02-ts?period_year={_TEST_YEAR}&period_month={_TEST_MONTH}",
            headers=headers,
        )
        return _parse_export_response(resp)

    # ── Metadata ──────────────────────────────────────────────────────────────

    @pytest.mark.asyncio
    async def test_vnpt_d02_ts_returns_xlsx_content_type(self, client: TestClient):
        headers = _admin(client)
        resp = client.get(
            f"{BASE}/change-events/export/vnpt-d02-ts?period_year={_TEST_YEAR}&period_month={_TEST_MONTH}",
            headers=headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers.get("content-type", "")

    @pytest.mark.asyncio
    async def test_vnpt_d02_ts_has_11_sheets_with_lookup_sheets_intact(self, client: TestClient):
        headers = _admin(client)
        wb = self._export(client, headers)

        assert len(wb.sheetnames) == 11, f"Kỳ vọng 11 sheets, nhận {len(wb.sheetnames)}: {wb.sheetnames}"

        expected_sheets = {"Dữ Liệu", "Phụ lục", "Bảng kê", "Dân tộc", "QuocTich", "BenhVien", "Tinh", "TinhBenhVien", "Xa", "Mức hưởng BHYT", "Tổng hợp "}
        assert set(wb.sheetnames) == expected_sheets

        # Lookup sheets phải còn nguyên dữ liệu
        assert wb["Dân tộc"].max_row >= 2, "Sheet 'Dân tộc' bị mất data"
        assert wb["QuocTich"].max_row >= 2, "Sheet 'QuocTich' bị mất data"
        assert wb["BenhVien"].max_row >= 2, "Sheet 'BenhVien' bị mất data"

    # ── Sheet "Dữ Liệu" (sheet index 0) ──────────────────────────────────────

    @pytest.mark.asyncio
    async def test_du_lieu_sheet_empty_period_returns_header_only(self, client: TestClient):
        """Không có events → sheet 'Dữ Liệu' chỉ có 3 header rows, không có data row."""
        headers = _admin(client)
        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        # Sau khi clear, max_row = 3 (header rows)
        assert ws.max_row == 3, f"Kỳ vọng 3 header rows khi không có data, nhận max_row={ws.max_row}"

    @pytest.mark.asyncio
    async def test_du_lieu_row_count_matches_event_count(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)
        self._create_decrease_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        data_rows = ws.max_row - 3  # 3 header rows
        assert data_rows == 2, f"Kỳ vọng 2 data rows, nhận {data_rows}"

    @pytest.mark.asyncio
    async def test_du_lieu_increase_sorts_before_decrease(self, client: TestClient):
        """Tăng phải đứng trước Giảm (change_type DESC)."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_decrease_event(client, headers, eid)
        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        row1_type = ws.cell(row=4, column=4).value  # TenLoaiKhaiBao
        row2_type = ws.cell(row=5, column=4).value
        assert row1_type == "Tăng lao động", f"Row đầu phải là 'Tăng lao động', nhận '{row1_type}'"
        assert row2_type == "Giảm lao động", f"Row thứ 2 phải là 'Giảm lao động', nhận '{row2_type}'"

    @pytest.mark.asyncio
    async def test_du_lieu_loai_is_1_for_increase_3_for_decrease(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)
        self._create_decrease_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        # Increase comes first (sorted DESC)
        assert ws.cell(row=4, column=5).value == 1, "Loai tăng phải là 1"
        assert ws.cell(row=5, column=5).value == 3, "Loai giảm phải là 3"

    @pytest.mark.asyncio
    async def test_du_lieu_pa_codes_correct_per_reason(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)   # new_hire → PA="TM"
        self._create_decrease_event(client, headers, eid)   # resignation → PA="GH"

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        assert ws.cell(row=4, column=7).value == "TM", "new_hire PA phải là TM"
        assert ws.cell(row=5, column=7).value == "GH", "resignation PA phải là GH"

    @pytest.mark.asyncio
    async def test_du_lieu_ten_phuong_an_correct(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        assert ws.cell(row=4, column=6).value == "Tăng mới"

    @pytest.mark.asyncio
    async def test_du_lieu_tyle_dong_no_dot_decimal(self, client: TestClient):
        """TyleDong phải dùng dấu phẩy (không dùng dấu chấm) hoặc là số nguyên."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        tyle = ws.cell(row=4, column=20).value
        assert tyle is not None, "TyleDong (col20) không được để trống"
        tyle_str = str(tyle)
        assert "." not in tyle_str, f"TyleDong không được dùng dấu chấm: '{tyle_str}'"

    @pytest.mark.asyncio
    async def test_du_lieu_mavung_ltt_at_column_26(self, client: TestClient):
        """MavungLTT phải ở đúng cột 26 (không phải cột 50)."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        mavung = ws.cell(row=4, column=26).value
        assert mavung is not None, "MavungLTT (col26) không được để trống"
        assert mavung.startswith("0"), f"MavungLTT phải bắt đầu bằng '0x', nhận '{mavung}'"
        # Cột 50 phải trống (không bị nhầm)
        assert ws.cell(row=4, column=50).value is None, "Cột 50 phải trống (không phải MavungLTT)"

    @pytest.mark.asyncio
    async def test_du_lieu_mavung_ltt_matches_company_region(self, client: TestClient):
        """MavungLTT phải khớp với region trong company_bhxh_region."""
        async with _make_session()() as s:
            r = await s.execute(text("SELECT region FROM company_bhxh_region WHERE effective_to IS NULL ORDER BY effective_from DESC LIMIT 1"))
            row = r.fetchone()
        expected = f"0{row[0]}" if row else "01"

        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        mavung = ws.cell(row=4, column=26).value
        assert mavung == expected, f"MavungLTT kỳ vọng '{expected}', nhận '{mavung}'"

    @pytest.mark.asyncio
    async def test_du_lieu_dan_toc_code_at_column_35(self, client: TestClient):
        """DanToc (mã dân tộc) phải ở đúng cột 35 (không phải cột 34)."""
        # Tìm employee có ethnicity với bhxh_code
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT e.id, eth.bhxh_code
                FROM employees e
                JOIN ethnicities eth ON eth.id = e.ethnicity_id
                JOIN employee_insurance_profiles eip ON eip.employee_id = e.id
                JOIN employee_contracts ec ON ec.employee_id = e.id AND ec.status = 'active'
                WHERE eth.bhxh_code IS NOT NULL
                  AND eip.participation_status = 'active'
                  AND ec.insurance_salary > 0
                ORDER BY e.id LIMIT 1
            """))
            row = r.fetchone()
        if row is None:
            pytest.skip("Không có employee với ethnicity bhxh_code")

        eid, expected_code = row[0], row[1]
        # Xóa test period events cho employee này
        async with _make_session()() as s:
            await s.execute(text("DELETE FROM insurance_change_events WHERE employee_id=:eid AND period_year=:y AND period_month=:m"), {"eid": eid, "y": _TEST_YEAR, "m": _TEST_MONTH})
            await s.commit()

        headers = _admin(client)
        resp = client.post(f"{BASE}/change-events", json={
            "employee_id": eid,
            "change_type": "increase",
            "change_reason": "new_hire",
            "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
        }, headers=headers)
        assert resp.status_code == 201

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        dan_toc_col35 = ws.cell(row=4, column=35).value
        dan_toc_col34 = ws.cell(row=4, column=34).value

        assert dan_toc_col35 == expected_code, f"DanToc code ở col35 kỳ vọng '{expected_code}', nhận '{dan_toc_col35}'"
        # col34 là TenDanToc (name) — chúng ta chưa fill nên phải là None
        assert dan_toc_col34 is None, f"TenDanToc (col34) phải trống, nhận '{dan_toc_col34}'"

    @pytest.mark.asyncio
    async def test_du_lieu_gender_male_is_1_female_is_0(self, client: TestClient):
        # Tìm employee male để test
        async with _make_session()() as s:
            r = await s.execute(text("""
                SELECT e.id FROM employees e
                JOIN employee_insurance_profiles eip ON eip.employee_id = e.id
                JOIN employee_contracts ec ON ec.employee_id = e.id AND ec.status = 'active'
                WHERE e.gender = 'male' AND eip.participation_status = 'active'
                  AND ec.insurance_salary > 0
                ORDER BY e.id LIMIT 1
            """))
            row = r.fetchone()
        if row is None:
            pytest.skip("Không có male employee")

        eid = row[0]
        async with _make_session()() as s:
            await s.execute(text("DELETE FROM insurance_change_events WHERE employee_id=:eid AND period_year=:y AND period_month=:m"), {"eid": eid, "y": _TEST_YEAR, "m": _TEST_MONTH})
            await s.commit()

        headers = _admin(client)
        client.post(f"{BASE}/change-events", json={
            "employee_id": eid,
            "change_type": "increase",
            "change_reason": "new_hire",
            "effective_date": f"{_TEST_YEAR}-{_TEST_MONTH:02d}-01",
        }, headers=headers)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        assert ws.cell(row=4, column=31).value == 1, "Male GioiTinh phải là 1"

    @pytest.mark.asyncio
    async def test_du_lieu_missing_clinic_code_does_not_crash(self, client: TestClient):
        """Employee không có bhyt_clinic_code → export vẫn thành công, MaBenhVien để trống."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        # Đảm bảo employee không có clinic code
        async with _make_session()() as s:
            await s.execute(
                text("UPDATE employee_insurance_profiles SET bhyt_initial_clinic_code=NULL, bhyt_initial_clinic_name=NULL WHERE employee_id=:eid"),
                {"eid": eid},
            )
            await s.commit()

        self._create_increase_event(client, headers, eid)

        resp = client.get(
            f"{BASE}/change-events/export/vnpt-d02-ts?period_year={_TEST_YEAR}&period_month={_TEST_MONTH}",
            headers=headers,
        )
        assert resp.status_code == 200, "Export phải thành công kể cả khi không có clinic code"
        wb = _parse_export_response(resp)
        ws = wb["Dữ Liệu"]
        # MaBenhVien (col39) để trống
        assert ws.cell(row=4, column=39).value is None

    @pytest.mark.asyncio
    async def test_du_lieu_stt_sequential_starting_from_1(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)
        self._create_decrease_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=5, column=1).value == 2

    @pytest.mark.asyncio
    async def test_du_lieu_basis_amount_written_as_integer(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        ev = self._create_increase_event(client, headers, eid)
        expected_basis = int(float(ev["basis_amount"]))

        wb = self._export(client, headers)
        ws = wb["Dữ Liệu"]
        assert ws.cell(row=4, column=10).value == expected_basis

    # ── Sheet "Phụ lục" (sheet index 1) ──────────────────────────────────────

    @pytest.mark.asyncio
    async def test_phu_luc_sheet_untouched_no_new_data_rows(self, client: TestClient):
        """Sheet 'Phụ lục' (thành viên hộ gia đình) không thuộc 6.3 → không được thêm data."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)
        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Phụ lục"]
        # Template có row 4 là ví dụ mẫu → sau export không được có rows mới hơn
        # (export service không chạm vào Phụ lục)
        assert "Phụ lục" in wb.sheetnames
        # Kiểm tra header row 3 còn nguyên (có 'STT' ở col1)
        assert ws.cell(row=3, column=1).value == "STT", "Header 'Phụ lục' bị mất"

    # ── Sheet "Bảng kê" (sheet index 2) ───────────────────────────────────────

    @pytest.mark.asyncio
    async def test_bang_ke_empty_period_has_no_data_rows(self, client: TestClient):
        headers = _admin(client)
        wb = self._export(client, headers)
        ws = wb["Bảng kê"]
        assert ws.max_row == 3, f"Bảng kê kỳ vọng 3 header rows khi trống, nhận {ws.max_row}"

    @pytest.mark.asyncio
    async def test_bang_ke_row_count_matches_event_count(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)
        self._create_decrease_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Bảng kê"]
        data_rows = ws.max_row - 3
        assert data_rows == 2, f"Bảng kê kỳ vọng 2 data rows, nhận {data_rows}"

    @pytest.mark.asyncio
    async def test_bang_ke_ten_loai_van_ban_is_hop_dong(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Bảng kê"]
        assert ws.cell(row=4, column=4).value == "Hợp đồng lao động"

    @pytest.mark.asyncio
    async def test_bang_ke_ho_ten_and_bhxh_code_match_du_lieu(self, client: TestClient):
        """HoTen và MasoBHXH ở Bảng kê phải khớp với Dữ Liệu."""
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)

        wb = self._export(client, headers)
        ws_dl = wb["Dữ Liệu"]
        ws_bk = wb["Bảng kê"]

        assert ws_bk.cell(row=4, column=2).value == ws_dl.cell(row=4, column=2).value
        assert ws_bk.cell(row=4, column=3).value == ws_dl.cell(row=4, column=3).value

    @pytest.mark.asyncio
    async def test_bang_ke_stt_sequential(self, client: TestClient):
        emp = await _get_employee_with_basis()
        eid = emp["id"]
        headers = _admin(client)

        self._create_increase_event(client, headers, eid)
        self._create_decrease_event(client, headers, eid)

        wb = self._export(client, headers)
        ws = wb["Bảng kê"]
        assert ws.cell(row=4, column=1).value == 1
        assert ws.cell(row=5, column=1).value == 2
