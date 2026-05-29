"""Dịch vụ xuất biểu mẫu BHXH — D02-LT (danh sách lao động) và D03-TS (tổng hợp mức đóng)."""
from __future__ import annotations

from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.models.employee import Employee
from app.models.employee_insurance import EmployeeInsuranceProfile
from app.models.employee_job import EmployeeJobRecord
from app.models.org import Department, JobTitle


# ── Helpers ───────────────────────────────────────────────────────────────────

_GENDER_MAP = {"male": "Nam", "female": "Nữ", "other": "Khác"}
_STATUS_MAP = {"active": "Đang tham gia", "paused": "Tạm dừng", "stopped": "Dừng"}

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT_WHITE = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def _fmt_date(d: object) -> str:
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")  # type: ignore[attr-defined]


def _set_col_widths(ws: object, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w  # type: ignore[index]


def _write_header_row(ws: object, row: int, headers: list[str]) -> None:
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)  # type: ignore[union-attr]
        cell.value = title
        cell.font = _HEADER_FONT_WHITE
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── D02-LT ────────────────────────────────────────────────────────────────────

_D02_HEADERS = [
    "STT", "Họ và tên", "Ngày sinh", "Giới tính", "Số CCCD/CMND",
    "Số sổ BHXH", "Mã bệnh viện KCB", "Chức danh",
    "Mức lương đóng BHXH", "Ngày tham gia", "Trạng thái",
]
_D02_WIDTHS = [6, 22, 14, 10, 18, 16, 18, 22, 20, 14, 14]


async def build_d02_lt(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
) -> bytes:
    """Xuất D02-LT — danh sách lao động tham gia BHXH, BHYT, BHTN."""

    # 1. Query
    stmt = (
        select(Employee, EmployeeInsuranceProfile, JobTitle, Department)
        .join(EmployeeInsuranceProfile, EmployeeInsuranceProfile.employee_id == Employee.id)
        .join(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
            isouter=True,
        )
        .join(JobTitle, JobTitle.id == EmployeeJobRecord.job_title_id, isouter=True)
        .join(Department, Department.id == EmployeeJobRecord.department_id, isouter=True)
        .where(EmployeeInsuranceProfile.participation_status == "active")
        .order_by(Employee.full_name)
    )
    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    rows = (await session.execute(stmt)).all()

    # 2. Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "D02-LT"

    # Row 1 — tiêu đề
    ws.merge_cells("A1:K1")
    c1 = ws["A1"]
    c1.value = "DANH SÁCH LAO ĐỘNG THAM GIA BHXH, BHYT, BHTN"
    c1.font = Font(bold=True, size=14)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — đơn vị / tháng
    ws["A2"].value = f"Đơn vị: {settings.COMPANY_NAME}"
    ws["G2"].value = f"Tháng {month}/{year}"

    # Row 3 — trống
    # Row 4 — header
    _write_header_row(ws, 4, _D02_HEADERS)

    # Row 5+ — data
    for idx, (emp, ins, jt, dept) in enumerate(rows, start=1):
        r = 4 + idx
        ws.cell(row=r, column=1).value = idx
        ws.cell(row=r, column=2).value = emp.full_name
        ws.cell(row=r, column=3).value = _fmt_date(emp.date_of_birth)
        ws.cell(row=r, column=4).value = _GENDER_MAP.get(emp.gender or "", "")
        ws.cell(row=r, column=5).value = emp.id_number
        ws.cell(row=r, column=6).value = ins.bhxh_code
        ws.cell(row=r, column=7).value = ins.bhyt_initial_clinic_code
        ws.cell(row=r, column=8).value = jt.name if jt else ""
        ws.cell(row=r, column=9).value = (
            int(ins.insurance_basis_amount) if ins.insurance_basis_amount is not None else 0
        )
        ws.cell(row=r, column=10).value = _fmt_date(ins.company_bhxh_joined_date)
        ws.cell(row=r, column=11).value = _STATUS_MAP.get(ins.participation_status, ins.participation_status)

    _set_col_widths(ws, _D02_WIDTHS)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── D03-TS ────────────────────────────────────────────────────────────────────

_D03_HEADERS = [
    "STT", "Họ và tên", "Mức lương đóng",
    "BHXH NLĐ (8%)", "BHYT NLĐ (1.5%)", "BHTN NLĐ (1%)",
    "BHXH NSDLĐ (17%)", "BHYT NSDLĐ (3%)", "BHTN NSDLĐ (1%)",
    "Tổng NLĐ đóng", "Tổng NSDLĐ đóng",
]
_D03_WIDTHS = [6, 22, 18, 16, 18, 14, 20, 16, 14, 18, 20]

_TWO = Decimal("0.01")


def _calc(basis: Decimal, rate_pct: Decimal) -> Decimal:
    return (basis * rate_pct / Decimal("100")).quantize(_TWO, rounding=ROUND_HALF_UP)


async def build_d03_ts(
    session: AsyncSession,
    *,
    year: int,
    month: int,
) -> bytes:
    """Xuất D03-TS — bảng tổng hợp mức đóng BHXH, BHYT, BHTN."""

    # 1. Query
    stmt = (
        select(Employee, EmployeeInsuranceProfile)
        .join(EmployeeInsuranceProfile, EmployeeInsuranceProfile.employee_id == Employee.id)
        .where(EmployeeInsuranceProfile.participation_status == "active")
        .order_by(Employee.full_name)
    )
    rows = (await session.execute(stmt)).all()

    # 2. Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "D03-TS"

    # Row 1
    ws.merge_cells("A1:K1")
    c1 = ws["A1"]
    c1.value = "BẢNG TỔNG HỢP MỨC ĐÓNG BHXH, BHYT, BHTN"
    c1.font = Font(bold=True, size=14)
    c1.alignment = Alignment(horizontal="center", vertical="center")

    # Row 2
    ws["A2"].value = f"Đơn vị: {settings.COMPANY_NAME}"
    ws["G2"].value = f"Tháng {month}/{year}"

    # Row 3 — trống
    # Row 4 — header
    _write_header_row(ws, 4, _D03_HEADERS)

    # Row 5+ — data
    totals = [Decimal("0")] * 9  # col 3..11 (0-indexed: 0..8)

    for idx, (emp, ins) in enumerate(rows, start=1):
        r = 4 + idx
        basis = ins.insurance_basis_amount if ins.insurance_basis_amount is not None else Decimal("0")

        bhxh_nld   = _calc(basis, Decimal("8"))
        bhyt_nld   = _calc(basis, Decimal("1.5"))
        bhtn_nld   = _calc(basis, Decimal("1"))
        bhxh_nsdld = _calc(basis, Decimal("17"))
        bhyt_nsdld = _calc(basis, Decimal("3"))
        bhtn_nsdld = _calc(basis, Decimal("1"))
        tong_nld   = bhxh_nld + bhyt_nld + bhtn_nld
        tong_nsdld = bhxh_nsdld + bhyt_nsdld + bhtn_nsdld

        values = [basis, bhxh_nld, bhyt_nld, bhtn_nld, bhxh_nsdld, bhyt_nsdld, bhtn_nsdld, tong_nld, tong_nsdld]

        ws.cell(row=r, column=1).value = idx
        ws.cell(row=r, column=2).value = emp.full_name
        for col_offset, val in enumerate(values):
            ws.cell(row=r, column=3 + col_offset).value = float(val)

        for i, val in enumerate(values):
            totals[i] += val

    # Tổng cộng row
    total_row = 4 + len(rows) + 1
    tc = ws.cell(row=total_row, column=2)
    tc.value = "TỔNG CỘNG"
    tc.font = _BOLD_FONT
    for col_offset, val in enumerate(totals):
        cell = ws.cell(row=total_row, column=3 + col_offset)
        cell.value = float(val)
        cell.font = _BOLD_FONT

    _set_col_widths(ws, _D03_WIDTHS)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
