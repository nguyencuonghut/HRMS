from __future__ import annotations

import io
from datetime import date
from typing import Optional

from fastapi import HTTPException, status
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import LeaveType
from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.leave_entitlement import LeaveEntitlement
from app.models.leave_record import LeaveRecord
from app.models.org import Department
from app.schemas.leave_report import (
    DepartmentSummaryReport,
    DepartmentSummaryRow,
    EmployeeSummaryReport,
    EmployeeSummaryRow,
    YearEndReport,
    YearEndRow,
)
from app.services.employee_service import compute_display_code


# ── remaining_days helper (mirrors LeaveEntitlementRead.remaining_days) ────────

def _remaining(allocated: float, carryover: float, used: float, expires: Optional[date]) -> float:
    if expires and date.today() > expires:
        used_from_regular = max(0.0, used - carryover)
        return allocated - used_from_regular
    return allocated + carryover - used


# ── Báo cáo A — Chi tiết nhân viên ────────────────────────────────────────────

async def get_employee_summary(
    session: AsyncSession,
    *,
    year: int,
    employee_id: Optional[int] = None,
    department_id: Optional[int] = None,
    leave_type_id: Optional[int] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
) -> EmployeeSummaryReport:
    # Subquery: đếm leave_records active per entitlement
    record_count_sq = (
        select(
            LeaveRecord.entitlement_id,
            func.count(LeaveRecord.id).label("cnt"),
        )
        .where(LeaveRecord.status == "active", LeaveRecord.entitlement_id.is_not(None))
        .group_by(LeaveRecord.entitlement_id)
        .subquery()
    )

    stmt = (
        select(
            LeaveEntitlement,
            Employee,
            LeaveType,
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            func.coalesce(record_count_sq.c.cnt, 0).label("record_count"),
        )
        .join(Employee, LeaveEntitlement.employee_id == Employee.id)
        .join(LeaveType, LeaveEntitlement.leave_type_id == LeaveType.id)
        .outerjoin(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(record_count_sq, record_count_sq.c.entitlement_id == LeaveEntitlement.id)
        .where(LeaveEntitlement.year == year)
    )

    filters = []
    if employee_id is not None:
        filters.append(LeaveEntitlement.employee_id == employee_id)
    if department_id is not None:
        filters.append(EmployeeJobRecord.department_id == department_id)
    if leave_type_id is not None:
        filters.append(LeaveEntitlement.leave_type_id == leave_type_id)
    if keyword:
        kw = f"%{keyword.strip()}%"
        filters.append(Employee.full_name.ilike(kw))

    if filters:
        stmt = stmt.where(*filters)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await session.execute(count_stmt)).scalar_one()

    stmt = (
        stmt
        .order_by(Employee.employee_seq, LeaveType.code)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await session.execute(stmt)).all()

    items: list[EmployeeSummaryRow] = []
    for ent, emp, lt, dept_id, dept_name, rc in rows:
        allocated = float(ent.allocated_days)
        carryover = float(ent.carryover_days)
        used      = float(ent.used_days)
        remaining = _remaining(allocated, carryover, used, ent.carryover_expires)
        items.append(EmployeeSummaryRow(
            employee_id=emp.id,
            employee_code=compute_display_code(emp.employee_seq),
            employee_name=emp.full_name,
            department_name=dept_name,
            leave_type_id=lt.id,
            leave_type_name=lt.name,
            leave_type_code=lt.code,
            allocated_days=allocated,
            carryover_days=carryover,
            carryover_expires=ent.carryover_expires,
            used_days=used,
            remaining_days=remaining,
            record_count=rc,
        ))

    return EmployeeSummaryReport(year=year, items=items, total=total, page=page, page_size=page_size)


# ── Báo cáo B — Tổng hợp phòng ban ───────────────────────────────────────────

async def get_department_summary(
    session: AsyncSession,
    *,
    year: int,
    month_from: Optional[int] = None,
    month_to: Optional[int] = None,
    department_id: Optional[int] = None,
    leave_type_id: Optional[int] = None,
) -> DepartmentSummaryReport:
    if month_from and month_to and month_from > month_to:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="month_from phải ≤ month_to",
        )

    stmt = (
        select(
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            LeaveType.id.label("lt_id"),
            LeaveType.name.label("lt_name"),
            func.count(LeaveRecord.id.distinct()).label("total_records"),
            func.count(LeaveRecord.employee_id.distinct()).label("employee_count"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_days"),
        )
        .join(Employee, LeaveRecord.employee_id == Employee.id)
        .outerjoin(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .join(LeaveType, LeaveRecord.leave_type_id == LeaveType.id)
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(Department.id, Department.name, LeaveType.id, LeaveType.name)
        .order_by(Department.name.nulls_last(), LeaveType.name)
    )

    filters = []
    if month_from is not None:
        filters.append(func.extract("month", LeaveRecord.start_date) >= month_from)
    if month_to is not None:
        filters.append(func.extract("month", LeaveRecord.start_date) <= month_to)
    if department_id is not None:
        filters.append(EmployeeJobRecord.department_id == department_id)
    if leave_type_id is not None:
        filters.append(LeaveRecord.leave_type_id == leave_type_id)

    if filters:
        stmt = stmt.where(*filters)

    rows = (await session.execute(stmt)).all()

    items: list[DepartmentSummaryRow] = []
    for dept_id, dept_name, lt_id, lt_name, total_records, emp_count, total_days in rows:
        total_days_f = float(total_days)
        avg = round(total_days_f / emp_count, 2) if emp_count else 0.0
        items.append(DepartmentSummaryRow(
            department_id=dept_id,
            department_name=dept_name,
            leave_type_id=lt_id,
            leave_type_name=lt_name,
            employee_count=emp_count,
            total_records=total_records,
            total_days_taken=total_days_f,
            avg_days_per_employee=avg,
        ))

    return DepartmentSummaryReport(
        year=year,
        month_from=month_from,
        month_to=month_to,
        items=items,
    )


# ── Báo cáo C — Tồn phép cuối năm ────────────────────────────────────────────

async def get_year_end(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 50,
) -> YearEndReport:
    stmt = (
        select(
            LeaveEntitlement,
            Employee,
            LeaveType,
            Department.name.label("dept_name"),
        )
        .join(Employee, LeaveEntitlement.employee_id == Employee.id)
        .join(LeaveType, LeaveEntitlement.leave_type_id == LeaveType.id)
        .outerjoin(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            LeaveEntitlement.year == year,
            LeaveType.carryover_allowed == True,
        )
    )

    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await session.execute(count_stmt)).scalar_one()

    stmt = (
        stmt
        .order_by(Employee.employee_seq, LeaveType.code)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await session.execute(stmt)).all()

    items: list[YearEndRow] = []
    for ent, emp, lt, dept_name in rows:
        allocated = float(ent.allocated_days)
        carryover = float(ent.carryover_days)
        used      = float(ent.used_days)
        remaining = _remaining(allocated, carryover, used, ent.carryover_expires)
        items.append(YearEndRow(
            employee_id=emp.id,
            employee_code=compute_display_code(emp.employee_seq),
            employee_name=emp.full_name,
            department_name=dept_name,
            leave_type_name=lt.name,
            leave_type_code=lt.code,
            allocated_days=allocated,
            carryover_days=carryover,
            used_days=used,
            remaining_days=remaining,
            will_carry=max(0.0, remaining),
        ))

    return YearEndReport(year=year, items=items, total=total, page=page, page_size=page_size)


# ── Export Excel ──────────────────────────────────────────────────────────────

def _build_xlsx(title: str, subtitle: str, headers: list[str], rows: list[list]) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"

    # Row 1 — tiêu đề
    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    # Row 2 — phụ đề
    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    sub_cell = ws["A2"]
    sub_cell.value = subtitle
    sub_cell.font = Font(italic=True, size=10, color="555555")
    sub_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Row 3 — headers
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 28

    # Data rows
    alt_fill = PatternFill("solid", fgColor="EEF2FF")
    for row_idx, row in enumerate(rows, start=4):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center")

    # Auto column width
    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(4, len(rows) + 4):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_employee_summary_xlsx(report: EmployeeSummaryReport) -> io.BytesIO:
    today = date.today().strftime("%d/%m/%Y")
    headers = [
        "Mã NV", "Họ và tên", "Phòng ban", "Loại phép",
        "Cấp phép", "Chuyển dư", "Đã dùng", "Còn lại", "Số lần",
    ]
    rows = [
        [
            r.employee_code, r.employee_name, r.department_name or "—", r.leave_type_name,
            r.allocated_days, r.carryover_days, r.used_days, r.remaining_days, r.record_count,
        ]
        for r in report.items
    ]
    return _build_xlsx(
        f"Báo cáo chi tiết nghỉ phép năm {report.year}",
        f"Năm: {report.year} · Xuất ngày: {today}",
        headers, rows,
    )


def export_department_summary_xlsx(report: DepartmentSummaryReport) -> io.BytesIO:
    today = date.today().strftime("%d/%m/%Y")
    period = f"{report.month_from or 1}–{report.month_to or 12}/{report.year}"
    headers = ["Phòng ban", "Loại phép", "Số NV", "Số lần", "Tổng ngày", "TB/NV"]
    rows = [
        [
            r.department_name or "—", r.leave_type_name,
            r.employee_count, r.total_records, r.total_days_taken, r.avg_days_per_employee,
        ]
        for r in report.items
    ]
    return _build_xlsx(
        f"Báo cáo tổng hợp nghỉ phép theo phòng ban — {period}",
        f"Kỳ: {period} · Xuất ngày: {today}",
        headers, rows,
    )


def export_year_end_xlsx(report: YearEndReport) -> io.BytesIO:
    today = date.today().strftime("%d/%m/%Y")
    headers = [
        "Mã NV", "Họ và tên", "Phòng ban", "Loại phép",
        "Cấp phép", "Chuyển dư", "Đã dùng", "Còn lại", "Sẽ chuyển",
    ]
    rows = [
        [
            r.employee_code, r.employee_name, r.department_name or "—", r.leave_type_name,
            r.allocated_days, r.carryover_days, r.used_days, r.remaining_days, r.will_carry,
        ]
        for r in report.items
    ]
    return _build_xlsx(
        f"Báo cáo tồn phép cuối năm {report.year}",
        f"Năm: {report.year} · Xuất ngày: {today}",
        headers, rows,
    )
