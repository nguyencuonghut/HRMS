"""Service báo cáo nhân sự (11.2)."""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, timedelta
from io import BytesIO
from math import ceil
from typing import Optional

import sqlalchemy as sa
from fastapi import HTTPException, status
from sqlalchemy import and_, func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import ContractCategory
from app.models.employee import Employee
from app.models.employee_contract import EmployeeContract
from app.models.employee_job import EmployeeJobRecord
from app.models.org import Department, JobTitle
from app.models.salary import RetirementAgePolicy, RetirementAgePolicyThreshold
from app.schemas.hr_report import (
    DepartmentNode,
    EmployeeListItem,
    EmployeeListResponse,
    JobTitleHeadcount,
    MovementPeriodRow,
    MovementReportResponse,
    OlderWorkerListItem,
    OlderWorkerReportResponse,
    OlderWorkerSummary,
    OrgStructureResponse,
    RetirementAgePoliciesRead,
    RetirementAgePolicyCreate,
    RetirementAgePolicyRead,
    RetirementAgeThresholdRead,
    RetirementAgeThresholdInput,
    RetirementAgePolicyUpdate,
    TenureGroup,
    TenureGroupDetail,
    TenureReportResponse,
)
from app.services import employee_code_service

TENURE_GROUPS: list[tuple[str, str, int, Optional[int]]] = [
    ("lt_1", "< 1 năm", 0, 1),
    ("1_3", "1–3 năm", 1, 3),
    ("3_5", "3–5 năm", 3, 5),
    ("5_10", "5–10 năm", 5, 10),
    ("gt_10", "> 10 năm", 10, None),
]

_OLDER_WORKER_SUPPORTED_GENDERS = {"male", "female"}


@dataclass
class _Period:
    start: date
    end: date


@dataclass(frozen=True)
class _RetirementAgeThresholdValue:
    years: int
    months: int


def _tenure_years(start_date: date, end_date: Optional[date] = None) -> int:
    effective_end = end_date or date.today()
    return max((effective_end - start_date).days // 365, 0)


def _round1(value: float) -> float:
    return round(value + 1e-9, 1)


def _save_workbook_bytes(workbook) -> BytesIO:
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _last_day_of_month(year: int, month: int) -> date:
    if month < 1 or month > 12:
        raise ValueError("month must be between 1 and 12")
    if month == 12:
        return date(year, 12, 31)
    return date(year, month + 1, 1) - timedelta(days=1)


def _add_months(value: date, months: int) -> date:
    total_month = (value.month - 1) + months
    year = value.year + total_month // 12
    month = total_month % 12 + 1
    next_month = date(year + (1 if month == 12 else 0), 1 if month == 12 else month + 1, 1)
    last_day = (next_month - timedelta(days=1)).day
    day = min(value.day, last_day)
    return date(year, month, day)


def _diff_years_months(start: date, end: date) -> tuple[int, int]:
    years = end.year - start.year
    months = end.month - start.month
    if end.day < start.day:
        months -= 1
    if months < 0:
        years -= 1
        months += 12
    return max(years, 0), max(months, 0)


def _diff_total_months(start: date, end: date) -> int:
    years, months = _diff_years_months(start, end)
    return years * 12 + months


def _threshold_label(value: _RetirementAgeThresholdValue | None) -> str | None:
    if value is None:
        return None
    return f"{value.years} tuổi {value.months} tháng"


def _normalize_thresholds(
    thresholds: list[RetirementAgeThresholdInput],
) -> list[RetirementAgeThresholdInput]:
    seen: set[tuple[str, int]] = set()
    normalized: list[RetirementAgeThresholdInput] = []
    for item in sorted(thresholds, key=lambda row: (row.gender, row.applicable_year)):
        key = (item.gender, item.applicable_year)
        if key in seen:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Trùng ngưỡng tuổi nghỉ hưu cho giới tính '{item.gender}' năm {item.applicable_year}",
            )
        seen.add(key)
        normalized.append(item)

    missing_genders = _OLDER_WORKER_SUPPORTED_GENDERS - {item.gender for item in normalized}
    if missing_genders:
        genders = ", ".join(sorted(missing_genders))
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Thiếu ngưỡng tuổi nghỉ hưu cho giới tính: {genders}",
        )
    return normalized


def _period_label(period_start: date, group_by: str) -> str:
    if group_by == "month":
        return period_start.strftime("%Y-%m")
    if group_by == "quarter":
        quarter = (period_start.month - 1) // 3 + 1
        return f"Q{quarter}/{period_start.year}"
    return str(period_start.year)


def _period_end(period_start: date, group_by: str) -> date:
    if group_by == "month":
        if period_start.month == 12:
            return date(period_start.year, 12, 31)
        return date(period_start.year, period_start.month + 1, 1) - timedelta(days=1)
    if group_by == "quarter":
        month = period_start.month + 3
        year = period_start.year
        if month > 12:
            month -= 12
            year += 1
        return date(year, month, 1) - timedelta(days=1)
    return date(period_start.year, 12, 31)


def _next_period(period_start: date, group_by: str) -> date:
    if group_by == "month":
        if period_start.month == 12:
            return date(period_start.year + 1, 1, 1)
        return date(period_start.year, period_start.month + 1, 1)
    if group_by == "quarter":
        month = period_start.month + 3
        year = period_start.year
        if month > 12:
            month -= 12
            year += 1
        return date(year, month, 1)
    return date(period_start.year + 1, 1, 1)


def _period_floor(value: date, group_by: str) -> date:
    if group_by == "month":
        return date(value.year, value.month, 1)
    if group_by == "quarter":
        quarter_month = ((value.month - 1) // 3) * 3 + 1
        return date(value.year, quarter_month, 1)
    return date(value.year, 1, 1)


def _iter_periods(start_date: date, end_date: date, group_by: str) -> list[_Period]:
    periods: list[_Period] = []
    cursor = _period_floor(start_date, group_by)
    limit = _period_floor(end_date, group_by)
    while cursor <= limit:
        periods.append(_Period(start=cursor, end=min(_period_end(cursor, group_by), end_date)))
        cursor = _next_period(cursor, group_by)
    return periods


async def _department_name(
    session: AsyncSession,
    department_id: Optional[int],
) -> Optional[str]:
    if department_id is None:
        return None
    return await session.scalar(
        select(Department.name).where(Department.id == department_id),
    )


async def _department_scope_ids(
    session: AsyncSession,
    department_id: Optional[int],
) -> Optional[set[int]]:
    if department_id is None:
        return None
    rows = await session.execute(
        sa.text(
            """
            WITH RECURSIVE subtree AS (
                SELECT id
                FROM departments
                WHERE id = :department_id AND is_active = TRUE
                UNION ALL
                SELECT d.id
                FROM departments d
                JOIN subtree s ON d.parent_id = s.id
                WHERE d.is_active = TRUE
            )
            SELECT id FROM subtree
            """
        ),
        {"department_id": department_id},
    )
    return {int(row[0]) for row in rows.fetchall()}


def _group_members(scope_ids: Optional[set[int]], direct_department_id: Optional[int]) -> bool:
    if scope_ids is None:
        return True
    if direct_department_id is None:
        return False
    return direct_department_id in scope_ids


async def _policy_to_read(
    session: AsyncSession,
    policy: RetirementAgePolicy,
) -> RetirementAgePolicyRead:
    rows = await session.execute(
        select(RetirementAgePolicyThreshold)
        .where(RetirementAgePolicyThreshold.policy_id == policy.id)
        .order_by(
            RetirementAgePolicyThreshold.gender.asc(),
            RetirementAgePolicyThreshold.applicable_year.asc(),
            RetirementAgePolicyThreshold.id.asc(),
        )
    )
    thresholds = [
        RetirementAgeThresholdRead(
            id=item.id,
            gender=item.gender,
            applicable_year=item.applicable_year,
            age_years=item.age_years,
            age_months=item.age_months,
        )
        for item in rows.scalars().all()
    ]
    return RetirementAgePolicyRead(
        id=policy.id,
        name=policy.name,
        legal_basis_summary=policy.legal_basis_summary,
        effective_from=policy.effective_from,
        effective_to=policy.effective_to,
        note=policy.note,
        thresholds=thresholds,
        created_at=policy.created_at,
        updated_at=policy.updated_at,
    )


async def _replace_policy_thresholds(
    session: AsyncSession,
    *,
    policy_id: int,
    thresholds: list[RetirementAgeThresholdInput],
) -> None:
    normalized = _normalize_thresholds(thresholds)
    await session.execute(
        sa.delete(RetirementAgePolicyThreshold).where(
            RetirementAgePolicyThreshold.policy_id == policy_id
        )
    )
    await session.flush()
    for item in normalized:
        session.add(
            RetirementAgePolicyThreshold(
                policy_id=policy_id,
                gender=item.gender,
                applicable_year=item.applicable_year,
                age_years=item.age_years,
                age_months=item.age_months,
            )
        )


async def get_retirement_age_policies(session: AsyncSession) -> RetirementAgePoliciesRead:
    rows = await session.execute(
        select(RetirementAgePolicy).order_by(
            RetirementAgePolicy.effective_from.desc(),
            RetirementAgePolicy.id.desc(),
        )
    )
    items = list(rows.scalars().all())
    history = [await _policy_to_read(session, item) for item in items]
    current = next((item for item in history if item.effective_to is None), None)
    return RetirementAgePoliciesRead(current=current, history=history)


async def create_retirement_age_policy(
    session: AsyncSession,
    payload: RetirementAgePolicyCreate,
) -> RetirementAgePoliciesRead:
    rows = await session.execute(
        select(RetirementAgePolicy)
        .where(RetirementAgePolicy.effective_to.is_(None))
        .order_by(RetirementAgePolicy.effective_from.desc(), RetirementAgePolicy.id.desc())
    )
    current = rows.scalars().first()
    if current and payload.effective_from <= current.effective_from:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Ngày hiệu lực policy mới phải lớn hơn policy đang hiệu lực",
        )

    if current:
        current.effective_to = payload.effective_from - timedelta(days=1)
        current.updated_at = _utcnow()

    policy = RetirementAgePolicy(
        name=payload.name,
        legal_basis_summary=payload.legal_basis_summary,
        effective_from=payload.effective_from,
        effective_to=None,
        note=payload.note,
    )
    session.add(policy)
    await session.flush()
    await _replace_policy_thresholds(
        session,
        policy_id=policy.id,
        thresholds=payload.thresholds,
    )
    await session.commit()
    return await get_retirement_age_policies(session)


async def update_retirement_age_policy(
    session: AsyncSession,
    policy_id: int,
    payload: RetirementAgePolicyUpdate,
) -> RetirementAgePoliciesRead:
    policy = await session.get(RetirementAgePolicy, policy_id)
    if not policy:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy policy tuổi nghỉ hưu")

    provided = payload.model_fields_set
    if "name" in provided and payload.name is not None:
        policy.name = payload.name
    if "legal_basis_summary" in provided:
        policy.legal_basis_summary = payload.legal_basis_summary
    if "note" in provided:
        policy.note = payload.note
    if "thresholds" in provided and payload.thresholds is not None:
        await _replace_policy_thresholds(
            session,
            policy_id=policy.id,
            thresholds=payload.thresholds,
        )
    policy.updated_at = _utcnow()
    session.add(policy)
    await session.commit()
    return await get_retirement_age_policies(session)


async def delete_retirement_age_policy(
    session: AsyncSession,
    policy_id: int,
) -> RetirementAgePoliciesRead:
    policy = await session.get(RetirementAgePolicy, policy_id)
    if not policy:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy policy tuổi nghỉ hưu")

    rows = await session.execute(
        select(RetirementAgePolicy).order_by(
            RetirementAgePolicy.effective_from.asc(),
            RetirementAgePolicy.id.asc(),
        )
    )
    items = list(rows.scalars().all())
    idx = next((i for i, item in enumerate(items) if item.id == policy.id), None)
    if idx is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy policy tuổi nghỉ hưu")

    previous = items[idx - 1] if idx > 0 else None
    following = items[idx + 1] if idx + 1 < len(items) else None
    if previous:
        previous.effective_to = (following.effective_from - timedelta(days=1)) if following else None
        session.add(previous)

    await session.delete(policy)
    await session.commit()
    return await get_retirement_age_policies(session)


async def _resolve_retirement_age_policy(
    session: AsyncSession,
    *,
    as_of_date: date,
) -> RetirementAgePolicy:
    row = await session.execute(
        select(RetirementAgePolicy)
        .where(
            RetirementAgePolicy.effective_from <= as_of_date,
            or_(
                RetirementAgePolicy.effective_to.is_(None),
                RetirementAgePolicy.effective_to >= as_of_date,
            ),
        )
        .order_by(RetirementAgePolicy.effective_from.desc(), RetirementAgePolicy.id.desc())
    )
    policy = row.scalars().first()
    if not policy:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"Chưa có policy tuổi nghỉ hưu hiệu lực tại ngày {as_of_date.isoformat()}",
        )
    return policy


async def _resolve_retirement_threshold_map(
    session: AsyncSession,
    *,
    policy_id: int,
    report_year: int,
) -> dict[str, _RetirementAgeThresholdValue]:
    rows = await session.execute(
        select(RetirementAgePolicyThreshold)
        .where(
            RetirementAgePolicyThreshold.policy_id == policy_id,
            RetirementAgePolicyThreshold.applicable_year <= report_year,
        )
        .order_by(
            RetirementAgePolicyThreshold.gender.asc(),
            RetirementAgePolicyThreshold.applicable_year.desc(),
            RetirementAgePolicyThreshold.id.desc(),
        )
    )
    result: dict[str, _RetirementAgeThresholdValue] = {}
    for row in rows.scalars().all():
        if row.gender not in result:
            result[row.gender] = _RetirementAgeThresholdValue(
                years=row.age_years,
                months=row.age_months,
            )

    missing = _OLDER_WORKER_SUPPORTED_GENDERS - set(result.keys())
    if missing:
        missing_gender = ", ".join(sorted(missing))
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "Policy tuổi nghỉ hưu chưa có đủ ngưỡng cho năm "
                f"{report_year}: {missing_gender}"
            ),
        )
    return result


async def get_employee_list(
    session: AsyncSession,
    *,
    page: int = 1,
    page_size: int = 20,
    department_id: Optional[int] = None,
    status: Optional[str] = None,
    gender: Optional[str] = None,
    document_kind: Optional[str] = None,
    start_date_from: Optional[date] = None,
    start_date_to: Optional[date] = None,
    tenure_min: Optional[int] = None,
    tenure_max: Optional[int] = None,
) -> EmployeeListResponse:
    page = max(page, 1)
    page_size = max(page_size, 1)

    scope_ids = await _department_scope_ids(session, department_id)
    stmt = (
        select(
            Employee,
            EmployeeJobRecord.department_id,
            Department.name.label("department_name"),
            JobTitle.name.label("job_title_name"),
            ContractCategory.name.label("contract_category_name"),
            EmployeeContract.document_kind.label("document_kind"),
        )
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(JobTitle, JobTitle.id == EmployeeJobRecord.job_title_id)
        .outerjoin(
            EmployeeContract,
            and_(
                EmployeeContract.employee_id == Employee.id,
                EmployeeContract.status == "active",
                EmployeeContract.parent_contract_id.is_(None),
            ),
        )
        .outerjoin(ContractCategory, ContractCategory.id == EmployeeContract.contract_category_id)
    )

    # tenure_years computed in SQL:
    # EXTRACT(year FROM AGE(COALESCE(resigned_date, CURRENT_DATE), start_date))
    tenure_expr = func.extract(
        "year",
        func.age(
            func.coalesce(Employee.resigned_date, func.current_date()),
            Employee.start_date,
        ),
    ).cast(sa.Integer)

    # Thêm tenure_years vào select
    stmt = stmt.add_columns(tenure_expr.label("tenure_years"))

    filters: list[sa.ColumnElement[bool]] = []
    if status:
        filters.append(Employee.status == status)
    if gender:
        filters.append(Employee.gender == gender)
    if document_kind:
        if document_kind == "probation":
            filters.append(ContractCategory.business_group == "probation")
        else:
            filters.append(EmployeeContract.document_kind == document_kind)
    if start_date_from:
        filters.append(Employee.start_date >= start_date_from)
    if start_date_to:
        filters.append(Employee.start_date <= start_date_to)
    if scope_ids is not None:
        if not scope_ids:
            filters.append(sa.false())
        else:
            filters.append(EmployeeJobRecord.department_id.in_(scope_ids))
    # Tenure filter trong SQL — tránh fetch toàn bộ rồi lọc Python
    if tenure_min is not None:
        filters.append(tenure_expr >= tenure_min)
    if tenure_max is not None:
        filters.append(tenure_expr < tenure_max)

    filtered_stmt = stmt.where(*filters).order_by(Employee.full_name, Employee.id)

    # Count query dùng subquery để giữ WHERE conditions
    count_q = select(func.count()).select_from(filtered_stmt.subquery())
    total: int = (await session.execute(count_q)).scalar_one()

    # Paginate trong SQL
    page_rows = (await session.execute(
        filtered_stmt.offset((page - 1) * page_size).limit(page_size)
    )).all()

    employees = [employee for employee, *_ in page_rows]
    employee_codes = await employee_code_service.batch_build_employee_display_codes(session, employees)

    items: list[EmployeeListItem] = [
        EmployeeListItem(
            id=employee.id,
            employee_code=employee_codes.get(employee.id, str(employee.employee_seq)),
            full_name=employee.full_name,
            gender=employee.gender,
            date_of_birth=employee.date_of_birth,
            status=employee.status,
            start_date=employee.start_date,
            resigned_date=employee.resigned_date,
            is_active=employee.is_active,
            department_id=dept_id,
            department_name=dept_name,
            job_title_name=job_title_name,
            contract_category_name=contract_category_name,
            document_kind=row_document_kind,
            tenure_years=tenure_yrs,
        )
        for employee, dept_id, dept_name, job_title_name, contract_category_name, row_document_kind, tenure_yrs
        in page_rows
    ]

    return EmployeeListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
        total_pages=max(ceil(total / page_size), 1),
    )


async def get_movement_report(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
    group_by: str = "month",
    department_id: Optional[int] = None,
) -> MovementReportResponse:
    if group_by not in {"month", "quarter", "year"}:
        raise ValueError("group_by must be month, quarter, or year")
    if end_date < start_date:
        raise ValueError("end_date must be >= start_date")

    periods = _iter_periods(start_date, end_date, group_by)
    scope_ids = await _department_scope_ids(session, department_id)
    period_map = {
        period.start: {
            "period_label": _period_label(period.start, group_by),
            "period_start": period.start,
            "period_end": period.end,
            "hired_count": 0,
            "resigned_count": 0,
            "transfer_count": 0,
        }
        for period in periods
    }

    job_join = and_(
        EmployeeJobRecord.employee_id == Employee.id,
        EmployeeJobRecord.is_current == True,  # noqa: E712
    )
    if scope_ids is not None:
        if not scope_ids:
            return MovementReportResponse(
                group_by=group_by,
                start_date=start_date,
                end_date=end_date,
                rows=[
                    MovementPeriodRow(
                        **row,
                        net_change=0,
                    )
                    for row in period_map.values()
                ],
                total_hired=0,
                total_resigned=0,
                total_transfers=0,
            )
        job_join = and_(job_join, EmployeeJobRecord.department_id.in_(scope_ids))

    hired_stmt = (
        select(func.date_trunc(group_by, Employee.start_date).label("period_start"), func.count(Employee.id))
        .join(EmployeeJobRecord, job_join)
        .where(Employee.start_date.between(start_date, end_date))
        .group_by("period_start")
    )
    for period_start_value, count in (await session.execute(hired_stmt)).all():
        period_start = period_start_value.date()
        if period_start in period_map:
            period_map[period_start]["hired_count"] = int(count or 0)

    resigned_stmt = (
        select(func.date_trunc(group_by, Employee.resigned_date).label("period_start"), func.count(Employee.id))
        .join(EmployeeJobRecord, job_join)
        .where(Employee.resigned_date.is_not(None), Employee.resigned_date.between(start_date, end_date))
        .group_by("period_start")
    )
    for period_start_value, count in (await session.execute(resigned_stmt)).all():
        period_start = period_start_value.date()
        if period_start in period_map:
            period_map[period_start]["resigned_count"] = int(count or 0)

    transfer_stmt = (
        select(
            func.date_trunc(group_by, EmployeeJobRecord.effective_from).label("period_start"),
            func.count(EmployeeJobRecord.id),
        )
        .join(Employee, Employee.id == EmployeeJobRecord.employee_id)
        .where(
            EmployeeJobRecord.effective_from.between(start_date, end_date),
            EmployeeJobRecord.effective_from != Employee.start_date,
            Employee.is_active == True,  # noqa: E712
        )
    )
    if scope_ids is not None:
        transfer_stmt = transfer_stmt.where(EmployeeJobRecord.department_id.in_(scope_ids))
    transfer_stmt = transfer_stmt.group_by("period_start")
    for period_start_value, count in (await session.execute(transfer_stmt)).all():
        period_start = period_start_value.date()
        if period_start in period_map:
            period_map[period_start]["transfer_count"] = int(count or 0)

    rows: list[MovementPeriodRow] = []
    total_hired = 0
    total_resigned = 0
    total_transfers = 0
    for period in periods:
        payload = period_map[period.start]
        net_change = payload["hired_count"] - payload["resigned_count"]
        total_hired += payload["hired_count"]
        total_resigned += payload["resigned_count"]
        total_transfers += payload["transfer_count"]
        rows.append(MovementPeriodRow(**payload, net_change=net_change))

    return MovementReportResponse(
        group_by=group_by,
        start_date=start_date,
        end_date=end_date,
        rows=rows,
        total_hired=total_hired,
        total_resigned=total_resigned,
        total_transfers=total_transfers,
    )


async def get_tenure_report(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
) -> TenureReportResponse:
    scope_ids = await _department_scope_ids(session, department_id)
    department_name = await _department_name(session, department_id)

    stmt = (
        select(Employee.id, Employee.full_name, Employee.start_date, Department.name, EmployeeJobRecord.department_id)
        .join(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .where(Employee.is_active == True, Employee.status != "resigned")  # noqa: E712
        .order_by(Employee.full_name, Employee.id)
    )
    rows = (await session.execute(stmt)).all()

    bucket_members: dict[str, list[TenureGroupDetail]] = {key: [] for key, _, _, _ in TENURE_GROUPS}
    bucket_tenures: dict[str, list[int]] = {key: [] for key, _, _, _ in TENURE_GROUPS}
    all_tenures: list[int] = []

    for employee_id, full_name, start_date_value, department_name_value, direct_department_id in rows:
        if not _group_members(scope_ids, direct_department_id):
            continue
        tenure_years = _tenure_years(start_date_value)
        all_tenures.append(tenure_years)
        detail = TenureGroupDetail(
            id=employee_id,
            full_name=full_name,
            department_name=department_name_value,
            start_date=start_date_value,
            tenure_years=tenure_years,
        )
        for key, _, min_years, max_years in TENURE_GROUPS:
            if tenure_years >= min_years and (max_years is None or tenure_years < max_years):
                bucket_members[key].append(detail)
                bucket_tenures[key].append(tenure_years)
                break

    total_active = len(all_tenures)
    groups: list[TenureGroup] = []
    for key, label, _, _ in TENURE_GROUPS:
        members = bucket_members[key]
        tenures = bucket_tenures[key]
        avg_tenure = _round1(sum(tenures) / len(tenures)) if tenures else 0.0
        percentage = _round1(len(members) / total_active * 100) if total_active else 0.0
        groups.append(
            TenureGroup(
                group_key=key,
                group_label=label,
                headcount=len(members),
                percentage=percentage,
                avg_tenure_years=avg_tenure,
                employees=members,
            )
        )

    overall_avg = _round1(sum(all_tenures) / total_active) if total_active else 0.0
    return TenureReportResponse(
        department_id=department_id,
        department_name=department_name,
        total_active=total_active,
        avg_tenure_years=overall_avg,
        groups=groups,
    )


async def get_org_structure(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
) -> OrgStructureResponse:
    scope_ids = await _department_scope_ids(session, department_id)
    dept_stmt = (
        select(Department.id, Department.name, Department.parent_id, Department.order_no)
        .where(Department.is_active == True)  # noqa: E712
        .order_by(Department.order_no, Department.name)
    )
    if scope_ids is not None:
        if not scope_ids:
            return OrgStructureResponse(total_headcount=0, department_count=0, tree=[])
        dept_stmt = dept_stmt.where(Department.id.in_(scope_ids))
    dept_rows = (await session.execute(dept_stmt)).all()
    if not dept_rows:
        return OrgStructureResponse(total_headcount=0, department_count=0, tree=[])

    dept_map = {
        row.id: {
            "department_id": row.id,
            "department_name": row.name,
            "parent_id": row.parent_id,
            "order_no": row.order_no or 0,
        }
        for row in dept_rows
    }
    children_map: dict[int, list[int]] = {dept_id: [] for dept_id in dept_map}
    roots: list[int] = []
    for dept_id, payload in dept_map.items():
        parent_id = payload["parent_id"]
        if parent_id in children_map:
            children_map[parent_id].append(dept_id)
        else:
            roots.append(dept_id)
    for child_ids in children_map.values():
        child_ids.sort(key=lambda item: (dept_map[item]["order_no"], dept_map[item]["department_name"]))
    roots.sort(key=lambda item: (dept_map[item]["order_no"], dept_map[item]["department_name"]))

    headcount_stmt = (
        select(
            EmployeeJobRecord.department_id,
            JobTitle.id,
            JobTitle.name,
            JobTitle.level,
            func.count(Employee.id),
        )
        .select_from(EmployeeJobRecord)
        .join(
            Employee,
            and_(
                Employee.id == EmployeeJobRecord.employee_id,
                Employee.is_active == True,  # noqa: E712
                Employee.status != "resigned",
            ),
        )
        .outerjoin(JobTitle, JobTitle.id == EmployeeJobRecord.job_title_id)
        .where(EmployeeJobRecord.is_current == True)  # noqa: E712
        .group_by(EmployeeJobRecord.department_id, JobTitle.id, JobTitle.name, JobTitle.level)
    )
    if scope_ids is not None:
        headcount_stmt = headcount_stmt.where(EmployeeJobRecord.department_id.in_(scope_ids))

    direct_totals = {dept_id: 0 for dept_id in dept_map}
    titles_by_dept: dict[int, list[JobTitleHeadcount]] = {dept_id: [] for dept_id in dept_map}
    for dept_id, job_title_id, job_title_name, job_level, headcount in (await session.execute(headcount_stmt)).all():
        if dept_id not in dept_map:
            continue
        value = int(headcount or 0)
        direct_totals[dept_id] += value
        titles_by_dept[dept_id].append(
            JobTitleHeadcount(
                job_title_id=job_title_id,
                job_title_name=job_title_name,
                job_level=job_level,
                headcount=value,
            )
        )
    for items in titles_by_dept.values():
        items.sort(key=lambda item: (item.job_level or 999, item.job_title_name or ""))

    def build_node(dept_id: int) -> DepartmentNode:
        children = [build_node(child_id) for child_id in children_map.get(dept_id, [])]
        total_headcount = direct_totals[dept_id] + sum(child.total_headcount for child in children)
        return DepartmentNode(
            department_id=dept_id,
            department_name=dept_map[dept_id]["department_name"],
            parent_id=dept_map[dept_id]["parent_id"],
            total_headcount=total_headcount,
            direct_headcount=direct_totals[dept_id],
            by_job_title=titles_by_dept[dept_id],
            children=children,
        )

    tree = [build_node(root_id) for root_id in roots]
    return OrgStructureResponse(
        total_headcount=sum(node.total_headcount for node in tree),
        department_count=len(dept_map),
        tree=tree,
    )


async def get_older_worker_report(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
    gender: Optional[str] = None,
) -> OlderWorkerReportResponse:
    normalized_gender = gender.strip().lower() if isinstance(gender, str) and gender.strip() else None
    if normalized_gender is not None and normalized_gender not in _OLDER_WORKER_SUPPORTED_GENDERS:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="gender phải là male hoặc female",
        )

    as_of_date = _last_day_of_month(year, month)
    department_name = await _department_name(session, department_id)
    scope_ids = await _department_scope_ids(session, department_id)
    policy = await _resolve_retirement_age_policy(session, as_of_date=as_of_date)
    threshold_map = await _resolve_retirement_threshold_map(
        session,
        policy_id=policy.id,
        report_year=year,
    )

    stmt = (
        select(
            Employee,
            EmployeeJobRecord.department_id,
            Department.name.label("department_name"),
            JobTitle.name.label("job_title_name"),
        )
        .join(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.effective_from <= as_of_date,
                or_(
                    EmployeeJobRecord.effective_to.is_(None),
                    EmployeeJobRecord.effective_to >= as_of_date,
                ),
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(JobTitle, JobTitle.id == EmployeeJobRecord.job_title_id)
        .where(
            Employee.date_of_birth.is_not(None),
            Employee.start_date <= as_of_date,
            or_(
                Employee.resigned_date.is_(None),
                Employee.resigned_date >= as_of_date,
            ),
            Employee.gender.in_(tuple(sorted(_OLDER_WORKER_SUPPORTED_GENDERS))),
        )
        .order_by(Department.name.asc(), JobTitle.name.asc(), Employee.full_name.asc(), Employee.id.asc())
    )
    if normalized_gender is not None:
        stmt = stmt.where(Employee.gender == normalized_gender)
    if scope_ids is not None:
        if not scope_ids:
            return OlderWorkerReportResponse(
                year=year,
                month=month,
                as_of_date=as_of_date,
                department_id=department_id,
                department_name=department_name,
                gender=normalized_gender,
                policy_id=policy.id,
                policy_name=policy.name,
                legal_basis_summary=policy.legal_basis_summary,
                male_threshold_label=_threshold_label(threshold_map.get("male")),
                female_threshold_label=_threshold_label(threshold_map.get("female")),
                summary=OlderWorkerSummary(total=0, male_count=0, female_count=0),
                items=[],
            )
        stmt = stmt.where(EmployeeJobRecord.department_id.in_(scope_ids))

    rows = (await session.execute(stmt)).all()
    employees = [employee for employee, *_ in rows]
    employee_codes = await employee_code_service.batch_build_employee_display_codes(session, employees)

    items: list[OlderWorkerListItem] = []
    male_count = 0
    female_count = 0
    for employee, dept_id, dept_name, job_title_name in rows:
        threshold = threshold_map.get(employee.gender)
        if threshold is None:
            continue
        retirement_date = _add_months(
            employee.date_of_birth,
            threshold.years * 12 + threshold.months,
        )
        if retirement_date > as_of_date:
            continue
        age_years, age_months = _diff_years_months(employee.date_of_birth, as_of_date)
        months_beyond_retirement = _diff_total_months(retirement_date, as_of_date)
        item = OlderWorkerListItem(
            id=employee.id,
            employee_code=employee_codes.get(employee.id, str(employee.employee_seq)),
            full_name=employee.full_name,
            gender=employee.gender,
            date_of_birth=employee.date_of_birth,
            start_date=employee.start_date,
            department_id=dept_id,
            department_name=dept_name,
            job_title_name=job_title_name,
            retirement_age_years=threshold.years,
            retirement_age_months=threshold.months,
            retirement_date=retirement_date,
            age_years=age_years,
            age_months=age_months,
            months_beyond_retirement=months_beyond_retirement,
        )
        items.append(item)
        if employee.gender == "male":
            male_count += 1
        elif employee.gender == "female":
            female_count += 1

    return OlderWorkerReportResponse(
        year=year,
        month=month,
        as_of_date=as_of_date,
        department_id=department_id,
        department_name=department_name,
        gender=normalized_gender,
        policy_id=policy.id,
        policy_name=policy.name,
        legal_basis_summary=policy.legal_basis_summary,
        male_threshold_label=_threshold_label(threshold_map.get("male")),
        female_threshold_label=_threshold_label(threshold_map.get("female")),
        summary=OlderWorkerSummary(
            total=len(items),
            male_count=male_count,
            female_count=female_count,
        ),
        items=items,
    )


async def export_employee_list_excel(
    session: AsyncSession,
    **filters,
) -> BytesIO:
    from openpyxl import Workbook

    normalized_filters = dict(filters)
    normalized_filters["page"] = 1
    normalized_filters["page_size"] = max(int(normalized_filters.get("page_size", 0) or 0), 10000)
    report = await get_employee_list(session, **normalized_filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách nhân sự"
    ws.append(
        [
            "Mã NV",
            "Họ tên",
            "Giới tính",
            "Trạng thái",
            "Ngày vào làm",
            "Ngày nghỉ việc",
            "Phòng ban",
            "Chức danh",
            "Loại hợp đồng",
            "Loại văn bản",
            "Thâm niên",
        ]
    )
    for item in report.items:
        ws.append(
            [
                item.employee_code,
                item.full_name,
                item.gender,
                item.status,
                item.start_date.isoformat(),
                item.resigned_date.isoformat() if item.resigned_date else "",
                item.department_name or "",
                item.job_title_name or "",
                item.contract_category_name or "",
                item.document_kind or "",
                item.tenure_years,
            ]
        )
    return _save_workbook_bytes(wb)


async def export_movement_excel(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
    group_by: str = "month",
    department_id: Optional[int] = None,
) -> BytesIO:
    from openpyxl import Workbook

    report = await get_movement_report(
        session,
        start_date=start_date,
        end_date=end_date,
        group_by=group_by,
        department_id=department_id,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Biến động nhân sự"
    ws.append(
        [
            "Kỳ",
            "Từ ngày",
            "Đến ngày",
            "Tuyển mới",
            "Thôi việc",
            "Chuyển bộ phận",
            "Biến động ròng",
        ]
    )
    for row in report.rows:
        ws.append(
            [
                row.period_label,
                row.period_start.isoformat(),
                row.period_end.isoformat(),
                row.hired_count,
                row.resigned_count,
                row.transfer_count,
                row.net_change,
            ]
    )
    return _save_workbook_bytes(wb)


async def export_older_worker_excel(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
    gender: Optional[str] = None,
) -> BytesIO:
    from openpyxl import Workbook

    report = await get_older_worker_report(
        session,
        year=year,
        month=month,
        department_id=department_id,
        gender=gender,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Lao động cao tuổi"
    ws.append(
        [
            "Mã NV",
            "Họ tên",
            "Giới tính",
            "Ngày sinh",
            "Phòng ban",
            "Chức danh",
            "Ngày vào làm",
            "Ngưỡng tuổi nghỉ hưu",
            "Ngày đủ tuổi nghỉ hưu",
            "Tuổi tại kỳ báo cáo",
            "Số tháng vượt ngưỡng",
        ]
    )
    for item in report.items:
        ws.append(
            [
                item.employee_code,
                item.full_name,
                item.gender,
                item.date_of_birth.isoformat(),
                item.department_name or "",
                item.job_title_name or "",
                item.start_date.isoformat(),
                f"{item.retirement_age_years} tuổi {item.retirement_age_months} tháng",
                item.retirement_date.isoformat(),
                f"{item.age_years} tuổi {item.age_months} tháng",
                item.months_beyond_retirement,
            ]
        )
    return _save_workbook_bytes(wb)


async def export_tenure_excel(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
) -> BytesIO:
    from openpyxl import Workbook

    report = await get_tenure_report(session, department_id=department_id)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Tổng hợp thâm niên"
    summary_ws.append(
        ["Nhóm", "Số nhân sự", "Tỷ lệ", "TB thâm niên"]
    )
    for group in report.groups:
        summary_ws.append(
            [group.group_label, group.headcount, group.percentage, group.avg_tenure_years]
        )

    detail_ws = wb.create_sheet("Chi tiết thâm niên")
    detail_ws.append(["Nhóm", "Họ tên", "Phòng ban", "Ngày vào làm", "Thâm niên"])
    for group in report.groups:
        for employee in group.employees:
            detail_ws.append(
                [
                    group.group_label,
                    employee.full_name,
                    employee.department_name or "",
                    employee.start_date.isoformat(),
                    employee.tenure_years,
                ]
            )
    return _save_workbook_bytes(wb)


async def export_org_structure_excel(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
) -> BytesIO:
    from openpyxl import Workbook

    report = await get_org_structure(session, department_id=department_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cơ cấu tổ chức"
    ws.append(
        [
            "Phòng ban",
            "Phòng ban cha",
            "Tổng nhân sự",
            "Nhân sự trực tiếp",
            "Chức danh",
            "Cấp bậc",
            "Số nhân sự theo chức danh",
        ]
    )

    def append_node(node: DepartmentNode, parent_name: str = "") -> None:
        if node.by_job_title:
            for title in node.by_job_title:
                ws.append(
                    [
                        node.department_name,
                        parent_name,
                        node.total_headcount,
                        node.direct_headcount,
                        title.job_title_name or "",
                        title.job_level or "",
                        title.headcount,
                    ]
                )
        else:
            ws.append(
                [
                    node.department_name,
                    parent_name,
                    node.total_headcount,
                    node.direct_headcount,
                    "",
                    "",
                    "",
                ]
            )
        for child in node.children:
            append_node(child, node.department_name)

    for root in report.tree:
        append_node(root)
    return _save_workbook_bytes(wb)
