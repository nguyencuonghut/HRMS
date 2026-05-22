from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import Nationality
from app.models.insurance import InsuranceChangeEvent, InsurancePeriodReport, InsuranceReportLineItem
from app.models.salary import CompanyBhxhRegion

TEMPLATE_PATH = Path(__file__).parent.parent.parent / "templates" / "FileMau_D02_TK1_VNPT.xlsx"

_CHANGE_REASON_PA_MAP: dict[tuple[str, str], tuple[int, str, str]] = {
    ("new_hire",                 "increase"): (1, "TM", "Tăng mới"),
    ("return_from_leave",        "increase"): (1, "ON", "Đi làm lại"),
    ("transfer_in",              "increase"): (1, "TD", "Tăng do chuyển đơn vị"),
    ("contract_renewal",         "increase"): (1, "TM", "Tăng mới"),
    ("manual_correction",        "increase"): (1, "TM", "Tăng mới"),
    ("resignation",              "decrease"): (3, "GH", "Giảm hẳn"),
    ("contract_end",             "decrease"): (3, "GH", "Giảm hẳn"),
    ("dismissal",                "decrease"): (3, "GH", "Giảm hẳn"),
    ("unpaid_leave",             "decrease"): (3, "KL", "Nghỉ không lương"),
    ("maternity_no_contribution","decrease"): (3, "TS", "Thai sản"),
    ("long_term_sick",           "decrease"): (3, "OF", "Nghỉ do ốm đau"),
    ("transfer_out",             "decrease"): (3, "GD", "Giảm do chuyển đơn vị"),
    ("manual_correction",        "decrease"): (3, "GH", "Giảm hẳn"),
}


def _fmt_date(d: object) -> Optional[str]:
    if d is None:
        return None
    return d.strftime("%d/%m/%Y")  # type: ignore[attr-defined]


def _fmt_month_year(d: object) -> Optional[str]:
    if d is None:
        return None
    return d.strftime("%m/%Y")  # type: ignore[attr-defined]


def _fmt_rate(employee: Decimal, employer: Decimal) -> str:
    total = float(employee) + float(employer)
    return f"{total:g}".replace(".", ",")


def _fill_row(
    ws_data: object,
    ws_bangke: object,
    *,
    row: int,
    idx: int,
    event: InsuranceChangeEvent,
    nat_map: dict[str, str],
    mavung_ltt: str,
    period_str: Optional[str],
) -> None:
    """Ghi 1 dòng vào sheet 'Dữ Liệu' và 'Bảng kê'. period_str = 'MM/YYYY' cho cột 16."""
    loai, pa, ten_pa = _CHANGE_REASON_PA_MAP.get(
        (event.change_reason, event.change_type),
        (1 if event.change_type == "increase" else 3, "TM", "Tăng mới"),
    )
    ten_loai = "Tăng lao động" if event.change_type == "increase" else "Giảm lao động"

    if pa == "TM":
        contract_part = event.contract_number_snapshot or ""
        clinic_part = event.bhyt_clinic_code_snapshot or ""
        ghichu: Optional[str] = f"{contract_part} - KCB: {clinic_part}".strip(" -") or None
    else:
        ghichu = event.note or None

    den_thang: Optional[str] = (
        _fmt_month_year(event.contract_to_snapshot) if pa == "GH" else None
    )

    nat_name = nat_map.get(event.nationality_code_snapshot or "", "Việt Nam")
    gender_val = 1 if event.gender_snapshot == "male" else 0

    # Sheet "Dữ Liệu"
    ws_data.cell(row=row, column=1).value  = idx  # type: ignore[union-attr]
    ws_data.cell(row=row, column=2).value  = event.employee_name_snapshot  # type: ignore[union-attr]
    ws_data.cell(row=row, column=3).value  = event.bhxh_code_snapshot  # type: ignore[union-attr]
    ws_data.cell(row=row, column=4).value  = ten_loai  # type: ignore[union-attr]
    ws_data.cell(row=row, column=5).value  = loai  # type: ignore[union-attr]
    ws_data.cell(row=row, column=6).value  = ten_pa  # type: ignore[union-attr]
    ws_data.cell(row=row, column=7).value  = pa  # type: ignore[union-attr]
    ws_data.cell(row=row, column=10).value = int(event.basis_amount)  # type: ignore[union-attr]
    ws_data.cell(row=row, column=11).value = int(event.allowances_amount)  # type: ignore[union-attr]
    ws_data.cell(row=row, column=16).value = period_str  # type: ignore[union-attr]
    ws_data.cell(row=row, column=17).value = den_thang  # type: ignore[union-attr]
    ws_data.cell(row=row, column=19).value = ghichu  # type: ignore[union-attr]
    ws_data.cell(row=row, column=20).value = _fmt_rate(  # type: ignore[union-attr]
        event.employee_rate_total_snapshot,
        event.employer_rate_total_snapshot,
    )
    ws_data.cell(row=row, column=26).value = mavung_ltt  # MavungLTT  # type: ignore[union-attr]
    ws_data.cell(row=row, column=28).value = event.identity_number_snapshot  # type: ignore[union-attr]
    ws_data.cell(row=row, column=30).value = _fmt_date(event.date_of_birth_snapshot)  # type: ignore[union-attr]
    ws_data.cell(row=row, column=31).value = gender_val  # type: ignore[union-attr]
    ws_data.cell(row=row, column=32).value = nat_name  # type: ignore[union-attr]
    ws_data.cell(row=row, column=33).value = event.nationality_code_snapshot  # type: ignore[union-attr]
    ws_data.cell(row=row, column=35).value = event.ethnicity_bhxh_code_snapshot  # DanToc code  # type: ignore[union-attr]
    ws_data.cell(row=row, column=38).value = event.bhyt_clinic_name_snapshot  # type: ignore[union-attr]
    ws_data.cell(row=row, column=39).value = event.bhyt_clinic_code_snapshot  # type: ignore[union-attr]

    # Sheet "Bảng kê"
    ws_bangke.cell(row=row, column=1).value = idx  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=2).value = event.employee_name_snapshot  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=3).value = event.bhxh_code_snapshot  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=4).value = "Hợp đồng lao động"  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=5).value = event.contract_number_snapshot  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=6).value = _fmt_date(event.contract_signed_date_snapshot)  # type: ignore[union-attr]
    ws_bangke.cell(row=row, column=7).value = _fmt_date(event.contract_from_snapshot)  # type: ignore[union-attr]


async def export_d02_ts_vnpt(
    session: AsyncSession,
    period_year: int,
    period_month: int,
) -> BytesIO:
    # 1. Query events — increase first, then decrease; stable id order within each type
    events = (
        await session.execute(
            select(InsuranceChangeEvent)
            .where(
                InsuranceChangeEvent.period_year == period_year,
                InsuranceChangeEvent.period_month == period_month,
            )
            .order_by(InsuranceChangeEvent.change_type.desc(), InsuranceChangeEvent.id.asc())
        )
    ).scalars().all()

    # 2. Current company region → MavungLTT
    region_row = (
        await session.execute(
            select(CompanyBhxhRegion)
            .where(CompanyBhxhRegion.effective_to.is_(None))
            .order_by(CompanyBhxhRegion.effective_from.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    mavung_ltt = f"0{region_row.region}" if region_row else "01"

    # 3. Nationality name lookup (iso2_code → name)
    nat_codes = {e.nationality_code_snapshot for e in events if e.nationality_code_snapshot}
    nat_map: dict[str, str] = {}
    if nat_codes:
        nat_rows = (
            await session.execute(
                select(Nationality).where(Nationality.iso2_code.in_(nat_codes))
            )
        ).scalars().all()
        nat_map = {n.iso2_code: n.name for n in nat_rows if n.iso2_code}

    # 4. Load template
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=False)
    ws_data = wb["Dữ Liệu"]
    ws_bangke = wb["Bảng kê"]

    # 5. Clear existing example rows (keep header rows 1–3)
    for ws in (ws_data, ws_bangke):
        if ws.max_row >= 4:
            ws.delete_rows(4, ws.max_row - 3)

    # 6. Fill rows
    for idx, event in enumerate(events, start=1):
        _fill_row(
            ws_data, ws_bangke,
            row=idx + 3,
            idx=idx,
            event=event,
            nat_map=nat_map,
            mavung_ltt=mavung_ltt,
            period_str=_fmt_month_year(event.effective_date),
        )

    # 7. Return as BytesIO
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def export_d02_ts_from_report(
    session: AsyncSession,
    report_id: int,
) -> tuple[BytesIO, str]:
    """Export D02-TS VNPT từ báo cáo đã được duyệt (approved).

    Trả về (BytesIO, filename).
    Thứ tự dòng theo sort_order của line item.
    Cột 16 (Kỳ kê khai) dùng declared_year/declared_month từ line item.
    """
    report = await session.get(InsurancePeriodReport, report_id)
    if not report:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy báo cáo")
    if report.status != "approved":
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Chỉ có thể xuất file từ báo cáo đã được duyệt (approved)",
        )

    # 1. Load line items + events ordered by sort_order
    rows = (
        await session.execute(
            select(InsuranceReportLineItem, InsuranceChangeEvent)
            .join(InsuranceChangeEvent, InsuranceChangeEvent.id == InsuranceReportLineItem.event_id)
            .where(InsuranceReportLineItem.report_id == report_id)
            .order_by(InsuranceReportLineItem.sort_order.asc(), InsuranceReportLineItem.id.asc())
        )
    ).all()

    # 2. Current company region → MavungLTT
    region_row = (
        await session.execute(
            select(CompanyBhxhRegion)
            .where(CompanyBhxhRegion.effective_to.is_(None))
            .order_by(CompanyBhxhRegion.effective_from.desc())
            .limit(1)
        )
    ).scalar_one_or_none()
    mavung_ltt = f"0{region_row.region}" if region_row else "01"

    # 3. Nationality name lookup
    nat_codes = {e.nationality_code_snapshot for _, e in rows if e.nationality_code_snapshot}
    nat_map: dict[str, str] = {}
    if nat_codes:
        nat_rows = (
            await session.execute(
                select(Nationality).where(Nationality.iso2_code.in_(nat_codes))
            )
        ).scalars().all()
        nat_map = {n.iso2_code: n.name for n in nat_rows if n.iso2_code}

    # 4. Load template
    wb = openpyxl.load_workbook(TEMPLATE_PATH, keep_vba=False)
    ws_data = wb["Dữ Liệu"]
    ws_bangke = wb["Bảng kê"]

    # 5. Clear existing example rows (keep header rows 1–3)
    for ws in (ws_data, ws_bangke):
        if ws.max_row >= 4:
            ws.delete_rows(4, ws.max_row - 3)

    # 6. Fill rows — col 16 = declared_month/year từ line item
    for idx, (item, event) in enumerate(rows, start=1):
        period_str = f"{item.declared_month:02d}/{item.declared_year}"
        _fill_row(
            ws_data, ws_bangke,
            row=idx + 3,
            idx=idx,
            event=event,
            nat_map=nat_map,
            mavung_ltt=mavung_ltt,
            period_str=period_str,
        )

    # 7. Return as BytesIO + filename
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = (
        f"D02-TS_BaoCao{report_id}"
        f"_T{report.period_month:02d}_{report.period_year}_VNPT.xlsx"
    )
    return buf, filename
