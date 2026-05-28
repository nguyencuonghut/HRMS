"""Báo cáo Onboarding & Thử việc (Plan 14.3)."""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.onboarding import OnboardingChecklist
from app.models.org import Department, JobTitle
from app.models.probation import ProbationEvaluation
from app.schemas.probation_report import (
    ActiveProbationReport,
    ActiveProbationRow,
    ChecklistCompletionReport,
    ChecklistCompletionRow,
    FailureCommentItem,
    FailureKeywordCount,
    FailureReasonReport,
    MonthlyProbationTrend,
    ProbationHistoryReport,
    ProbationHistoryRow,
    ProbationPassRateReport,
    ProbationPassRateStat,
)
from app.services import employee_code_service

_FAILURE_KEYWORDS = [
    "thái độ", "năng lực", "chuyên môn", "kpi", "văn hóa",
    "kỷ luật", "giao tiếp", "teamwork", "trễ hạn", "nghỉ nhiều",
]


# ── A: Danh sách nhân viên đang thử việc ─────────────────────────────────────

async def get_active_probation(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
    keyword: Optional[str] = None,
) -> ActiveProbationReport:
    stmt = (
        select(
            Employee,
            EmployeeJobRecord,
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            OnboardingChecklist.status.label("oc_status"),
            OnboardingChecklist.completion_pct.label("oc_pct"),
            ProbationEvaluation.result.label("eval_result"),
        )
        .where(Employee.status == "probation", Employee.is_active == True)
        .join(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(OnboardingChecklist, OnboardingChecklist.employee_id == Employee.id)
        .outerjoin(
            ProbationEvaluation,
            (ProbationEvaluation.employee_id == Employee.id)
            & (ProbationEvaluation.job_record_id == EmployeeJobRecord.id),
        )
    )

    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    rows = (await session.execute(stmt)).all()

    employees = [emp for emp, *_ in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees)

    today = date.today()
    items: list[ActiveProbationRow] = []
    for emp, ejr, dept_id, dept_name, oc_status, oc_pct, eval_result in rows:
        # Compute days_remaining
        days_remaining: Optional[int] = None
        if ejr.probation_end_date is not None:
            days_remaining = (ejr.probation_end_date - today).days

        # urgency
        if days_remaining is None:
            urgency = "normal"
        elif days_remaining <= 7:
            urgency = "critical"
        elif days_remaining <= 15:
            urgency = "warning"
        else:
            urgency = "normal"

        # evaluation_result
        if eval_result is None:
            evaluation_result = "not_started"
        else:
            evaluation_result = eval_result

        items.append(ActiveProbationRow(
            employee_id=emp.id,
            employee_name=emp.full_name,
            employee_code=code_map.get(emp.id, ""),
            department_id=dept_id,
            department_name=dept_name,
            probation_start_date=ejr.probation_start_date,
            probation_end_date=ejr.probation_end_date,
            days_remaining=days_remaining,
            urgency=urgency,
            onboarding_status=oc_status,
            completion_pct=float(oc_pct) if oc_pct is not None else None,
            evaluation_result=evaluation_result,
        ))

    # Filter by keyword (name or employee_code)
    if keyword:
        kw = keyword.strip().lower()
        items = [
            r for r in items
            if kw in r.employee_name.lower() or kw in r.employee_code.lower()
        ]

    # Sort by days_remaining ASC NULLS LAST
    items.sort(key=lambda r: (r.days_remaining is None, r.days_remaining or 0))

    return ActiveProbationReport(items=items, total=len(items))


# ── A2: Lịch sử thử việc trong kỳ (bao gồm cả đã kết thúc) ──────────────────

async def get_probation_history(
    session: AsyncSession,
    *,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    department_id: Optional[int] = None,
    keyword: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> ProbationHistoryReport:
    """Trả về tất cả nhân viên có probation_start_date trong khoảng [start_date, end_date].
    Nếu không có start_date/end_date, trả về tất cả.
    Bao gồm cả những người đã kết thúc thử việc (official, resigned, …).
    """
    base_conditions = [Employee.is_active == True]  # noqa: E712
    if start_date is not None:
        base_conditions.append(EmployeeJobRecord.probation_start_date >= start_date)
    if end_date is not None:
        base_conditions.append(EmployeeJobRecord.probation_start_date <= end_date)

    stmt = (
        select(
            Employee,
            EmployeeJobRecord,
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            OnboardingChecklist.status.label("oc_status"),
            OnboardingChecklist.completion_pct.label("oc_pct"),
            ProbationEvaluation.result.label("eval_result"),
            ProbationEvaluation.status.label("eval_status"),
        )
        .where(*base_conditions)
        .join(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),  # noqa: E712
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(OnboardingChecklist, OnboardingChecklist.employee_id == Employee.id)
        .outerjoin(
            ProbationEvaluation,
            (ProbationEvaluation.employee_id == Employee.id)
            & (ProbationEvaluation.job_record_id == EmployeeJobRecord.id),
        )
    )

    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    rows = (await session.execute(stmt)).all()

    employees = [emp for emp, *_ in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees)

    today = date.today()
    items: list[ProbationHistoryRow] = []
    for emp, ejr, dept_id, dept_name, oc_status, oc_pct, eval_result, eval_status in rows:
        # days_remaining chỉ tính nếu còn đang probation
        days_remaining: Optional[int] = None
        if emp.status == "probation" and ejr.probation_end_date is not None:
            days_remaining = (ejr.probation_end_date - today).days

        # evaluation_result
        evaluation_result = eval_result if eval_result is not None else "not_started"

        items.append(ProbationHistoryRow(
            employee_id=emp.id,
            employee_name=emp.full_name,
            employee_code=code_map.get(emp.id, ""),
            employee_status=emp.status,
            department_id=dept_id,
            department_name=dept_name,
            probation_start_date=ejr.probation_start_date,
            probation_end_date=ejr.probation_end_date,
            days_remaining=days_remaining,
            onboarding_status=oc_status,
            completion_pct=float(oc_pct) if oc_pct is not None else None,
            evaluation_result=evaluation_result,
            evaluation_status=eval_status,
        ))

    # Filter by keyword
    if keyword:
        kw = keyword.strip().lower()
        items = [
            r for r in items
            if kw in r.employee_name.lower() or kw in r.employee_code.lower()
        ]

    # Sort: đang probation lên trước (theo days_remaining), đã xong xuống dưới
    items.sort(key=lambda r: (
        r.employee_status != "probation",
        r.days_remaining is None,
        r.days_remaining or 0,
    ))

    total = len(items)
    offset = (page - 1) * page_size
    items = items[offset: offset + page_size]

    return ProbationHistoryReport(
        period_start=start_date,
        period_end=end_date,
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


# ── B: Tỷ lệ hoàn thành Checklist theo phòng ban ─────────────────────────────

async def get_checklist_completion(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
) -> ChecklistCompletionReport:
    raw_stmt = (
        select(
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            OnboardingChecklist.status,
            OnboardingChecklist.completion_pct,
        )
        .join(Employee, OnboardingChecklist.employee_id == Employee.id)
        .join(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
        )
        .join(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            Employee.start_date >= start_date,
            Employee.start_date <= end_date,
        )
    )
    if department_id is not None:
        raw_stmt = raw_stmt.where(EmployeeJobRecord.department_id == department_id)

    raw_rows = (await session.execute(raw_stmt)).all()

    # Aggregate in Python
    from collections import defaultdict
    dept_agg: dict[int, dict] = defaultdict(lambda: {
        "name": "", "total": 0, "completed": 0, "pct_sum": 0.0
    })
    for dept_id, dept_name, status, completion_pct in raw_rows:
        agg = dept_agg[dept_id]
        agg["name"] = dept_name
        agg["total"] += 1
        if status == "completed":
            agg["completed"] += 1
        agg["pct_sum"] += float(completion_pct) if completion_pct is not None else 0.0

    items: list[ChecklistCompletionRow] = []
    for dept_id, agg in sorted(dept_agg.items(), key=lambda x: x[1]["name"]):
        total = agg["total"]
        completed = agg["completed"]
        completion_rate = round(completed / total * 100, 2) if total > 0 else 0.0
        avg_pct = round(agg["pct_sum"] / total, 2) if total > 0 else 0.0
        items.append(ChecklistCompletionRow(
            department_id=dept_id,
            department_name=agg["name"],
            total_checklists=total,
            completed_count=completed,
            completion_rate=completion_rate,
            avg_completion_pct=avg_pct,
        ))

    return ChecklistCompletionReport(
        period_start=start_date,
        period_end=end_date,
        items=items,
    )


# ── C: Tỷ lệ vượt thử việc ────────────────────────────────────────────────────

async def get_pass_rate(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
) -> ProbationPassRateReport:
    base_stmt = (
        select(
            ProbationEvaluation,
            EmployeeJobRecord.department_id,
            EmployeeJobRecord.job_title_id,
        )
        .join(EmployeeJobRecord, ProbationEvaluation.job_record_id == EmployeeJobRecord.id)
        .where(
            ProbationEvaluation.status == "approved",
            ProbationEvaluation.evaluation_date >= start_date,
            ProbationEvaluation.evaluation_date <= end_date,
            ProbationEvaluation.result != "pending",
        )
    )
    if department_id is not None:
        base_stmt = base_stmt.where(EmployeeJobRecord.department_id == department_id)

    eval_rows = (await session.execute(base_stmt)).all()

    # Overall stats
    passed = sum(1 for pe, *_ in eval_rows if pe.result == "passed")
    failed = sum(1 for pe, *_ in eval_rows if pe.result == "failed")
    extended = sum(1 for pe, *_ in eval_rows if pe.result == "extended")
    total_decided = passed + failed + extended
    overall = ProbationPassRateStat(
        group_id=None,
        group_name="Toàn công ty",
        passed=passed,
        failed=failed,
        extended=extended,
        total_decided=total_decided,
        pass_rate=round(passed / total_decided * 100, 2) if total_decided > 0 else None,
    )

    # By department
    from collections import defaultdict
    dept_stats: dict[int, dict] = defaultdict(lambda: {"passed": 0, "failed": 0, "extended": 0})
    jt_stats: dict[int, dict] = defaultdict(lambda: {"passed": 0, "failed": 0, "extended": 0})
    monthly_stats: dict[tuple, dict] = defaultdict(lambda: {"passed": 0, "failed": 0, "extended": 0})

    for pe, dept_id, jt_id in eval_rows:
        if dept_id is not None:
            dept_stats[dept_id][pe.result] += 1
        if jt_id is not None:
            jt_stats[jt_id][pe.result] += 1
        ym = (pe.evaluation_date.year, pe.evaluation_date.month)
        monthly_stats[ym][pe.result] += 1

    # Fetch department names
    dept_ids = list(dept_stats.keys())
    dept_name_map: dict[int, str] = {}
    if dept_ids:
        dept_name_rows = (
            await session.execute(select(Department.id, Department.name).where(Department.id.in_(dept_ids)))
        ).all()
        dept_name_map = {r.id: r.name for r in dept_name_rows}

    # Fetch job title names
    jt_ids = list(jt_stats.keys())
    jt_name_map: dict[int, str] = {}
    if jt_ids:
        jt_name_rows = (
            await session.execute(select(JobTitle.id, JobTitle.name).where(JobTitle.id.in_(jt_ids)))
        ).all()
        jt_name_map = {r.id: r.name for r in jt_name_rows}

    by_department: list[ProbationPassRateStat] = []
    for dept_id, stats in sorted(dept_stats.items(), key=lambda x: dept_name_map.get(x[0], "")):
        p, f, e = stats["passed"], stats["failed"], stats["extended"]
        td = p + f + e
        by_department.append(ProbationPassRateStat(
            group_id=dept_id,
            group_name=dept_name_map.get(dept_id, f"Phòng ban {dept_id}"),
            passed=p,
            failed=f,
            extended=e,
            total_decided=td,
            pass_rate=round(p / td * 100, 2) if td > 0 else None,
        ))

    by_position: list[ProbationPassRateStat] = []
    for jt_id, stats in sorted(jt_stats.items(), key=lambda x: jt_name_map.get(x[0], "")):
        p, f, e = stats["passed"], stats["failed"], stats["extended"]
        td = p + f + e
        by_position.append(ProbationPassRateStat(
            group_id=jt_id,
            group_name=jt_name_map.get(jt_id, f"Chức danh {jt_id}"),
            passed=p,
            failed=f,
            extended=e,
            total_decided=td,
            pass_rate=round(p / td * 100, 2) if td > 0 else None,
        ))

    monthly_trend: list[MonthlyProbationTrend] = []
    for (yr, mo), stats in sorted(monthly_stats.items()):
        p, f, e = stats["passed"], stats["failed"], stats["extended"]
        monthly_trend.append(MonthlyProbationTrend(
            year=yr,
            month=mo,
            passed=p,
            failed=f,
            extended=e,
            total=p + f + e,
        ))

    return ProbationPassRateReport(
        period_start=start_date,
        period_end=end_date,
        overall=overall,
        by_department=by_department,
        by_position=by_position,
        monthly_trend=monthly_trend,
    )


# ── D: Phân tích lý do không đạt ─────────────────────────────────────────────

async def get_failure_reasons(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
) -> FailureReasonReport:
    stmt = (
        select(
            ProbationEvaluation,
            Employee.id.label("emp_id"),
            Employee.full_name.label("emp_name"),
        )
        .join(Employee, ProbationEvaluation.employee_id == Employee.id)
        .where(
            ProbationEvaluation.status == "approved",
            ProbationEvaluation.result == "failed",
            ProbationEvaluation.evaluation_date >= start_date,
            ProbationEvaluation.evaluation_date <= end_date,
        )
    )
    rows = (await session.execute(stmt)).all()

    total_failed = len(rows)
    comments = [(pe.manager_comment or "").lower() for pe, *_ in rows]

    keyword_counts: list[FailureKeywordCount] = []
    for kw in _FAILURE_KEYWORDS:
        cnt = sum(1 for c in comments if kw in c)
        pct = round(cnt / total_failed * 100, 2) if total_failed > 0 else 0.0
        keyword_counts.append(FailureKeywordCount(keyword=kw, count=cnt, pct=pct))

    keyword_counts.sort(key=lambda k: k.count, reverse=True)

    raw_comments: list[FailureCommentItem] = [
        FailureCommentItem(
            employee_id=emp_id,
            employee_name=emp_name,
            evaluation_date=pe.evaluation_date,
            manager_comment=pe.manager_comment,
        )
        for pe, emp_id, emp_name in rows
    ]

    return FailureReasonReport(
        total_failed=total_failed,
        reasons=keyword_counts,
        raw_comments=raw_comments,
    )


# ── Export Excel ──────────────────────────────────────────────────────────────

async def export_excel(
    session: AsyncSession,
    *,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
) -> io.BytesIO:
    if (end_date - start_date).days > 366:
        raise ValueError("Phạm vi tối đa 1 năm")

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Fetch all data
    active_report = await get_active_probation(session, department_id=department_id)
    checklist_report = await get_checklist_completion(
        session, start_date=start_date, end_date=end_date, department_id=department_id
    )
    pass_rate_report = await get_pass_rate(
        session, start_date=start_date, end_date=end_date, department_id=department_id
    )
    failure_report = await get_failure_reasons(session, start_date=start_date, end_date=end_date)

    wb = Workbook()

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    alt_fill = PatternFill("solid", fgColor="EEF2FF")

    def _write_sheet(ws, title: str, headers: list[str], data_rows: list[list]) -> None:
        ws.title = title
        # Header row
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

        for row_idx, row in enumerate(data_rows, start=2):
            for col_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(vertical="center")

        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(headers[col_idx - 1]))
            for row_idx in range(2, len(data_rows) + 2):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    # Sheet 1: Đang thử việc
    ws1 = wb.active
    _write_sheet(
        ws1,
        "Đang thử việc",
        ["Mã NV", "Họ tên", "Phòng ban", "Ngày bắt đầu TV", "Ngày kết thúc TV",
         "Còn (ngày)", "Mức độ", "Trạng thái OB", "Checklist %", "Kết quả đánh giá"],
        [
            [
                r.employee_code, r.employee_name, r.department_name or "—",
                str(r.probation_start_date) if r.probation_start_date else "",
                str(r.probation_end_date) if r.probation_end_date else "",
                r.days_remaining if r.days_remaining is not None else "",
                r.urgency, r.onboarding_status or "—",
                r.completion_pct if r.completion_pct is not None else "",
                r.evaluation_result,
            ]
            for r in active_report.items
        ],
    )

    # Sheet 2: Checklist
    ws2 = wb.create_sheet()
    _write_sheet(
        ws2,
        "Checklist",
        ["Phòng ban", "Tổng checklist", "Hoàn thành", "Tỷ lệ HT (%)", "TB tiến độ (%)"],
        [
            [r.department_name, r.total_checklists, r.completed_count,
             r.completion_rate, r.avg_completion_pct]
            for r in checklist_report.items
        ],
    )

    # Sheet 3: Tỷ lệ vượt TV
    ws3 = wb.create_sheet()
    _write_sheet(
        ws3,
        "Tỷ lệ vượt TV",
        ["Nhóm", "Đạt", "Không đạt", "Kéo dài", "Tổng", "Tỷ lệ đạt (%)"],
        [
            [r.group_name, r.passed, r.failed, r.extended, r.total_decided,
             r.pass_rate if r.pass_rate is not None else "N/A"]
            for r in [pass_rate_report.overall] + pass_rate_report.by_department
        ],
    )

    # Sheet 4: Lý do không đạt
    ws4 = wb.create_sheet()
    _write_sheet(
        ws4,
        "Lý do không đạt",
        ["Từ khóa", "Số lần", "Tỷ lệ (%)"],
        [
            [r.keyword, r.count, r.pct]
            for r in failure_report.reasons
        ],
    )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
