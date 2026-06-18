"""Service Leave Analytics (11.3) — tách hoàn toàn với leave_report_service."""
from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Optional, Sequence

import sqlalchemy as sa
from sqlalchemy import and_, false, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import LeaveType
from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.leave_entitlement import LeaveEntitlement
from app.models.leave_record import LeaveRecord
from app.models.org import Department
from app.schemas.leave_analytics import (
    DeptComparisonReport,
    DeptComparisonRow,
    ExpiringBalanceReport,
    ExpiringBalanceRow,
    HeatmapDeptRow,
    LeaveAnalyticsSummary,
    LeaveByTypeReport,
    LeaveTypeStatRow,
    MonthlyHeatmapReport,
    MonthlyTrendPoint,
    TopEmployeeRow,
    TopEmployeesReport,
)
from app.services import employee_code_service


def _to_float(v: object) -> float:
    if v is None:
        return 0.0
    return float(v)


def _calc_remaining(ent: LeaveEntitlement) -> float:
    return _to_float(ent.allocated_days) + _to_float(ent.carryover_days) - _to_float(ent.used_days)


async def get_analytics_summary(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
    leave_type_id: Optional[int] = None,
    allowed_department_ids: Optional[Sequence[int]] = None,
) -> LeaveAnalyticsSummary:
    today = date.today()

    # ── KPI 1: Tổng ngày nghỉ YTD ─────────────────────────────────────────────
    ytd_stmt = (
        select(func.coalesce(func.sum(LeaveRecord.total_days), 0))
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
    )
    if department_id is not None:
        ytd_stmt = (
            ytd_stmt
            .join(Employee, Employee.id == LeaveRecord.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            ytd_stmt = ytd_stmt.where(false())
        else:
            ytd_stmt = (
                ytd_stmt
                .join(Employee, Employee.id == LeaveRecord.employee_id)
                .join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,  # noqa: E712
                    ),
                )
                .where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
            )
    if leave_type_id is not None:
        ytd_stmt = ytd_stmt.where(LeaveRecord.leave_type_id == leave_type_id)

    total_days_ytd = _to_float((await session.execute(ytd_stmt)).scalar())

    # ── KPI 2: Tỷ lệ sử dụng phép bình quân ──────────────────────────────────
    usage_stmt = select(
        func.coalesce(func.sum(LeaveEntitlement.used_days), 0).label("used"),
        func.coalesce(
            func.sum(LeaveEntitlement.allocated_days + LeaveEntitlement.carryover_days), 0
        ).label("allocated"),
    ).where(LeaveEntitlement.year == year)
    if department_id is not None:
        usage_stmt = (
            usage_stmt
            .join(Employee, Employee.id == LeaveEntitlement.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            usage_stmt = usage_stmt.where(false())
        else:
            usage_stmt = (
                usage_stmt
                .join(Employee, Employee.id == LeaveEntitlement.employee_id)
                .join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,  # noqa: E712
                    ),
                )
                .where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
            )
    usage_row = (await session.execute(usage_stmt)).one()
    used_total = _to_float(usage_row.used)
    alloc_total = _to_float(usage_row.allocated)
    avg_usage_rate = (used_total / alloc_total * 100) if alloc_total > 0 else 0.0

    # ── KPI 3: Số NV chưa dùng phép ──────────────────────────────────────────
    not_taken_stmt = (
        select(func.count(sa.distinct(LeaveEntitlement.employee_id)))
        .join(LeaveType, LeaveType.id == LeaveEntitlement.leave_type_id)
        .where(
            LeaveEntitlement.year == year,
            LeaveEntitlement.used_days == 0,
            LeaveType.carryover_allowed == True,  # noqa: E712
        )
    )
    if department_id is not None:
        not_taken_stmt = (
            not_taken_stmt
            .join(Employee, Employee.id == LeaveEntitlement.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            not_taken_stmt = not_taken_stmt.where(false())
        else:
            not_taken_stmt = (
                not_taken_stmt
                .join(Employee, Employee.id == LeaveEntitlement.employee_id)
                .join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,  # noqa: E712
                    ),
                )
                .where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
            )
    employees_not_taken: int = (await session.execute(not_taken_stmt)).scalar() or 0

    # ── KPI 4: Ngày tồn sắp hết hạn (30 ngày tới) ────────────────────────────
    expire_cutoff = today + timedelta(days=30)
    expiring_stmt = select(
        func.coalesce(
            func.sum(
                LeaveEntitlement.allocated_days
                + LeaveEntitlement.carryover_days
                - LeaveEntitlement.used_days
            ),
            0,
        )
    ).where(
        LeaveEntitlement.year == year,
        LeaveEntitlement.carryover_expires.between(today, expire_cutoff),
        (LeaveEntitlement.allocated_days + LeaveEntitlement.carryover_days - LeaveEntitlement.used_days) > 0,
    )
    days_expiring_30d = _to_float((await session.execute(expiring_stmt)).scalar())

    # ── Trend 12 tháng ────────────────────────────────────────────────────────
    trend_base = (
        select(
            func.extract("month", LeaveRecord.start_date).cast(sa.Integer).label("month"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_days"),
            func.count(LeaveRecord.id).label("total_records"),
            func.count(sa.distinct(LeaveRecord.employee_id)).label("employee_count"),
        )
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(sa.text("1"))
        .order_by(sa.text("1"))
    )
    if department_id is not None:
        trend_base = (
            trend_base
            .join(Employee, Employee.id == LeaveRecord.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            trend_base = trend_base.where(false())
        else:
            trend_base = (
                trend_base
                .join(Employee, Employee.id == LeaveRecord.employee_id)
                .join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,  # noqa: E712
                    ),
                )
                .where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
            )
    if leave_type_id is not None:
        trend_base = trend_base.where(LeaveRecord.leave_type_id == leave_type_id)

    trend_rows = (await session.execute(trend_base)).all()
    trend_map = {r.month: r for r in trend_rows}
    monthly_trend = [
        MonthlyTrendPoint(
            month=m,
            total_days=_to_float(trend_map[m].total_days) if m in trend_map else 0.0,
            total_records=trend_map[m].total_records if m in trend_map else 0,
            employee_count=trend_map[m].employee_count if m in trend_map else 0,
        )
        for m in range(1, 13)
    ]

    return LeaveAnalyticsSummary(
        year=year,
        department_id=department_id,
        total_days_ytd=total_days_ytd,
        avg_usage_rate=round(avg_usage_rate, 2),
        employees_not_taken=employees_not_taken,
        days_expiring_30d=days_expiring_30d,
        monthly_trend=monthly_trend,
    )


async def get_by_type(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
    allowed_department_ids: Optional[Sequence[int]] = None,
) -> LeaveByTypeReport:
    stmt = (
        select(
            LeaveType.id.label("lt_id"),
            LeaveType.name.label("lt_name"),
            LeaveType.color_tag,
            func.count(LeaveRecord.id).label("record_count"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_days"),
            func.count(sa.distinct(LeaveRecord.employee_id)).label("unique_employees"),
        )
        .join(LeaveRecord, LeaveRecord.leave_type_id == LeaveType.id)
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(LeaveType.id, LeaveType.name, LeaveType.color_tag)
        .order_by(sa.desc("total_days"))
    )
    if department_id is not None:
        stmt = (
            stmt
            .join(Employee, Employee.id == LeaveRecord.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            stmt = stmt.where(false())
        else:
            stmt = (
                stmt
                .join(Employee, Employee.id == LeaveRecord.employee_id)
                .join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,  # noqa: E712
                    ),
                )
                .where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
            )

    rows = (await session.execute(stmt)).all()
    grand_total = sum(_to_float(r.total_days) for r in rows)

    items = [
        LeaveTypeStatRow(
            leave_type_id=r.lt_id,
            leave_type_name=r.lt_name,
            color_tag=r.color_tag,
            record_count=r.record_count,
            total_days=_to_float(r.total_days),
            unique_employees=r.unique_employees,
            percentage=round(_to_float(r.total_days) / grand_total * 100, 2) if grand_total > 0 else 0.0,
        )
        for r in rows
    ]
    return LeaveByTypeReport(
        year=year,
        department_id=department_id,
        items=items,
        grand_total_days=grand_total,
    )


async def get_monthly_heatmap(
    session: AsyncSession,
    *,
    year: int,
    allowed_department_ids: Optional[Sequence[int]] = None,
) -> MonthlyHeatmapReport:
    stmt = (
        select(
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            func.extract("month", LeaveRecord.start_date).cast(sa.Integer).label("month"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_days"),
        )
        .join(Employee, Employee.id == LeaveRecord.employee_id)
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(Department.id, Department.name, sa.text("3"))
        .order_by(Department.name.nulls_last(), sa.text("3"))
    )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            stmt = stmt.where(false())
        else:
            stmt = stmt.where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
    rows = (await session.execute(stmt)).all()

    # Pivot thành dict[dept_id][month]
    dept_map: dict[int | None, dict] = {}
    for r in rows:
        key = r.dept_id
        if key not in dept_map:
            dept_map[key] = {"dept_name": r.dept_name, "months": {}}
        dept_map[key]["months"][r.month] = _to_float(r.total_days)

    departments = []
    company_monthly: dict[int, float] = {m: 0.0 for m in range(1, 13)}

    for dept_id, info in dept_map.items():
        monthly_days: dict[int, float] = {}
        for m in range(1, 13):
            val = info["months"].get(m, 0.0)
            monthly_days[m] = val
            company_monthly[m] += val
        departments.append(HeatmapDeptRow(
            dept_id=dept_id,
            dept_name=info["dept_name"],
            monthly_days=monthly_days,
            annual_total=sum(monthly_days.values()),
        ))

    return MonthlyHeatmapReport(
        year=year,
        departments=departments,
        company_monthly=company_monthly,
    )


async def get_top_employees(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
    leave_type_id: Optional[int] = None,
    allowed_department_ids: Optional[Sequence[int]] = None,
    limit: int = 10,
) -> TopEmployeesReport:
    stmt = (
        select(
            Employee,
            Department.name.label("dept_name"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_days_taken"),
            func.count(LeaveRecord.id).label("record_count"),
            func.coalesce(
                func.max(LeaveEntitlement.allocated_days + LeaveEntitlement.carryover_days), 0
            ).label("total_entitled"),
        )
        .join(LeaveRecord, LeaveRecord.employee_id == Employee.id)
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .outerjoin(
            LeaveEntitlement,
            and_(
                LeaveEntitlement.employee_id == Employee.id,
                LeaveEntitlement.year == year,
            ),
        )
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(Employee.id, Department.name)
        .order_by(sa.desc("total_days_taken"))
        .limit(min(limit, 50))
    )
    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            stmt = stmt.where(false())
        else:
            stmt = stmt.where(EmployeeJobRecord.department_id.in_(allowed_department_ids))
    if leave_type_id is not None:
        stmt = stmt.where(LeaveRecord.leave_type_id == leave_type_id)

    rows = (await session.execute(stmt)).all()

    employees_list = [r[0] for r in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees_list)

    items = []
    for idx, row in enumerate(rows):
        emp = row[0]
        total_entitled = _to_float(row.total_entitled)
        total_taken = _to_float(row.total_days_taken)
        items.append(TopEmployeeRow(
            rank=idx + 1,
            employee_id=emp.id,
            employee_code=code_map.get(emp.id, ""),
            employee_name=emp.full_name,
            dept_name=row.dept_name,
            total_days_taken=total_taken,
            record_count=row.record_count,
            total_entitled=total_entitled,
            usage_rate=round(total_taken / total_entitled * 100, 2) if total_entitled > 0 else 0.0,
        ))

    return TopEmployeesReport(year=year, department_id=department_id, items=items)


async def get_expiring_balance(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
    allowed_department_ids: Optional[Sequence[int]] = None,
    expire_days: int = 30,
) -> ExpiringBalanceReport:
    today = date.today()
    cutoff = today + timedelta(days=expire_days)

    stmt = (
        select(
            Employee,
            Department.name.label("dept_name"),
            LeaveType.name.label("lt_name"),
            LeaveEntitlement,
        )
        .join(Employee, Employee.id == LeaveEntitlement.employee_id)
        .join(LeaveType, LeaveType.id == LeaveEntitlement.leave_type_id)
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            LeaveEntitlement.year == year,
            LeaveEntitlement.carryover_expires.isnot(None),
            LeaveEntitlement.carryover_expires > today,
            LeaveEntitlement.carryover_expires <= cutoff,
            (LeaveEntitlement.allocated_days + LeaveEntitlement.carryover_days - LeaveEntitlement.used_days) > 0,
        )
        .order_by(LeaveEntitlement.carryover_expires)
    )
    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            stmt = stmt.where(false())
        else:
            stmt = stmt.where(EmployeeJobRecord.department_id.in_(allowed_department_ids))

    rows = (await session.execute(stmt)).all()

    employees_list = [r[0] for r in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees_list)

    items = []
    total_expiring = 0.0
    for row in rows:
        emp: Employee = row[0]
        ent: LeaveEntitlement = row[3]
        remaining = _calc_remaining(ent)
        total_expiring += remaining
        days_until = (ent.carryover_expires - today).days if ent.carryover_expires else 0
        items.append(ExpiringBalanceRow(
            employee_id=emp.id,
            employee_code=code_map.get(emp.id, ""),
            employee_name=emp.full_name,
            dept_name=row.dept_name,
            leave_type_name=row.lt_name,
            carryover_days=_to_float(ent.carryover_days),
            remaining_days=remaining,
            carryover_expires=ent.carryover_expires,
            days_until_expire=days_until,
        ))

    return ExpiringBalanceReport(
        expire_days=expire_days,
        year=year,
        department_id=department_id,
        items=items,
        total_expiring_days=round(total_expiring, 2),
    )


async def get_department_comparison(
    session: AsyncSession,
    *,
    year: int,
    allowed_department_ids: Optional[Sequence[int]] = None,
) -> DeptComparisonReport:
    # Lấy tổng ngày nghỉ và số NV theo phòng ban
    taken_stmt = (
        select(
            Department.id.label("dept_id"),
            Department.name.label("dept_name"),
            func.count(sa.distinct(LeaveRecord.employee_id)).label("employee_count"),
            func.coalesce(func.sum(LeaveRecord.total_days), 0).label("total_taken"),
        )
        .join(Employee, Employee.id == LeaveRecord.employee_id)
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            LeaveRecord.status == "active",
            func.extract("year", LeaveRecord.start_date) == year,
        )
        .group_by(Department.id, Department.name)
    )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            taken_stmt = taken_stmt.where(false())
        else:
            taken_stmt = taken_stmt.where(EmployeeJobRecord.department_id.in_(allowed_department_ids))

    # Lấy tổng phép được cấp theo phòng ban
    alloc_stmt = (
        select(
            EmployeeJobRecord.department_id.label("dept_id"),
            func.coalesce(
                func.sum(LeaveEntitlement.allocated_days + LeaveEntitlement.carryover_days), 0
            ).label("allocated_total"),
        )
        .join(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == LeaveEntitlement.employee_id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .where(LeaveEntitlement.year == year)
        .group_by(EmployeeJobRecord.department_id)
    )
    if allowed_department_ids is not None:
        if not allowed_department_ids:
            alloc_stmt = alloc_stmt.where(false())
        else:
            alloc_stmt = alloc_stmt.where(EmployeeJobRecord.department_id.in_(allowed_department_ids))

    taken_rows = (await session.execute(taken_stmt)).all()
    alloc_rows = (await session.execute(alloc_stmt)).all()
    alloc_map = {r.dept_id: _to_float(r.allocated_total) for r in alloc_rows}

    items = []
    for r in taken_rows:
        alloc = alloc_map.get(r.dept_id, 0.0)
        taken = _to_float(r.total_taken)
        avg = taken / r.employee_count if r.employee_count > 0 else 0.0
        usage = round(taken / alloc * 100, 2) if alloc > 0 else 0.0
        items.append(DeptComparisonRow(
            dept_id=r.dept_id,
            dept_name=r.dept_name,
            employee_count=r.employee_count,
            total_days_taken=taken,
            avg_days_per_employee=round(avg, 2),
            allocated_total=alloc,
            usage_rate=usage,
        ))

    items.sort(key=lambda x: x.avg_days_per_employee, reverse=True)
    return DeptComparisonReport(year=year, items=items)


def build_analytics_xlsx(
    summary: LeaveAnalyticsSummary,
    by_type: LeaveByTypeReport,
    heatmap: MonthlyHeatmapReport,
    top: TopEmployeesReport,
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def _header_style(cell: object) -> None:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center")

    def _autofit(ws: object) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Tổng quan KPI ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng quan"
    kpi_headers = ["Chỉ số", "Giá trị"]
    ws1.append(kpi_headers)
    for cell in ws1[1]:
        _header_style(cell)
    ws1.append(["Năm báo cáo", summary.year])
    ws1.append(["Tổng ngày nghỉ YTD", summary.total_days_ytd])
    ws1.append(["Tỷ lệ sử dụng phép (%)", summary.avg_usage_rate])
    ws1.append(["Nhân viên chưa nghỉ (NV)", summary.employees_not_taken])
    ws1.append(["Ngày tồn sắp hết hạn (30 ngày)", summary.days_expiring_30d])
    ws1.append([])
    ws1.append(["Xu hướng theo tháng"])
    ws1.append(["Tháng", "Tổng ngày", "Số lần", "Số NV"])
    for cell in ws1[ws1.max_row]:
        _header_style(cell)
    for pt in summary.monthly_trend:
        ws1.append([pt.month, pt.total_days, pt.total_records, pt.employee_count])
    _autofit(ws1)

    # ── Sheet 2: Theo loại phép ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Theo loại phép")
    ws2.append(["Loại phép", "Số lần", "Tổng ngày", "Số NV", "Tỷ lệ (%)"])
    for cell in ws2[1]:
        _header_style(cell)
    for item in by_type.items:
        ws2.append([
            item.leave_type_name, item.record_count,
            item.total_days, item.unique_employees, item.percentage,
        ])
    _autofit(ws2)

    # ── Sheet 3: Heatmap tháng ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Heatmap tháng")
    month_headers = ["Phòng ban"] + [f"T{m}" for m in range(1, 13)] + ["Tổng năm"]
    ws3.append(month_headers)
    for cell in ws3[1]:
        _header_style(cell)
    for dept in heatmap.departments:
        row = [dept.dept_name or "Không xác định"]
        for m in range(1, 13):
            row.append(dept.monthly_days.get(m, 0.0))
        row.append(dept.annual_total)
        ws3.append(row)
    # Dòng tổng công ty
    total_row = ["TỔNG CÔNG TY"]
    for m in range(1, 13):
        total_row.append(heatmap.company_monthly.get(m, 0.0))
    total_row.append(sum(heatmap.company_monthly.values()))
    ws3.append(total_row)
    for cell in ws3[ws3.max_row]:
        cell.font = Font(bold=True)
    _autofit(ws3)

    # ── Sheet 4: Top nhân viên ────────────────────────────────────────────────
    ws4 = wb.create_sheet("Top nhân viên")
    ws4.append(["#", "Mã NV", "Họ tên", "Phòng ban", "Tổng ngày", "Số lần", "Được cấp", "Tỷ lệ (%)"])
    for cell in ws4[1]:
        _header_style(cell)
    for item in top.items:
        ws4.append([
            item.rank, item.employee_code, item.employee_name,
            item.dept_name or "—", item.total_days_taken,
            item.record_count, item.total_entitled, item.usage_rate,
        ])
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
