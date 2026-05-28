"""Service Insurance Analytics (11.4)."""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Optional

import sqlalchemy as sa
from sqlalchemy import and_, func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.employee_insurance import EmployeeInsuranceProfile
from app.models.insurance import InsuranceChangeEvent
from app.models.org import Department
from app.schemas.insurance_analytics import (
    InsuranceDashboardKPI,
    MonthlyChangePoint,
    InsuranceMonthlyChangesResponse,
    PayrollFundPoint,
    InsurancePayrollFundResponse,
    NonParticipantRow,
    InsuranceNonParticipantsResponse,
    DepartmentBreakdownRow,
    InsuranceDepartmentBreakdownResponse,
    EmployeeHistoryPoint,
    InsuranceEmployeeHistoryResponse,
)
from app.services import employee_code_service


async def get_dashboard(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
) -> InsuranceDashboardKPI:
    # ── KPI 1 & 2: Số nhân sự đang tham gia & Tổng số nhân sự active ─────────
    total_active_stmt = select(func.count(Employee.id)).where(Employee.is_active == True)
    
    participating_stmt = (
        select(func.count(Employee.id))
        .join(EmployeeInsuranceProfile, EmployeeInsuranceProfile.employee_id == Employee.id)
        .where(
            Employee.is_active == True,
            EmployeeInsuranceProfile.participation_status == "active",
        )
    )

    if department_id is not None:
        total_active_stmt = (
            total_active_stmt.join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            ).where(EmployeeJobRecord.department_id == department_id)
        )
        participating_stmt = (
            participating_stmt.join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            ).where(EmployeeJobRecord.department_id == department_id)
        )

    total_active_employees = (await session.execute(total_active_stmt)).scalar() or 0
    participating_count = (await session.execute(participating_stmt)).scalar() or 0
    participation_rate = (
        round(participating_count / total_active_employees * 100, 2)
        if total_active_employees > 0
        else 0.0
    )

    # ── KPI 3: Tổng quỹ lương BHXH ──────────────────────────────
    total_basis_stmt = (
        select(func.coalesce(func.sum(EmployeeInsuranceProfile.insurance_basis_amount), 0))
        .join(Employee, Employee.id == EmployeeInsuranceProfile.employee_id)
        .where(
            Employee.is_active == True,
            EmployeeInsuranceProfile.participation_status == "active",
            EmployeeInsuranceProfile.insurance_basis_amount.isnot(None),
        )
    )

    if department_id is not None:
        total_basis_stmt = (
            total_basis_stmt.join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            ).where(EmployeeJobRecord.department_id == department_id)
        )

    total_basis_amount = (await session.execute(total_basis_stmt)).scalar() or Decimal("0")

    # ── KPI 4 & 5: Số tăng/giảm biến động tháng ────────────────────────────────────
    increased_stmt = select(func.count(InsuranceChangeEvent.id)).where(
        InsuranceChangeEvent.change_type == "increase",
        func.extract("year", InsuranceChangeEvent.effective_date) == year,
        func.extract("month", InsuranceChangeEvent.effective_date) == month,
    )
    
    decreased_stmt = select(func.count(InsuranceChangeEvent.id)).where(
        InsuranceChangeEvent.change_type == "decrease",
        func.extract("year", InsuranceChangeEvent.effective_date) == year,
        func.extract("month", InsuranceChangeEvent.effective_date) == month,
    )

    if department_id is not None:
        increased_stmt = (
            increased_stmt.join(Employee, Employee.id == InsuranceChangeEvent.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )
        decreased_stmt = (
            decreased_stmt.join(Employee, Employee.id == InsuranceChangeEvent.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )

    increased_count = (await session.execute(increased_stmt)).scalar() or 0
    decreased_count = (await session.execute(decreased_stmt)).scalar() or 0

    return InsuranceDashboardKPI(
        year=year,
        month=month,
        department_id=department_id,
        participating_count=participating_count,
        total_active_employees=total_active_employees,
        participation_rate=participation_rate,
        total_basis_amount=total_basis_amount,
        increased_count=increased_count,
        decreased_count=decreased_count,
        net_change=increased_count - decreased_count,
    )


async def get_monthly_changes(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
) -> InsuranceMonthlyChangesResponse:
    stmt = (
        select(
            func.extract("month", InsuranceChangeEvent.effective_date).cast(sa.Integer).label("month"),
            func.coalesce(func.sum(sa.case((InsuranceChangeEvent.change_type == "increase", 1), else_=0)), 0).label("increased"),
            func.coalesce(func.sum(sa.case((InsuranceChangeEvent.change_type == "decrease", 1), else_=0)), 0).label("decreased"),
        )
        .where(func.extract("year", InsuranceChangeEvent.effective_date) == year)
        .group_by(sa.text("1"))
    )

    if department_id is not None:
        stmt = (
            stmt.join(Employee, Employee.id == InsuranceChangeEvent.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )

    rows = (await session.execute(stmt)).all()
    month_map = {m: {"increased": 0, "decreased": 0} for m in range(1, 13)}
    
    for r in rows:
        month_map[r.month] = {
            "increased": int(r.increased or 0),
            "decreased": int(r.decreased or 0),
        }

    data = [
        MonthlyChangePoint(
            month=m,
            increased=month_map[m]["increased"],
            decreased=month_map[m]["decreased"],
            net=month_map[m]["increased"] - month_map[m]["decreased"],
        )
        for m in range(1, 13)
    ]

    return InsuranceMonthlyChangesResponse(
        year=year,
        department_id=department_id,
        data=data,
    )


async def get_payroll_fund(
    session: AsyncSession,
    *,
    year: int,
    department_id: Optional[int] = None,
) -> InsurancePayrollFundResponse:
    stmt = (
        select(
            func.extract("month", InsuranceChangeEvent.effective_date).cast(sa.Integer).label("month"),
            func.coalesce(func.sum(sa.case((InsuranceChangeEvent.change_type == "increase", InsuranceChangeEvent.basis_amount), else_=0)), 0).label("added_amount"),
            func.coalesce(func.sum(sa.case((InsuranceChangeEvent.change_type == "decrease", InsuranceChangeEvent.basis_amount), else_=0)), 0).label("removed_amount"),
        )
        .where(func.extract("year", InsuranceChangeEvent.effective_date) == year)
        .group_by(sa.text("1"))
    )

    if department_id is not None:
        stmt = (
            stmt.join(Employee, Employee.id == InsuranceChangeEvent.employee_id)
            .join(
                EmployeeJobRecord,
                and_(
                    EmployeeJobRecord.employee_id == Employee.id,
                    EmployeeJobRecord.is_current == True,
                ),
            )
            .where(EmployeeJobRecord.department_id == department_id)
        )

    rows = (await session.execute(stmt)).all()
    month_map = {m: {"added": Decimal("0"), "removed": Decimal("0")} for m in range(1, 13)}
    
    for r in rows:
        month_map[r.month] = {
            "added": Decimal(r.added_amount or 0),
            "removed": Decimal(r.removed_amount or 0),
        }

    # Tính snapshot tổng quỹ hiện hành cho tháng hiện tại nếu truy vấn năm hiện tại
    today = date.today()
    current_month_snapshot = None
    if year == today.year:
        snap_stmt = (
            select(func.coalesce(func.sum(EmployeeInsuranceProfile.insurance_basis_amount), 0))
            .join(Employee, Employee.id == EmployeeInsuranceProfile.employee_id)
            .where(
                Employee.is_active == True,
                EmployeeInsuranceProfile.participation_status == "active",
                EmployeeInsuranceProfile.insurance_basis_amount.isnot(None),
            )
        )
        if department_id is not None:
            snap_stmt = (
                snap_stmt.join(
                    EmployeeJobRecord,
                    and_(
                        EmployeeJobRecord.employee_id == Employee.id,
                        EmployeeJobRecord.is_current == True,
                    ),
                ).where(EmployeeJobRecord.department_id == department_id)
            )
        current_month_snapshot = (await session.execute(snap_stmt)).scalar() or Decimal("0")

    data = []
    for m in range(1, 13):
        snap_val = (
            current_month_snapshot
            if (year == today.year and m == today.month)
            else None
        )
        data.append(
            PayrollFundPoint(
                month=m,
                added_amount=month_map[m]["added"],
                removed_amount=month_map[m]["removed"],
                snapshot_amount=snap_val,
            )
        )

    return InsurancePayrollFundResponse(
        year=year,
        department_id=department_id,
        data=data,
    )


async def get_non_participants(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 10,
) -> InsuranceNonParticipantsResponse:
    # Tìm nhân viên active nhưng không tham gia BHXH hoặc chưa có hồ sơ
    base_query = (
        select(
            Employee,
            Department.name.label("department_name"),
            EmployeeInsuranceProfile,
        )
        .outerjoin(EmployeeInsuranceProfile, EmployeeInsuranceProfile.employee_id == Employee.id)
        .join(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,
            ),
        )
        .join(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            Employee.is_active == True,
            or_(
                EmployeeInsuranceProfile.id.is_(None),
                EmployeeInsuranceProfile.participation_status != "active",
            ),
        )
    )

    if department_id is not None:
        base_query = base_query.where(EmployeeJobRecord.department_id == department_id)

    # Đếm tổng số lượng cho phân trang
    count_stmt = select(func.count()).select_from(base_query.subquery())
    total = (await session.execute(count_stmt)).scalar() or 0

    # Phân trang và sắp xếp
    stmt = (
        base_query.order_by(Department.name, Employee.full_name)
        .offset((page - 1) * page_size)
        .limit(page_size)
    )
    rows = (await session.execute(stmt)).all()

    employees_list = [r[0] for r in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees_list)

    items = []
    for r in rows:
        emp: Employee = r[0]
        dept_name: str = r[1]
        profile: Optional[EmployeeInsuranceProfile] = r[2]
        
        items.append(
            NonParticipantRow(
                employee_id=emp.id,
                employee_code=code_map.get(emp.id, ""),
                full_name=emp.full_name,
                department_name=dept_name,
                participation_status=profile.participation_status if profile else None,
                status_effective_from=profile.status_effective_from if profile else None,
                status_note=profile.status_note if profile else None,
                company_bhxh_joined_date=profile.company_bhxh_joined_date if profile else None,
            )
        )

    return InsuranceNonParticipantsResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


async def get_department_breakdown(
    session: AsyncSession,
    *,
    year: int,
    month: int,
) -> InsuranceDepartmentBreakdownResponse:
    stmt = (
        select(
            Department.id.label("department_id"),
            Department.name.label("department_name"),
            func.count(Employee.id)
            .filter(EmployeeInsuranceProfile.participation_status == "active")
            .label("participating_count"),
            func.count(Employee.id).label("total_count"),
            func.coalesce(
                func.sum(EmployeeInsuranceProfile.insurance_basis_amount).filter(
                    EmployeeInsuranceProfile.participation_status == "active"
                ),
                0,
            ).label("total_basis_amount"),
        )
        .select_from(Department)
        .join(EmployeeJobRecord, EmployeeJobRecord.department_id == Department.id)
        .join(
            Employee,
            and_(
                Employee.id == EmployeeJobRecord.employee_id,
                Employee.is_active == True,
                EmployeeJobRecord.is_current == True,
            ),
        )
        .outerjoin(EmployeeInsuranceProfile, EmployeeInsuranceProfile.employee_id == Employee.id)
        .group_by(Department.id, Department.name)
        .order_by(Department.name)
    )

    rows = (await session.execute(stmt)).all()
    items = []
    for r in rows:
        total_c = r.total_count or 0
        part_c = r.participating_count or 0
        rate = round(part_c / total_c * 100, 2) if total_c > 0 else 0.0
        items.append(
            DepartmentBreakdownRow(
                department_id=r.department_id,
                department_name=r.department_name,
                participating_count=part_c,
                total_count=total_c,
                participation_rate=rate,
                total_basis_amount=Decimal(r.total_basis_amount or 0),
            )
        )

    return InsuranceDepartmentBreakdownResponse(
        year=year,
        month=month,
        items=items,
    )


async def get_employee_history(
    session: AsyncSession,
    *,
    employee_id: int,
    year: Optional[int] = None,
) -> InsuranceEmployeeHistoryResponse:
    emp_stmt = select(Employee).where(Employee.id == employee_id)
    emp = (await session.execute(emp_stmt)).scalar_one_or_none()
    if not emp:
        raise ValueError("Nhân viên không tồn tại")

    prof_stmt = select(EmployeeInsuranceProfile).where(
        EmployeeInsuranceProfile.employee_id == employee_id
    )
    profile = (await session.execute(prof_stmt)).scalar_one_or_none()

    stmt = select(InsuranceChangeEvent).where(InsuranceChangeEvent.employee_id == employee_id)
    if year is not None:
        stmt = stmt.where(func.extract("year", InsuranceChangeEvent.effective_date) == year)
    stmt = stmt.order_by(InsuranceChangeEvent.effective_date.desc())
    events = (await session.execute(stmt)).scalars().all()

    history = [
        EmployeeHistoryPoint(
            effective_date=ev.effective_date,
            change_type=ev.change_type,
            change_reason=ev.change_reason,
            basis_amount=ev.basis_amount,
            participation_status_after=ev.new_status,
        )
        for ev in events
    ]

    return InsuranceEmployeeHistoryResponse(
        employee_id=employee_id,
        full_name=emp.full_name,
        current_participation_status=profile.participation_status if profile else None,
        current_basis_amount=profile.insurance_basis_amount if profile else None,
        history=history,
    )


async def export_analytics_xlsx(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Lấy dữ liệu báo cáo
    dashboard = await get_dashboard(session, year=year, month=month, department_id=department_id)
    breakdown = await get_department_breakdown(session, year=year, month=month)
    # Lấy toàn bộ danh sách không tham gia (không phân trang bằng page_size lớn)
    non_participants = await get_non_participants(
        session, department_id=department_id, page=1, page_size=10000
    )

    wb = Workbook()

    def _header_style(cell: object) -> None:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A8A")  # Xanh dương đậm sang trọng
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autofit(ws: object) -> None:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ── Sheet 1: Tổng quan KPI ──
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws1.append(["BÁO CÁO TỔNG QUAN BẢO HIỂM", ""])
    ws1.cell(1, 1).font = Font(bold=True, size=14)
    ws1.append(["Kỳ báo cáo", f"Tháng {month}/{year}"])
    ws1.append([])

    ws1.append(["Chỉ số", "Giá trị"])
    for cell in ws1[ws1.max_row]:
        _header_style(cell)

    ws1.append(["Nhân sự tham gia BHXH", dashboard.participating_count])
    ws1.append(["Tổng số nhân sự active", dashboard.total_active_employees])
    ws1.append(["Tỷ lệ tham gia (%)", f"{dashboard.participation_rate}%"])
    ws1.append(["Tổng quỹ lương BHXH (VND)", dashboard.total_basis_amount])
    ws1.append(["Số lượng tăng trong tháng", dashboard.increased_count])
    ws1.append(["Số lượng giảm trong tháng", dashboard.decreased_count])
    ws1.append(["Biến động ròng", dashboard.net_change])
    
    # Format số tiền cho quỹ lương
    ws1.cell(ws1.max_row - 3, 2).number_format = "#,##0"
    _autofit(ws1)

    # ── Sheet 2: Theo phòng ban ──
    ws2 = wb.create_sheet("Cơ cấu phòng ban")
    ws2.append(["Mã phòng ban", "Tên phòng ban", "Số NV tham gia", "Tổng số NV", "Tỷ lệ tham gia (%)", "Tổng quỹ lương BHXH (VND)"])
    for cell in ws2[1]:
        _header_style(cell)
    
    for dept in breakdown.items:
        ws2.append([
            dept.department_id,
            dept.department_name,
            dept.participating_count,
            dept.total_count,
            f"{dept.participation_rate}%",
            dept.total_basis_amount,
        ])
        ws2.cell(ws2.max_row, 6).number_format = "#,##0"
    _autofit(ws2)

    # ── Sheet 3: Không tham gia BHXH ──
    ws3 = wb.create_sheet("Không tham gia BHXH")
    ws3.append(["Mã NV", "Họ và tên", "Phòng ban", "Trạng thái bảo hiểm", "Ngày hiệu lực", "Ngày tham gia BHXH công ty", "Ghi chú"])
    for cell in ws3[1]:
        _header_style(cell)
        
    for row in non_participants.items:
        ws3.append([
            row.employee_code,
            row.full_name,
            row.department_name,
            row.participation_status or "Chưa có hồ sơ",
            row.status_effective_from.strftime("%d/%m/%Y") if row.status_effective_from else "—",
            row.company_bhxh_joined_date.strftime("%d/%m/%Y") if row.company_bhxh_joined_date else "—",
            row.status_note or "—",
        ])
    _autofit(ws3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
