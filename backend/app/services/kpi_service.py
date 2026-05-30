"""Service quản lý KPI tháng nhân viên (10.1)."""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from decimal import Decimal
from typing import List, Optional

from fastapi import HTTPException, status
from sqlalchemy import String, and_, func, or_, select
from sqlalchemy import cast as sa_cast
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auth import User
from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.org import Department
from app.models.performance import EmployeeKpiMonthly
from app.schemas.performance import (
    KpiMonthlyCreate,
    KpiMonthlyListPage,
    KpiMonthlyRead,
    KpiMonthlyUpdate,
)
from app.services import employee_code_service

log = logging.getLogger(__name__)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


async def _get_or_404(session: AsyncSession, kpi_id: int) -> EmployeeKpiMonthly:
    kpi = await session.get(EmployeeKpiMonthly, kpi_id)
    if not kpi:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy bản ghi KPI")
    return kpi


async def _build_read(session: AsyncSession, kpi: EmployeeKpiMonthly) -> KpiMonthlyRead:
    emp = await session.get(Employee, kpi.employee_id)
    creator: Optional[User] = await session.get(User, kpi.created_by_id) if kpi.created_by_id else None

    emp_code = ""
    dept_name: Optional[str] = None
    if emp:
        emp_code = await employee_code_service.build_employee_display_code(session, emp)
        jr = (
            await session.execute(
                select(EmployeeJobRecord)
                .where(
                    EmployeeJobRecord.employee_id == emp.id,
                    EmployeeJobRecord.is_current == True,  # noqa: E712
                )
                .limit(1)
            )
        ).scalar_one_or_none()
        if jr and jr.department_id:
            dept = await session.get(Department, jr.department_id)
            dept_name = dept.name if dept else None

    return KpiMonthlyRead(
        id=kpi.id,
        employee_id=kpi.employee_id,
        employee_code=emp_code,
        employee_name=emp.full_name if emp else "",
        department_name=dept_name,
        year=kpi.year,
        month=kpi.month,
        score=kpi.score,
        note=kpi.note,
        created_by_name=getattr(creator, "full_name", None) or (creator.email if creator else None),
        created_at=kpi.created_at,
        updated_at=kpi.updated_at,
    )


async def get_kpi_list(
    session: AsyncSession,
    *,
    year: Optional[int] = None,
    month: Optional[int] = None,
    department_id: Optional[int] = None,
    search: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> KpiMonthlyListPage:
    stmt = (
        select(EmployeeKpiMonthly)
        .join(Employee, Employee.id == EmployeeKpiMonthly.employee_id)
        .outerjoin(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,  # noqa: E712
            ),
        )
        .outerjoin(Department, Department.id == EmployeeJobRecord.department_id)
    )

    if year is not None:
        stmt = stmt.where(EmployeeKpiMonthly.year == year)
    if month is not None:
        stmt = stmt.where(EmployeeKpiMonthly.month == month)
    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    if search:
        from app.services.administrative_import_service import normalize_text
        kw = f"%{search.strip()}%"
        norm_kw = f"%{normalize_text(search.strip())}%"
        dept_prefix = func.coalesce(
            func.nullif(func.btrim(Department.display_prefix), ""),
            Department.code,
        )
        generated_code = dept_prefix + func.lpad(
            sa_cast(Employee.employee_seq, String), 4, "0"
        )
        stmt = stmt.where(
            or_(
                Employee.full_name.ilike(kw),
                Employee.normalized_name.ilike(norm_kw),
                generated_code.ilike(kw),
            )
        )

    total_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await session.execute(total_stmt)).scalar_one()

    stmt = stmt.order_by(
        EmployeeKpiMonthly.year.desc(),
        EmployeeKpiMonthly.month.desc(),
        EmployeeKpiMonthly.id.desc(),
    )
    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    rows = (await session.execute(stmt)).scalars().all()

    items: List[KpiMonthlyRead] = []
    for row in rows:
        items.append(await _build_read(session, row))

    return KpiMonthlyListPage(items=items, total=total, page=page, page_size=page_size)


async def get_kpi(session: AsyncSession, kpi_id: int) -> KpiMonthlyRead:
    kpi = await _get_or_404(session, kpi_id)
    return await _build_read(session, kpi)


async def create_kpi(
    session: AsyncSession,
    data: KpiMonthlyCreate,
    created_by_id: int,
) -> KpiMonthlyRead:
    emp = await session.get(Employee, data.employee_id)
    if not emp:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy nhân viên")

    existing = (
        await session.execute(
            select(EmployeeKpiMonthly).where(
                EmployeeKpiMonthly.employee_id == data.employee_id,
                EmployeeKpiMonthly.year == data.year,
                EmployeeKpiMonthly.month == data.month,
            )
        )
    ).scalar_one_or_none()
    if existing:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            detail=f"Đã tồn tại điểm KPI tháng {data.month}/{data.year} của nhân viên này",
        )

    kpi = EmployeeKpiMonthly(
        employee_id=data.employee_id,
        year=data.year,
        month=data.month,
        score=data.score,
        note=data.note,
        created_by_id=created_by_id,
        created_at=_utcnow(),
        updated_at=_utcnow(),
    )
    session.add(kpi)
    await session.flush()
    return await _build_read(session, kpi)


async def update_kpi(
    session: AsyncSession,
    kpi_id: int,
    data: KpiMonthlyUpdate,
) -> KpiMonthlyRead:
    kpi = await _get_or_404(session, kpi_id)

    if data.score is not None:
        kpi.score = data.score
    if data.note is not None:
        kpi.note = data.note

    kpi.updated_at = _utcnow()
    session.add(kpi)
    await session.flush()
    return await _build_read(session, kpi)


async def delete_kpi(session: AsyncSession, kpi_id: int) -> None:
    kpi = await _get_or_404(session, kpi_id)
    await session.delete(kpi)


async def import_kpi_excel(
    session: AsyncSession,
    file_bytes: bytes,
    created_by_id: int,
) -> "KpiImportResult":
    from io import BytesIO

    import openpyxl
    from sqlalchemy.dialects.postgresql import insert as pg_insert

    from app.schemas.performance import KpiImportResult

    # Build employee_code → id map (one query)
    from app.models.employee import Employee
    emp_rows = (await session.execute(select(Employee.id, Employee.employee_seq, Employee.employee_code_sequence_id))).all()

    # Build code map using the same display-code logic: need prefix from dept
    # Simpler: build a raw-code map using employee_seq padded approach won't work
    # since codes come from sequences. Use employee_code_service for each code.
    # For perf on import, pre-build code→id using a generated column query.
    from sqlalchemy import String, cast, func, text
    from app.models.employee_job import EmployeeJobRecord
    from app.models.org import Department
    from app.services import employee_code_service

    # Query all active employees with their codes in one pass
    all_emps = (await session.execute(select(Employee))).scalars().all()
    code_to_id: dict[str, int] = {}
    for emp in all_emps:
        code = await employee_code_service.build_employee_display_code(session, emp)
        code_to_id[code.strip().upper()] = emp.id

    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        log.warning("kpi_excel_parse_error", extra={"error": str(exc)})
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="File không hợp lệ hoặc không phải định dạng .xlsx") from exc

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = updated = skipped = 0
    errors: list[str] = []
    upsert_rows: list[dict] = []

    for i, row in enumerate(rows, start=2):
        if not row or all(v is None for v in row):
            continue

        raw_code = str(row[0]).strip() if row[0] is not None else ""
        raw_year = row[2]
        raw_month = row[3]
        raw_score = row[4]
        raw_note = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None

        if not raw_code:
            errors.append(f"Hàng {i}: Mã nhân viên trống")
            skipped += 1
            continue

        emp_id = code_to_id.get(raw_code.upper())
        if emp_id is None:
            errors.append(f"Hàng {i}: Không tìm thấy nhân viên '{raw_code}'")
            skipped += 1
            continue

        try:
            year = int(raw_year)
            if not (2000 <= year <= 2100):
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Hàng {i}: Năm không hợp lệ '{raw_year}'")
            skipped += 1
            continue

        try:
            month = int(raw_month)
            if not (1 <= month <= 12):
                raise ValueError()
        except (TypeError, ValueError):
            errors.append(f"Hàng {i}: Tháng không hợp lệ '{raw_month}'")
            skipped += 1
            continue

        try:
            raw_score_str = str(raw_score).replace(",", ".").strip() if raw_score is not None else ""
            score = Decimal(raw_score_str)
            if not (Decimal("0") <= score <= Decimal("100")):
                raise ValueError()
        except (TypeError, ValueError, Exception):
            errors.append(f"Hàng {i}: Điểm KPI không hợp lệ '{raw_score}'")
            skipped += 1
            continue

        upsert_rows.append({
            "employee_id": emp_id,
            "year": year,
            "month": month,
            "score": score,
            "note": raw_note,
            "created_by_id": created_by_id,
            "created_at": _utcnow(),
            "updated_at": _utcnow(),
        })

    if upsert_rows:
        # Check which (employee_id, year, month) already exist
        existing_keys: set[tuple] = set()
        existing_stmt = select(
            EmployeeKpiMonthly.employee_id,
            EmployeeKpiMonthly.year,
            EmployeeKpiMonthly.month,
        ).where(
            func.row(
                EmployeeKpiMonthly.employee_id,
                EmployeeKpiMonthly.year,
                EmployeeKpiMonthly.month,
            ).in_([(r["employee_id"], r["year"], r["month"]) for r in upsert_rows])
        )
        for row_res in (await session.execute(existing_stmt)).all():
            existing_keys.add((row_res.employee_id, row_res.year, row_res.month))

        stmt = pg_insert(EmployeeKpiMonthly).values(upsert_rows)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_employee_kpi_year_month",
            set_={
                "score": stmt.excluded.score,
                "note": stmt.excluded.note,
                "updated_at": stmt.excluded.updated_at,
            },
        )
        await session.execute(stmt)

        for r in upsert_rows:
            key = (r["employee_id"], r["year"], r["month"])
            if key in existing_keys:
                updated += 1
            else:
                created += 1

    return KpiImportResult(created=created, updated=updated, skipped=skipped, errors=errors)


def get_kpi_template() -> bytes:
    from io import BytesIO

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI_Tháng"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    sample_font = Font(color="808080", italic=True)

    headers = [
        "Mã nhân viên (*)",
        "Họ và tên (tham khảo)",
        "Năm (*)",
        "Tháng (*)",
        "Điểm KPI (*)",
        "Ghi chú",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sample = ["HH001", "Nguyễn Văn A", 2026, 1, 85.5, ""]
    for col, v in enumerate(sample, start=1):
        cell = ws.cell(row=2, column=col, value=v)
        cell.font = sample_font

    col_widths = [20, 25, 10, 10, 15, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 20

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
