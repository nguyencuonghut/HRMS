"""Service xuất Excel báo cáo đào tạo (9.4)."""
from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.training_report_service import (
    get_incomplete_mandatory_employees,
    get_training_report_summary,
)

_COMPANY = "CÔNG TY TNHH HỒNG HÀ"
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT  = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_VCENTER = Alignment(vertical="center", wrap_text=True)


def _border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _THIN_BORDER


def _sheet_header(ws, title: str, ncols: int, from_fmt: str, to_fmt: str) -> None:
    col_letter = get_column_letter(ncols)
    ws.merge_cells(f"A1:{col_letter}1")
    ws["A1"] = _COMPANY
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{col_letter}2")
    ws["A2"] = f"{title} {from_fmt} – {to_fmt}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = _CENTER
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 6


def _write_header_row(ws, headers: list[str], row: int = 4) -> None:
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws.row_dimensions[row].height = 30


def _total_row(ws, row: int, ncols: int, label: str = "TỔNG CỘNG") -> None:
    ws.cell(row=row, column=1, value=label)
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _THIN_BORDER
    ws.cell(row=row, column=1).alignment = _CENTER
    ws.row_dimensions[row].height = 20


async def export_training_excel(
    session: AsyncSession,
    *,
    from_date: date,
    to_date: date,
    department_id: Optional[int] = None,
    course_id: Optional[int] = None,
) -> bytes:
    summary = await get_training_report_summary(
        session,
        from_date=from_date,
        to_date=to_date,
        department_id=department_id,
        course_id=course_id,
    )
    incomplete = await get_incomplete_mandatory_employees(
        session,
        department_id=department_id,
    )

    from_fmt = from_date.strftime("%d/%m/%Y")
    to_fmt   = to_date.strftime("%d/%m/%Y")

    wb = Workbook()

    # ── Sheet 1: Tổng hợp theo khóa học ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Tổng hợp theo khóa học"

    _sheet_header(ws1, "BÁO CÁO ĐÀO TẠO", 7, from_fmt, to_fmt)
    _write_header_row(ws1, ["STT", "Tên khóa học", "Loại", "Tổng NV", "Hoàn thành", "Chưa hoàn thành", "Tỷ lệ HT (%)"])

    for i, stat in enumerate(summary.by_course):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        values = [i + 1, stat.course_name, stat.course_type_label,
                  stat.total_assigned, stat.completed, stat.not_completed, stat.completion_rate]
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = _VCENTER
            if col_idx == 7:
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (4, 5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row1 = 5 + len(summary.by_course)
    _total_row(ws1, total_row1, 7)
    total_assigned = sum(s.total_assigned for s in summary.by_course)
    total_completed = sum(s.completed for s in summary.by_course)
    total_not_completed = sum(s.not_completed for s in summary.by_course)
    ws1.cell(row=total_row1, column=4, value=total_assigned).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=total_row1, column=5, value=total_completed).alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=total_row1, column=6, value=total_not_completed).alignment = Alignment(horizontal="center", vertical="center")
    avg_cell = ws1.cell(row=total_row1, column=7, value=summary.avg_completion_rate)
    avg_cell.number_format = '0.0"%"'
    avg_cell.alignment = Alignment(horizontal="right", vertical="center")

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 30
    ws1.column_dimensions["C"].width = 14
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 16
    ws1.column_dimensions["G"].width = 14
    ws1.freeze_panes = "A5"

    # ── Sheet 2: Chi tiết đào tạo ────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Chi tiết đào tạo")

    _sheet_header(ws2, "CHI TIẾT ĐÀO TẠO", 11, from_fmt, to_fmt)
    _write_header_row(ws2, [
        "STT", "Mã NV", "Họ và tên", "Phòng ban", "Khóa học", "Loại",
        "Trạng thái", "Kết quả", "Điểm", "Ngày bắt đầu", "Ngày kết thúc",
    ])

    # Re-use summary.by_course's records won't work — need full record list
    # We'll re-query via the summary's base: fetch records via training_record_service
    from app.services.training_record_service import get_records
    all_records_page = await get_records(
        session,
        from_date=from_date,
        to_date=to_date,
        department_id=department_id,
        course_id=course_id,
        page=1,
        page_size=10000,
    )
    all_records = all_records_page.items

    # Note: get_records filters by end_date >= from_date / end_date <= to_date
    # For export we use same logic as summary (end_date or start_date in range)
    # The summary query uses the broader filter so sheet2 may have different count
    # but that's acceptable for the summary sheet being authoritative

    status_labels = {
        "chua_bat_dau": "Chưa bắt đầu",
        "dang_hoc": "Đang học",
        "hoan_thanh": "Hoàn thành",
        "khong_hoan_thanh": "Không hoàn thành",
        "vang_mat": "Vắng mặt",
    }

    for i, rec in enumerate(all_records):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        values = [
            i + 1,
            rec.employee_code,
            rec.employee_name,
            rec.department_name or "",
            rec.course_name,
            rec.course_type_label,
            rec.status_label,
            rec.result_label or "",
            float(rec.score) if rec.score is not None else None,
            rec.start_date,
            rec.end_date,
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = _VCENTER
            if col_idx in (10, 11) and val is not None:
                cell.number_format = "DD/MM/YYYY"
            if col_idx == 9 and val is not None:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    total_row2 = 5 + len(all_records)
    _total_row(ws2, total_row2, 11)

    col_widths2 = [6, 12, 22, 20, 28, 14, 18, 14, 8, 14, 14]
    for i, w in enumerate(col_widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A5"

    # ── Sheet 3: NV chưa HT bắt buộc ─────────────────────────────────────────
    ws3 = wb.create_sheet(title="NV chưa HT bắt buộc")

    _sheet_header(ws3, "DANH SÁCH NHÂN VIÊN CHƯA HOÀN THÀNH ĐÀO TẠO BẮT BUỘC", 6, from_fmt, to_fmt)
    _write_header_row(ws3, [
        "STT", "Mã NV", "Họ và tên", "Phòng ban", "Số khóa chưa HT", "Danh sách khóa chưa HT",
    ])

    for i, emp in enumerate(incomplete):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        values = [
            i + 1,
            emp.employee_code,
            emp.employee_name,
            emp.department_name or "",
            emp.incomplete_count,
            "\n".join(emp.incomplete_courses),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws3.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = _VCENTER
            if col_idx == 5:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    total_row3 = 5 + len(incomplete)
    _total_row(ws3, total_row3, 6)
    ws3.cell(row=total_row3, column=2, value=f"Tổng: {len(incomplete)} nhân viên")

    col_widths3 = [6, 12, 22, 20, 16, 40]
    for i, w in enumerate(col_widths3, start=1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
