"""Service Xuất Excel Báo cáo Hợp đồng (11.5)."""
from __future__ import annotations

import io
from datetime import date, timedelta
from typing import Optional

import sqlalchemy as sa
from sqlalchemy import and_, select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.employee_contract import EmployeeContract
from app.models.catalog import ContractCategory
from app.models.org import Department
from app.services import employee_code_service


async def export_contracts_xlsx(
    session: AsyncSession,
    *,
    department_id: Optional[int] = None,
    status: str = "active",
    days_ahead: Optional[int] = None,
) -> io.BytesIO:
    """Xuất danh sách hợp đồng gốc ra file Excel có định dạng màu sắc cảnh báo."""
    today = date.today()

    # Base query: chỉ lấy HĐ gốc của nhân viên active
    stmt = (
        select(
            EmployeeContract,
            Employee,
            Department.name.label("department_name"),
            ContractCategory.name.label("category_name"),
        )
        .join(Employee, Employee.id == EmployeeContract.employee_id)
        .join(ContractCategory, ContractCategory.id == EmployeeContract.contract_category_id)
        .join(
            EmployeeJobRecord,
            and_(
                EmployeeJobRecord.employee_id == Employee.id,
                EmployeeJobRecord.is_current == True,
            ),
        )
        .join(Department, Department.id == EmployeeJobRecord.department_id)
        .where(
            EmployeeContract.document_kind == "labor_contract",
            EmployeeContract.parent_contract_id.is_(None),
            Employee.is_active == True,
        )
    )

    if department_id is not None:
        stmt = stmt.where(EmployeeJobRecord.department_id == department_id)

    # Lọc theo trạng thái
    if status == "active":
        stmt = stmt.where(EmployeeContract.status == "active")
    elif status == "expired":
        stmt = stmt.where(EmployeeContract.status == "expired")
    # "all" không lọc status

    # Lọc theo số ngày sắp hết hạn (nếu có)
    if days_ahead is not None:
        limit_date = today + timedelta(days=days_ahead)
        stmt = stmt.where(
            EmployeeContract.status == "active",
            EmployeeContract.effective_to.isnot(None),
            EmployeeContract.effective_to >= today,
            EmployeeContract.effective_to <= limit_date,
        )

    # Sắp xếp
    if days_ahead is not None:
        stmt = stmt.order_by(EmployeeContract.effective_to.asc(), EmployeeContract.id.asc())
    else:
        stmt = stmt.order_by(EmployeeContract.effective_from.desc(), EmployeeContract.id.asc())

    rows = (await session.execute(stmt)).all()

    # Build display codes
    employees_list = [r.Employee for r in rows]
    code_map = await employee_code_service.batch_build_employee_display_codes(session, employees_list)

    # Khởi tạo Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Hợp đồng"

    # Định nghĩa style
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A5F")  # Màu xanh đậm navy sang trọng
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")

    critical_fill = PatternFill("solid", fgColor="FFCCCC")  # Đỏ nhạt
    warning_fill = PatternFill("solid", fgColor="FFE5CC")   # Cam nhạt
    notice_fill = PatternFill("solid", fgColor="FFFFCC")    # Vàng nhạt

    # Đặt tiêu đề cột
    headers = [
        "STT", "Mã NV", "Họ tên", "Phòng ban", "Loại HĐ", "Số HĐ",
        "Ngày ký", "Hiệu lực từ", "Hiệu lực đến", "Còn lại (ngày)",
        "Mức lương BHXH", "Trạng thái"
    ]
    ws.append(headers)

    # Style cho header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # Điền dữ liệu
    for stt, r in enumerate(rows, 1):
        ec: EmployeeContract = r.EmployeeContract
        emp: Employee = r.Employee
        dept_name: str = r.department_name
        cat_name: str = r.category_name

        # Tính số ngày còn lại và xác định style màu nếu là active và có ngày kết thúc
        days_rem = None
        fill_to_apply = None

        if ec.effective_to is not None:
            days_rem = (ec.effective_to - today).days
            if ec.status == "active":
                if days_rem <= 7:
                    fill_to_apply = critical_fill
                elif days_rem <= 30:
                    fill_to_apply = warning_fill
                elif days_rem <= 90:
                    fill_to_apply = notice_fill

        # Map các giá trị
        status_text = {
            "active": "Đang hiệu lực",
            "expired": "Hết hiệu lực",
            "terminated": "Đã chấm dứt",
            "draft": "Nháp"
        }.get(ec.status, ec.status)

        row_data = [
            stt,
            code_map.get(emp.id, ""),
            emp.full_name,
            dept_name,
            cat_name,
            ec.contract_number,
            ec.signed_date.strftime("%d/%m/%Y") if ec.signed_date else "",
            ec.effective_from.strftime("%d/%m/%Y") if ec.effective_from else "",
            ec.effective_to.strftime("%d/%m/%Y") if ec.effective_to else "Vô thời hạn",
            days_rem if days_rem is not None else "",
            float(ec.insurance_salary) if ec.insurance_salary is not None else "",
            status_text
        ]
        ws.append(row_data)

        # Style cho hàng dữ liệu hiện tại
        row_idx = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.font = Font(name="Arial", size=10)

            # Áp dụng màu cảnh báo hết hạn
            if fill_to_apply:
                cell.fill = fill_to_apply

            # Căn lề và format số
            if col_idx in [1, 2, 7, 8, 9, 12]:
                cell.alignment = center_align
            elif col_idx in [10]:
                cell.alignment = right_align
            elif col_idx in [11]:
                cell.alignment = right_align
                if cell.value != "":
                    cell.number_format = "#,##0"
            else:
                cell.alignment = left_align

    # Freeze row đầu tiên (header)
    ws.freeze_panes = "A2"

    # Tự động căn chỉnh độ rộng cột
    for col in ws.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        # Cộng thêm padding khoảng cách
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
