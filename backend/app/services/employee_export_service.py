"""Export nhân viên ra file Excel (3.7)."""

from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import Bank
from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.org import Department, JobTitle
from app.schemas.employee_import import EXPORT_MAX_ROWS
from app.services import (
    employee_code_service,
    employee_education_service,
    employee_insurance_service,
    employee_job_service,
    employee_relative_service,
    employee_service,
)


# ── Style helpers ─────────────────────────────────────────────────────────────

_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="EBF3FB")
_THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def _write_header(ws, headers: list[str], widths: list[int] | None = None) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _THIN_BORDER
    if widths:
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
    ws.row_dimensions[1].height = 28


def _fmt_date(d) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


def _gender_label(g: str) -> str:
    return {"male": "Nam", "female": "Nữ", "other": "Khác"}.get(g, g)


def _status_label(s: str) -> str:
    return {
        "probation":  "Thử việc",
        "official":   "Chính thức",
        "long_leave": "Nghỉ dài hạn",
        "resigned":   "Đã nghỉ",
    }.get(s, s)


# ── Export list ───────────────────────────────────────────────────────────────

async def export_employee_list(
    session: AsyncSession,
    *,
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    is_active: Optional[bool] = None,
) -> bytes:
    result = await employee_service.list_employees_page(
        session,
        keyword=keyword,
        status=status,
        is_active=is_active,
        page=1,
        page_size=EXPORT_MAX_ROWS + 1,
    )
    employees: list[Employee] = result["items"]
    total: int = result["total"]

    if total > EXPORT_MAX_ROWS:
        raise ValueError(
            f"Danh sách có {total} nhân viên, vượt giới hạn {EXPORT_MAX_ROWS}. "
            "Hãy dùng bộ lọc để thu hẹp kết quả."
        )

    # Batch fetch current job (dept + job_title) để tránh N+1
    emp_ids = [e.id for e in employees]
    job_map: dict[int, EmployeeJobRecord] = {}
    dept_map: dict[int, Department] = {}
    jt_map: dict[int, JobTitle] = {}

    if emp_ids:
        rows = await session.execute(
            select(EmployeeJobRecord)
            .where(EmployeeJobRecord.employee_id.in_(emp_ids), EmployeeJobRecord.is_current.is_(True))
        )
        for rec in rows.scalars().all():
            job_map[rec.employee_id] = rec

        dept_ids = {r.department_id for r in job_map.values()}
        jt_ids   = {r.job_title_id  for r in job_map.values() if r.job_title_id}

        if dept_ids:
            drows = await session.execute(select(Department).where(Department.id.in_(dept_ids)))
            for d in drows.scalars().all():
                dept_map[d.id] = d
        if jt_ids:
            jrows = await session.execute(select(JobTitle).where(JobTitle.id.in_(jt_ids)))
            for j in jrows.scalars().all():
                jt_map[j.id] = j

    display_codes = await employee_code_service.batch_build_employee_display_codes(session, employees)

    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách nhân viên"

    headers = [
        "Mã NV", "Họ và tên", "Ngày sinh", "Giới tính",
        "Số CCCD/CMND", "Điện thoại", "Email",
        "Trạng thái", "Ngày vào làm",
        "Phòng ban", "Chức danh", "Ngày nghỉ việc",
    ]
    widths = [12, 24, 14, 10, 16, 14, 26, 14, 14, 22, 26, 14]
    _write_header(ws, headers, widths)

    for row_idx, emp in enumerate(employees, start=2):
        code     = display_codes[emp.id]
        job_rec  = job_map.get(emp.id)
        dept_name = dept_map.get(job_rec.department_id).name if job_rec and job_rec.department_id in dept_map else ""
        jt_name   = jt_map.get(job_rec.job_title_id).name  if job_rec and job_rec.job_title_id  in jt_map  else ""

        row_data = [
            code,
            emp.full_name,
            _fmt_date(emp.date_of_birth),
            _gender_label(emp.gender),
            emp.id_number,
            emp.phone_number or "",
            emp.personal_email or "",
            _status_label(emp.status),
            _fmt_date(emp.start_date),
            dept_name,
            jt_name,
            _fmt_date(emp.resigned_date),
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col, value=val)
            if row_idx % 2 == 0:
                c.fill = _ALT_FILL

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Export full profile ───────────────────────────────────────────────────────

async def export_employee_profile(session: AsyncSession, employee_id: int) -> bytes:
    emp = await session.get(Employee, employee_id)
    if not emp:
        raise ValueError("Không tìm thấy nhân viên")
    insurance_profile = await employee_insurance_service.get_employee_insurance_profile(session, employee_id)
    bhxh_code = (
        insurance_profile.bhxh_code
        if insurance_profile and insurance_profile.bhxh_code is not None
        else emp.bhxh_code
    )

    wb = Workbook()

    # ── Sheet 1: Thông tin cá nhân ────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Thông tin cá nhân"

    def kv(ws, key: str, val: str, row: int) -> None:
        k = ws.cell(row=row, column=1, value=key)
        k.font = Font(bold=True)
        k.fill = PatternFill("solid", fgColor="D6E4F0")
        ws.cell(row=row, column=2, value=val)

    personal_rows = [
        ("Họ và tên",       emp.full_name),
        ("Ngày sinh",       _fmt_date(emp.date_of_birth)),
        ("Giới tính",       _gender_label(emp.gender)),
        ("Số CCCD/CMND",    emp.id_number),
        ("Ngày cấp",        _fmt_date(emp.id_issued_on)),
        ("Nơi cấp",         emp.id_issued_by or ""),
        ("Hộ chiếu",        emp.passport_number or ""),
        ("Giấy phép LĐ",   emp.work_permit_number or ""),
        ("Điện thoại",      emp.phone_number or ""),
        ("Email cá nhân",   emp.personal_email or ""),
        ("Mã số thuế",      emp.personal_tax_code or ""),
        ("Số BHXH",         bhxh_code or ""),
        ("Trạng thái",      _status_label(emp.status)),
        ("Ngày vào làm",    _fmt_date(emp.start_date)),
        ("Ngày nghỉ việc",  _fmt_date(emp.resigned_date) if emp.resigned_date else ""),
    ]
    for i, (k, v) in enumerate(personal_rows, start=1):
        kv(ws1, k, v, i)
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 30

    # ── Sheet 2: Lịch sử công việc ────────────────────────────────────
    ws2 = wb.create_sheet("Công việc")
    job_headers = ["Phòng ban", "Chức danh", "Ngày hiệu lực", "Ngày kết thúc", "Thử việc từ", "Thử việc đến", "Ngày chính thức", "Hiện tại"]
    _write_header(ws2, job_headers, [22, 22, 14, 14, 14, 14, 14, 10])
    job_records = await employee_job_service.get_job_records(session, employee_id)
    for ri, rec in enumerate(job_records, start=2):
        dept = await session.get(Department, rec.department_id)
        jt   = await session.get(JobTitle, rec.job_title_id) if rec.job_title_id else None
        row  = [
            dept.name if dept else "",
            jt.name   if jt   else "",
            _fmt_date(rec.effective_from),
            _fmt_date(rec.effective_to),
            _fmt_date(rec.probation_start_date),
            _fmt_date(rec.probation_end_date),
            _fmt_date(rec.official_date),
            "✅" if rec.is_current else "",
        ]
        for ci, v in enumerate(row, start=1):
            ws2.cell(row=ri, column=ci, value=v)

    # ── Sheet 3: Học vấn ──────────────────────────────────────────────
    ws3 = wb.create_sheet("Học vấn")
    edu_headers = ["Trường", "Ngành", "Trình độ", "Năm tốt nghiệp", "Loại bằng", "Học vấn chính"]
    _write_header(ws3, edu_headers, [28, 24, 16, 18, 18, 14])
    edu_list = await employee_education_service.get_education_histories(session, employee_id)
    for ri, edu in enumerate(edu_list, start=2):
        row = [
            edu.institution_name or "",
            edu.major_name or "",
            edu.education_level_name or "",
            str(edu.graduation_year or ""),
            edu.diploma_type or "",
            "✅" if edu.is_main_education else "",
        ]
        for ci, v in enumerate(row, start=1):
            ws3.cell(row=ri, column=ci, value=v)

    # ── Sheet 4: Người thân ───────────────────────────────────────────
    ws4 = wb.create_sheet("Người thân")
    rel_headers = ["Họ tên", "Quan hệ", "Ngày sinh", "Điện thoại", "Liên hệ khẩn cấp"]
    _write_header(ws4, rel_headers, [24, 16, 14, 14, 16])
    relatives = await employee_relative_service.get_relatives(session, employee_id)
    for ri, rel in enumerate(relatives, start=2):
        row = [
            rel.full_name,
            rel.relationship_type or "",
            _fmt_date(rel.date_of_birth) if rel.date_of_birth else "",
            rel.phone_number or "",
            "✅" if rel.is_emergency_contact else "",
        ]
        for ci, v in enumerate(row, start=1):
            ws4.cell(row=ri, column=ci, value=v)

    # ── Sheet 5: Tài khoản ngân hàng ─────────────────────────────────
    ws5 = wb.create_sheet("Tài khoản ngân hàng")
    bank_headers = ["Số tài khoản", "Tên tài khoản", "Ngân hàng", "Chi nhánh", "Tài khoản chính"]
    _write_header(ws5, bank_headers, [20, 24, 24, 24, 14])
    raw_accounts = await employee_service.get_employee_bank_accounts(session, employee_id)
    bank_id_map: dict[int, Bank] = {}
    for acc in raw_accounts:
        if acc.bank_id not in bank_id_map:
            b = await session.get(Bank, acc.bank_id)
            if b:
                bank_id_map[acc.bank_id] = b
    for ri, acc in enumerate(raw_accounts, start=2):
        bank = bank_id_map.get(acc.bank_id)
        row = [
            acc.account_number,
            acc.account_name or "",
            bank.name if bank else "",
            acc.branch_name or "",
            "✅" if acc.is_primary else "",
        ]
        for ci, v in enumerate(row, start=1):
            ws5.cell(row=ri, column=ci, value=v)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def export_comprehensive_employee_list(session: AsyncSession) -> bytes:
    from app.models.employee import EmployeeAddress, EmployeeBankAccount
    from app.models.employee_insurance import EmployeeInsuranceProfile
    from app.models.employee_relative import EmployeeRelative
    from app.models.employee_education import EmployeeEducationHistory
    from app.models.employee_contract import EmployeeContract
    from app.models.reward import EmployeeReward
    from app.models.discipline import EmployeeDiscipline
    from app.models.employee_asset import EmployeeAsset
    from app.models.recruitment import EmployeeDocumentChecklist, DocumentChecklistType
    from app.models.catalog import AdministrativeUnit, Ethnicity, ContractCategory, EducationLevel

    employees = (
        await session.execute(
            select(Employee).order_by(Employee.id)
        )
    ).scalars().all()

    emp_ids = [e.id for e in employees]

    # Get display codes
    display_codes = await employee_code_service.batch_build_employee_display_codes(session, employees)

    # 1. Job records
    job_records = []
    if emp_ids:
        job_records = (
            await session.execute(
                select(EmployeeJobRecord)
                .where(EmployeeJobRecord.employee_id.in_(emp_ids))
                .order_by(EmployeeJobRecord.employee_id, EmployeeJobRecord.effective_from.desc())
            )
        ).scalars().all()

    from collections import defaultdict
    job_rec_map = defaultdict(list)
    for j in job_records:
        job_rec_map[j.employee_id].append(j)

    depts = (await session.execute(select(Department))).scalars().all()
    dept_name_map = {d.id: d.name for d in depts}

    jts = (await session.execute(select(JobTitle))).scalars().all()
    jt_name_map = {j.id: j.name for j in jts}

    # 2. Permanent addresses
    addresses = []
    if emp_ids:
        addresses = (
            await session.execute(
                select(EmployeeAddress)
                .where(EmployeeAddress.employee_id.in_(emp_ids), EmployeeAddress.address_type == "permanent")
            )
        ).scalars().all()
    addr_map = {a.employee_id: a for a in addresses}

    unit_ids = set()
    for addr in addresses:
        for uid in [addr.new_province_unit_id, addr.new_district_unit_id, addr.new_ward_unit_id,
                    addr.old_province_unit_id, addr.old_district_unit_id, addr.old_ward_unit_id]:
            if uid:
                unit_ids.add(uid)
    unit_names = {}
    if unit_ids:
        rows = (await session.execute(
            select(AdministrativeUnit.id, AdministrativeUnit.name)
            .where(AdministrativeUnit.id.in_(unit_ids))
        )).all()
        unit_names = {row.id: row.name for row in rows}

    # 3. Ethnicities
    eths = (await session.execute(select(Ethnicity))).scalars().all()
    eth_map = {e.id: e.name for e in eths}

    # 4. Contracts
    contracts = []
    if emp_ids:
        contracts = (
            await session.execute(
                select(EmployeeContract)
                .where(EmployeeContract.employee_id.in_(emp_ids))
                .order_by(EmployeeContract.employee_id, EmployeeContract.signed_date.asc())
            )
        ).scalars().all()
    contract_map = defaultdict(list)
    for c in contracts:
        contract_map[c.employee_id].append(c)

    categories = (await session.execute(select(ContractCategory))).scalars().all()
    cat_map = {c.id: c.name for c in categories}

    max_contracts = max((len(lst) for lst in contract_map.values()), default=0)

    # 5. Insurance profiles
    profiles = []
    if emp_ids:
        profiles = (
            await session.execute(
                select(EmployeeInsuranceProfile)
                .where(EmployeeInsuranceProfile.employee_id.in_(emp_ids))
            )
        ).scalars().all()
    profile_map = {p.employee_id: p for p in profiles}

    # 6. Education histories
    edu_list = []
    if emp_ids:
        edu_list = (
            await session.execute(
                select(EmployeeEducationHistory)
                .where(EmployeeEducationHistory.employee_id.in_(emp_ids))
            )
        ).scalars().all()
    edu_map = defaultdict(list)
    for edu in edu_list:
        edu_map[edu.employee_id].append(edu)

    # 7. Bank accounts
    bank_accounts = []
    if emp_ids:
        bank_accounts = (
            await session.execute(
                select(EmployeeBankAccount)
                .where(EmployeeBankAccount.employee_id.in_(emp_ids))
            )
        ).scalars().all()
    bank_acc_map = defaultdict(list)
    for acc in bank_accounts:
        bank_acc_map[acc.employee_id].append(acc)

    banks = (await session.execute(select(Bank))).scalars().all()
    bank_name_map = {b.id: b.name for b in banks}

    edu_levels = (await session.execute(select(EducationLevel))).scalars().all()
    edu_level_map = {el.id: el.name for el in edu_levels}

    # 8. Relatives
    relatives = []
    if emp_ids:
        relatives = (
            await session.execute(
                select(EmployeeRelative)
                .where(EmployeeRelative.employee_id.in_(emp_ids))
            )
        ).scalars().all()
    relative_map = defaultdict(list)
    for r in relatives:
        relative_map[r.employee_id].append(r)

    # 9. Rewards
    rewards = []
    if emp_ids:
        rewards = (
            await session.execute(
                select(EmployeeReward)
                .where(EmployeeReward.employee_id.in_(emp_ids))
                .order_by(EmployeeReward.reward_date.desc())
            )
        ).scalars().all()
    reward_map = defaultdict(list)
    for r in rewards:
        reward_map[r.employee_id].append(r)

    # 10. Disciplines
    disciplines = []
    if emp_ids:
        disciplines = (
            await session.execute(
                select(EmployeeDiscipline)
                .where(EmployeeDiscipline.employee_id.in_(emp_ids))
                .order_by(EmployeeDiscipline.effective_date.desc())
            )
        ).scalars().all()
    discipline_map = defaultdict(list)
    for d in disciplines:
        discipline_map[d.employee_id].append(d)

    # 11. Assets
    assets = []
    if emp_ids:
        assets = (
            await session.execute(
                select(EmployeeAsset)
                .where(EmployeeAsset.employee_id.in_(emp_ids), EmployeeAsset.status == "allocated")
                .order_by(EmployeeAsset.handover_date.desc())
            )
        ).scalars().all()
    asset_map = defaultdict(list)
    for a in assets:
        asset_map[a.employee_id].append(a)

    # 12. Checklist documents
    checklist_rows = []
    if emp_ids:
        checklist_rows = (
            await session.execute(
                select(EmployeeDocumentChecklist, DocumentChecklistType)
                .join(DocumentChecklistType, EmployeeDocumentChecklist.document_type_id == DocumentChecklistType.id)
                .where(
                    EmployeeDocumentChecklist.employee_id.in_(emp_ids),
                    DocumentChecklistType.is_required == True,
                    DocumentChecklistType.is_active == True,
                )
            )
        ).all()
    chk_map = defaultdict(list)
    for item, dtype in checklist_rows:
        chk_map[item.employee_id].append((item, dtype))

    # Helper formatting functions
    def format_permanent_address(addr) -> str:
        if not addr:
            return ""
        if addr.new_address_line or addr.new_ward_unit_id or addr.new_province_unit_id:
            parts = [
                (addr.new_address_line or "").strip(),
                unit_names.get(addr.new_ward_unit_id, "") if addr.new_ward_unit_id else "",
                unit_names.get(addr.new_province_unit_id, "") if addr.new_province_unit_id else ""
            ]
            joined = ", ".join(p for p in parts if p)
            if joined:
                return joined
        if addr.old_address_line or addr.old_ward_unit_id or addr.old_district_unit_id or addr.old_province_unit_id:
            parts = [
                (addr.old_address_line or "").strip(),
                unit_names.get(addr.old_ward_unit_id, "") if addr.old_ward_unit_id else "",
                unit_names.get(addr.old_district_unit_id, "") if addr.old_district_unit_id else "",
                unit_names.get(addr.old_province_unit_id, "") if addr.old_province_unit_id else ""
            ]
            joined = ", ".join(p for p in parts if p)
            if joined:
                return joined
        return addr.full_address_text or ""

    resigned_reason_map = {
        "1": "Không hài lòng với quản lý trực tiếp",
        "2": "Không hài lòng về thu nhập",
        "3": "Không phù hợp với môi trường, văn hóa công ty",
        "4": "Không thích công việc được giao",
        "5": "Nghỉ việc do kết quả công việc yếu, vi phạm kỷ luật",
        "6": "Khác",
    }

    # Initialize Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Dữ liệu nhân sự tổng hợp"

    headers = [
        "STT",
        "Mã nhân viên",
        "Họ và tên",
        "Vị trí",
        "Phòng/ban",
        "Giới tính",
        "Dân tộc",
        "Ngày sinh",
        "Số CCCD",
        "Ngày cấp CCCD",
        "Nơi cấp CCCD",
        "Số điện thoại",
        "Địa chỉ thường trú",
        "Ngày vào",
        "Ngày ký hợp đồng mới nhất",
        "Đến ngày của hợp đồng mới nhất",
        "Số hợp đồng mới nhất (đang áp dụng)",
        "Loại hợp đồng hiệu lực",
    ]
    for i in range(1, max_contracts + 1):
        headers.extend([
            f"Thông tin hợp đồng lần {i}: Ngày ký",
            f"Thông tin hợp đồng lần {i}: Đến ngày",
            f"Thông tin hợp đồng lần {i}: Số hợp đồng",
            f"Thông tin hợp đồng lần {i}: Loại hợp đồng",
        ])
    headers.extend([
        "Tháng tham gia BHXH lần đầu",
        "Mã số BHXH",
        "Bậc đóng BHXH",
        "Lương đóng BHXH mới nhất",
        "Loại BHXH khác tham gia",
        "Tham gia BH khác cho Cá nhân/ gia đình",
        "Số thẻ bảo hiểm CSSK/ Bảo hiểm 24/24",
        "Trình độ học vấn",
        "Trường tốt nghiệp",
        "Chuyên ngành",
        "MST",
        "Số tài khoản ngân hàng",
        "Ngân hàng",
        "Chủ tài khoản",
        "Email",
        "Người thân",
        "Số điện thoại liên hệ khẩn cấp",
        "Ngày bổ nhiệm/thuyên chuyển công tác",
        "Khen thưởng",
        "Kỷ luật",
        "Tài sản được cấp",
        "Cam kết bảo mật thông tin",
        "Giấy tờ còn thiếu",
        "Ngày nghỉ việc",
        "Lý do nghỉ việc",
    ])

    widths = [8, 15, 25, 25, 25, 12, 15, 14, 18, 14, 25, 15, 40, 14, 14, 14, 20, 25]
    for _ in range(max_contracts * 4):
        widths.append(20)
    widths.extend([15, 15, 12, 20, 25, 25, 30, 20, 30, 25, 18, 20, 25, 25, 30, 30, 20, 25, 35, 35, 30, 15, 35, 14, 25])

    _write_header(ws, headers, widths)

    for idx, emp in enumerate(employees):
        row_idx = idx + 2
        code = display_codes[emp.id]
        
        emp_jobs = job_rec_map.get(emp.id, [])
        curr_job = next((j for j in emp_jobs if j.is_current), None)
        pos_name = jt_name_map.get(curr_job.job_title_id, "") if curr_job else ""
        dept_name = dept_name_map.get(curr_job.department_id, "") if curr_job else ""
        
        addr = addr_map.get(emp.id)
        addr_str = format_permanent_address(addr)
        
        emp_contracts = contract_map.get(emp.id, [])
        latest_contract = emp_contracts[-1] if emp_contracts else None
        
        latest_contract_sign = _fmt_date(latest_contract.signed_date) if latest_contract else ""
        latest_contract_expired = _fmt_date(latest_contract.effective_to) if latest_contract and latest_contract.effective_to else ("Vô thời hạn" if latest_contract else "")
        latest_contract_no = latest_contract.contract_number if latest_contract else ""
        latest_contract_cat = cat_map.get(latest_contract.contract_category_id, "") if latest_contract else ""

        # Construct basic row data
        row_data = [
            idx + 1,
            code,
            emp.full_name,
            pos_name,
            dept_name,
            _gender_label(emp.gender),
            eth_map.get(emp.ethnicity_id, ""),
            _fmt_date(emp.date_of_birth),
            emp.id_number,
            _fmt_date(emp.id_issued_on),
            emp.id_issued_by,
            emp.phone_number or "",
            addr_str,
            _fmt_date(emp.start_date),
            latest_contract_sign,
            latest_contract_expired,
            latest_contract_no,
            latest_contract_cat,
        ]

        # Historical contracts
        for i in range(max_contracts):
            if i < len(emp_contracts):
                c = emp_contracts[i]
                row_data.extend([
                    _fmt_date(c.signed_date),
                    _fmt_date(c.effective_to) if c.effective_to else "Vô thời hạn",
                    c.contract_number,
                    cat_map.get(c.contract_category_id, ""),
                ])
            else:
                row_data.extend(["", "", "", ""])

        # Insurance details
        profile = profile_map.get(emp.id)
        joined_month = profile.company_bhxh_joined_date.strftime("%m/%Y") if profile and profile.company_bhxh_joined_date else ""
        bhxh_code_str = (profile.bhxh_code if profile and profile.bhxh_code is not None else emp.bhxh_code) or ""
        
        grade_no = latest_contract.insurance_salary_grade_no if latest_contract else None
        
        # Basis salary
        basis_amount = None
        if profile:
            if profile.insurance_basis_source == "contract" and latest_contract:
                basis_amount = latest_contract.insurance_salary
            else:
                basis_amount = profile.insurance_basis_amount
        if basis_amount is None and latest_contract:
            basis_amount = latest_contract.insurance_salary
        basis_amount_val = float(basis_amount) if basis_amount is not None else ""

        # Other insurance types
        other_bh = []
        if profile and profile.accident_insurance_code:
            other_bh.append("Bảo hiểm 24/24")
        if profile and profile.health_care_insurance_code:
            other_bh.append("Bảo hiểm CSSK")
        other_bh_str = ", ".join(other_bh)

        # Other insurance scope
        bh_scope = []
        if profile and profile.accident_insurance_code:
            bh_scope.append("Cá nhân")
        if profile and profile.health_care_insurance_code:
            has_family = profile.health_care_family_participation
            if not has_family:
                emp_relatives = relative_map.get(emp.id, [])
                if any(r.participates_in_health_care_insurance for r in emp_relatives):
                    has_family = True
            bh_scope.append("Gia đình" if has_family else "Cá nhân")
        bh_scope_str = ", ".join(list(dict.fromkeys(bh_scope)))

        # Card numbers
        card_codes = []
        if profile and profile.health_care_insurance_code:
            card_codes.append(profile.health_care_insurance_code)
        if profile and profile.accident_insurance_code:
            card_codes.append(profile.accident_insurance_code)
        card_codes_str = ", ".join(card_codes)

        # Education
        emp_edus = edu_map.get(emp.id, [])
        main_edu = next((edu for edu in emp_edus if edu.is_main_education), None)
        if not main_edu and emp_edus:
            main_edu = emp_edus[0]
        edu_level = edu_level_map.get(main_edu.education_level_id, "") if main_edu else ""
        edu_inst = main_edu.institution_name if main_edu else ""
        edu_major = main_edu.major_name if main_edu else ""

        # Bank account
        emp_banks = bank_acc_map.get(emp.id, [])
        primary_bank = next((acc for acc in emp_banks if acc.is_primary), None)
        if not primary_bank and emp_banks:
            primary_bank = emp_banks[0]
        bank_acc = primary_bank.account_number if primary_bank else ""
        bank_name = bank_name_map.get(primary_bank.bank_id, "") if primary_bank else ""
        bank_holder = primary_bank.account_name if primary_bank else ""

        # Relatives list
        emp_relatives = relative_map.get(emp.id, [])
        rel_str = ",\n".join(r.full_name for r in emp_relatives)
        
        # Emergency contact
        emergency_phones = [r.phone_number for r in emp_relatives if r.is_emergency_contact and r.phone_number]
        emergency_phones_str = ", ".join(emergency_phones)

        # Job appointment record dates
        job_dates = [_fmt_date(j.effective_from) for j in emp_jobs if j.effective_from]
        job_dates_str = ",\n".join(job_dates)

        # Rewards & Disciplines
        emp_rewards = reward_map.get(emp.id, [])
        rewards_str = "\n".join(f"- {r.title} ({_fmt_date(r.reward_date)})" for r in emp_rewards)

        emp_disciplines = discipline_map.get(emp.id, [])
        disciplines_str = "\n".join(f"- {d.title} ({_fmt_date(d.effective_date)})" for d in emp_disciplines)

        # Assigned assets
        emp_assets = asset_map.get(emp.id, [])
        assets_str = ",\n".join(a.asset_name for a in emp_assets)

        # NDA & checklist
        emp_chk = chk_map.get(emp.id, [])
        NDA_item = next((item for item, dtype in emp_chk if dtype.code == "cam_ket_bao_mat_thong_tin"), None)
        nda_status = "Đã ký" if NDA_item and NDA_item.status == "submitted" else "Chưa ký"

        missing_papers = [dtype.name for item, dtype in emp_chk if item.status == "not_submitted"]
        missing_papers_str = ",\n".join(missing_papers)

        resigned_reason_parts = []
        r_type_label = resigned_reason_map.get(emp.resigned_reason_type, emp.resigned_reason_type or "")
        if r_type_label:
            resigned_reason_parts.append(r_type_label)
        if emp.resigned_reason_note:
            resigned_reason_parts.append(emp.resigned_reason_note)
        resigned_reason_val = " - ".join(resigned_reason_parts)

        row_data.extend([
            joined_month,
            bhxh_code_str,
            f"Bậc {grade_no}" if grade_no is not None else "",
            basis_amount_val,
            other_bh_str,
            bh_scope_str,
            card_codes_str,
            edu_level,
            edu_inst,
            edu_major,
            emp.personal_tax_code or "",
            bank_acc,
            bank_name,
            bank_holder,
            emp.personal_email or "",
            rel_str,
            emergency_phones_str,
            job_dates_str,
            rewards_str,
            disciplines_str,
            assets_str,
            nda_status,
            missing_papers_str,
            _fmt_date(emp.resigned_date),
            resigned_reason_val,
        ])

        for col_idx, val in enumerate(row_data, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = _THIN_BORDER
            if row_idx % 2 == 0:
                c.fill = _ALT_FILL
            if isinstance(val, str) and "\n" in val:
                c.alignment = Alignment(wrap_text=True, vertical="center")
            else:
                c.alignment = Alignment(vertical="center")

    # Enable grid lines visibility
    if ws.views.sheetView:
        ws.views.sheetView[0].showGridLines = True

    # Auto-adjust column widths dynamically based on max length of content
    for col in range(1, ws.max_column + 1):
        col_letter = ws.cell(row=1, column=col).column_letter
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = str(ws.cell(row=row, column=col).value or "")
            lines = val.split("\n")
            max_len = max(max_len, max(len(l) for l in lines))
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 12)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
