"""Upload đính kèm & import Excel ứng viên (13.3)."""
from __future__ import annotations

import uuid
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core import storage
from app.models.recruitment import Candidate, CandidateAttachment
from app.schemas.recruitment import (
    AttachmentTypeLabels,
    CandidateAttachmentRead,
    CandidateCreate,
    ImportResult,
)
from app.services import candidate_service

_MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB
_ALLOWED_MIME = {
    "application/pdf",
    "application/msword",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "image/jpeg",
    "image/png",
}

# ── MinIO attachment ──────────────────────────────────────────────────────────


async def save_candidate_attachment(
    session: AsyncSession,
    candidate_id: int,
    upload: UploadFile,
    attachment_type: str,
    note: Optional[str],
    uploaded_by_id: int,
) -> CandidateAttachmentRead:
    await candidate_service._get_candidate_or_404(session, candidate_id)

    content = await upload.read()
    if len(content) > _MAX_FILE_SIZE:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="File quá lớn. Giới hạn 20MB.",
        )

    mime = upload.content_type or "application/octet-stream"
    if mime not in _ALLOWED_MIME:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Định dạng file không được hỗ trợ: {mime}",
        )

    safe_name = Path(upload.filename or "file").name
    object_name = f"recruitment/candidates/{candidate_id}/{attachment_type}/{uuid.uuid4().hex[:8]}_{safe_name}"

    from app.core.storage import _client, bucket_name

    _client().put_object(
        bucket_name=bucket_name(),
        object_name=object_name,
        data=BytesIO(content),
        length=len(content),
        content_type=mime,
    )

    att = CandidateAttachment(
        candidate_id=candidate_id,
        attachment_type=attachment_type,
        file_path=object_name,
        file_name=safe_name,
        file_size=len(content),
        mime_type=mime,
        note=note,
        uploaded_by_id=uploaded_by_id,
    )
    session.add(att)
    await session.flush()

    return CandidateAttachmentRead(
        id=att.id,
        candidate_id=att.candidate_id,
        attachment_type=att.attachment_type,
        attachment_type_label=AttachmentTypeLabels.get(att.attachment_type, att.attachment_type),
        file_name=att.file_name,
        file_size=att.file_size,
        mime_type=att.mime_type,
        note=att.note,
        uploaded_at=att.uploaded_at,
        download_url=f"/api/v1/recruitment/candidates/{att.candidate_id}/attachments/{att.id}/download",
    )


async def delete_candidate_attachment(
    session: AsyncSession, candidate_id: int, att_id: int
) -> None:
    att = await session.get(CandidateAttachment, att_id)
    if not att or att.candidate_id != candidate_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy tài liệu")
    storage.delete_attachment(att.file_path)
    await session.delete(att)
    await session.flush()


async def get_attachment_download(
    session: AsyncSession, candidate_id: int, att_id: int
) -> StreamingResponse:
    att = await session.get(CandidateAttachment, att_id)
    if not att or att.candidate_id != candidate_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy tài liệu")

    return StreamingResponse(
        storage.get_object_stream(att.file_path),
        media_type=att.mime_type or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{att.file_name}"'},
    )


# ── Excel import template ─────────────────────────────────────────────────────

_IMPORT_COLUMNS = [
    "Họ và tên (*)",
    "Ngày sinh (dd/mm/yyyy)",
    "Giới tính (male/female/other)",
    "Số CCCD/Hộ chiếu",
    "Điện thoại",
    "Email cá nhân",
    "Địa chỉ",
    "Công ty hiện tại",
    "Vị trí hiện tại",
    "Lương kỳ vọng",
    "Ghi chú nguồn",
    "Ghi chú nội bộ",
    "Quốc tịch (text)",
]


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Ứng viên"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    req_fill = PatternFill("solid", fgColor="D6E4F0")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(_IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill if "(*)" in col_name else req_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sample row
    sample = [
        "Nguyễn Văn A", "01/01/1995", "male", "012345678901",
        "0901234567", "vana@example.com", "123 Đường ABC, Hà Nội",
        "Công ty XYZ", "Kỹ sư phần mềm", "15000000", "", "", "Việt Nam",
    ]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    widths = [22, 18, 18, 18, 14, 24, 28, 22, 22, 14, 22, 22, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 36

    guide = wb.create_sheet("Hướng dẫn")
    guide_data = [
        ["Cột", "Bắt buộc", "Mô tả", "Ví dụ"],
        ["Họ và tên (*)", "✅", "Họ tên đầy đủ", "Nguyễn Văn A"],
        ["Ngày sinh", "", "Định dạng dd/mm/yyyy", "01/01/1995"],
        ["Giới tính", "", "male | female | other", "male"],
        ["Số CCCD/Hộ chiếu", "", "Số định danh cá nhân", "012345678901"],
        ["Điện thoại", "", "Số điện thoại", "0901234567"],
        ["Email cá nhân", "", "Địa chỉ email cá nhân", "vana@example.com"],
        ["Địa chỉ", "", "Địa chỉ thường trú", "123 Đường ABC, Hà Nội"],
        ["Công ty hiện tại", "", "Nơi đang công tác", "Công ty XYZ"],
        ["Vị trí hiện tại", "", "Chức danh hiện tại", "Kỹ sư phần mềm"],
        ["Lương kỳ vọng", "", "Số tiền (VNĐ)", "15000000"],
        ["Ghi chú nguồn", "", "Nguồn tiếp nhận hồ sơ", ""],
        ["Ghi chú nội bộ", "", "Ghi chú nội bộ", ""],
        ["Quốc tịch (text)", "", "Dùng để auto-map sang catalog quốc tịch", "Việt Nam"],
    ]
    for row_data in guide_data:
        guide.append(row_data)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Excel import ──────────────────────────────────────────────────────────────


def _parse_date(val):
    if not val:
        return None
    from datetime import datetime
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except Exception:
            pass
    return None


async def import_candidates_excel(
    session: AsyncSession,
    content: bytes,
    created_by_id: int,
) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status.HTTP_422_UNPROCESSABLE_ENTITY, detail="File Excel không hợp lệ")

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = updated = skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows, start=2):
        if not row or not row[0]:
            skipped += 1
            continue

        def _col(idx):
            return row[idx] if len(row) > idx else None

        full_name = str(_col(0)).strip() if _col(0) else None
        if not full_name:
            errors.append(f"Dòng {row_num}: Họ và tên không được để trống")
            skipped += 1
            continue

        dob_raw = str(_col(1)).strip() if _col(1) else None
        gender_raw = str(_col(2)).strip().lower() if _col(2) else None
        id_number = str(_col(3)).strip() if _col(3) else None
        phone = str(_col(4)).strip() if _col(4) else None
        personal_email = str(_col(5)).strip().lower() if _col(5) else None
        address = str(_col(6)).strip() if _col(6) else None
        current_company = str(_col(7)).strip() if _col(7) else None
        current_position = str(_col(8)).strip() if _col(8) else None
        salary_raw = _col(9)
        source_note = str(_col(10)).strip() if _col(10) else None
        internal_note = str(_col(11)).strip() if _col(11) else None
        raw_nationality_text = str(_col(12)).strip() if _col(12) else None

        dob = _parse_date(dob_raw)
        gender = gender_raw if gender_raw in ("male", "female", "other") else None

        expected_salary = None
        if salary_raw is not None:
            try:
                expected_salary = float(str(salary_raw).replace(",", "").replace(".", ""))
            except Exception:
                pass

        # Upsert by personal_email
        existing = None
        if personal_email:
            q = await session.execute(
                select(Candidate).where(Candidate.personal_email == personal_email, Candidate.is_active == True)  # noqa: E712
            )
            existing = q.scalars().first()

        if existing:
            existing.full_name = full_name
            if not existing.last_name or not existing.first_name:
                last_name, first_name = candidate_service._split_full_name(full_name)
                existing.last_name = existing.last_name or last_name
                existing.first_name = existing.first_name or first_name
            if dob:
                existing.date_of_birth = dob
            if gender:
                existing.gender = gender
            if id_number:
                existing.id_number = id_number
            if phone:
                existing.phone_number = phone
            if personal_email:
                existing.personal_email = personal_email
            if address:
                existing.address = address
            if current_company:
                existing.current_company = current_company
            if current_position:
                existing.current_position = current_position
            if raw_nationality_text:
                existing.raw_nationality_text = raw_nationality_text
                existing.nationality_id = await candidate_service._resolve_nationality_id(
                    session,
                    existing.nationality_id,
                    raw_nationality_text,
                )
            if expected_salary is not None:
                existing.expected_salary = expected_salary
            if source_note:
                existing.source_note = source_note
            if internal_note:
                existing.internal_note = internal_note
            from app.services.candidate_service import _utcnow
            existing.updated_at = _utcnow()
            await session.flush()
            updated += 1
        else:
            last_name, first_name = candidate_service._split_full_name(full_name)
            c = Candidate(
                full_name=full_name,
                last_name=last_name,
                first_name=first_name,
                date_of_birth=dob,
                gender=gender,
                id_number=id_number,
                phone_number=phone,
                personal_email=personal_email or None,
                raw_nationality_text=raw_nationality_text,
                nationality_id=await candidate_service._resolve_nationality_id(session, None, raw_nationality_text),
                address=address,
                current_company=current_company,
                current_position=current_position,
                expected_salary=expected_salary,
                source_note=source_note,
                internal_note=internal_note,
                created_by_id=created_by_id,
            )
            session.add(c)
            await session.flush()
            created += 1

    return ImportResult(created=created, updated=updated, skipped=skipped, errors=errors)
