"""Import phòng ban hàng loạt từ file Excel."""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org import Department
from app.schemas.department import DepartmentCreate, DepartmentUpdate, DeptType
from app.schemas.employee_import import ImportResult, ImportRowError
from app.services.import_excel_helper import extract_header_and_non_blank_rows
from app.services import department_service

IMPORT_MAX_ROWS = 1000

COLUMNS = [
    "Mã phòng ban",
    "Tên phòng ban",
    "Tên viết tắt",
    "Mã phòng ban cha",
    "Tiền tố hiển thị",
    "Loại đơn vị",
    "Thứ tự",
    "Kích hoạt",
]
REQUIRED_COLUMNS = {"Mã phòng ban", "Tên phòng ban"}
VALID_DEPT_TYPES = {item.value for item in DeptType}


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Phòng ban"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")
    opt_font = Font(color="1F2937", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font if col_name in REQUIRED_COLUMNS else opt_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample = ["HCNS", "Phòng Hành chính nhân sự", "HCNS", "", "HC", "PHONG", "10", "1"]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    widths = [18, 32, 18, 20, 18, 16, 12, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    ws.row_dimensions[1].height = 36

    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Ví dụ"],
        ["Mã phòng ban", "✅", "Mã duy nhất của phòng ban", "HCNS"],
        ["Tên phòng ban", "✅", "Tên hiển thị đầy đủ", "Phòng Hành chính nhân sự"],
        ["Tên viết tắt", "", "Tên rút gọn", "HCNS"],
        ["Mã phòng ban cha", "", "Để trống nếu là đơn vị cấp gốc", "HC"],
        ["Tiền tố hiển thị", "", "Tiền tố mã nhân viên hiển thị", "HC"],
        ["Loại đơn vị", "", "PHONG | BAN | BO_PHAN | NHOM | TO", "PHONG"],
        ["Thứ tự", "", "Số thứ tự hiển thị", "10"],
        ["Kích hoạt", "", "1/true/yes = hoạt động; 0/false/no = khóa", "1"],
        [],
        ["LƯU Ý", "", "Có thể import theo kiểu upsert: trùng mã sẽ cập nhật record hiện có.", ""],
        ["", "", f"Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["C"].width = 42
    guide.column_dimensions["D"].width = 32
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    value = row_vals[idx] if idx < len(row_vals) else None
    return str(value).strip() if value is not None else ""


def _parse_bool(raw: str, row: int, col: str, errors: list[ImportRowError]) -> Optional[bool]:
    if not raw:
        return None
    value = raw.strip().lower()
    if value in {"1", "true", "yes", "y", "co", "có"}:
        return True
    if value in {"0", "false", "no", "n", "khong", "không"}:
        return False
    errors.append(ImportRowError(row=row, column=col, message=f"Giá trị không hợp lệ: '{raw}'"))
    return None


def _parse_int(raw: str, row: int, col: str, errors: list[ImportRowError]) -> Optional[int]:
    if not raw:
        return None
    try:
        return int(raw)
    except ValueError:
        errors.append(ImportRowError(row=row, column=col, message=f"Giá trị phải là số nguyên: '{raw}'"))
        return None


async def _get_department_by_code(session: AsyncSession, code: str) -> Optional[Department]:
    result = await session.execute(
        select(Department).where(Department.code == code.strip().upper(), Department.deleted_at.is_(None))
    )
    return result.scalar_one_or_none()


async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    header_row, data_rows = extract_header_and_non_blank_rows(ws)
    if not header_row:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    col_map = {header: idx for idx, header in enumerate(header_row)}
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    pending: list[tuple[int, dict[str, object]]] = []
    total = 0
    failed = 0
    errors: list[ImportRowError] = []

    for excel_row, row_vals in data_rows:
        total += 1
        row_errors: list[ImportRowError] = []
        code = _cell(row_vals, "Mã phòng ban", col_map).upper()
        name = _cell(row_vals, "Tên phòng ban", col_map)
        short_name = _cell(row_vals, "Tên viết tắt", col_map) or None
        parent_code = _cell(row_vals, "Mã phòng ban cha", col_map).upper() or None
        display_prefix = _cell(row_vals, "Tiền tố hiển thị", col_map) or None
        dept_type = (_cell(row_vals, "Loại đơn vị", col_map).upper() or "PHONG")
        order_no = _parse_int(_cell(row_vals, "Thứ tự", col_map), excel_row, "Thứ tự", row_errors)
        is_active = _parse_bool(_cell(row_vals, "Kích hoạt", col_map), excel_row, "Kích hoạt", row_errors)

        if not code:
            row_errors.append(ImportRowError(row=excel_row, column="Mã phòng ban", message="Trường bắt buộc không được để trống"))
        if not name:
            row_errors.append(ImportRowError(row=excel_row, column="Tên phòng ban", message="Trường bắt buộc không được để trống"))
        if dept_type not in VALID_DEPT_TYPES:
            row_errors.append(ImportRowError(
                row=excel_row,
                column="Loại đơn vị",
                message=f"Giá trị phải là một trong: {', '.join(sorted(VALID_DEPT_TYPES))}",
            ))

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        pending.append((
            excel_row,
            {
                "code": code,
                "name": name,
                "short_name": short_name,
                "parent_code": parent_code,
                "display_prefix": display_prefix,
                "dept_type": dept_type,
                "order_no": order_no if order_no is not None else 0,
                "is_active": is_active,
            },
        ))

    created_ids: list[int] = []
    success = 0
    remaining = pending[:]
    while remaining:
        progressed = False
        next_round: list[tuple[int, dict[str, object]]] = []
        for excel_row, item in remaining:
            row_errors: list[ImportRowError] = []
            parent_id = None
            parent_code = item["parent_code"]
            if parent_code:
                parent = await _get_department_by_code(session, str(parent_code))
                if not parent:
                    next_round.append((excel_row, item))
                    continue
                parent_id = parent.id

            existing = await _get_department_by_code(session, str(item["code"]))
            try:
                if existing:
                    updated = await department_service.update(
                        session,
                        existing.id,
                        DepartmentUpdate(
                            name=str(item["name"]),
                            short_name=item["short_name"],
                            display_prefix=item["display_prefix"],
                            parent_id=parent_id,
                            dept_type=DeptType(str(item["dept_type"])),
                            order_no=int(item["order_no"]),
                            is_active=item["is_active"],
                        ),
                    )
                    created_ids.append(updated.id)
                else:
                    created = await department_service.create(
                        session,
                        DepartmentCreate(
                            code=str(item["code"]),
                            name=str(item["name"]),
                            short_name=item["short_name"],
                            display_prefix=item["display_prefix"],
                            parent_id=parent_id,
                            dept_type=DeptType(str(item["dept_type"])),
                            order_no=int(item["order_no"]),
                        ),
                    )
                    if item["is_active"] is False:
                        created = await department_service.update(
                            session,
                            created.id,
                            DepartmentUpdate(is_active=False),
                        )
                    created_ids.append(created.id)
                success += 1
                progressed = True
            except Exception as exc:
                row_errors.append(ImportRowError(row=excel_row, column="Mã phòng ban", message=str(exc)))

            if row_errors:
                errors.extend(row_errors)
                failed += 1

        if progressed:
            remaining = next_round
            continue

        for excel_row, item in next_round:
            parent_code = item["parent_code"]
            errors.append(ImportRowError(
                row=excel_row,
                column="Mã phòng ban cha",
                message=f"Không tìm thấy phòng ban cha với mã '{parent_code}'",
            ))
            failed += 1
        break

    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
