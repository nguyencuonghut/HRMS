"""Service cho Document Checklist nhân viên (Plan 13.6)."""
from __future__ import annotations

import uuid
from datetime import date, datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import HTTPException, UploadFile, status
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel, ConfigDict
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.storage import _client, bucket_name, delete_attachment
from app.models.employee import Employee
from app.models.employee_job import EmployeeJobRecord
from app.models.org import Department, JobPosition
from app.models.recruitment import DocumentChecklistType, EmployeeDocumentChecklist


# ── Internal Pydantic schemas ─────────────────────────────────────────────────


class ChecklistItemRead(BaseModel):
    id: int
    document_type_id: int
    document_type_name: str
    document_type_code: str
    is_required: bool
    has_expiry: bool
    status: str
    submitted_at: Optional[date]
    expires_at: Optional[date]
    days_until_expiry: Optional[int]
    is_expiring_soon: bool
    waived_reason: Optional[str]
    has_file: bool
    file_name: Optional[str]
    note: Optional[str]
    updated_at: datetime

    model_config = ConfigDict(from_attributes=True)


class ChecklistItemUpdate(BaseModel):
    status: Optional[str] = None
    submitted_at: Optional[date] = None
    expires_at: Optional[date] = None
    note: Optional[str] = None


class EmployeeChecklistSummary(BaseModel):
    employee_id: int
    employee_code: str
    employee_name: str
    department_name: Optional[str]
    total_required: int
    submitted_count: int
    missing_count: int
    expiring_count: int
    completion_rate: float


# ── Helpers ───────────────────────────────────────────────────────────────────


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _today() -> date:
    return date.today()


def _to_read(item: EmployeeDocumentChecklist, dtype: DocumentChecklistType) -> ChecklistItemRead:
    today = _today()

    # Compute expiry info
    days_until_expiry: Optional[int] = None
    if item.expires_at is not None:
        days_until_expiry = (item.expires_at - today).days

    is_expiring_soon = days_until_expiry is not None and days_until_expiry <= 30

    # Auto-mark expired in response only (no DB write)
    effective_status = item.status
    if item.expires_at is not None and item.expires_at < today and item.status == "submitted":
        effective_status = "expired"

    return ChecklistItemRead(
        id=item.id,
        document_type_id=item.document_type_id,
        document_type_name=dtype.name,
        document_type_code=dtype.code,
        is_required=dtype.is_required,
        has_expiry=dtype.has_expiry,
        status=effective_status,
        submitted_at=item.submitted_at,
        expires_at=item.expires_at,
        days_until_expiry=days_until_expiry,
        is_expiring_soon=is_expiring_soon,
        waived_reason=item.waived_reason,
        has_file=item.file_path is not None,
        file_name=item.file_name,
        note=item.note,
        updated_at=item.updated_at,
    )


async def _get_item_with_dtype(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
) -> tuple[EmployeeDocumentChecklist, DocumentChecklistType]:
    item = await session.get(EmployeeDocumentChecklist, item_id)
    if not item or item.employee_id != employee_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy mục checklist")
    dtype = await session.get(DocumentChecklistType, item.document_type_id)
    if not dtype:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy loại tài liệu")
    return item, dtype


# ── Public service functions ──────────────────────────────────────────────────


async def get_employee_checklist(
    session: AsyncSession,
    employee_id: int,
) -> list[ChecklistItemRead]:
    """Lấy danh sách checklist tài liệu của một nhân viên."""
    q = (
        select(EmployeeDocumentChecklist, DocumentChecklistType)
        .join(
            DocumentChecklistType,
            EmployeeDocumentChecklist.document_type_id == DocumentChecklistType.id,
        )
        .where(EmployeeDocumentChecklist.employee_id == employee_id)
        .order_by(DocumentChecklistType.sort_order, DocumentChecklistType.name)
    )
    rows = (await session.execute(q)).all()
    return [_to_read(item, dtype) for item, dtype in rows]


async def update_checklist_item(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
    data: ChecklistItemUpdate,
    user_id: int,
) -> ChecklistItemRead:
    """Cập nhật trạng thái / thông tin mục checklist."""
    item, dtype = await _get_item_with_dtype(session, employee_id, item_id)
    today = _today()

    if data.status is not None:
        item.status = data.status
    if data.submitted_at is not None:
        item.submitted_at = data.submitted_at
    if data.expires_at is not None:
        item.expires_at = data.expires_at
    if data.note is not None:
        item.note = data.note

    # Validate: submitted + has_expiry but no expires_at
    if item.status == "submitted" and dtype.has_expiry and item.expires_at is None:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Loại tài liệu này yêu cầu ngày hết hạn khi trạng thái là 'submitted'",
        )

    # Auto-expire
    if item.expires_at is not None and item.expires_at < today:
        item.status = "expired"

    item.updated_by_id = user_id
    item.updated_at = _utcnow()
    await session.flush()
    return _to_read(item, dtype)


async def upload_document_file(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
    upload: UploadFile,
    user_id: int,
) -> ChecklistItemRead:
    """Upload file tài liệu lên MinIO và gắn vào mục checklist."""
    item, dtype = await _get_item_with_dtype(session, employee_id, item_id)

    content = await upload.read()
    safe_name = Path(upload.filename or "file").name
    object_name = (
        f"recruitment/employee-documents/{employee_id}"
        f"/{dtype.code}/{uuid.uuid4().hex[:8]}_{safe_name}"
    )
    content_type = upload.content_type or "application/octet-stream"

    _client().put_object(
        bucket_name=bucket_name(),
        object_name=object_name,
        data=BytesIO(content),
        length=len(content),
        content_type=content_type,
    )

    # If there was an old file, delete it from MinIO
    if item.file_path and item.file_path != object_name:
        delete_attachment(item.file_path)

    item.file_path = object_name
    item.file_name = safe_name
    item.file_size = len(content)
    item.mime_type = content_type

    # Auto-submit if currently not submitted
    if item.status == "not_submitted":
        item.status = "submitted"

    item.updated_by_id = user_id
    item.updated_at = _utcnow()
    await session.flush()
    return _to_read(item, dtype)


async def get_document_download_stream(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
) -> tuple[str, str, str]:
    """Trả về (file_path, file_name, mime_type) để caller dùng với get_object_stream."""
    item, _ = await _get_item_with_dtype(session, employee_id, item_id)
    if not item.file_path:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Mục này chưa có file đính kèm")
    return (
        item.file_path,
        item.file_name or "document",
        item.mime_type or "application/octet-stream",
    )


async def delete_document_file(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
    user_id: int,
) -> ChecklistItemRead:
    """Xóa file đính kèm khỏi MinIO và mục checklist."""
    item, dtype = await _get_item_with_dtype(session, employee_id, item_id)
    if not item.file_path:
        raise HTTPException(
            status.HTTP_422_UNPROCESSABLE_CONTENT,
            detail="Mục này không có file đính kèm để xóa",
        )

    delete_attachment(item.file_path)

    item.file_path = None
    item.file_name = None
    item.file_size = None
    item.mime_type = None

    if item.status == "submitted":
        item.status = "not_submitted"

    item.updated_by_id = user_id
    item.updated_at = _utcnow()
    await session.flush()
    return _to_read(item, dtype)


async def waive_item(
    session: AsyncSession,
    employee_id: int,
    item_id: int,
    reason: str,
    user_id: int,
) -> ChecklistItemRead:
    """Miễn loại tài liệu này cho nhân viên."""
    item, dtype = await _get_item_with_dtype(session, employee_id, item_id)
    item.status = "waived"
    item.waived_reason = reason
    item.updated_by_id = user_id
    item.updated_at = _utcnow()
    await session.flush()
    return _to_read(item, dtype)


async def get_missing_documents_report(
    session: AsyncSession,
    status_filter: Optional[str] = None,
    department_id: Optional[int] = None,
    search: Optional[str] = None,
) -> list[EmployeeChecklistSummary]:
    """Báo cáo tình trạng nộp hồ sơ của các nhân viên (chỉ loại bắt buộc)."""
    today = _today()
    soon_threshold = today + timedelta(days=30)

    # Load active employees with their current job record + department
    emp_q = (
        select(Employee, EmployeeJobRecord, Department)
        .join(
            EmployeeJobRecord,
            (EmployeeJobRecord.employee_id == Employee.id) & (EmployeeJobRecord.is_current == True),
            isouter=True,
        )
        .join(
            Department,
            Department.id == EmployeeJobRecord.department_id,
            isouter=True,
        )
        .where(Employee.status.in_(["probation", "official"]))
    )
    if department_id is not None:
        emp_q = emp_q.where(EmployeeJobRecord.department_id == department_id)
    if search:
        kw = f"%{search.strip()}%"
        emp_q = emp_q.where(Employee.full_name.ilike(kw))

    emp_rows = (await session.execute(emp_q)).all()

    if not emp_rows:
        return []

    employee_ids = [emp.id for emp, _, _ in emp_rows]

    # Build display codes (e.g. HC0001, CNT0009)
    from app.services import employee_code_service
    employees_only = [emp for emp, _, _ in emp_rows]
    display_codes: dict[int, str] = await employee_code_service.batch_build_employee_display_codes(
        session, employees_only
    )

    # Load all required checklist items for these employees in one query
    checklist_q = (
        select(EmployeeDocumentChecklist, DocumentChecklistType)
        .join(
            DocumentChecklistType,
            EmployeeDocumentChecklist.document_type_id == DocumentChecklistType.id,
        )
        .where(
            EmployeeDocumentChecklist.employee_id.in_(employee_ids),
            DocumentChecklistType.is_required == True,
        )
    )
    checklist_rows = (await session.execute(checklist_q)).all()

    # Group checklist items by employee_id
    from collections import defaultdict
    items_by_employee: dict[int, list[tuple[EmployeeDocumentChecklist, DocumentChecklistType]]] = defaultdict(list)
    for item, dtype in checklist_rows:
        items_by_employee[item.employee_id].append((item, dtype))

    # Build summaries
    summaries: list[EmployeeChecklistSummary] = []
    for emp, job_record, dept in emp_rows:
        emp_items = items_by_employee.get(emp.id, [])
        total_required = len(emp_items)
        submitted_count = 0
        missing_count = 0
        expiring_count = 0

        for item, dtype in emp_items:
            # Determine effective status
            effective_status = item.status
            if (
                item.expires_at is not None
                and item.expires_at < today
                and item.status == "submitted"
            ):
                effective_status = "expired"

            if effective_status == "submitted":
                submitted_count += 1
            elif effective_status == "not_submitted":
                missing_count += 1

            # Expiring soon: submitted, not yet expired, expires within 30 days
            if (
                effective_status == "submitted"
                and item.expires_at is not None
                and today <= item.expires_at <= soon_threshold
            ):
                expiring_count += 1

        completion_rate = (submitted_count / total_required * 100) if total_required > 0 else 0.0

        employee_code = display_codes.get(emp.id, str(emp.employee_seq))

        summary = EmployeeChecklistSummary(
            employee_id=emp.id,
            employee_code=employee_code,
            employee_name=emp.full_name,
            department_name=dept.name if dept else None,
            total_required=total_required,
            submitted_count=submitted_count,
            missing_count=missing_count,
            expiring_count=expiring_count,
            completion_rate=round(completion_rate, 1),
        )

        # Apply status filter on the summary level
        if status_filter == "complete" and completion_rate < 100:
            continue
        if status_filter == "incomplete" and completion_rate >= 100:
            continue
        if status_filter == "expiring" and expiring_count == 0:
            continue

        summaries.append(summary)

    return summaries


# ── Excel export ──────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_THIN = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


def _cell(ws, row: int, col: int, value=None, *, font=None, fill=None,
          alignment=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if border:
        c.border = border
    return c


async def export_labor_report_excel(
    session: AsyncSession,
    year: int,
    month: int,
) -> bytes:
    """Xuất danh sách lao động mới trong tháng ra Excel (theo TT23/2014)."""
    # Query employees whose probation started in the given month/year
    q = (
        select(Employee, EmployeeJobRecord, Department, JobPosition)
        .join(EmployeeJobRecord, EmployeeJobRecord.employee_id == Employee.id)
        .join(Department, Department.id == EmployeeJobRecord.department_id, isouter=True)
        .join(JobPosition, JobPosition.id == EmployeeJobRecord.job_position_id, isouter=True)
        .where(
            func.extract("year", EmployeeJobRecord.probation_start_date) == year,
            func.extract("month", EmployeeJobRecord.probation_start_date) == month,
            EmployeeJobRecord.is_current == True,
        )
        .order_by(Employee.full_name)
    )
    rows = (await session.execute(q)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = f"T{month:02d}-{year}"

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws.cell(row=1, column=1, value=f"DANH SÁCH LAO ĐỘNG MỚI THÁNG {month}/{year}")
    title_cell.font = Font(bold=True, size=13)
    title_cell.alignment = _CENTER

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        ("STT", 6),
        ("Họ và tên", 28),
        ("Ngày sinh", 14),
        ("Giới tính", 12),
        ("Số CCCD", 16),
        ("Vị trí công việc", 24),
        ("Phòng ban", 22),
        ("Ngày vào làm", 15),
        ("Ghi chú", 20),
    ]
    for col_idx, (header_text, col_width) in enumerate(headers, start=1):
        _cell(
            ws, 2, col_idx, header_text,
            font=_HEADER_FONT,
            fill=_HEADER_FILL,
            alignment=_CENTER,
            border=_THIN_BORDER,
        )
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = col_width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    gender_map = {"male": "Nam", "female": "Nữ", "other": "Khác"}

    for row_offset, (emp, job, dept, pos) in enumerate(rows):
        row_num = row_offset + 3
        fill = _ALT_FILL if row_offset % 2 == 1 else None

        dob_str = emp.date_of_birth.strftime("%d/%m/%Y") if emp.date_of_birth else ""
        start_str = job.probation_start_date.strftime("%d/%m/%Y") if job.probation_start_date else ""
        gender_str = gender_map.get(emp.gender, emp.gender)
        dept_name = dept.name if dept else ""
        pos_name = pos.name if pos else ""

        row_data = [
            row_offset + 1,   # STT
            emp.full_name,
            dob_str,
            gender_str,
            emp.id_number,
            pos_name,
            dept_name,
            start_str,
            "",               # Ghi chú — để trống
        ]

        for col_idx, value in enumerate(row_data, start=1):
            alignment = _CENTER if col_idx in (1, 3, 4, 8) else _LEFT
            _cell(
                ws, row_num, col_idx, value,
                font=_NORMAL_FONT,
                fill=fill,
                alignment=alignment,
                border=_THIN_BORDER,
            )

        ws.row_dimensions[row_num].height = 18

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Internal initialiser ──────────────────────────────────────────────────────


async def init_employee_checklist(
    session: AsyncSession,
    employee_id: int,
    user_id: int,
    is_foreign_worker: bool = False,
) -> list[ChecklistItemRead]:
    """Khởi tạo checklist cho nhân viên cũ (không qua convert_to_employee).
    Bỏ qua các loại đã có (upsert-safe).
    """
    applies_to_values = ["all"]
    if is_foreign_worker:
        applies_to_values.append("foreign_worker")

    q = (
        select(DocumentChecklistType)
        .where(
            DocumentChecklistType.is_active == True,
            DocumentChecklistType.applies_to.in_(applies_to_values),
        )
        .order_by(DocumentChecklistType.sort_order)
    )
    dtypes = (await session.execute(q)).scalars().all()

    existing_type_ids = {
        r for (r,) in (await session.execute(
            select(EmployeeDocumentChecklist.document_type_id)
            .where(EmployeeDocumentChecklist.employee_id == employee_id)
        )).all()
    }

    for dtype in dtypes:
        if dtype.id not in existing_type_ids:
            session.add(EmployeeDocumentChecklist(
                employee_id=employee_id,
                document_type_id=dtype.id,
                status="not_submitted",
                created_by_id=user_id,
                updated_by_id=user_id,
            ))

    await session.flush()
    return await get_employee_checklist(session, employee_id)


async def add_checklist_item(
    session: AsyncSession,
    employee_id: int,
    document_type_id: int,
    user_id: int,
) -> ChecklistItemRead:
    """Thêm thủ công một loại giấy tờ vào checklist nhân viên."""
    dtype = await session.get(DocumentChecklistType, document_type_id)
    if not dtype:
        raise HTTPException(status.HTTP_404_NOT_FOUND, detail="Không tìm thấy loại giấy tờ")

    existing = (await session.execute(
        select(EmployeeDocumentChecklist)
        .where(
            EmployeeDocumentChecklist.employee_id == employee_id,
            EmployeeDocumentChecklist.document_type_id == document_type_id,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, detail="Loại giấy tờ này đã có trong checklist")

    item = EmployeeDocumentChecklist(
        employee_id=employee_id,
        document_type_id=document_type_id,
        status="not_submitted",
        created_by_id=user_id,
        updated_by_id=user_id,
    )
    session.add(item)
    await session.flush()
    return _to_read(item, dtype)


async def _init_document_checklist(
    session: AsyncSession,
    employee_id: int,
    is_foreign_worker: bool = False,
) -> None:
    """Khởi tạo checklist tài liệu cho nhân viên mới (gọi sau convert_to_employee).

    Không gọi flush — caller chịu trách nhiệm flush/commit.
    """
    applies_to_values = ["all"]
    if is_foreign_worker:
        applies_to_values.append("foreign_worker")

    q = (
        select(DocumentChecklistType)
        .where(
            DocumentChecklistType.is_active == True,
            DocumentChecklistType.applies_to.in_(applies_to_values),
        )
        .order_by(DocumentChecklistType.sort_order, DocumentChecklistType.name)
    )
    dtype_rows = (await session.execute(q)).scalars().all()

    for dtype in dtype_rows:
        checklist_item = EmployeeDocumentChecklist(
            employee_id=employee_id,
            document_type_id=dtype.id,
            status="not_submitted",
        )
        session.add(checklist_item)
