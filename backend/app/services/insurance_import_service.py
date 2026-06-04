"""Import hồ sơ bảo hiểm hàng loạt từ file Excel (12.1 — Slice 2)."""
from __future__ import annotations

import structlog

logger = structlog.get_logger(__name__)
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.bhyt_clinic import BhytClinic
from app.models.employee_contract import EmployeeContract
from app.models.employee_insurance import EmployeeInsuranceProfile
from app.schemas.employee_import import ImportResult, ImportRowError
from app.services.import_employee_lookup_service import EmployeeImportLookup

IMPORT_MAX_ROWS = 1000
VALID_BASIS_SOURCES = {"manual_fixed", "contract", "computed"}

COLUMNS = [
    "Mã nhân viên",
    "Hệ mã nhân viên",
    "Mã BHXH",
    "Ngày tham gia",
    "Mức lương đóng",
    "Nguồn mức lương đóng",
    "Mã bệnh viện KCB",
    "Trạng thái",
]
REQUIRED_COLUMNS = {"Mã nhân viên", "Ngày tham gia", "Mức lương đóng"}

VALID_STATUSES = {"active", "paused", "stopped"}


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


# ── Template ──────────────────────────────────────────────────────────────────

def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bảo hiểm"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample = ["1", "SYS1", "BHXH1234567", "01/01/2026", "5000000", "manual_fixed", "01003", "active"]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    widths = [16, 16, 18, 14, 18, 18, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 36

    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Giá trị hợp lệ / Ví dụ"],
        ["Mã nhân viên", "✅", "Mã hiển thị hoặc phần số nhân viên", "1 / 001 / NV001"],
        ["Hệ mã nhân viên", "", "Bắt buộc khi công ty dùng nhiều hệ mã", "SYS1 / SYS2 / SYS3"],
        ["Mã BHXH", "", "Số sổ BHXH (tùy chọn)", "BHXH1234567"],
        ["Ngày tham gia", "✅", "Ngày bắt đầu tham gia BHXH (dd/mm/yyyy)", "01/01/2026"],
        ["Mức lương đóng", "✅*", "Mức lương đóng BHXH (VNĐ, > 0)", "5000000"],
        ["Nguồn mức lương đóng", "", "Để trống = giữ theo profile/hợp đồng hiện có; manual_fixed / contract / computed", "manual_fixed"],
        ["Mã bệnh viện KCB", "", "Mã bệnh viện khám chữa bệnh ban đầu", "01003"],
        ["Trạng thái", "", "Trạng thái tham gia", "active / paused / stopped (mặc định: active)"],
        [],
        ["LƯU Ý:", "", "Nếu nguồn = manual_fixed thì Mức lương đóng là bắt buộc.", ""],
        ["", "", "Nếu nguồn = contract/computed thì importer sẽ dùng mức từ hợp đồng active mới nhất; nếu điền tay thì phải khớp với hợp đồng.", ""],
        ["", "", "Nếu để trống nguồn, importer sẽ ưu tiên giữ source hiện tại của profile; nếu chưa có profile thì suy ra từ hợp đồng active.", ""],
        ["", "", f"Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["C"].width = 46
    guide.column_dimensions["D"].width = 36
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val: object, col: str, row: int, errors: list[ImportRowError]) -> Optional[date]:
    raw = str(val).strip() if val is not None else ""
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    errors.append(ImportRowError(row=row, column=col, message=f"Định dạng ngày không hợp lệ: '{raw}' (cần dd/mm/yyyy)"))
    return None


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    v = row_vals[idx] if idx < len(row_vals) else None
    return str(v).strip() if v is not None else ""

async def _find_bhyt_clinic(session: AsyncSession, code: str) -> Optional[BhytClinic]:
    result = await session.execute(select(BhytClinic).where(BhytClinic.code == code.strip()))
    return result.scalar_one_or_none()


async def _find_existing_profile(session: AsyncSession, employee_id: int) -> Optional[EmployeeInsuranceProfile]:
    result = await session.execute(
        select(EmployeeInsuranceProfile).where(EmployeeInsuranceProfile.employee_id == employee_id)
    )
    return result.scalar_one_or_none()


async def _find_active_contract_for_basis(session: AsyncSession, employee_id: int) -> Optional[EmployeeContract]:
    result = await session.execute(
        select(EmployeeContract)
        .where(
            EmployeeContract.employee_id == employee_id,
            EmployeeContract.status == "active",
            EmployeeContract.insurance_salary.is_not(None),
        )
        .order_by(EmployeeContract.effective_from.desc(), EmployeeContract.id.desc())
        .limit(1)
    )
    return result.scalar_one_or_none()


# ── Main import ───────────────────────────────────────────────────────────────

async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("excel_read_error", error=str(exc))
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    header_row = [str(c).strip() if c else "" for c in all_rows[0]]
    col_map = {h: i for i, h in enumerate(header_row)}

    data_rows = all_rows[1:]
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = success = failed = 0
    errors: list[ImportRowError] = []
    created_ids: list[int] = []
    employee_lookup = await EmployeeImportLookup.build(session)

    for rel_idx, row_vals in enumerate(data_rows):
        excel_row = rel_idx + 2
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        total += 1
        row_errors: list[ImportRowError] = []

        def get(col_name: str) -> str:
            return _cell(row_vals, col_name, col_map)

        emp_code     = get("Mã nhân viên")
        emp_sequence = get("Hệ mã nhân viên")
        bhxh_code    = get("Mã BHXH") or None
        clinic_code  = get("Mã bệnh viện KCB") or None
        basis_source_raw = get("Nguồn mức lương đóng").lower()
        status_raw   = get("Trạng thái").lower() or "active"

        if not emp_code:
            row_errors.append(ImportRowError(row=excel_row, column="Mã nhân viên", message="Trường bắt buộc không được để trống"))

        participation_date = _parse_date(get("Ngày tham gia"), "Ngày tham gia", excel_row, row_errors)
        if not participation_date and "Ngày tham gia" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày tham gia", message="Trường bắt buộc không được để trống"))

        # Mức lương đóng
        basis_amount: Optional[Decimal] = None
        salary_raw = get("Mức lương đóng")
        if salary_raw:
            try:
                basis_amount = Decimal(salary_raw.replace(",", "").replace(".", ""))
                if basis_amount <= 0:
                    row_errors.append(ImportRowError(row=excel_row, column="Mức lương đóng", message="Mức lương đóng phải > 0"))
                    basis_amount = None
            except InvalidOperation:
                row_errors.append(ImportRowError(row=excel_row, column="Mức lương đóng", message=f"Giá trị không hợp lệ: '{salary_raw}'"))

        if basis_source_raw and basis_source_raw not in VALID_BASIS_SOURCES:
            row_errors.append(ImportRowError(
                row=excel_row,
                column="Nguồn mức lương đóng",
                message="Nguồn mức lương đóng phải là manual_fixed, contract hoặc computed",
            ))

        if status_raw and status_raw not in VALID_STATUSES:
            row_errors.append(ImportRowError(
                row=excel_row, column="Trạng thái",
                message=f"Giá trị phải là active/paused/stopped, nhận được: '{status_raw}'"
            ))

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        # DB lookups
        employee_lookup_result = employee_lookup.resolve(
            employee_code_raw=emp_code,
            sequence_code_raw=emp_sequence,
        )
        employee = employee_lookup_result.employee
        if not employee:
            errors.append(ImportRowError(
                row=excel_row,
                column="Mã nhân viên",
                message=employee_lookup_result.error or f"Không tìm thấy nhân viên với mã '{emp_code}'",
            ))
            failed += 1
            continue

        if not employee.is_active:
            errors.append(ImportRowError(row=excel_row, column="Mã nhân viên", message=f"Nhân viên '{emp_code}' không còn hoạt động"))
            failed += 1
            continue

        clinic: Optional[BhytClinic] = None
        if clinic_code:
            clinic = await _find_bhyt_clinic(session, clinic_code)
            if not clinic:
                errors.append(ImportRowError(row=excel_row, column="Mã bệnh viện KCB", message=f"Không tìm thấy bệnh viện KCB với mã '{clinic_code}'"))
                failed += 1
                continue

        existing = await _find_existing_profile(session, employee.id)
        active_contract = await _find_active_contract_for_basis(session, employee.id)
        existing_meaningful_source = (
            existing.insurance_basis_source
            if existing is not None and existing.insurance_basis_amount is not None
            else ""
        )
        resolved_basis_source = (
            basis_source_raw
            or existing_meaningful_source
            or (
                "computed"
                if active_contract and active_contract.insurance_salary_mode == "computed_by_position_group"
                else "contract"
                if active_contract
                else "manual_fixed"
            )
        )

        resolved_basis_amount: Optional[Decimal] = None
        if resolved_basis_source == "manual_fixed":
            if basis_amount is None:
                errors.append(ImportRowError(
                    row=excel_row,
                    column="Mức lương đóng",
                    message="Nguồn manual_fixed yêu cầu nhập Mức lương đóng > 0",
                ))
                failed += 1
                continue
            resolved_basis_amount = basis_amount
        else:
            contract_amount = (
                Decimal(str(active_contract.insurance_salary))
                if active_contract and active_contract.insurance_salary is not None
                else None
            )
            if contract_amount is not None:
                if basis_amount is not None and basis_amount != contract_amount:
                    errors.append(ImportRowError(
                        row=excel_row,
                        column="Mức lương đóng",
                        message="Mức lương đóng không khớp với hợp đồng active đang dùng cho nguồn contract/computed",
                    ))
                    failed += 1
                    continue
                resolved_basis_amount = contract_amount
            elif (
                not basis_source_raw
                and existing is not None
                and existing.insurance_basis_source == resolved_basis_source
                and existing.insurance_basis_amount is not None
            ):
                resolved_basis_amount = existing.insurance_basis_amount
            else:
                errors.append(ImportRowError(
                    row=excel_row,
                    column="Nguồn mức lương đóng",
                    message="Nguồn contract/computed yêu cầu nhân viên có hợp đồng active với mức lương BHXH, hoặc profile hiện có cùng source",
                ))
                failed += 1
                continue

        # Upsert by employee_id
        try:
            if existing:
                # Update
                existing.bhxh_code = bhxh_code if bhxh_code else existing.bhxh_code
                existing.company_bhxh_joined_date = participation_date
                existing.insurance_basis_amount = resolved_basis_amount
                existing.insurance_basis_source = resolved_basis_source
                existing.participation_status = status_raw
                if clinic:
                    existing.bhyt_initial_clinic_code = clinic.code
                    existing.bhyt_initial_clinic_name = clinic.name
                existing.updated_at = _utcnow()
                await session.flush()
                created_ids.append(existing.id)
            else:
                # Insert
                profile = EmployeeInsuranceProfile(
                    employee_id=employee.id,
                    bhxh_code=bhxh_code,
                    company_bhxh_joined_date=participation_date,
                    insurance_basis_amount=resolved_basis_amount,
                    insurance_basis_source=resolved_basis_source,
                    participation_status=status_raw,
                    bhyt_initial_clinic_code=clinic.code if clinic else None,
                    bhyt_initial_clinic_name=clinic.name if clinic else None,
                )
                session.add(profile)
                await session.flush()
                created_ids.append(profile.id)
            success += 1
        except Exception as exc:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="—", message=f"Lỗi hệ thống: {exc}"))
            failed += 1

    await session.commit()
    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
