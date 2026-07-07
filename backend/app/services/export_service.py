from __future__ import annotations

import io
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Awaitable, Callable, Optional

import structlog
from fastapi import HTTPException, status
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.auth import User
from app.models.export import ExportJob, ReportTemplate
from app.core.export_catalog import (
    EXPORT_FORMAT_LABELS,
    EXPORT_FORMAT_ORDER,
    EXPORT_STATUS_LABELS,
    EXPORT_STATUS_ORDER,
    EXPORT_STATUS_SEVERITIES,
    REPORT_TYPE_LABELS,
    REPORT_TYPE_ORDER,
)
from app.schemas.export import (
    ExportFormat,
    ExportHistoryResponse,
    ExportJobRequest,
    ExportJobResponse,
    ExportJobStatusResponse,
    ReportTemplateCreate,
    ReportTemplateResponse,
    ReportTemplateUpdate,
)
from app.services import (
    auth_service,
    contract_export_service,
    contract_report_service,
    dashboard_service,
    employee_export_service,
    hr_report_service,
    insurance_analytics_service,
    leave_analytics_service,
    leave_report_service,
    probation_report_service,
    recruitment_report_service,
    salary_export_service,
)
from app.utils.excel_style import ExcelStyler

logger = structlog.get_logger(__name__)

ASYNC_THRESHOLD_ROWS = 1_000
EXPORT_DIR = Path("/tmp/hrms_exports")
_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_PDF_MIME = "application/pdf"
_PDF_FONT_NAME = "NotoSans"
_PDF_FONT_PATH = Path(__file__).resolve().parent.parent / "assets" / "fonts" / "NotoSans-Regular.ttf"

@dataclass(frozen=True)
class ReportHandler:
    permissions: tuple[str, ...]
    estimate_rows: Callable[[AsyncSession, dict[str, Any]], Awaitable[int]]
    export_xlsx: Callable[[AsyncSession, dict[str, Any]], Awaitable[io.BytesIO]]


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    return str(value)


def _sheet_rows(sheet: Any) -> list[list[str]]:
    rows: list[list[str]] = []
    max_len = 0
    for row in sheet.iter_rows(values_only=True):
        values = [_stringify_cell(cell).strip() for cell in row]
        while values and values[-1] == "":
            values.pop()
        if not values:
            continue
        rows.append([value if value else "—" for value in values])
        max_len = max(max_len, len(values))
    for row in rows:
        row.extend(["—"] * (max_len - len(row)))
    return rows


def _table_column_widths(rows: list[list[str]], usable_width: float) -> list[float]:
    if not rows:
        return [usable_width]
    raw_widths: list[float] = []
    for col_index in range(len(rows[0])):
        longest = max(len(row[col_index]) for row in rows)
        raw_widths.append(float(min(max(longest, 10), 28)))
    scale = usable_width / sum(raw_widths)
    return [width * scale for width in raw_widths]


def _ensure_pdf_font() -> None:
    if _PDF_FONT_NAME in pdfmetrics.getRegisteredFontNames():
        return
    if not _PDF_FONT_PATH.exists():
        raise FileNotFoundError(f"Thiếu font PDF Unicode: {_PDF_FONT_PATH}")
    pdfmetrics.registerFont(TTFont(_PDF_FONT_NAME, str(_PDF_FONT_PATH)))


def _report_type_label(report_type: str) -> str:
    return REPORT_TYPE_LABELS.get(report_type, report_type)


def _workbook_pdf(report_type: str, filters: dict[str, Any], row_count: int, workbook_bytes: bytes) -> bytes:
    _ensure_pdf_font()
    workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
    is_landscape = any((sheet.max_column or 0) >= 8 for sheet in workbook.worksheets)
    page_size = landscape(A4) if is_landscape else A4
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=24,
        pageCompression=0,
        title=_report_type_label(report_type),
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ExportTitle",
        parent=styles["Heading2"],
        fontName=_PDF_FONT_NAME,
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1E3A5F"),
        spaceAfter=8,
    )
    meta_style = ParagraphStyle(
        "ExportMeta",
        parent=styles["BodyText"],
        fontName=_PDF_FONT_NAME,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=10,
    )
    cell_style = ParagraphStyle(
        "ExportCell",
        parent=styles["BodyText"],
        fontName=_PDF_FONT_NAME,
        fontSize=8 if is_landscape else 9,
        leading=10 if is_landscape else 11,
        textColor=colors.HexColor("#111827"),
    )
    header_cell_style = ParagraphStyle(
        "ExportHeaderCell",
        parent=cell_style,
        fontName=_PDF_FONT_NAME,
        fontSize=8 if is_landscape else 9,
        leading=10 if is_landscape else 11,
        textColor=colors.white,
    )
    story: list[Any] = []

    for index, sheet in enumerate(workbook.worksheets):
        rows = _sheet_rows(sheet)
        if not rows:
            rows = [["Không có dữ liệu"]]

        if index > 0:
            story.append(PageBreak())

        usable_width = doc.width
        widths = _table_column_widths(rows, usable_width=usable_width)
        story.append(Paragraph(f"Báo cáo: {_report_type_label(report_type)}", title_style))
        if index == 0:
            header = f"Sheet: {sheet.title} | Số dòng: {row_count} | Bộ lọc: {filters or '{}'}"
        else:
            header = f"Sheet: {sheet.title}"
        story.append(Paragraph(header, meta_style))

        table_data: list[list[Paragraph]] = []
        for row_index, row in enumerate(rows):
            style = header_cell_style if row_index == 0 else cell_style
            table_data.append(
                [
                    Paragraph(
                        cell.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"),
                        style,
                    )
                    for cell in row
                ]
            )
        table = Table(table_data, colWidths=widths, repeatRows=1, splitByRow=1)
        table.setStyle(
            TableStyle(
                [
                    ("FONTNAME", (0, 0), (-1, -1), _PDF_FONT_NAME),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ("LEFTPADDING", (0, 0), (-1, -1), 5),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 8))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def _coerce_filter_value(key: str, value: Any) -> Any:
    if value is None:
        return None
    if key in {"start_date", "end_date"} and isinstance(value, str):
        return date.fromisoformat(value)
    return value


def _normalize_filters(filters: dict[str, Any]) -> dict[str, Any]:
    return {key: _coerce_filter_value(key, value) for key, value in filters.items()}


async def _estimate_dashboard_rows(session: AsyncSession, filters: dict[str, Any]) -> int:
    data = await dashboard_service.get_headcount_by_dept(session, department_id=filters.get("department_id"))
    return max(len(data), 1)


async def _export_dashboard_xlsx(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    year = filters.get("year")
    month = filters.get("month")
    department_id = filters.get("department_id")
    summary = await dashboard_service.get_summary(session, year=year, month=month, department_id=department_id)
    headcount = await dashboard_service.get_headcount_by_dept(session, department_id=department_id)
    trend = await dashboard_service.get_monthly_trend(session, year=year, department_id=department_id)
    structure = await dashboard_service.get_structure(session, department_id=department_id)

    rows: list[list[object]] = [
        ["Tổng nhân sự", summary.total_headcount],
        ["Mới trong tháng", summary.new_hires_this_month],
        ["Nghỉ việc trong tháng", summary.resigned_this_month],
        ["Số nhân sự đầu tháng", summary.headcount_start_of_month],
        ["Tỷ lệ nghỉ việc", summary.turnover_rate],
        [],
        ["Phòng ban", "Số nhân sự"],
    ]
    rows.extend([[item.department_name, item.headcount] for item in headcount])
    rows.append([])
    rows.append(["Tháng", "Tuyển mới", "Nghỉ việc", "Biến động ròng"])
    rows.extend([[item.month, item.new_hires, item.resigned_count, item.net_change] for item in trend.monthly])
    rows.append([])
    rows.append(["Cơ cấu", "Giá trị"])
    rows.extend([
        [item.label, item.count]
        for item in structure.gender + structure.age_group + structure.education_level + structure.tenure_group
    ])
    return ExcelStyler.build_table_workbook("Dashboard nhân sự", ["Chỉ mục", "Giá trị"], rows, "Dashboard")


async def _estimate_hr_employee_list(session: AsyncSession, filters: dict[str, Any]) -> int:
    normalized = dict(filters)
    normalized.pop("page", None)
    normalized.pop("page_size", None)
    report = await hr_report_service.get_employee_list(session, page=1, page_size=1, **normalized)
    return report.total


async def _export_hr_employee_list(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await hr_report_service.export_employee_list_excel(session, **filters)


async def _estimate_hr_movement(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await hr_report_service.get_movement_report(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        group_by=filters.get("group_by", "month"),
        department_id=filters.get("department_id"),
    )
    return len(report.rows)


async def _export_hr_movement(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await hr_report_service.export_movement_excel(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        group_by=filters.get("group_by", "month"),
        department_id=filters.get("department_id"),
    )


async def _estimate_hr_tenure(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await hr_report_service.get_tenure_report(session, department_id=filters.get("department_id"))
    return report.total_active


async def _export_hr_tenure(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await hr_report_service.export_tenure_excel(session, department_id=filters.get("department_id"))


async def _estimate_hr_org_structure(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await hr_report_service.get_org_structure(session, department_id=filters.get("department_id"))
    return report.department_count


async def _export_hr_org_structure(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await hr_report_service.export_org_structure_excel(session, department_id=filters.get("department_id"))


async def _estimate_leaves(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await leave_analytics_service.get_top_employees(
        session,
        year=filters["year"],
        department_id=filters.get("department_id"),
    )
    return len(report.items)


async def _export_leaves(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    summary = await leave_analytics_service.get_analytics_summary(
        session,
        year=filters["year"],
        department_id=filters.get("department_id"),
        leave_type_id=filters.get("leave_type_id"),
    )
    by_type = await leave_analytics_service.get_by_type(
        session,
        year=filters["year"],
        department_id=filters.get("department_id"),
    )
    heatmap = await leave_analytics_service.get_monthly_heatmap(session, year=filters["year"])
    top = await leave_analytics_service.get_top_employees(
        session,
        year=filters["year"],
        department_id=filters.get("department_id"),
        leave_type_id=filters.get("leave_type_id"),
    )
    return leave_analytics_service.build_analytics_xlsx(summary, by_type, heatmap, top)


async def _estimate_insurance(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await insurance_analytics_service.get_non_participants(
        session,
        department_id=filters.get("department_id"),
        page=1,
        page_size=1,
    )
    return report.total


async def _export_insurance(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await insurance_analytics_service.export_analytics_xlsx(
        session,
        year=filters["year"],
        month=filters["month"],
        department_id=filters.get("department_id"),
    )


async def _estimate_contracts(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await contract_report_service.get_expiring(
        session,
        department_id=filters.get("department_id"),
        days_ahead=filters.get("days_ahead", 90),
        page=1,
        page_size=1,
    )
    return report.total


async def _export_contracts(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await contract_export_service.export_contracts_xlsx(
        session,
        department_id=filters.get("department_id"),
        status=filters.get("status", "active"),
        days_ahead=filters.get("days_ahead"),
    )


async def _estimate_recruitment(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await recruitment_report_service.get_funnel(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        department_id=filters.get("department_id"),
    )
    return len(report.stages)


async def _export_recruitment(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await recruitment_report_service.export_excel(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        department_id=filters.get("department_id"),
    )


async def _estimate_probation(session: AsyncSession, filters: dict[str, Any]) -> int:
    report = await probation_report_service.get_probation_history(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        department_id=filters.get("department_id"),
        keyword=filters.get("keyword"),
        page=1,
        page_size=1,
    )
    return report.total


async def _export_probation(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return await probation_report_service.export_excel(
        session,
        start_date=filters["start_date"],
        end_date=filters["end_date"],
        department_id=filters.get("department_id"),
    )


async def _estimate_comprehensive_employee_list(session: AsyncSession, filters: dict[str, Any]) -> int:
    from app.models.employee import Employee
    result = await session.execute(select(func.count(Employee.id)).where(Employee.is_active == True))
    return result.scalar() or 0


async def _export_comprehensive_employee_list(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    return io.BytesIO(await employee_export_service.export_comprehensive_employee_list(session))


async def _estimate_salary_summary(session: AsyncSession, filters: dict[str, Any]) -> int:
    from app.models.employee import Employee
    result = await session.execute(select(func.count(Employee.id)).where(Employee.is_active == True))
    return result.scalar() or 0


async def _export_salary_summary(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    excel_bytes = await salary_export_service.export_salary_summary_excel(
        session,
        year=int(filters["year"]),
        month=int(filters["month"]),
        department_id=filters.get("department_id"),
    )
    return io.BytesIO(excel_bytes)


async def _estimate_leave_employee_summary(session: AsyncSession, filters: dict[str, Any]) -> int:
    from app.models.employee import Employee
    result = await session.execute(select(func.count(Employee.id)).where(Employee.is_active == True))
    return result.scalar() or 0


async def _export_leave_employee_summary(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    report = await leave_report_service.get_employee_summary(
        session,
        year=int(filters["year"]),
        employee_id=filters.get("employee_id"),
        department_id=filters.get("department_id"),
        leave_type_id=filters.get("leave_type_id"),
        keyword=filters.get("keyword"),
        page=1,
        page_size=10000,
    )
    return leave_report_service.export_employee_summary_xlsx(report)


async def _estimate_leave_department_summary(session: AsyncSession, filters: dict[str, Any]) -> int:
    return 100


async def _export_leave_department_summary(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    report = await leave_report_service.get_department_summary(
        session,
        year=int(filters["year"]),
        month_from=filters.get("month_from"),
        month_to=filters.get("month_to"),
        department_id=filters.get("department_id"),
        leave_type_id=filters.get("leave_type_id"),
    )
    return leave_report_service.export_department_summary_xlsx(report)


async def _estimate_leave_year_end(session: AsyncSession, filters: dict[str, Any]) -> int:
    from app.models.employee import Employee
    result = await session.execute(select(func.count(Employee.id)).where(Employee.is_active == True))
    return result.scalar() or 0


async def _export_leave_year_end(session: AsyncSession, filters: dict[str, Any]) -> io.BytesIO:
    report = await leave_report_service.get_year_end(
        session,
        year=int(filters["year"]),
        department_id=filters.get("department_id"),
        page=1,
        page_size=10000,
    )
    return leave_report_service.export_year_end_xlsx(report)


REPORT_HANDLERS: dict[str, ReportHandler] = {
    "dashboard": ReportHandler(("employees:view",), _estimate_dashboard_rows, _export_dashboard_xlsx),
    "hr-employee-list": ReportHandler(("employees:view",), _estimate_hr_employee_list, _export_hr_employee_list),
    "hr-movement": ReportHandler(("employees:view",), _estimate_hr_movement, _export_hr_movement),
    "hr-tenure": ReportHandler(("employees:view",), _estimate_hr_tenure, _export_hr_tenure),
    "hr-org-structure": ReportHandler(("employees:view",), _estimate_hr_org_structure, _export_hr_org_structure),
    "leaves": ReportHandler(("leaves:view",), _estimate_leaves, _export_leaves),
    "insurance": ReportHandler(("insurance:view",), _estimate_insurance, _export_insurance),
    "contracts": ReportHandler(("contracts:view",), _estimate_contracts, _export_contracts),
    "recruitment": ReportHandler(("recruitment:view",), _estimate_recruitment, _export_recruitment),
    "probation": ReportHandler(("employees:view",), _estimate_probation, _export_probation),
    "comprehensive-employee-list": ReportHandler(("employees:export",), _estimate_comprehensive_employee_list, _export_comprehensive_employee_list),
    "salary-summary": ReportHandler(("insurance:view",), _estimate_salary_summary, _export_salary_summary),
    "leave-employee-summary": ReportHandler(("leaves:view",), _estimate_leave_employee_summary, _export_leave_employee_summary),
    "leave-department-summary": ReportHandler(("leaves:view",), _estimate_leave_department_summary, _export_leave_department_summary),
    "leave-year-end": ReportHandler(("leaves:view",), _estimate_leave_year_end, _export_leave_year_end),
}


def _filename_with_extension(filename: Optional[str], report_type: str, fmt: ExportFormat) -> str:
    if filename:
        base = filename.rsplit(".", 1)[0]
    else:
        base = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    return f"{base}.{fmt}"


def _build_download_url(job: ExportJob) -> Optional[str]:
    if job.status != "done":
        return None
    return f"/api/v1/reports/export/{job.id}/download"


class ExportService:
    def __init__(self, session: AsyncSession):
        self.session = session

    async def create_export_job(self, request: ExportJobRequest, user: User) -> ExportJob:
        await self._assert_permission(user, request.report_type)

        job = ExportJob(
            user_id=user.id,
            report_type=request.report_type,
            format=request.format,
            filters=request.filters,
            filename=_filename_with_extension(request.filename, request.report_type, request.format),
            status="pending",
            expires_at=_utcnow() + timedelta(days=7),
        )
        self.session.add(job)
        await self.session.flush()

        job.row_count = await self._estimate_rows(request.report_type, request.filters)
        if job.row_count <= ASYNC_THRESHOLD_ROWS:
            await self._run_sync(job)
        else:
            from app.workers.export_tasks import run_export_task

            task = run_export_task.delay(str(job.id), request.model_dump(mode="json"))
            job.celery_task_id = task.id

        await self.session.commit()
        await self.session.refresh(job)
        return job

    async def get_history(self, user: User, *, page: int, page_size: int) -> ExportHistoryResponse:
        stmt = select(ExportJob).where(ExportJob.user_id == user.id).order_by(ExportJob.created_at.desc())
        total = (await self.session.execute(select(func.count()).select_from(stmt.subquery()))).scalar_one()
        rows = (await self.session.execute(stmt.offset((page - 1) * page_size).limit(page_size))).scalars().all()
        return ExportHistoryResponse(
            items=[self._to_response(row) for row in rows],
            total=total,
            page=page,
            page_size=page_size,
        )

    async def list_templates(self, user: User) -> list[ReportTemplateResponse]:
        rows = (
            await self.session.execute(
                select(ReportTemplate)
                .where(ReportTemplate.user_id == user.id)
                .order_by(ReportTemplate.is_default.desc(), ReportTemplate.updated_at.desc(), ReportTemplate.id.desc())
            )
        ).scalars().all()
        return [self._to_template_response(row) for row in rows]

    async def create_template(self, payload: ReportTemplateCreate, user: User) -> ReportTemplateResponse:
        await self._assert_permission(user, payload.report_type)
        if payload.is_default:
            await self._clear_default_template(user.id, payload.report_type)

        template = ReportTemplate(
            user_id=user.id,
            name=payload.name.strip(),
            description=payload.description,
            report_type=payload.report_type,
            format=payload.format,
            filters=payload.filters,
            is_default=payload.is_default,
            created_at=_utcnow(),
            updated_at=_utcnow(),
        )
        self.session.add(template)
        await self.session.commit()
        await self.session.refresh(template)
        return self._to_template_response(template)

    async def update_template(self, template_id: int, payload: ReportTemplateUpdate, user: User) -> ReportTemplateResponse:
        template = await self._get_template(template_id, user)
        if payload.name is not None:
            template.name = payload.name.strip()
        if payload.description is not None:
            template.description = payload.description
        if payload.filters is not None:
            template.filters = payload.filters
        if payload.format is not None:
            template.format = payload.format
        if payload.is_default is not None:
            if payload.is_default:
                await self._clear_default_template(user.id, template.report_type, exclude_id=template.id)
            template.is_default = payload.is_default
        template.updated_at = _utcnow()
        self.session.add(template)
        await self.session.commit()
        await self.session.refresh(template)
        return self._to_template_response(template)

    async def delete_template(self, template_id: int, user: User) -> None:
        template = await self._get_template(template_id, user)
        await self.session.delete(template)
        await self.session.commit()

    async def get_status(self, job_id: uuid.UUID, user: User) -> ExportJobStatusResponse:
        job = await self._get_job(job_id, user)
        return ExportJobStatusResponse(
            id=job.id,
            status=job.status,
            status_label=EXPORT_STATUS_LABELS.get(job.status, job.status),
            status_order=EXPORT_STATUS_ORDER.get(job.status, 999),
            status_severity=EXPORT_STATUS_SEVERITIES.get(job.status, "secondary"),
            error_message=job.error_message,
            download_url=_build_download_url(job),
        )

    async def delete_job(self, job_id: uuid.UUID, user: User) -> None:
        job = await self._get_job(job_id, user)
        if job.file_path:
            Path(job.file_path).unlink(missing_ok=True)
        await self.session.delete(job)
        await self.session.commit()

    async def get_download_job(self, job_id: uuid.UUID, user: User) -> ExportJob:
        job = await self._get_job(job_id, user)
        if job.status != "done" or not job.file_path:
            raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="Export job chưa sẵn sàng để tải")
        if not Path(job.file_path).exists():
            raise HTTPException(status_code=status.HTTP_410_GONE, detail="File export đã bị xóa hoặc hết hạn")
        return job

    async def _get_job(self, job_id: uuid.UUID, user: User) -> ExportJob:
        job = (await self.session.execute(
            select(ExportJob).where(ExportJob.id == job_id, ExportJob.user_id == user.id)
        )).scalar_one_or_none()
        if not job:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy export job")
        return job

    async def _get_template(self, template_id: int, user: User) -> ReportTemplate:
        template = (
            await self.session.execute(
                select(ReportTemplate).where(ReportTemplate.id == template_id, ReportTemplate.user_id == user.id)
            )
        ).scalar_one_or_none()
        if not template:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Không tìm thấy mẫu báo cáo")
        return template

    async def _clear_default_template(self, user_id: int, report_type: str, exclude_id: int | None = None) -> None:
        stmt = select(ReportTemplate).where(
            ReportTemplate.user_id == user_id,
            ReportTemplate.report_type == report_type,
            ReportTemplate.is_default == True,  # noqa: E712
        )
        if exclude_id is not None:
            stmt = stmt.where(ReportTemplate.id != exclude_id)
        rows = (await self.session.execute(stmt)).scalars().all()
        for row in rows:
            row.is_default = False
            row.updated_at = _utcnow()
            self.session.add(row)

    async def _assert_permission(self, user: User, report_type: str) -> None:
        handler = REPORT_HANDLERS.get(report_type)
        if handler is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Loại báo cáo chưa được hỗ trợ")
        if user.is_superuser:
            return
        permissions = await auth_service.get_user_permissions(self.session, user.id)
        if not any(code in permissions for code in handler.permissions):
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Không có quyền export báo cáo này")

    async def _estimate_rows(self, report_type: str, filters: dict[str, Any]) -> int:
        handler = REPORT_HANDLERS[report_type]
        normalized_filters = _normalize_filters(filters)
        try:
            return max(int(await handler.estimate_rows(self.session, normalized_filters)), 0)
        except Exception as exc:
            logger.warning("export_size_estimate_error", error=str(exc))
            return 0

    async def _run_sync(self, job: ExportJob) -> None:
        job.status = "processing"
        job.started_at = _utcnow()
        try:
            file_bytes, row_count = await self._build_file(job.report_type, job.format, job.filters)
            job.file_path = self._save_temp(job.id, job.format, file_bytes)
            job.file_size_bytes = len(file_bytes)
            job.row_count = row_count
            job.status = "done"
            job.error_message = None
        except Exception as exc:
            job.status = "failed"
            job.error_message = str(exc)
        finally:
            job.completed_at = _utcnow()

    async def _build_file(self, report_type: str, fmt: ExportFormat, filters: dict[str, Any]) -> tuple[bytes, int]:
        handler = REPORT_HANDLERS[report_type]
        normalized_filters = _normalize_filters(filters)
        workbook = await handler.export_xlsx(self.session, normalized_filters)
        xlsx_bytes = workbook.getvalue()
        row_count = await self._estimate_rows(report_type, normalized_filters)
        if fmt == "xlsx":
            return xlsx_bytes, row_count
        return _workbook_pdf(report_type, normalized_filters, row_count, xlsx_bytes), row_count

    @staticmethod
    def _save_temp(job_id: uuid.UUID, fmt: str, payload: bytes) -> str:
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        path = EXPORT_DIR / f"{job_id}.{fmt}"
        path.write_bytes(payload)
        return str(path)

    @staticmethod
    def _to_response(job: ExportJob) -> ExportJobResponse:
        return ExportJobResponse(
            id=job.id,
            report_type=job.report_type,
            report_type_label=REPORT_TYPE_LABELS.get(job.report_type, job.report_type),
            report_type_order=REPORT_TYPE_ORDER.get(job.report_type, 999),
            format=job.format,
            format_label=EXPORT_FORMAT_LABELS.get(job.format, job.format.upper()),
            format_order=EXPORT_FORMAT_ORDER.get(job.format, 999),
            status=job.status,
            status_label=EXPORT_STATUS_LABELS.get(job.status, job.status),
            status_order=EXPORT_STATUS_ORDER.get(job.status, 999),
            status_severity=EXPORT_STATUS_SEVERITIES.get(job.status, "secondary"),
            filename=job.filename,
            file_size_bytes=job.file_size_bytes,
            row_count=job.row_count,
            error_message=job.error_message,
            created_at=job.created_at,
            started_at=job.started_at,
            completed_at=job.completed_at,
            expires_at=job.expires_at,
            download_url=_build_download_url(job),
        )

    @staticmethod
    def _to_template_response(template: ReportTemplate) -> ReportTemplateResponse:
        return ReportTemplateResponse(
            id=template.id,
            name=template.name,
            description=template.description,
            report_type=template.report_type,
            report_type_label=REPORT_TYPE_LABELS.get(template.report_type, template.report_type),
            report_type_order=REPORT_TYPE_ORDER.get(template.report_type, 999),
            format=template.format,
            format_label=EXPORT_FORMAT_LABELS.get(template.format, template.format.upper()),
            format_order=EXPORT_FORMAT_ORDER.get(template.format, 999),
            filters=template.filters,
            is_default=template.is_default,
            created_at=template.created_at,
            updated_at=template.updated_at,
        )


async def cleanup_expired_exports_async() -> dict[str, int]:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine

    from app.core.config import settings

    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    SessionLocal = async_sessionmaker(engine, expire_on_commit=False)
    deleted_files = 0
    deleted_jobs = 0
    try:
        async with SessionLocal() as session:
            rows = (
                await session.execute(
                    select(ExportJob).where(ExportJob.expires_at < _utcnow())
                )
            ).scalars().all()
            for row in rows:
                if row.file_path and Path(row.file_path).exists():
                    Path(row.file_path).unlink(missing_ok=True)
                    deleted_files += 1
                await session.delete(row)
                deleted_jobs += 1
            await session.commit()
    finally:
        await engine.dispose()
    return {"deleted_jobs": deleted_jobs, "deleted_files": deleted_files}


async def run_export_job_by_id(job_id: uuid.UUID, request_dict: dict[str, Any]) -> None:
    from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
    from celery.exceptions import SoftTimeLimitExceeded
    from app.core.config import settings

    engine = create_async_engine(settings.DATABASE_URL, connect_args={"ssl": False})
    SessionLocal = async_sessionmaker(engine, expire_on_commit=False)
    try:
        async with SessionLocal() as session:
            job = (await session.execute(select(ExportJob).where(ExportJob.id == job_id))).scalar_one_or_none()
            if not job:
                raise ValueError(f"ExportJob with id {job_id} not found")
            
            try:
                service = ExportService(session)
                request = ExportJobRequest.model_validate(request_dict)
                job.report_type = request.report_type
                job.format = request.format
                job.filters = request.filters
                job.filename = _filename_with_extension(request.filename or job.filename, request.report_type, request.format)
                await service._run_sync(job)
                await session.commit()
            except SoftTimeLimitExceeded as exc:
                job.status = "failed"
                job.error_message = "Xuất báo cáo bị quá thời hạn (Soft Time Limit Exceeded)"
                job.completed_at = _utcnow()
                await session.commit()
                raise exc
            except Exception as exc:
                job.status = "failed"
                job.error_message = str(exc)
                job.completed_at = _utcnow()
                await session.commit()
                raise exc
    finally:
        await engine.dispose()
