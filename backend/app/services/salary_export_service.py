"""Export bảng tổng hợp lương BHXH ra Excel (Plan 7.3 Slice 2)."""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.schemas.salary import SalarySummaryRates, SalarySummaryRow, SalarySummaryTotals
from app.services import salary_service

# ── Styles ────────────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")
_HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")
_ALT_FILL = PatternFill("solid", fgColor="F2F2F2")

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)

_THIN = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_NUM_FMT = "#,##0"
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


def _cell(ws, row: int, col: int, value=None, *, font=None, fill=None,
          alignment=None, number_format=None, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if alignment:
        c.alignment = alignment
    if number_format:
        c.number_format = number_format
    if border:
        c.border = border
    return c


def _apply_border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = _THIN_BORDER


# ── Header writer ─────────────────────────────────────────────────────────────

def _write_header(ws, year: int, month: int, company_name: str,
                  rates: SalarySummaryRates) -> None:
    # Hàng 1: Tên công ty
    ws.merge_cells("A1:N1")
    _cell(ws, 1, 1, company_name,
          font=Font(bold=True, size=14),
          alignment=_CENTER)

    # Hàng 2: Tiêu đề bảng
    ws.merge_cells("A2:N2")
    _cell(ws, 2, 1, f"BẢNG TỔNG HỢP LƯƠNG BHXH THÁNG {month:02d}/{year}",
          font=Font(bold=True, size=12),
          alignment=_CENTER)

    # Hàng 3: Thông tin tỷ lệ
    ws.merge_cells("A3:N3")
    def _fmt_rate(r: Decimal) -> str:
        return f"{float(r):g}"

    rate_text = (
        f"Tỷ lệ: NLĐ: BHXH {_fmt_rate(rates.bhxh_employee_rate)}% + "
        f"BHYT {_fmt_rate(rates.bhyt_employee_rate)}% + BHTN {_fmt_rate(rates.bhtn_employee_rate)}%   |   "
        f"NSDLĐ: BHXH {_fmt_rate(rates.bhxh_employer_rate)}% + "
        f"BHYT {_fmt_rate(rates.bhyt_employer_rate)}% + BHTN {_fmt_rate(rates.bhtn_employer_rate)}%"
    )
    _cell(ws, 3, 1, rate_text, font=Font(italic=True, size=9), alignment=_CENTER)

    # Hàng 4: trống
    ws.row_dimensions[4].height = 6

    # Hàng 5-6: Header nhóm cột
    # Merge cells STT, Mã NV, Họ tên, Phòng ban, Lương BHXH, Tổng cộng (span 2 hàng)
    for col_range in ("A5:A6", "B5:B6", "C5:C6", "D5:D6", "E5:E6", "N5:N6"):
        ws.merge_cells(col_range)

    ws.merge_cells("F5:I5")
    ws.merge_cells("J5:M5")

    headers_row5 = [
        (1, "STT"), (2, "Mã NV"), (3, "Họ và tên"), (4, "Phòng ban"),
        (5, "Mức lương BHXH"),
        (6, "NGƯỜI LAO ĐỘNG ĐÓNG"),
        (10, "NGƯỜI SỬ DỤNG LAO ĐỘNG ĐÓNG"),
        (14, "Tổng cộng"),
    ]
    for col, text in headers_row5:
        _cell(ws, 5, col, text,
              font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER, border=_THIN_BORDER)

    bhxh_emp_pct = f"BHXH ({_fmt_rate(rates.bhxh_employee_rate)}%)"
    bhyt_emp_pct = f"BHYT ({_fmt_rate(rates.bhyt_employee_rate)}%)"
    bhtn_emp_pct = f"BHTN ({_fmt_rate(rates.bhtn_employee_rate)}%)"
    bhxh_er_pct = f"BHXH ({_fmt_rate(rates.bhxh_employer_rate)}%)"
    bhyt_er_pct = f"BHYT ({_fmt_rate(rates.bhyt_employer_rate)}%)"
    bhtn_er_pct = f"BHTN ({_fmt_rate(rates.bhtn_employer_rate)}%)"

    headers_row6 = [
        (6, bhxh_emp_pct), (7, bhyt_emp_pct), (8, bhtn_emp_pct), (9, "Tổng NLĐ"),
        (10, bhxh_er_pct), (11, bhyt_er_pct), (12, bhtn_er_pct), (13, "Tổng NSDLĐ"),
    ]
    for col, text in headers_row6:
        _cell(ws, 6, col, text,
              font=_HEADER_FONT, fill=_HEADER_FILL, alignment=_CENTER, border=_THIN_BORDER)

    # Fill màu cho E5:E6 và N5:N6 (highlight)
    for r in (5, 6):
        for c in (5, 14):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="2E75B6")

    ws.row_dimensions[5].height = 28
    ws.row_dimensions[6].height = 28


# ── Data rows writer ──────────────────────────────────────────────────────────

def _write_data_rows(ws, rows: list[SalarySummaryRow], start_row: int = 7) -> int:
    for i, row in enumerate(rows):
        r = start_row + i
        fill = _ALT_FILL if i % 2 == 0 else None

        values = [
            row.stt, row.employee_code, row.full_name, row.department_name or "—",
            int(row.basis_amount),
            int(row.bhxh_employee), int(row.bhyt_employee),
            int(row.bhtn_employee), int(row.total_employee),
            int(row.bhxh_employer), int(row.bhyt_employer),
            int(row.bhtn_employer), int(row.total_employer),
            int(row.grand_total),
        ]
        alignments = [_CENTER, _LEFT, _LEFT, _LEFT] + [_RIGHT] * 10
        for col, (val, align) in enumerate(zip(values, alignments), start=1):
            kw = dict(font=_NORMAL_FONT, alignment=align, border=_THIN_BORDER)
            if fill:
                kw["fill"] = fill
            if col >= 5:
                kw["number_format"] = _NUM_FMT
            # Bold cho cột Tổng NLĐ (9), Tổng NSDLĐ (13), Tổng cộng (14)
            if col in (9, 13, 14):
                kw["font"] = _BOLD_FONT
            if col in (5, 14):
                kw["fill"] = _HIGHLIGHT_FILL if not fill else _HIGHLIGHT_FILL
            _cell(ws, r, col, val, **kw)

        ws.row_dimensions[r].height = 18
    return start_row + len(rows)


# ── Totals row writer ─────────────────────────────────────────────────────────

def _write_totals_row(ws, totals: SalarySummaryTotals, row_idx: int) -> None:
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    _cell(ws, row_idx, 1,
          f"TỔNG CỘNG ({totals.total_employees} nhân viên)",
          font=_BOLD_FONT, fill=_TOTAL_FILL, alignment=_CENTER, border=_THIN_BORDER)

    total_values = [
        int(totals.sum_basis),
        int(totals.sum_bhxh_employee), int(totals.sum_bhyt_employee),
        int(totals.sum_bhtn_employee), int(totals.sum_total_employee),
        int(totals.sum_bhxh_employer), int(totals.sum_bhyt_employer),
        int(totals.sum_bhtn_employer), int(totals.sum_total_employer),
        int(totals.sum_grand_total),
    ]
    for i, val in enumerate(total_values):
        col = 5 + i
        _cell(ws, row_idx, col, val,
              font=_BOLD_FONT, fill=_TOTAL_FILL,
              alignment=_RIGHT, number_format=_NUM_FMT, border=_THIN_BORDER)

    ws.row_dimensions[row_idx].height = 20


# ── Signature row ─────────────────────────────────────────────────────────────

def _write_signature_row(ws, row_idx: int) -> None:
    ws.merge_cells(f"A{row_idx}:D{row_idx}")
    _cell(ws, row_idx, 1, "Người lập bảng: _______________",
          font=_NORMAL_FONT, alignment=_CENTER)

    ws.merge_cells(f"F{row_idx}:I{row_idx}")
    _cell(ws, row_idx, 6, "Kế toán trưởng: _______________",
          font=_NORMAL_FONT, alignment=_CENTER)

    ws.merge_cells(f"K{row_idx}:N{row_idx}")
    _cell(ws, row_idx, 11, "Giám đốc: _______________",
          font=_NORMAL_FONT, alignment=_CENTER)

    ws.row_dimensions[row_idx].height = 28


# ── Column widths ─────────────────────────────────────────────────────────────

def _set_column_widths(ws) -> None:
    widths = [6, 10, 25, 18, 15, 14, 14, 14, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Main export function ──────────────────────────────────────────────────────

async def export_salary_summary_excel(
    session: AsyncSession,
    *,
    year: int,
    month: int,
    department_id: Optional[int] = None,
    company_name: str = "CÔNG TY TNHH HỒNG HÀ",
) -> bytes:
    rows, rates = await salary_service.get_salary_summary_all(
        session, year=year, month=month, department_id=department_id
    )

    zero = Decimal("0")
    totals = SalarySummaryTotals(
        total_employees=len(rows),
        sum_basis=sum((x.basis_amount for x in rows), zero),
        sum_bhxh_employee=sum((x.bhxh_employee for x in rows), zero),
        sum_bhyt_employee=sum((x.bhyt_employee for x in rows), zero),
        sum_bhtn_employee=sum((x.bhtn_employee for x in rows), zero),
        sum_total_employee=sum((x.total_employee for x in rows), zero),
        sum_bhxh_employer=sum((x.bhxh_employer for x in rows), zero),
        sum_bhyt_employer=sum((x.bhyt_employer for x in rows), zero),
        sum_bhtn_employer=sum((x.bhtn_employer for x in rows), zero),
        sum_total_employer=sum((x.total_employer for x in rows), zero),
        sum_grand_total=sum((x.grand_total for x in rows), zero),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Tổng hợp lương BHXH"

    _write_header(ws, year, month, company_name, rates)
    next_row = _write_data_rows(ws, rows, start_row=7)
    _write_totals_row(ws, totals, next_row)
    _write_signature_row(ws, next_row + 2)
    _set_column_widths(ws)

    ws.freeze_panes = "A7"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
