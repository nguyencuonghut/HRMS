"""Recruitment analytics service — Plan 13.8."""
from __future__ import annotations

import io
from datetime import date, datetime, timezone
from typing import Optional

from sqlalchemy import case, extract, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org import Department
from app.models.recruitment import (
    CandidateApplication,
    HiringDecision,
    JobPosting,
    JobRequisition,
    Offer,
    RecruitmentBudgetItem,
    RecruitmentChannel,
)
from app.schemas.recruitment import (
    ChannelEffectivenessItem,
    DepartmentRecruitmentStat,
    FunnelReport,
    FunnelStage,
    MonthlyTimeMetric,
    RecruitmentSummaryReport,
    TimeMetricsReport,
)

# Stage display order and labels for funnel
_FUNNEL_ORDER = ["new", "screening", "test", "interview", "offer", "hired", "rejected", "withdrawn"]
_STAGE_LABELS = {
    "new": "Mới nộp",
    "screening": "Sàng lọc hồ sơ",
    "test": "Kiểm tra",
    "interview": "Phỏng vấn",
    "offer": "Đề nghị làm việc",
    "hired": "Đã tuyển",
    "rejected": "Đã loại",
    "withdrawn": "Đã rút đơn",
}


def _utcnow() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


async def get_summary(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
) -> RecruitmentSummaryReport:
    # Base JR filter
    jr_q = select(JobRequisition.id).where(
        func.date(JobRequisition.created_at) >= start_date,
        func.date(JobRequisition.created_at) <= end_date,
    )
    if department_id:
        jr_q = jr_q.where(JobRequisition.department_id == department_id)

    jr_ids_res = await session.execute(jr_q)
    jr_ids = [row[0] for row in jr_ids_res.fetchall()]

    total_jr = len(jr_ids)

    if not jr_ids:
        return RecruitmentSummaryReport(
            period_start=start_date,
            period_end=end_date,
            total_jr=0,
            total_applications=0,
            total_screened=0,
            total_interviewed=0,
            total_offered=0,
            total_hired=0,
        )

    # Application counts
    def _app_count(stages: list[str]):
        return select(func.count()).select_from(CandidateApplication).where(
            CandidateApplication.job_requisition_id.in_(jr_ids),
            CandidateApplication.current_stage.in_(stages),
        )

    total_apps_res = await session.execute(
        select(func.count()).select_from(CandidateApplication).where(
            CandidateApplication.job_requisition_id.in_(jr_ids)
        )
    )
    total_applications = total_apps_res.scalar() or 0

    screened_res = await session.execute(_app_count(["screening"]))
    total_screened = screened_res.scalar() or 0

    interviewed_res = await session.execute(_app_count(["interview"]))
    total_interviewed = interviewed_res.scalar() or 0

    # Offers for these JRs
    offered_res = await session.execute(
        select(func.count()).select_from(Offer).where(
            Offer.job_requisition_id.in_(jr_ids),
            Offer.status.in_(["sent", "waiting", "accepted", "rejected", "negotiating", "expired"]),
        )
    )
    total_offered = offered_res.scalar() or 0

    # Hired = converted hiring decisions for these JRs
    hired_res = await session.execute(
        select(func.count()).select_from(HiringDecision).where(
            HiringDecision.job_requisition_id.in_(jr_ids),
            HiringDecision.status == "converted",
        )
    )
    total_hired = hired_res.scalar() or 0

    # avg_time_to_hire: apply_date → hiring_decision.created_at (converted)
    tth_res = await session.execute(
        select(
            func.avg(
                func.extract("epoch", HiringDecision.created_at)
                - func.extract("epoch", CandidateApplication.applied_date)
            )
            / 86400.0
        )
        .select_from(HiringDecision)
        .join(Offer, Offer.id == HiringDecision.offer_id)
        .join(CandidateApplication, CandidateApplication.id == Offer.application_id)
        .where(
            HiringDecision.job_requisition_id.in_(jr_ids),
            HiringDecision.status == "converted",
        )
    )
    avg_tth_raw = tth_res.scalar()
    avg_time_to_hire = round(float(avg_tth_raw), 1) if avg_tth_raw is not None else None

    # avg_time_to_fill: job_posting opened_at → jr.updated_at (completed)
    ttf_res = await session.execute(
        select(
            func.avg(
                func.extract("epoch", JobRequisition.updated_at)
                - func.extract("epoch", JobPosting.opened_at)
            )
            / 86400.0
        )
        .select_from(JobRequisition)
        .join(JobPosting, JobPosting.job_requisition_id == JobRequisition.id)
        .where(
            JobRequisition.id.in_(jr_ids),
            JobRequisition.status == "completed",
            JobPosting.opened_at.isnot(None),
        )
    )
    avg_ttf_raw = ttf_res.scalar()
    avg_time_to_fill = round(float(avg_ttf_raw), 1) if avg_ttf_raw is not None else None

    # Offer acceptance rate
    oar_res = await session.execute(
        select(
            func.count().filter(Offer.status == "accepted"),
            func.count().filter(Offer.status.in_(["sent", "waiting", "accepted", "rejected", "expired"])),
        )
        .select_from(Offer)
        .where(Offer.job_requisition_id.in_(jr_ids))
    )
    oar_row = oar_res.first()
    accepted_count, sent_count = (oar_row[0] or 0, oar_row[1] or 0)
    offer_acceptance_rate = (
        round(accepted_count * 100.0 / sent_count, 1) if sent_count > 0 else None
    )

    # Cost per hire
    cost_res = await session.execute(
        select(func.sum(RecruitmentBudgetItem.actual_amount))
        .select_from(RecruitmentBudgetItem)
        .where(RecruitmentBudgetItem.job_requisition_id.in_(jr_ids))
    )
    total_cost = cost_res.scalar()
    cost_per_hire = (
        round(float(total_cost) / total_hired, 0) if total_cost and total_hired > 0 else None
    )

    return RecruitmentSummaryReport(
        period_start=start_date,
        period_end=end_date,
        total_jr=total_jr,
        total_applications=total_applications,
        total_screened=total_screened,
        total_interviewed=total_interviewed,
        total_offered=total_offered,
        total_hired=total_hired,
        avg_time_to_hire=avg_time_to_hire,
        avg_time_to_fill=avg_time_to_fill,
        offer_acceptance_rate=offer_acceptance_rate,
        cost_per_hire=cost_per_hire,
        probation_pass_rate=None,  # Module 14 not yet available
    )


async def get_funnel(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
    job_requisition_id: Optional[int] = None,
) -> FunnelReport:
    q = select(
        CandidateApplication.current_stage,
        func.count().label("cnt"),
    ).select_from(CandidateApplication)

    if job_requisition_id:
        q = q.where(CandidateApplication.job_requisition_id == job_requisition_id)
    else:
        jr_q = select(JobRequisition.id).where(
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        if department_id:
            jr_q = jr_q.where(JobRequisition.department_id == department_id)
        q = q.where(CandidateApplication.job_requisition_id.in_(jr_q))

    q = q.group_by(CandidateApplication.current_stage)
    res = await session.execute(q)
    counts: dict[str, int] = {row.current_stage: row.cnt for row in res.fetchall()}

    stages: list[FunnelStage] = []
    prev_count: Optional[int] = None

    for stage_key in _FUNNEL_ORDER:
        count = counts.get(stage_key, 0)
        conversion_rate: Optional[float] = None
        if prev_count and prev_count > 0 and stage_key not in ("rejected", "withdrawn"):
            conversion_rate = round(count * 100.0 / prev_count, 1)
        stages.append(
            FunnelStage(
                stage=stage_key,
                stage_label=_STAGE_LABELS.get(stage_key, stage_key),
                count=count,
                conversion_rate=conversion_rate,
            )
        )
        if stage_key not in ("rejected", "withdrawn"):
            prev_count = count if count > 0 else prev_count

    return FunnelReport(stages=stages)


async def get_channel_effectiveness(
    session: AsyncSession,
    start_date: date,
    end_date: date,
) -> list[ChannelEffectivenessItem]:
    # All channels
    channels_res = await session.execute(select(RecruitmentChannel).order_by(RecruitmentChannel.name))
    channels = channels_res.scalars().all()

    # Applications per channel
    apps_res = await session.execute(
        select(
            CandidateApplication.source_channel_id,
            func.count().label("total"),
        )
        .select_from(CandidateApplication)
        .join(JobRequisition, JobRequisition.id == CandidateApplication.job_requisition_id)
        .where(
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
            CandidateApplication.source_channel_id.isnot(None),
        )
        .group_by(CandidateApplication.source_channel_id)
    )
    apps_by_channel: dict[int, int] = {row.source_channel_id: row.total for row in apps_res.fetchall()}

    # Hired per channel (via hiring decisions → offers → applications)
    hired_res = await session.execute(
        select(
            CandidateApplication.source_channel_id,
            func.count().label("hired"),
        )
        .select_from(HiringDecision)
        .join(Offer, Offer.id == HiringDecision.offer_id)
        .join(CandidateApplication, CandidateApplication.id == Offer.application_id)
        .join(JobRequisition, JobRequisition.id == CandidateApplication.job_requisition_id)
        .where(
            HiringDecision.status == "converted",
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
            CandidateApplication.source_channel_id.isnot(None),
        )
        .group_by(CandidateApplication.source_channel_id)
    )
    hired_by_channel: dict[int, int] = {row.source_channel_id: row.hired for row in hired_res.fetchall()}

    result: list[ChannelEffectivenessItem] = []
    for ch in channels:
        total = apps_by_channel.get(ch.id, 0)
        hired = hired_by_channel.get(ch.id, 0)
        if total == 0 and hired == 0:
            continue  # Skip channels with no activity
        hire_rate = round(hired * 100.0 / total, 1) if total > 0 else 0.0
        result.append(
            ChannelEffectivenessItem(
                channel_id=ch.id,
                channel_name=ch.name,
                total_candidates=total,
                hired_count=hired,
                hire_rate=hire_rate,
                total_cost=None,
                cost_per_hire=None,
            )
        )

    result.sort(key=lambda x: (-x.hired_count, -x.total_candidates))
    return result


async def get_department_breakdown(
    session: AsyncSession,
    start_date: date,
    end_date: date,
) -> list[DepartmentRecruitmentStat]:
    # JR counts per department
    jr_res = await session.execute(
        select(
            JobRequisition.department_id,
            func.count().label("total_jr"),
            func.count().filter(
                JobRequisition.status.in_(["approved", "in_progress"])
            ).label("open_jr"),
        )
        .where(
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        .group_by(JobRequisition.department_id)
    )
    jr_by_dept: dict[int, dict] = {
        row.department_id: {"total": row.total_jr, "open": row.open_jr}
        for row in jr_res.fetchall()
    }

    if not jr_by_dept:
        return []

    dept_ids = list(jr_by_dept.keys())

    # Hired per department
    hired_res = await session.execute(
        select(
            HiringDecision.department_id,
            func.count().label("hired"),
        )
        .select_from(HiringDecision)
        .join(JobRequisition, JobRequisition.id == HiringDecision.job_requisition_id)
        .where(
            HiringDecision.status == "converted",
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        .group_by(HiringDecision.department_id)
    )
    hired_by_dept: dict[int, int] = {row.department_id: row.hired for row in hired_res.fetchall()}

    # avg_time_to_hire per department
    tth_res = await session.execute(
        select(
            HiringDecision.department_id,
            func.avg(
                func.extract("epoch", HiringDecision.created_at)
                - func.extract("epoch", CandidateApplication.applied_date)
            ) / 86400.0,
        )
        .select_from(HiringDecision)
        .join(Offer, Offer.id == HiringDecision.offer_id)
        .join(CandidateApplication, CandidateApplication.id == Offer.application_id)
        .join(JobRequisition, JobRequisition.id == HiringDecision.job_requisition_id)
        .where(
            HiringDecision.status == "converted",
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        .group_by(HiringDecision.department_id)
    )
    tth_by_dept: dict[int, float] = {}
    for row in tth_res.fetchall():
        if row[1] is not None:
            tth_by_dept[row[0]] = round(float(row[1]), 1)

    # Budget per department
    budget_res = await session.execute(
        select(
            JobRequisition.department_id,
            func.sum(RecruitmentBudgetItem.actual_amount).label("budget"),
        )
        .select_from(RecruitmentBudgetItem)
        .join(JobRequisition, JobRequisition.id == RecruitmentBudgetItem.job_requisition_id)
        .where(
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        .group_by(JobRequisition.department_id)
    )
    budget_by_dept: dict[int, float] = {}
    for row in budget_res.fetchall():
        if row[1] is not None:
            budget_by_dept[row[0]] = float(row[1])

    # OAR per department
    oar_res = await session.execute(
        select(
            Offer.department_id,
            func.count().filter(Offer.status == "accepted").label("accepted"),
            func.count().filter(
                Offer.status.in_(["sent", "waiting", "accepted", "rejected", "expired"])
            ).label("sent"),
        )
        .select_from(Offer)
        .join(JobRequisition, JobRequisition.id == Offer.job_requisition_id)
        .where(
            func.date(JobRequisition.created_at) >= start_date,
            func.date(JobRequisition.created_at) <= end_date,
        )
        .group_by(Offer.department_id)
    )
    oar_by_dept: dict[int, float] = {}
    for row in oar_res.fetchall():
        if row.sent > 0:
            oar_by_dept[row.department_id] = round(row.accepted * 100.0 / row.sent, 1)

    # Load department names
    dept_res = await session.execute(
        select(Department.id, Department.name).where(Department.id.in_(dept_ids))
    )
    dept_names: dict[int, str] = {row.id: row.name for row in dept_res.fetchall()}

    result: list[DepartmentRecruitmentStat] = []
    for dept_id in dept_ids:
        hired = hired_by_dept.get(dept_id, 0)
        budget = budget_by_dept.get(dept_id)
        result.append(
            DepartmentRecruitmentStat(
                department_id=dept_id,
                department_name=dept_names.get(dept_id, f"Phòng ban #{dept_id}"),
                total_jr=jr_by_dept[dept_id]["total"],
                open_jr=jr_by_dept[dept_id]["open"],
                hired_count=hired,
                avg_time_to_hire=tth_by_dept.get(dept_id),
                offer_acceptance_rate=oar_by_dept.get(dept_id),
                budget_used=budget,
                cost_per_hire=round(budget / hired, 0) if budget and hired > 0 else None,
            )
        )

    result.sort(key=lambda x: (-x.hired_count, -x.total_jr))
    return result


async def get_time_metrics_monthly(
    session: AsyncSession,
    year: int,
    department_id: Optional[int] = None,
) -> TimeMetricsReport:
    # Base JR filter
    jr_q_base = select(JobRequisition.id).where(
        extract("year", JobRequisition.created_at) == year
    )
    if department_id:
        jr_q_base = jr_q_base.where(JobRequisition.department_id == department_id)
    jr_ids_res = await session.execute(jr_q_base)
    jr_ids = [row[0] for row in jr_ids_res.fetchall()]

    # Applications count per month
    apps_res = await session.execute(
        select(
            extract("month", CandidateApplication.created_at).label("month"),
            func.count().label("cnt"),
        )
        .where(CandidateApplication.job_requisition_id.in_(jr_ids))
        .group_by(extract("month", CandidateApplication.created_at))
    )
    apps_by_month: dict[int, int] = {int(row.month): row.cnt for row in apps_res.fetchall()}

    # Hired count + avg_time_to_hire per month
    hired_res = await session.execute(
        select(
            extract("month", HiringDecision.created_at).label("month"),
            func.count().label("hired"),
            func.avg(
                func.extract("epoch", HiringDecision.created_at)
                - func.extract("epoch", CandidateApplication.applied_date)
            ) / 86400.0,
        )
        .select_from(HiringDecision)
        .join(Offer, Offer.id == HiringDecision.offer_id)
        .join(CandidateApplication, CandidateApplication.id == Offer.application_id)
        .where(
            HiringDecision.job_requisition_id.in_(jr_ids),
            HiringDecision.status == "converted",
            extract("year", HiringDecision.created_at) == year,
        )
        .group_by(extract("month", HiringDecision.created_at))
    )
    hired_data: dict[int, tuple[int, Optional[float]]] = {}
    for row in hired_res.fetchall():
        m = int(row.month)
        tth = round(float(row[2]), 1) if row[2] is not None else None
        hired_data[m] = (row.hired, tth)

    # avg_time_to_fill per month (by jr.updated_at month)
    ttf_res = await session.execute(
        select(
            extract("month", JobRequisition.updated_at).label("month"),
            func.avg(
                func.extract("epoch", JobRequisition.updated_at)
                - func.extract("epoch", JobPosting.opened_at)
            ) / 86400.0,
        )
        .select_from(JobRequisition)
        .join(JobPosting, JobPosting.job_requisition_id == JobRequisition.id)
        .where(
            JobRequisition.id.in_(jr_ids),
            JobRequisition.status == "completed",
            JobPosting.opened_at.isnot(None),
            extract("year", JobRequisition.updated_at) == year,
        )
        .group_by(extract("month", JobRequisition.updated_at))
    )
    ttf_by_month: dict[int, float] = {}
    for row in ttf_res.fetchall():
        if row[1] is not None:
            ttf_by_month[int(row.month)] = round(float(row[1]), 1)

    monthly: list[MonthlyTimeMetric] = []
    for m in range(1, 13):
        hired_count, avg_tth = hired_data.get(m, (0, None))
        monthly.append(
            MonthlyTimeMetric(
                month=m,
                year=year,
                avg_time_to_hire=avg_tth,
                avg_time_to_fill=ttf_by_month.get(m),
                hired_count=hired_count,
                applications_count=apps_by_month.get(m, 0),
            )
        )

    return TimeMetricsReport(year=year, monthly=monthly)


def _fmt_vnd(val: Optional[float]) -> str:
    if val is None:
        return "—"
    return f"{int(val):,} ₫"


def _fmt_pct(val: Optional[float]) -> str:
    if val is None:
        return "—"
    return f"{val}%"


def _fmt_days(val: Optional[float]) -> str:
    if val is None:
        return "—"
    return f"{val} ngày"


async def export_excel(
    session: AsyncSession,
    start_date: date,
    end_date: date,
    department_id: Optional[int] = None,
) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    summary = await get_summary(session, start_date, end_date, department_id)
    funnel = await get_funnel(session, start_date, end_date, department_id)
    channels = await get_channel_effectiveness(session, start_date, end_date)
    depts = await get_department_breakdown(session, start_date, end_date)
    year = start_date.year
    time_report = await get_time_metrics_monthly(session, year, department_id)

    HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="EBF5FB")
    period_label = f"{start_date.strftime('%d/%m/%Y')} – {end_date.strftime('%d/%m/%Y')}"

    wb = Workbook()

    def _setup_ws(ws, title: str, headers: list[str], rows: list[list]) -> None:
        ws.title = title
        n = len(headers)
        # Title row
        ws.merge_cells(f"A1:{get_column_letter(n)}1")
        c = ws["A1"]
        c.value = f"BÁO CÁO TUYỂN DỤNG — {title.upper()}"
        c.font = Font(bold=True, size=13, color="1B4F72")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        # Period row
        ws.merge_cells(f"A2:{get_column_letter(n)}2")
        c2 = ws["A2"]
        c2.value = f"Kỳ: {period_label}"
        c2.font = Font(italic=True, size=10, color="555555")
        c2.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16

        # Headers
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[3].height = 28

        # Data
        for ri, row in enumerate(rows, 4):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                if ri % 2 == 0:
                    cell.fill = ALT_FILL
                cell.alignment = Alignment(vertical="center")

        # Column widths
        for ci in range(1, n + 1):
            max_len = len(str(headers[ci - 1]))
            for ri in range(4, len(rows) + 4):
                v = ws.cell(row=ri, column=ci).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 40)

    # Sheet 1 — Tổng quan
    ws1 = wb.active
    ws1.title = "Tổng quan"
    ws1["A1"] = "BÁO CÁO TUYỂN DỤNG — TỔNG QUAN"
    ws1["A1"].font = Font(bold=True, size=14, color="1B4F72")
    ws1["A1"].alignment = Alignment(horizontal="center")
    ws1.merge_cells("A1:B1")
    ws1.row_dimensions[1].height = 24

    ws1["A2"] = f"Kỳ: {period_label}"
    ws1["A2"].font = Font(italic=True, size=10, color="555555")
    ws1.merge_cells("A2:B2")
    ws1.row_dimensions[2].height = 16

    kpi_rows = [
        ("Tổng JR", summary.total_jr),
        ("Tổng lượt ứng tuyển", summary.total_applications),
        ("Đang sàng lọc hồ sơ", summary.total_screened),
        ("Đang phỏng vấn", summary.total_interviewed),
        ("Đã gửi offer", summary.total_offered),
        ("Đã tuyển dụng", summary.total_hired),
        ("Thời gian tuyển TB (ngày)", _fmt_days(summary.avg_time_to_hire)),
        ("Thời gian lấp chỗ TB (ngày)", _fmt_days(summary.avg_time_to_fill)),
        ("Tỷ lệ chấp nhận offer", _fmt_pct(summary.offer_acceptance_rate)),
        ("Chi phí mỗi tuyển dụng", _fmt_vnd(summary.cost_per_hire)),
    ]

    ws1["A3"] = "Chỉ số"
    ws1["B3"] = "Giá trị"
    for c in ["A3", "B3"]:
        ws1[c].fill = HEADER_FILL
        ws1[c].font = HEADER_FONT
        ws1[c].alignment = Alignment(horizontal="center")
    ws1.row_dimensions[3].height = 24

    for ri, (label, val) in enumerate(kpi_rows, 4):
        ws1.cell(row=ri, column=1, value=label).alignment = Alignment(vertical="center")
        ws1.cell(row=ri, column=2, value=val).alignment = Alignment(vertical="center")
        if ri % 2 == 0:
            for ci in range(1, 3):
                ws1.cell(row=ri, column=ci).fill = ALT_FILL

    ws1.column_dimensions["A"].width = 36
    ws1.column_dimensions["B"].width = 22

    # Sheet 2 — Funnel
    funnel_headers = ["Giai đoạn", "Số lượt ứng tuyển", "Tỷ lệ chuyển đổi"]
    funnel_rows = [
        [s.stage_label, s.count, _fmt_pct(s.conversion_rate)]
        for s in funnel.stages
    ]
    _setup_ws(wb.create_sheet("Funnel tuyển dụng"), "Funnel tuyển dụng", funnel_headers, funnel_rows)

    # Sheet 3 — Kênh tuyển dụng
    ch_headers = ["Kênh", "Ứng viên", "Đã tuyển", "Tỷ lệ tuyển"]
    ch_rows = [[c.channel_name, c.total_candidates, c.hired_count, _fmt_pct(c.hire_rate)] for c in channels]
    _setup_ws(wb.create_sheet("Theo kênh tuyển dụng"), "Theo kênh tuyển dụng", ch_headers, ch_rows)

    # Sheet 4 — Theo phòng ban
    dept_headers = ["Phòng ban", "Tổng JR", "Đang tuyển", "Đã tuyển", "T/g tuyển TB", "OAR", "Chi phí/người"]
    dept_rows = [
        [
            d.department_name, d.total_jr, d.open_jr, d.hired_count,
            _fmt_days(d.avg_time_to_hire), _fmt_pct(d.offer_acceptance_rate),
            _fmt_vnd(d.cost_per_hire),
        ]
        for d in depts
    ]
    _setup_ws(wb.create_sheet("Theo phòng ban"), "Theo phòng ban", dept_headers, dept_rows)

    # Sheet 5 — Xu hướng tháng
    trend_headers = ["Tháng", "Lượt ứng tuyển", "Đã tuyển", "T/g tuyển TB (ngày)", "T/g lấp chỗ TB (ngày)"]
    trend_rows = [
        [
            f"T{m.month:02d}/{m.year}",
            m.applications_count, m.hired_count,
            m.avg_time_to_hire or "—", m.avg_time_to_fill or "—",
        ]
        for m in time_report.monthly
    ]
    _setup_ws(wb.create_sheet("Xu hướng tháng"), "Xu hướng tháng", trend_headers, trend_rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
