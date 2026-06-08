"""Import chức danh hàng loạt từ file Excel."""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org import JobTitle
from app.schemas.employee_import import ImportResult, ImportRowError
from app.schemas.job_title import JobTitleCreate, JobTitleUpdate
from app.services.import_excel_helper import extract_header_and_non_blank_rows
from app.services import job_title_service

IMPORT_MAX_ROWS = 1000
COLUMNS = ["Mã chức danh", "Tên chức danh", "Cấp bậc", "Kích hoạt"]
REQUIRED_COLUMNS = {"Mã chức danh", "Tên chức danh"}


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Chức danh"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")
    opt_font = Font(color="1F2937", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font if col_name in REQUIRED_COLUMNS else opt_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample = ["TPKD", "Trưởng phòng kinh doanh", "4", "1"]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)
    widths = [18, 32, 14, 12]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Ví dụ"],
        ["Mã chức danh", "✅", "Mã duy nhất của chức danh", "TPKD"],
        ["Tên chức danh", "✅", "Tên chức danh", "Trưởng phòng kinh doanh"],
        ["Cấp bậc", "", "Số nguyên >= 1. 1 là cao nhất", "4"],
        ["Kích hoạt", "", "1/true/yes = hoạt động; 0/false/no = khóa", "1"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 20
    guide.column_dimensions["C"].width = 42
    guide.column_dimensions["D"].width = 28
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    value = row_vals[idx] if idx < len(row_vals) else None
    return str(value).strip() if value is not None else ""


def _parse_bool(raw: str, row: int, col: str, errors: list[ImportRowError]) -> Optional[bool]:
    if not raw:
        return None
    value = raw.strip().lower()
    if value in {"1", "true", "yes", "y", "co", "có"}:
        return True
    if value in {"0", "false", "no", "n", "khong", "không"}:
        return False
    errors.append(ImportRowError(row=row, column=col, message=f"Giá trị không hợp lệ: '{raw}'"))
    return None


def _parse_level(raw: str, row: int, errors: list[ImportRowError]) -> Optional[int]:
    if not raw:
        return 1
    try:
        level = int(raw)
    except ValueError:
        errors.append(ImportRowError(row=row, column="Cấp bậc", message=f"Giá trị phải là số nguyên: '{raw}'"))
        return None
    if level < 1:
        errors.append(ImportRowError(row=row, column="Cấp bậc", message="Cấp bậc phải >= 1"))
        return None
    return level


async def _get_by_code(session: AsyncSession, code: str) -> Optional[JobTitle]:
    result = await session.execute(
        select(JobTitle).where(JobTitle.code == code.strip().upper(), JobTitle.deleted_at.is_(None))
    )
    return result.scalar_one_or_none()


async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    header_row, data_rows = extract_header_and_non_blank_rows(ws)
    if not header_row:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    col_map = {header: idx for idx, header in enumerate(header_row)}
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = success = failed = 0
    errors: list[ImportRowError] = []
    created_ids: list[int] = []

    for excel_row, row_vals in data_rows:
        total += 1
        row_errors: list[ImportRowError] = []

        code = _cell(row_vals, "Mã chức danh", col_map).upper()
        name = _cell(row_vals, "Tên chức danh", col_map)
        level = _parse_level(_cell(row_vals, "Cấp bậc", col_map), excel_row, row_errors)
        is_active = _parse_bool(_cell(row_vals, "Kích hoạt", col_map), excel_row, "Kích hoạt", row_errors)

        if not code:
            row_errors.append(ImportRowError(row=excel_row, column="Mã chức danh", message="Trường bắt buộc không được để trống"))
        if not name:
            row_errors.append(ImportRowError(row=excel_row, column="Tên chức danh", message="Trường bắt buộc không được để trống"))
        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        existing = await _get_by_code(session, code)
        try:
            if existing:
                row = await job_title_service.update(
                    session,
                    existing.id,
                    JobTitleUpdate(name=name, level=level, is_active=is_active),
                )
            else:
                row = await job_title_service.create(
                    session,
                    JobTitleCreate(code=code, name=name, level=level or 1),
                )
                if is_active is False:
                    row = await job_title_service.update(session, row.id, JobTitleUpdate(is_active=False))
            created_ids.append(row.id)
            success += 1
        except Exception as exc:
            errors.append(ImportRowError(row=excel_row, column="Mã chức danh", message=str(exc)))
            failed += 1

    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
