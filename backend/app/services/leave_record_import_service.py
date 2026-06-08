"""Import nghỉ phép hàng loạt từ file Excel (12.1 — Slice 2)."""
from __future__ import annotations

import structlog

logger = structlog.get_logger(__name__)
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import and_, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import LeaveType
from app.models.leave_record import LeaveRecord
from app.schemas.employee_import import ImportResult, ImportRowError
from app.services import leave_entitlement_service
from app.services.import_excel_helper import extract_header_and_non_blank_rows
from app.services.import_employee_lookup_service import EmployeeImportLookup

IMPORT_MAX_ROWS = 1000

COLUMNS = [
    "Mã nhân viên",
    "Hệ mã nhân viên",
    "Mã loại phép",
    "Ngày bắt đầu",
    "Ngày kết thúc",
    "Ghi chú",
]
REQUIRED_COLUMNS = {"Mã nhân viên", "Mã loại phép", "Ngày bắt đầu", "Ngày kết thúc"}


# ── Template ──────────────────────────────────────────────────────────────────

def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nghỉ phép"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")
    opt_font = Font(color="1F2937", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font if col_name in REQUIRED_COLUMNS else opt_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample = ["1", "SYS1", "annual_leave", "01/06/2026", "03/06/2026", "Nghỉ phép năm"]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    widths = [16, 16, 18, 14, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 36

    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Giá trị hợp lệ / Ví dụ"],
        ["Mã nhân viên", "✅", "Mã hiển thị hoặc phần số nhân viên", "1 / 001 / NV001"],
        ["Hệ mã nhân viên", "", "Bắt buộc khi công ty dùng nhiều hệ mã", "SYS1 / SYS2 / SYS3"],
        ["Mã loại phép", "✅", "Mã loại phép nghỉ", "annual_leave / sick_leave / maternity_leave"],
        ["Ngày bắt đầu", "✅", "Định dạng dd/mm/yyyy", "01/06/2026"],
        ["Ngày kết thúc", "✅", "Định dạng dd/mm/yyyy (≥ ngày bắt đầu)", "03/06/2026"],
        ["Ghi chú", "", "Tùy chọn, tối đa 500 ký tự", "Nghỉ phép năm"],
        [],
        ["LƯU Ý:", "", "Số ngày nghỉ được tính tự động từ ngày bắt đầu đến kết thúc.", ""],
        ["", "", f"Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.", ""],
        ["", "", "Nếu NV đã có record trùng ngày → cảnh báo nhưng vẫn tạo.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 18
    guide.column_dimensions["C"].width = 46
    guide.column_dimensions["D"].width = 38
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val: object, col: str, row: int, errors: list[ImportRowError]) -> Optional[date]:
    from datetime import datetime
    raw = str(val).strip() if val is not None else ""
    if not raw:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    errors.append(ImportRowError(row=row, column=col, message=f"Định dạng ngày không hợp lệ: '{raw}' (cần dd/mm/yyyy)"))
    return None


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    v = row_vals[idx] if idx < len(row_vals) else None
    return str(v).strip() if v is not None else ""

async def _find_leave_type(session: AsyncSession, code: str) -> Optional[LeaveType]:
    result = await session.execute(select(LeaveType).where(LeaveType.code == code.strip()))
    return result.scalar_one_or_none()


async def _has_overlapping_record(
    session: AsyncSession, employee_id: int, start_date: date, end_date: date
) -> bool:
    result = await session.execute(
        select(LeaveRecord).where(
            and_(
                LeaveRecord.employee_id == employee_id,
                LeaveRecord.status == "active",
                LeaveRecord.start_date <= end_date,
                LeaveRecord.end_date >= start_date,
            )
        ).limit(1)
    )
    return result.scalar_one_or_none() is not None


# ── Main import ───────────────────────────────────────────────────────────────

async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("excel_read_error", error=str(exc))
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    header_row, data_rows = extract_header_and_non_blank_rows(ws)
    if not header_row:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    col_map = {h: i for i, h in enumerate(header_row)}
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = success = failed = 0
    errors: list[ImportRowError] = []
    warnings: list[ImportRowError] = []
    created_ids: list[int] = []
    employee_lookup = await EmployeeImportLookup.build(session)

    for excel_row, row_vals in data_rows:
        total += 1
        row_errors: list[ImportRowError] = []

        def get(col_name: str) -> str:
            return _cell(row_vals, col_name, col_map)

        emp_code      = get("Mã nhân viên")
        emp_sequence  = get("Hệ mã nhân viên")
        leave_type_code = get("Mã loại phép")
        notes         = get("Ghi chú")[:500] if get("Ghi chú") else None

        for field, val in [("Mã nhân viên", emp_code), ("Mã loại phép", leave_type_code)]:
            if not val:
                row_errors.append(ImportRowError(row=excel_row, column=field, message="Trường bắt buộc không được để trống"))

        start_date = _parse_date(get("Ngày bắt đầu"), "Ngày bắt đầu", excel_row, row_errors)
        end_date   = _parse_date(get("Ngày kết thúc"), "Ngày kết thúc", excel_row, row_errors)

        if not start_date and "Ngày bắt đầu" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày bắt đầu", message="Trường bắt buộc không được để trống"))
        if not end_date and "Ngày kết thúc" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày kết thúc", message="Trường bắt buộc không được để trống"))

        if start_date and end_date and end_date < start_date:
            row_errors.append(ImportRowError(
                row=excel_row, column="Ngày kết thúc",
                message="Ngày kết thúc phải sau hoặc bằng ngày bắt đầu"
            ))

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        # DB lookups
        employee_lookup_result = employee_lookup.resolve(
            employee_code_raw=emp_code,
            sequence_code_raw=emp_sequence,
        )
        employee = employee_lookup_result.employee
        if not employee:
            errors.append(ImportRowError(
                row=excel_row,
                column="Mã nhân viên",
                message=employee_lookup_result.error or f"Không tìm thấy nhân viên với mã '{emp_code}'",
            ))
            failed += 1
            continue

        leave_type = await _find_leave_type(session, leave_type_code)
        if not leave_type:
            errors.append(ImportRowError(row=excel_row, column="Mã loại phép", message=f"Không tìm thấy loại phép '{leave_type_code}'"))
            failed += 1
            continue

        # Cảnh báo trùng lịch (không chặn)
        if await _has_overlapping_record(session, employee.id, start_date, end_date):
            warnings.append(ImportRowError(
                row=excel_row, column="Ngày bắt đầu",
                message=f"Nhân viên đã có record nghỉ phép trùng với khoảng {start_date}–{end_date}"
            ))

        # Tính total_days = số ngày nghỉ (ngày làm việc kể cả đầu và cuối)
        total_days = Decimal((end_date - start_date).days + 1)

        try:
            entitlement = await leave_entitlement_service.ensure_entitlement_for_import(
                session,
                employee=employee,
                leave_type=leave_type,
                year=start_date.year,
            )
            entitlement.used_days += total_days
            session.add(entitlement)
            record = LeaveRecord(
                employee_id=employee.id,
                leave_type_id=leave_type.id,
                entitlement_id=entitlement.id,
                start_date=start_date,
                end_date=end_date,
                total_days=total_days,
                status="active",
                notes=notes,
            )
            session.add(record)
            await session.flush()
            created_ids.append(record.id)
            success += 1
        except IntegrityError:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="—", message="Lỗi ràng buộc dữ liệu (IntegrityError)"))
            failed += 1
        except Exception as exc:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="—", message=f"Lỗi hệ thống: {exc}"))
            failed += 1

    await session.commit()

    # Gộp warnings vào errors list (với prefix để FE phân biệt)
    all_errors = errors + [
        ImportRowError(row=w.row, column=w.column, message=f"[CẢNH BÁO] {w.message}")
        for w in warnings
    ]
    return ImportResult(
        total=total, success=success, failed=failed,
        errors=all_errors, created_ids=created_ids,
    )
