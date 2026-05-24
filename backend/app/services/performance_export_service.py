"""Service xuất Excel báo cáo hiệu suất (10.4)."""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.services.performance_report_service import (
    get_department_kpi_stats,
    get_rating_distribution,
)

_COMPANY     = "CÔNG TY TNHH HỒNG HÀ"
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT  = Font(bold=True)
_THIN        = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center")


def _sheet_title(ws, subtitle: str, ncols: int) -> None:
    col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{col}1")
    ws["A1"] = _COMPANY
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER
    ws.row_dimensions[1].height = 22

    ws.merge_cells(f"A2:{col}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = _CENTER
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6


def _header_row(ws, headers: list[str], row: int = 4) -> None:
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _THIN_BORDER
    ws.row_dimensions[row].height = 30


def _border_range(ws, min_row, max_row, min_col, max_col) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _THIN_BORDER


async def export_performance_excel(session: AsyncSession, year: int) -> bytes:
    dist   = await get_rating_distribution(session, year)
    dept_stats = await get_department_kpi_stats(session, year)

    wb = Workbook()

    # ── Sheet 1: Phân phối xếp loại ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Phân phối xếp loại"
    NCOLS1 = 3

    _sheet_title(ws1, f"PHÂN PHỐI XẾP LOẠI HIỆU SUẤT NĂM {year}", NCOLS1)
    _header_row(ws1, ["Xếp loại", "Số nhân viên", "Tỷ lệ (%)"])

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 14
    ws1.freeze_panes = "A5"

    for i, rc in enumerate(dist.distribution):
        r = 5 + i
        fill = _ALT_FILL if i % 2 else _WHITE_FILL
        ws1.cell(row=r, column=1, value=rc.rating_label).fill = fill
        ws1.cell(row=r, column=2, value=rc.count).fill = fill
        ws1.cell(row=r, column=3, value=rc.percentage).fill = fill
        ws1.cell(row=r, column=2).alignment = _CENTER
        ws1.cell(row=r, column=3).number_format = "0.0"
        ws1.cell(row=r, column=3).alignment = _CENTER
    _border_range(ws1, 5, 5 + len(dist.distribution) - 1, 1, NCOLS1)

    # Total row
    total_row = 5 + len(dist.distribution)
    ws1.cell(row=total_row, column=1, value="TỔNG CỘNG").font = _TOTAL_FONT
    ws1.cell(row=total_row, column=2, value=dist.total_reviewed).font = _TOTAL_FONT
    ws1.cell(row=total_row, column=3, value=100.0 if dist.total_reviewed else 0.0).font = _TOTAL_FONT
    ws1.cell(row=total_row, column=3).number_format = "0.0"
    for ci in range(1, NCOLS1 + 1):
        ws1.cell(row=total_row, column=ci).fill = _TOTAL_FILL
        ws1.cell(row=total_row, column=ci).border = _THIN_BORDER
    ws1.cell(row=total_row, column=1).alignment = _CENTER
    ws1.cell(row=total_row, column=2).alignment = _CENTER
    ws1.cell(row=total_row, column=3).alignment = _CENTER
    ws1.row_dimensions[total_row].height = 20

    # footnote
    fn_row = total_row + 2
    ws1.cell(row=fn_row, column=1,
             value=f"Tổng NV active: {dist.total_employees}   |   Tỷ lệ có đánh giá: {dist.coverage_rate}%")
    ws1.cell(row=fn_row, column=1).font = Font(italic=True)

    # ── Sheet 2: KPI theo phòng ban ───────────────────────────────────────────
    ws2 = wb.create_sheet("KPI theo phòng ban")
    NCOLS2 = 7

    _sheet_title(ws2, f"ĐIỂM KPI TRUNG BÌNH THEO PHÒNG BAN NĂM {year}", NCOLS2)
    _header_row(ws2, ["STT", "Phòng ban", "Số NV có KPI", "Điểm TB", "Điểm thấp nhất", "Điểm cao nhất", "Tổng lượt nhập"])

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 16
    ws2.column_dimensions["G"].width = 14
    ws2.freeze_panes = "A5"

    total_emp = total_months = 0
    for i, stat in enumerate(dept_stats):
        r = 5 + i
        fill = _ALT_FILL if i % 2 else _WHITE_FILL

        def _v(v): return float(v) if v is not None else None  # noqa: E731

        ws2.cell(row=r, column=1, value=i + 1).fill = fill
        ws2.cell(row=r, column=2, value=stat.department_name or "Chưa phân công").fill = fill
        ws2.cell(row=r, column=3, value=stat.employee_count).fill = fill
        ws2.cell(row=r, column=4, value=_v(stat.avg_score)).fill = fill
        ws2.cell(row=r, column=5, value=_v(stat.min_score)).fill = fill
        ws2.cell(row=r, column=6, value=_v(stat.max_score)).fill = fill
        ws2.cell(row=r, column=7, value=stat.months_data_count).fill = fill

        for ci in [1, 3, 4, 5, 6, 7]:
            ws2.cell(row=r, column=ci).alignment = _CENTER
        for ci in [4, 5, 6]:
            ws2.cell(row=r, column=ci).number_format = "0.0"

        total_emp    += stat.employee_count
        total_months += stat.months_data_count
    if dept_stats:
        _border_range(ws2, 5, 5 + len(dept_stats) - 1, 1, NCOLS2)

    # Total row
    t2_row = 5 + len(dept_stats)
    ws2.cell(row=t2_row, column=1, value="TỔNG CỘNG").font = _TOTAL_FONT
    ws2.cell(row=t2_row, column=3, value=total_emp).font = _TOTAL_FONT
    ws2.cell(row=t2_row, column=7, value=total_months).font = _TOTAL_FONT
    for ci in range(1, NCOLS2 + 1):
        ws2.cell(row=t2_row, column=ci).fill = _TOTAL_FILL
        ws2.cell(row=t2_row, column=ci).border = _THIN_BORDER
    ws2.cell(row=t2_row, column=1).alignment = _CENTER
    ws2.cell(row=t2_row, column=3).alignment = _CENTER
    ws2.cell(row=t2_row, column=7).alignment = _CENTER
    ws2.row_dimensions[t2_row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
