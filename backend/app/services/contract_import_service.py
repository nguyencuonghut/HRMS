"""Import hợp đồng hàng loạt từ file Excel (12.1 — Slice 1)."""
from __future__ import annotations

import structlog

logger = structlog.get_logger(__name__)
from datetime import date
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.catalog import ContractCategory
from app.models.employee_contract import EmployeeContract
from app.models.salary import BhxhPositionGroup
from app.schemas.employee_contract import ContractCreate
from app.schemas.employee_import import ImportResult, ImportRowError
from app.services import employee_contract_service
from app.services.import_employee_lookup_service import EmployeeImportLookup

IMPORT_MAX_ROWS = 1000
VALID_INSURANCE_SALARY_MODES = {"computed_by_position_group", "fixed_manual"}

COLUMNS = [
    "Mã nhân viên",
    "Hệ mã nhân viên",
    "Số hợp đồng",
    "Loại văn bản",
    "Mã loại hợp đồng",
    "Ngày ký",
    "Ngày hiệu lực",
    "Ngày hết hạn",
    "Mode lương BHXH",
    "Mã nhóm vị trí BHXH",
    "Bậc hệ số BHXH",
    "Ngày bắt đầu tính thâm niên BHXH",
    "Mức lương BHXH",
]
REQUIRED_COLUMNS = {
    "Mã nhân viên", "Số hợp đồng", "Loại văn bản",
    "Mã loại hợp đồng", "Ngày ký", "Ngày hiệu lực",
}

VALID_DOCUMENT_KINDS = {"labor_contract", "contract_appendix"}


# ── Template generation ───────────────────────────────────────────────────────

def generate_template() -> bytes:
    wb = Workbook()

    ws = wb.active
    ws.title = "Hợp đồng"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sample row
    sample = [
        "1", "SYS1", "HDLD/01/2026-001", "labor_contract", "labor_indefinite",
        "01/01/2026", "01/01/2026", "", "fixed_manual", "", "", "", "5000000",
    ]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    widths = [16, 16, 22, 20, 22, 14, 14, 14, 22, 22, 16, 24, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.row_dimensions[1].height = 36

    # Sheet Hướng dẫn
    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Giá trị hợp lệ / Ví dụ"],
        ["Mã nhân viên", "✅", "Mã hiển thị hoặc phần số nhân viên", "1 / 001 / NV001"],
        ["Hệ mã nhân viên", "", "Bắt buộc khi công ty dùng nhiều hệ mã", "SYS1 / SYS2 / SYS3"],
        ["Số hợp đồng", "✅", "Số hợp đồng duy nhất trong hệ thống", "HDLD/01/2026-001"],
        ["Loại văn bản", "✅", "Phân loại tài liệu hợp đồng", "labor_contract / contract_appendix"],
        ["Mã loại hợp đồng", "✅", "Mã danh mục loại hợp đồng", "labor_indefinite / labor_definite / probation_agreement"],
        ["Ngày ký", "✅", "Ngày ký kết hợp đồng (dd/mm/yyyy)", "01/01/2026"],
        ["Ngày hiệu lực", "✅", "Ngày bắt đầu có hiệu lực (dd/mm/yyyy)", "01/01/2026"],
        ["Ngày hết hạn", "", "Để trống = vô thời hạn (dd/mm/yyyy)", "31/12/2026"],
        ["Mode lương BHXH", "", "Mode tính lương BHXH", "fixed_manual / computed_by_position_group"],
        ["Mã nhóm vị trí BHXH", "", "Bắt buộc khi mode = computed_by_position_group", "OFFICE_STAFF / EXEC_COMPANY"],
        ["Bậc hệ số BHXH", "", "Bậc gốc 1..7, bắt buộc khi mode = computed_by_position_group", "1"],
        ["Ngày bắt đầu tính thâm niên BHXH", "", "Nếu để trống, hệ thống tự lấy theo employee.start_date", "01/01/2023"],
        ["Mức lương BHXH", "", "Bắt buộc khi mode = fixed_manual; để trống khi mode = computed_by_position_group", "5000000"],
        [],
        ["LƯU Ý:", "", "Cột header màu xanh đậm = bắt buộc.", ""],
        ["", "", f"Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.", ""],
        ["", "", "Dòng trống sẽ bị bỏ qua.", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["C"].width = 42
    guide.column_dimensions["D"].width = 38
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val: object, col: str, row: int, errors: list[ImportRowError]) -> Optional[date]:
    raw = str(val).strip() if val is not None else ""
    if not raw:
        return None
    from datetime import datetime
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            pass
    errors.append(ImportRowError(row=row, column=col, message=f"Định dạng ngày không hợp lệ: '{raw}' (cần dd/mm/yyyy)"))
    return None


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    v = row_vals[idx] if idx < len(row_vals) else None
    return str(v).strip() if v is not None else ""

async def _find_contract_category(session: AsyncSession, code: str) -> Optional[ContractCategory]:
    result = await session.execute(
        select(ContractCategory).where(ContractCategory.code == code.strip())
    )
    return result.scalar_one_or_none()


async def _find_bhxh_position_group(session: AsyncSession, code: str) -> Optional[BhxhPositionGroup]:
    result = await session.execute(
        select(BhxhPositionGroup).where(BhxhPositionGroup.code == code.strip())
    )
    return result.scalar_one_or_none()


# ── Main import ───────────────────────────────────────────────────────────────

async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("excel_read_error", error=str(exc))
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    header_row = [str(c).strip() if c else "" for c in all_rows[0]]
    col_map: dict[str, int] = {h: i for i, h in enumerate(header_row)}

    data_rows = all_rows[1:]
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = success = failed = 0
    errors: list[ImportRowError] = []
    created_ids: list[int] = []
    employee_lookup = await EmployeeImportLookup.build(session)

    for rel_idx, row_vals in enumerate(data_rows):
        excel_row = rel_idx + 2
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        total += 1
        row_errors: list[ImportRowError] = []

        def get(col_name: str) -> str:
            return _cell(row_vals, col_name, col_map)

        # ── Required string fields ────────────────────────────────────
        emp_code      = get("Mã nhân viên")
        emp_sequence  = get("Hệ mã nhân viên")
        contract_num  = get("Số hợp đồng")
        doc_kind      = get("Loại văn bản")
        category_code = get("Mã loại hợp đồng")

        for field, val in [
            ("Mã nhân viên", emp_code),
            ("Số hợp đồng", contract_num),
            ("Loại văn bản", doc_kind),
            ("Mã loại hợp đồng", category_code),
        ]:
            if not val:
                row_errors.append(ImportRowError(row=excel_row, column=field, message="Trường bắt buộc không được để trống"))

        if doc_kind and doc_kind not in VALID_DOCUMENT_KINDS:
            row_errors.append(ImportRowError(
                row=excel_row, column="Loại văn bản",
                message=f"Giá trị phải là labor_contract hoặc contract_appendix, nhận được: '{doc_kind}'"
            ))

        # ── Dates ────────────────────────────────────────────────────
        signed_date   = _parse_date(get("Ngày ký"),         "Ngày ký",        excel_row, row_errors)
        effective_from = _parse_date(get("Ngày hiệu lực"),  "Ngày hiệu lực",  excel_row, row_errors)
        effective_to   = _parse_date(get("Ngày hết hạn"),   "Ngày hết hạn",   excel_row, row_errors)

        if not signed_date and "Ngày ký" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày ký", message="Trường bắt buộc không được để trống"))
        if not effective_from and "Ngày hiệu lực" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày hiệu lực", message="Trường bắt buộc không được để trống"))

        if effective_from and effective_to and effective_to < effective_from:
            row_errors.append(ImportRowError(
                row=excel_row, column="Ngày hết hạn",
                message="Ngày hết hạn phải sau hoặc bằng ngày hiệu lực"
            ))

        # ── Optional: insurance/BHXH metadata ────────────────────────
        insurance_salary_mode_raw = get("Mode lương BHXH")
        insurance_salary_mode = insurance_salary_mode_raw or None
        if insurance_salary_mode and insurance_salary_mode not in VALID_INSURANCE_SALARY_MODES:
            row_errors.append(ImportRowError(
                row=excel_row,
                column="Mode lương BHXH",
                message="Mode lương BHXH phải là fixed_manual hoặc computed_by_position_group",
            ))

        bhxh_group_code = get("Mã nhóm vị trí BHXH")
        grade_raw = get("Bậc hệ số BHXH")
        insurance_salary_grade_no: Optional[int] = None
        if grade_raw:
            try:
                insurance_salary_grade_no = int(str(grade_raw))
            except ValueError:
                row_errors.append(ImportRowError(
                    row=excel_row,
                    column="Bậc hệ số BHXH",
                    message="Bậc hệ số BHXH phải là số nguyên từ 1 đến 7",
                ))
            else:
                if insurance_salary_grade_no < 1 or insurance_salary_grade_no > 7:
                    row_errors.append(ImportRowError(
                        row=excel_row,
                        column="Bậc hệ số BHXH",
                        message="Bậc hệ số BHXH phải từ 1 đến 7",
                    ))

        bhxh_seniority_start_date = _parse_date(
            get("Ngày bắt đầu tính thâm niên BHXH"),
            "Ngày bắt đầu tính thâm niên BHXH",
            excel_row,
            row_errors,
        )

        insurance_salary: Optional[Decimal] = None
        salary_raw = get("Mức lương BHXH")
        if salary_raw:
            try:
                insurance_salary = Decimal(salary_raw.replace(",", "").replace(".", ""))
                if insurance_salary <= 0:
                    row_errors.append(ImportRowError(row=excel_row, column="Mức lương BHXH", message="Mức lương BHXH phải > 0"))
                    insurance_salary = None
            except InvalidOperation:
                row_errors.append(ImportRowError(row=excel_row, column="Mức lương BHXH", message=f"Giá trị không hợp lệ: '{salary_raw}'"))

        if insurance_salary_mode is None:
            if bhxh_group_code or insurance_salary_grade_no is not None or bhxh_seniority_start_date is not None:
                insurance_salary_mode = "computed_by_position_group"
            elif insurance_salary is not None:
                insurance_salary_mode = "fixed_manual"

        if insurance_salary_mode == "computed_by_position_group":
            if not bhxh_group_code:
                row_errors.append(ImportRowError(
                    row=excel_row,
                    column="Mã nhóm vị trí BHXH",
                    message="Mode computed_by_position_group yêu cầu nhập Mã nhóm vị trí BHXH",
                ))
            if insurance_salary_grade_no is None:
                row_errors.append(ImportRowError(
                    row=excel_row,
                    column="Bậc hệ số BHXH",
                    message="Mode computed_by_position_group yêu cầu nhập Bậc hệ số BHXH",
                ))
            if salary_raw:
                row_errors.append(ImportRowError(
                    row=excel_row,
                    column="Mức lương BHXH",
                    message="Mode computed_by_position_group không nhận nhập tay Mức lương BHXH; hãy để trống để hệ thống tự tính",
                ))
        elif insurance_salary_mode == "fixed_manual" and insurance_salary is None and salary_raw == "":
            row_errors.append(ImportRowError(
                row=excel_row,
                column="Mức lương BHXH",
                message="Mode fixed_manual yêu cầu nhập Mức lương BHXH > 0",
            ))

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        # ── DB lookups ────────────────────────────────────────────────
        employee_lookup_result = employee_lookup.resolve(
            employee_code_raw=emp_code,
            sequence_code_raw=emp_sequence,
        )
        employee = employee_lookup_result.employee
        if not employee:
            errors.append(ImportRowError(
                row=excel_row,
                column="Mã nhân viên",
                message=employee_lookup_result.error or f"Không tìm thấy nhân viên với mã '{emp_code}'",
            ))
            failed += 1
            continue

        category = await _find_contract_category(session, category_code)
        if not category:
            errors.append(ImportRowError(row=excel_row, column="Mã loại hợp đồng", message=f"Không tìm thấy loại hợp đồng với mã '{category_code}'"))
            failed += 1
            continue

        # Validate document_kind matches category
        if category.document_kind != doc_kind:
            errors.append(ImportRowError(
                row=excel_row, column="Loại văn bản",
                message=f"Loại văn bản '{doc_kind}' không khớp với loại HĐ '{category_code}' (phải là '{category.document_kind}')"
            ))
            failed += 1
            continue

        bhxh_group: Optional[BhxhPositionGroup] = None
        if bhxh_group_code:
            bhxh_group = await _find_bhxh_position_group(session, bhxh_group_code)
            if not bhxh_group:
                errors.append(ImportRowError(
                    row=excel_row,
                    column="Mã nhóm vị trí BHXH",
                    message=f"Không tìm thấy nhóm vị trí BHXH với mã '{bhxh_group_code}'",
                ))
                failed += 1
                continue

        # Check contract_number unique
        existing = await session.execute(
            select(EmployeeContract).where(EmployeeContract.contract_number == contract_num)
        )
        if existing.scalar_one_or_none():
            errors.append(ImportRowError(row=excel_row, column="Số hợp đồng", message=f"Số hợp đồng '{contract_num}' đã tồn tại"))
            failed += 1
            continue

        # ── Create ────────────────────────────────────────────────────
        try:
            has_bhxh_config = (
                insurance_salary_mode is not None
                or insurance_salary is not None
                or bhxh_group is not None
                or insurance_salary_grade_no is not None
                or bhxh_seniority_start_date is not None
            )
            if has_bhxh_config:
                created = await employee_contract_service.create_contract(
                    session,
                    employee.id,
                    ContractCreate(
                        contract_category_id=category.id,
                        contract_number=contract_num,
                        signed_date=signed_date,
                        effective_from=effective_from,
                        effective_to=effective_to,
                        insurance_salary=insurance_salary if insurance_salary_mode == "fixed_manual" else None,
                        insurance_salary_mode=insurance_salary_mode,
                        bhxh_position_group_id=bhxh_group.id if bhxh_group else None,
                        insurance_salary_grade_no=insurance_salary_grade_no,
                        bhxh_seniority_start_date=bhxh_seniority_start_date,
                        insurance_salary_fixed_amount=insurance_salary if insurance_salary_mode == "fixed_manual" else None,
                    ),
                )
                created_ids.append(created.id)
            else:
                contract = EmployeeContract(
                    employee_id=employee.id,
                    contract_category_id=category.id,
                    document_kind=category.document_kind,
                    contract_number=contract_num,
                    signed_date=signed_date,
                    effective_from=effective_from,
                    effective_to=effective_to,
                    insurance_salary=None,
                    status="active",
                )
                session.add(contract)
                await session.flush()
                created_ids.append(contract.id)
            success += 1
        except HTTPException as exc:
            await session.rollback()
            errors.append(ImportRowError(
                row=excel_row,
                column="Mức lương BHXH",
                message=str(exc.detail),
            ))
            failed += 1
        except IntegrityError:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="Số hợp đồng", message=f"Số hợp đồng '{contract_num}' bị trùng (race condition)"))
            failed += 1
        except Exception as exc:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="—", message=f"Lỗi hệ thống: {exc}"))
            failed += 1

    await session.commit()
    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
