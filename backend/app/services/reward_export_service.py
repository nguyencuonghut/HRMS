"""Service xuất Excel khen thưởng – kỷ luật (8.3)."""
from __future__ import annotations

import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.ext.asyncio import AsyncSession

from app.services import discipline_service, reward_service

# ── Color constants ───────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL    = PatternFill("solid", fgColor="F2F2F2")
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TOTAL_FILL  = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT  = Font(bold=True)
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _apply_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = _THIN_BORDER


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


async def export_reward_discipline_excel(
    session: AsyncSession,
    *,
    from_date: date,
    to_date: date,
    department_id: Optional[int] = None,
    allowed_department_ids: Optional[list[int] | set[int] | tuple[int, ...]] = None,
    company_name: str = "CÔNG TY TNHH HỒNG HÀ",
) -> bytes:
    # Fetch ALL records (no pagination)
    reward_result = await reward_service.list_rewards(
        session,
        from_date=from_date,
        to_date=to_date,
        department_id=department_id,
        allowed_department_ids=allowed_department_ids,
        page=1,
        page_size=10000,
    )
    discipline_result = await discipline_service.list_disciplines(
        session,
        from_date=from_date,
        to_date=to_date,
        department_id=department_id,
        allowed_department_ids=allowed_department_ids,
        page=1,
        page_size=10000,
    )

    from_date_fmt = from_date.strftime("%d/%m/%Y")
    to_date_fmt   = to_date.strftime("%d/%m/%Y")

    wb = Workbook()

    # ── Sheet 1: Khen thưởng ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Khen thưởng"

    # Row 1 – company name
    ws1.merge_cells("A1:J1")
    ws1["A1"] = company_name
    ws1["A1"].font = Font(bold=True, size=14)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 22

    # Row 2 – title
    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"DANH SÁCH KHEN THƯỞNG {from_date_fmt} – {to_date_fmt}"
    ws1["A2"].font = Font(bold=True, size=12)
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Row 3 – empty
    ws1.row_dimensions[3].height = 6

    # Row 4 – headers
    headers1 = [
        "STT", "Mã NV", "Họ và tên", "Phòng ban",
        "Loại khen thưởng", "Tiêu đề / Nội dung", "Số quyết định",
        "Ngày khen thưởng", "Giá trị (VND)", "Đơn vị khen thưởng",
    ]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=4, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws1.row_dimensions[4].height = 30

    # Data rows (start row 5)
    rewards = reward_result.items
    for i, item in enumerate(rewards):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        values = [
            i + 1,
            item.employee_code,
            item.employee_name,
            item.department_name or "",
            item.reward_type_name,
            item.title,
            item.decision_number or "",
            item.reward_date,
            item.value,
            item.issued_by or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws1.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx == 8 and val is not None:  # date column
                cell.number_format = "DD/MM/YYYY"
            if col_idx == 9 and val is not None:  # value column
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    # Total row
    total_row1 = 5 + len(rewards)
    ws1.cell(row=total_row1, column=1, value="TỔNG CỘNG").font = _TOTAL_FONT
    ws1.cell(row=total_row1, column=1).fill = _TOTAL_FILL
    ws1.cell(row=total_row1, column=1).border = _THIN_BORDER
    ws1.cell(row=total_row1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(2, 11):
        cell = ws1.cell(row=total_row1, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _THIN_BORDER
        if col_idx == 9 and len(rewards) > 0:
            start_cell = f"I5"
            end_cell   = f"I{total_row1 - 1}"
            cell.value = f"=SUM({start_cell}:{end_cell})"
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right", vertical="center")

    ws1.row_dimensions[total_row1].height = 20

    _set_col_widths(ws1, [6, 12, 22, 20, 22, 30, 16, 14, 16, 20])
    ws1.freeze_panes = "A5"

    # ── Sheet 2: Kỷ luật ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Kỷ luật")

    # Row 1 – company name
    ws2.merge_cells("A1:K1")
    ws2["A1"] = company_name
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    # Row 2 – title
    ws2.merge_cells("A2:K2")
    ws2["A2"] = f"DANH SÁCH KỶ LUẬT {from_date_fmt} – {to_date_fmt}"
    ws2["A2"].font = Font(bold=True, size=12)
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 20

    # Row 3 – empty
    ws2.row_dimensions[3].height = 6

    # Row 4 – headers
    headers2 = [
        "STT", "Mã NV", "Họ và tên", "Phòng ban",
        "Hình thức kỷ luật", "Tiêu đề / Vi phạm", "Số quyết định",
        "Ngày vi phạm", "Ngày hiệu lực", "Ngày hết hiệu lực", "Đơn vị ký",
    ]
    for col_idx, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _THIN_BORDER
    ws2.row_dimensions[4].height = 30

    # Data rows
    disciplines = discipline_result.items
    for i, item in enumerate(disciplines):
        row = 5 + i
        fill = _ALT_FILL if i % 2 == 0 else _WHITE_FILL
        values = [
            i + 1,
            item.employee_code,
            item.employee_name,
            item.department_name or "",
            item.discipline_form_label,
            item.title,
            item.decision_number or "",
            item.violation_date,
            item.effective_date,
            item.end_date,
            item.issued_by or "",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.border = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if col_idx in (8, 9, 10) and val is not None:  # date columns
                cell.number_format = "DD/MM/YYYY"

    # Total row
    total_row2 = 5 + len(disciplines)
    total_cell = ws2.cell(row=total_row2, column=1, value="TỔNG CỘNG")
    total_cell.font = _TOTAL_FONT
    total_cell.fill = _TOTAL_FILL
    total_cell.border = _THIN_BORDER
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(2, 12):
        cell = ws2.cell(row=total_row2, column=col_idx)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _THIN_BORDER
    ws2.row_dimensions[total_row2].height = 20

    _set_col_widths(ws2, [6, 12, 22, 20, 25, 30, 16, 14, 14, 16, 20])
    ws2.freeze_panes = "A5"

    # ── Serialize to bytes ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
