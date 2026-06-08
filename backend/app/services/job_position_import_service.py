"""Import vị trí công việc hàng loạt từ file Excel."""
from __future__ import annotations

from io import BytesIO
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.org import Department, JobPosition, JobTitle
from app.schemas.employee_import import ImportResult, ImportRowError
from app.schemas.job_position import JobPositionCreate, JobPositionUpdate
from app.services.import_excel_helper import extract_header_and_non_blank_rows
from app.services import job_position_service

IMPORT_MAX_ROWS = 1000
COLUMNS = [
    "Mã vị trí",
    "Tên vị trí",
    "Mã phòng ban",
    "Mã chức danh",
    "Bậc mặc định",
    "Phụ cấp BHXH",
    "Phụ cấp ngoài BHXH",
    "Kích hoạt",
    "Mô tả",
    "Yêu cầu",
]
REQUIRED_COLUMNS = {"Mã vị trí", "Tên vị trí", "Mã phòng ban"}


def generate_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vị trí công việc"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")
    opt_font = Font(color="1F2937", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font if col_name in REQUIRED_COLUMNS else opt_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sample = ["KD_NV_01", "Nhân viên kinh doanh miền Bắc", "KD", "NVKD", "1", "500000", "300000", "1", "Chăm sóc đại lý", "Kỹ năng bán hàng"]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)
    widths = [18, 30, 16, 16, 14, 16, 18, 12, 28, 28]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Ví dụ"],
        ["Mã vị trí", "✅", "Mã duy nhất của vị trí công việc", "KD_NV_01"],
        ["Tên vị trí", "✅", "Tên hiển thị của vị trí", "Nhân viên kinh doanh miền Bắc"],
        ["Mã phòng ban", "✅", "Phải tồn tại trước khi import vị trí", "KD"],
        ["Mã chức danh", "", "Phải tồn tại trước khi import vị trí", "NVKD"],
        ["Bậc mặc định", "", "Số nguyên >= 1", "1"],
        ["Phụ cấp BHXH", "", "Số tiền VND >= 0", "500000"],
        ["Phụ cấp ngoài BHXH", "", "Số tiền VND >= 0", "300000"],
        ["Kích hoạt", "", "1/true/yes = hoạt động; 0/false/no = khóa", "1"],
        ["Mô tả", "", "Mô tả công việc", "Chăm sóc đại lý"],
        ["Yêu cầu", "", "Yêu cầu tuyển dụng", "Kỹ năng bán hàng"],
    ]
    for row in guide_rows:
        guide.append(row)
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["C"].width = 42
    guide.column_dimensions["D"].width = 28
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cell(row_vals: tuple, col_name: str, col_map: dict[str, int]) -> str:
    idx = col_map.get(col_name)
    if idx is None:
        return ""
    value = row_vals[idx] if idx < len(row_vals) else None
    return str(value).strip() if value is not None else ""


def _parse_bool(raw: str, row: int, col: str, errors: list[ImportRowError]) -> Optional[bool]:
    if not raw:
        return None
    value = raw.strip().lower()
    if value in {"1", "true", "yes", "y", "co", "có"}:
        return True
    if value in {"0", "false", "no", "n", "khong", "không"}:
        return False
    errors.append(ImportRowError(row=row, column=col, message=f"Giá trị không hợp lệ: '{raw}'"))
    return None


def _parse_int(raw: str, row: int, col: str, errors: list[ImportRowError], *, minimum: int = 0, default: Optional[int] = None) -> Optional[int]:
    if not raw:
        return default
    try:
        value = int(str(raw).replace(",", "").strip())
    except ValueError:
        errors.append(ImportRowError(row=row, column=col, message=f"Giá trị phải là số nguyên: '{raw}'"))
        return None
    if value < minimum:
        errors.append(ImportRowError(row=row, column=col, message=f"Giá trị phải >= {minimum}"))
        return None
    return value


async def _get_department_by_code(session: AsyncSession, code: str) -> Optional[Department]:
    result = await session.execute(
        select(Department).where(Department.code == code.strip().upper(), Department.deleted_at.is_(None))
    )
    return result.scalar_one_or_none()


async def _get_job_title_by_code(session: AsyncSession, code: str) -> Optional[JobTitle]:
    result = await session.execute(
        select(JobTitle).where(JobTitle.code == code.strip().upper(), JobTitle.deleted_at.is_(None))
    )
    return result.scalar_one_or_none()


async def _get_position_by_code(session: AsyncSession, code: str) -> Optional[JobPosition]:
    result = await session.execute(
        select(JobPosition).where(JobPosition.code == code.strip().upper(), JobPosition.deleted_at.is_(None))
    )
    return result.scalar_one_or_none()


async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    header_row, data_rows = extract_header_and_non_blank_rows(ws)
    if not header_row:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    col_map = {header: idx for idx, header in enumerate(header_row)}
    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = success = failed = 0
    errors: list[ImportRowError] = []
    created_ids: list[int] = []

    for excel_row, row_vals in data_rows:
        total += 1
        row_errors: list[ImportRowError] = []

        code = _cell(row_vals, "Mã vị trí", col_map).upper()
        name = _cell(row_vals, "Tên vị trí", col_map)
        department_code = _cell(row_vals, "Mã phòng ban", col_map).upper()
        job_title_code = _cell(row_vals, "Mã chức danh", col_map).upper()
        default_grade = _parse_int(_cell(row_vals, "Bậc mặc định", col_map), excel_row, "Bậc mặc định", row_errors, minimum=1, default=1)
        bhxh_allowance = _parse_int(_cell(row_vals, "Phụ cấp BHXH", col_map), excel_row, "Phụ cấp BHXH", row_errors, minimum=0, default=0)
        non_bhxh_allowance = _parse_int(_cell(row_vals, "Phụ cấp ngoài BHXH", col_map), excel_row, "Phụ cấp ngoài BHXH", row_errors, minimum=0, default=0)
        is_active = _parse_bool(_cell(row_vals, "Kích hoạt", col_map), excel_row, "Kích hoạt", row_errors)
        description = _cell(row_vals, "Mô tả", col_map) or None
        requirements = _cell(row_vals, "Yêu cầu", col_map) or None

        if not code:
            row_errors.append(ImportRowError(row=excel_row, column="Mã vị trí", message="Trường bắt buộc không được để trống"))
        if not name:
            row_errors.append(ImportRowError(row=excel_row, column="Tên vị trí", message="Trường bắt buộc không được để trống"))
        if not department_code:
            row_errors.append(ImportRowError(row=excel_row, column="Mã phòng ban", message="Trường bắt buộc không được để trống"))

        department = None
        if department_code:
            department = await _get_department_by_code(session, department_code)
            if not department:
                row_errors.append(ImportRowError(row=excel_row, column="Mã phòng ban", message=f"Không tìm thấy phòng ban với mã '{department_code}'"))

        job_title = None
        if job_title_code:
            job_title = await _get_job_title_by_code(session, job_title_code)
            if not job_title:
                row_errors.append(ImportRowError(row=excel_row, column="Mã chức danh", message=f"Không tìm thấy chức danh với mã '{job_title_code}'"))

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        existing = await _get_position_by_code(session, code)
        try:
            if existing:
                row = await job_position_service.update(
                    session,
                    existing.id,
                    JobPositionUpdate(
                        name=name,
                        department_id=department.id,
                        job_title_id=job_title.id if job_title else None,
                        default_grade=default_grade,
                        bhxh_allowance=bhxh_allowance,
                        non_bhxh_allowance=non_bhxh_allowance,
                        is_active=is_active,
                        description=description,
                        requirements=requirements,
                    ),
                )
            else:
                row = await job_position_service.create(
                    session,
                    JobPositionCreate(
                        code=code,
                        name=name,
                        department_id=department.id,
                        job_title_id=job_title.id if job_title else None,
                        default_grade=default_grade or 1,
                        bhxh_allowance=bhxh_allowance or 0,
                        non_bhxh_allowance=non_bhxh_allowance or 0,
                        description=description,
                        requirements=requirements,
                    ),
                )
                if is_active is False:
                    row = await job_position_service.update(session, row.id, JobPositionUpdate(is_active=False))
            created_ids.append(row.id)
            success += 1
        except Exception as exc:
            errors.append(ImportRowError(row=excel_row, column="Mã vị trí", message=str(exc)))
            failed += 1

    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
