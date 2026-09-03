"""Import nhân viên hàng loạt từ file Excel (3.7)."""

from __future__ import annotations

import re
import structlog

logger = structlog.get_logger(__name__)
import unicodedata
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.core.encryption import hash_sensitive
from app.models.catalog import AdministrativeHierarchy, AdministrativeUnit, Ethnicity, Nationality, Religion
from app.models.employee import Employee
from app.models.employee_code import EmployeeCodeSequence
from app.models.org import Department, DepartmentJobPosition, JobPosition, JobTitle
from app.schemas.employee import EmployeeAddressWrite, EmployeeCreate
from app.schemas.employee_import import (
    GENDER_MAP,
    IMPORT_COLUMNS,
    IMPORT_MAX_ROWS,
    REQUIRED_COLUMNS,
    STATUS_MAP,
    ImportResult,
    ImportRowError,
)
from app.services import administrative_unit_service, employee_service
from app.services.employee_code_service import compute_employee_display_code
from app.services.import_excel_helper import extract_header_and_non_blank_rows


# ── Template generation ───────────────────────────────────────────────────────

async def _fetch_new_system_rows(session: AsyncSession) -> list[tuple[str, str]]:
    Province = aliased(AdministrativeUnit)
    Ward = aliased(AdministrativeUnit)
    q = (
        select(Province.name, Ward.name)
        .select_from(AdministrativeHierarchy)
        .join(Province, Province.id == AdministrativeHierarchy.parent_unit_id)
        .join(Ward, Ward.id == AdministrativeHierarchy.child_unit_id)
        .where(
            AdministrativeHierarchy.system_type == "new",
            AdministrativeHierarchy.is_active == True,
            Province.is_active == True,
            Ward.is_active == True,
        )
        .order_by(Province.name, Ward.name)
    )
    return list((await session.execute(q)).all())


async def _fetch_old_system_rows(session: AsyncSession) -> list[tuple[str, str, str]]:
    h_province_district = aliased(AdministrativeHierarchy)
    h_district_ward = aliased(AdministrativeHierarchy)
    Province = aliased(AdministrativeUnit)
    District = aliased(AdministrativeUnit)
    Ward = aliased(AdministrativeUnit)
    q = (
        select(Province.name, District.name, Ward.name)
        .select_from(h_province_district)
        .join(h_district_ward, h_district_ward.parent_unit_id == h_province_district.child_unit_id)
        .join(Province, Province.id == h_province_district.parent_unit_id)
        .join(District, District.id == h_province_district.child_unit_id)
        .join(Ward, Ward.id == h_district_ward.child_unit_id)
        .where(
            h_province_district.system_type == "old",
            h_district_ward.system_type == "old",
            h_province_district.is_active == True,
            h_district_ward.is_active == True,
            Province.is_active == True,
            District.is_active == True,
            Ward.is_active == True,
        )
        .order_by(Province.name, District.name, Ward.name)
    )
    return list((await session.execute(q)).all())


def _build_reference_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple]) -> None:
    ws = wb.create_sheet(title)
    header_fill = PatternFill("solid", fgColor="4B6B8A")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
    for row in rows:
        ws.append(list(row))
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = 28
    ws.freeze_panes = "A2"


async def generate_template(session: AsyncSession) -> bytes:
    wb = Workbook()

    # Sheet 1: Data
    ws = wb.active
    ws.title = "Nhân viên"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    opt_fill = PatternFill("solid", fgColor="D6E4F0")
    opt_font = Font(color="1F2937", bold=True)

    for col_idx, col_name in enumerate(IMPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font if col_name in REQUIRED_COLUMNS else opt_font
        cell.fill = header_fill if col_name in REQUIRED_COLUMNS else opt_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Sample row
    sample = [
        "Nguyễn Văn A", "Nguyễn", "Văn A",
        "01/01/1990", "nam", "123456789012",
        "01/01/2020", "Cục Cảnh sát ĐKQLCƯ",
        "probation", "01/01/2026",
        "0901234567", "vana@email.com",
        "1234567890", "BHXH123456", "Việt Nam", "Kinh", "Không",
        "HC", "Chuyên viên nhân sự", "", "SYS1", "1024", "HC1024",
        "01/01/2026", "31/03/2026",
        "thôn Vạc, Xã Đường An, Thành phố Hải Phòng",
        "thôn Vạc, Xã Thái Học, Huyện Bình Giang, Tỉnh Hải Dương",
        "Số 12, Xã Đường An, Thành phố Hải Phòng",
        "",
    ]
    for col_idx, val in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Column widths
    widths = [20, 12, 12, 14, 10, 16, 14, 24, 12, 14, 14, 22, 14, 14, 16, 16, 16, 14, 22, 22, 18, 16, 18, 20, 20, 45, 55, 45, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.row_dimensions[1].height = 36

    # Sheet 2: Hướng dẫn
    guide = wb.create_sheet("Hướng dẫn")
    guide_rows = [
        ["Cột", "Bắt buộc", "Mô tả", "Giá trị hợp lệ / Ví dụ"],
        ["Họ và tên", "✅", "Họ tên đầy đủ", "Nguyễn Văn A"],
        ["Họ", "✅", "Họ (last name)", "Nguyễn"],
        ["Tên", "✅", "Tên đệm + tên (first name)", "Văn A"],
        ["Ngày sinh", "✅", "Định dạng dd/mm/yyyy", "01/01/1990"],
        ["Giới tính", "✅", "Chữ thường", "nam / nữ / khác"],
        ["Số CCCD/CMND", "✅", "Số giấy tờ tùy thân", "123456789012"],
        ["Ngày cấp CCCD", "✅", "Định dạng dd/mm/yyyy", "01/01/2020"],
        ["Nơi cấp CCCD", "✅", "Cơ quan cấp", "Cục Cảnh sát ĐKQLCƯ"],
        ["Trạng thái", "✅", "Trạng thái nhân viên", "probation / official / long_leave"],
        ["Ngày vào làm", "✅", "Định dạng dd/mm/yyyy", "01/01/2026"],
        ["Số điện thoại", "", "Tùy chọn", "0901234567"],
        ["Email cá nhân", "", "Tùy chọn", "email@example.com"],
        ["Mã số thuế", "", "Mã số thuế thu nhập cá nhân", "1234567890"],
        ["Số BHXH", "", "Mã BHXH", "BHXH123456"],
        ["Quốc tịch", "", "Tên, mã hoặc ISO quốc tịch để auto-map sang catalog", "Việt Nam / VN"],
        ["Dân tộc", "", "Tên hoặc mã dân tộc để auto-map sang catalog", "Kinh / KINH"],
        ["Tôn giáo", "", "Tên hoặc mã tôn giáo để auto-map sang catalog", "Không / KHONG"],
        ["Phòng ban", "", "Mã hoặc tên phòng ban (tìm kiếm tương đối)", "HC"],
        ["Chức danh", "", "Tên chức danh (tìm kiếm tương đối)", "Chuyên viên nhân sự"],
        ["Vị trí công việc", "", "Tên vị trí công việc, nên đi cùng phòng ban", "Nhân viên nhân sự tổng hợp"],
        [
            "Hệ mã nhân viên",
            "",
            "Hệ mã để xác định employee code; bắt buộc nếu nhập mã nhân viên cũ",
            "SYS1 = Hệ 1 (mặc định toàn công ty); SYS2 = Hệ 2 (công nhân bốc xếp / ra cám / tạp vụ); SYS3 = Hệ 3 (công nhân / bảo vệ thuộc Phòng trại)",
        ],
        ["Số thứ tự mã NV", "", "Seq cũ nếu cần giữ nguyên mã nhân viên hiện hữu", "1024"],
        ["Mã NV hiện hữu", "", "Dùng để đối chiếu. Hệ thống sẽ báo lỗi nếu mã tính ra không khớp", "HC1024"],
        ["Ngày bắt đầu thử việc", "", "Định dạng dd/mm/yyyy", "01/01/2026"],
        ["Ngày kết thúc thử việc", "", "Định dạng dd/mm/yyyy", "31/03/2026"],
        [
            "Địa chỉ thường trú (Hệ mới 2 cấp)",
            "✅",
            "Các phần cách nhau bởi dấu phẩy, theo thứ tự: chi tiết (số nhà/thôn xóm), Xã/Phường, Tỉnh/Thành. "
            "PHẢI copy đúng chính tả tên Xã/Phường và Tỉnh/Thành từ sheet 'DM Hệ mới'.",
            "thôn Vạc, Xã Đường An, Thành phố Hải Phòng",
        ],
        [
            "Địa chỉ thường trú (Hệ cũ 3 cấp)",
            "",
            "Tùy chọn — để trống nếu không có địa chỉ hệ cũ. Nếu điền thì phải đủ 3 cấp, cách nhau bởi dấu phẩy: "
            "chi tiết, Xã/Phường, Huyện/Quận, Tỉnh/Thành. PHẢI copy đúng chính tả từ sheet 'DM Hệ cũ'.",
            "thôn Vạc, Xã Thái Học, Huyện Bình Giang, Tỉnh Hải Dương",
        ],
        [
            "Địa chỉ liên lạc (Hệ mới 2 cấp)",
            "✅",
            "Địa chỉ liên lạc/tạm trú — cùng định dạng và quy tắc với 'Địa chỉ thường trú (Hệ mới 2 cấp)'.",
            "Số 12, Xã Đường An, Thành phố Hải Phòng",
        ],
        [
            "Địa chỉ liên lạc (Hệ cũ 3 cấp)",
            "",
            "Địa chỉ liên lạc/tạm trú hệ cũ — tùy chọn, cùng quy tắc với 'Địa chỉ thường trú (Hệ cũ 3 cấp)'.",
            "",
        ],
        [],
        ["LƯU Ý:", "", "Cột header màu xanh đậm = bắt buộc; xanh nhạt = không bắt buộc. Dòng trống sẽ bị bỏ qua.", ""],
        ["", "", f"Tối đa {IMPORT_MAX_ROWS} dòng dữ liệu mỗi lần import.", ""],
        ["", "", "Tên Tỉnh/Huyện/Xã phải khớp chính xác (không phân biệt hoa/thường, có dấu/không dấu) với danh mục ở sheet 'DM Hệ mới' và 'DM Hệ cũ'.", ""],
    ]
    for row in guide_rows:
        guide.append(row)

    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["C"].width = 40
    guide.column_dimensions["D"].width = 30
    for cell in guide["1"]:
        cell.font = Font(bold=True)

    # Sheet 3: Danh mục địa chỉ hệ mới (tham chiếu)
    new_rows = await _fetch_new_system_rows(session)
    _build_reference_sheet(wb, "DM Hệ mới", ["Tỉnh/Thành phố", "Xã/Phường"], new_rows)

    # Sheet 4: Danh mục địa chỉ hệ cũ (tham chiếu)
    old_rows = await _fetch_old_system_rows(session)
    _build_reference_sheet(wb, "DM Hệ cũ", ["Tỉnh/Thành phố", "Huyện/Quận", "Xã/Phường"], old_rows)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Parse helpers ─────────────────────────────────────────────────────────────

def _norm(s: str) -> str:
    """Chuẩn hoá Unicode NFC, bỏ dấu, lowercase — dùng để tra cứu fuzzy."""
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()


# ── Địa chỉ (thường trú / liên lạc) ────────────────────────────────────────────

def _split_address_text(raw: str, admin_levels: int) -> tuple[str, list[str]]:
    """Tách chuỗi địa chỉ tự do theo dấu phẩy, lấy N phần cuối làm địa danh.

    Tách từ phải sang trái để phần chi tiết (số nhà/thôn xóm) có chứa dấu phẩy
    không làm lệch vị trí các cấp hành chính.
    """
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    if len(parts) < admin_levels:
        raise ValueError(
            f"Địa chỉ cần tối thiểu {admin_levels} phần cách nhau bởi dấu phẩy, "
            f"chỉ thấy {len(parts)} phần: '{raw}'"
        )
    admin_parts = parts[-admin_levels:]
    detail = ", ".join(parts[:-admin_levels])
    return detail, admin_parts


async def _find_province_exact(
    session: AsyncSession, system_type: str, name: str
) -> Optional[AdministrativeUnit]:
    candidates = await administrative_unit_service.list_provinces(session, system_type=system_type)
    target = _norm(name)
    for c in candidates:
        if _norm(c.name) == target:
            return c
    return None


async def _find_child_exact(
    session: AsyncSession, system_type: str, parent_id: int, name: str
) -> Optional[AdministrativeUnit]:
    candidates = await administrative_unit_service.list_children(
        session, system_type=system_type, parent_id=parent_id
    )
    target = _norm(name)
    for c in candidates:
        if _norm(c.name) == target:
            return c
    return None


async def _resolve_address_column(
    session: AsyncSession,
    raw_text: str,
    *,
    system_type: str,
    admin_levels: int,
    column_name: str,
    excel_row: int,
    errors: list[ImportRowError],
) -> Optional[dict]:
    """Parse + resolve 1 cột địa chỉ tự do thành các field id/text của EmployeeAddress."""
    raw_text = (raw_text or "").strip()
    if not raw_text:
        return None

    try:
        detail, admin_parts = _split_address_text(raw_text, admin_levels)
    except ValueError as exc:
        errors.append(ImportRowError(row=excel_row, column=column_name, message=str(exc)))
        return None

    sheet_name = "DM Hệ mới" if system_type == "new" else "DM Hệ cũ"

    if admin_levels == 2:
        ward_name, province_name = admin_parts
        province = await _find_province_exact(session, system_type, province_name)
        if not province:
            errors.append(ImportRowError(
                row=excel_row, column=column_name,
                message=f"Không tìm thấy tỉnh/thành '{province_name}' — chuỗi gốc: '{raw_text}'. Đối chiếu sheet '{sheet_name}'.",
            ))
            return None
        ward = await _find_child_exact(session, system_type, province.id, ward_name)
        if not ward:
            errors.append(ImportRowError(
                row=excel_row, column=column_name,
                message=f"Không tìm thấy xã/phường '{ward_name}' thuộc tỉnh '{province.name}' — chuỗi gốc: '{raw_text}'. Đối chiếu sheet '{sheet_name}'.",
            ))
            return None
        return {
            "province_unit_id": province.id,
            "ward_unit_id": ward.id,
            "address_line": detail or None,
        }

    # Hệ cũ — 3 cấp: xã, huyện, tỉnh
    ward_name, district_name, province_name = admin_parts
    province = await _find_province_exact(session, system_type, province_name)
    if not province:
        errors.append(ImportRowError(
            row=excel_row, column=column_name,
            message=f"Không tìm thấy tỉnh/thành '{province_name}' — chuỗi gốc: '{raw_text}'. Đối chiếu sheet '{sheet_name}'.",
        ))
        return None
    district = await _find_child_exact(session, system_type, province.id, district_name)
    if not district:
        errors.append(ImportRowError(
            row=excel_row, column=column_name,
            message=f"Không tìm thấy huyện/quận '{district_name}' thuộc tỉnh '{province.name}' — chuỗi gốc: '{raw_text}'. Đối chiếu sheet '{sheet_name}'.",
        ))
        return None
    ward = await _find_child_exact(session, system_type, district.id, ward_name)
    if not ward:
        errors.append(ImportRowError(
            row=excel_row, column=column_name,
            message=f"Không tìm thấy xã/phường '{ward_name}' thuộc huyện '{district.name}' — chuỗi gốc: '{raw_text}'. Đối chiếu sheet '{sheet_name}'.",
        ))
        return None
    return {
        "province_unit_id": province.id,
        "district_unit_id": district.id,
        "ward_unit_id": ward.id,
        "address_line": detail or None,
    }


async def _build_address_payload(
    session: AsyncSession,
    *,
    address_type: str,
    new_raw: str,
    old_raw: str,
    new_column_name: str,
    old_column_name: str,
    excel_row: int,
    errors: list[ImportRowError],
) -> Optional[EmployeeAddressWrite]:
    new_resolved = await _resolve_address_column(
        session, new_raw, system_type="new", admin_levels=2,
        column_name=new_column_name, excel_row=excel_row, errors=errors,
    )
    if not (new_raw or "").strip():
        errors.append(ImportRowError(row=excel_row, column=new_column_name, message="Trường bắt buộc không được để trống"))
    if new_resolved is None:
        return None

    old_resolved = await _resolve_address_column(
        session, old_raw, system_type="old", admin_levels=3,
        column_name=old_column_name, excel_row=excel_row, errors=errors,
    )

    return EmployeeAddressWrite(
        address_type=address_type,
        new_province_unit_id=new_resolved["province_unit_id"],
        new_ward_unit_id=new_resolved["ward_unit_id"],
        new_address_line=new_resolved["address_line"],
        old_province_unit_id=old_resolved["province_unit_id"] if old_resolved else None,
        old_district_unit_id=old_resolved["district_unit_id"] if old_resolved else None,
        old_ward_unit_id=old_resolved["ward_unit_id"] if old_resolved else None,
        old_address_line=old_resolved["address_line"] if old_resolved else None,
    )


def _parse_date(val: str, col: str, row: int, errors: list[ImportRowError]) -> Optional[date]:
    val = (val or "").strip()
    if not val:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return date.fromisoformat(val) if fmt == "%Y-%m-%d" else date(
                *[int(x) for x in re.split(r"[/\-]", val)][::-1]
                if fmt == "%d/%m/%Y" else [int(x) for x in re.split(r"[/\-]", val)]
            )
        except (ValueError, TypeError):
            pass
    # Try strptime fallback
    from datetime import datetime
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val, fmt).date()
        except ValueError:
            pass
    errors.append(ImportRowError(row=row, column=col, message=f"Định dạng ngày không hợp lệ: '{val}' (cần dd/mm/yyyy)"))
    return None


def _cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


# ── Department/JobTitle lookup ────────────────────────────────────────────────

async def _find_department(session: AsyncSession, name_or_code: str) -> Optional[int]:
    """Tra cứu department theo code (exact) rồi theo name (normalized)."""
    if not name_or_code:
        return None
    val = name_or_code.strip()
    # exact code
    row = await session.execute(select(Department).where(Department.code == val))
    dept = row.scalar_one_or_none()
    if dept:
        return dept.id
    # normalized name
    norm_val = _norm(val)
    rows = await session.execute(select(Department))
    for d in rows.scalars().all():
        if _norm(d.name) == norm_val or _norm(d.short_name or "") == norm_val:
            return d.id
    return None


async def _find_job_title(session: AsyncSession, name: str) -> Optional[int]:
    if not name:
        return None
    norm_val = _norm(name.strip())
    rows = await session.execute(select(JobTitle))
    for jt in rows.scalars().all():
        if _norm(jt.name) == norm_val:
            return jt.id
    return None


async def _find_job_position(
    session: AsyncSession,
    name_or_code: str,
    *,
    department_id: int | None = None,
) -> Optional[JobPosition]:
    if not name_or_code:
        return None

    raw = name_or_code.strip()
    query = select(JobPosition).where(JobPosition.deleted_at.is_(None))
    if department_id is not None:
        query = query.join(
            DepartmentJobPosition,
            (DepartmentJobPosition.job_position_id == JobPosition.id)
            & (DepartmentJobPosition.is_active.is_(True)),
        ).where(DepartmentJobPosition.department_id == department_id)

    exact = (await session.execute(query.where(JobPosition.code == raw))).scalar_one_or_none()
    if exact:
        return exact

    norm_val = _norm(raw)
    rows = await session.execute(query)
    for position in rows.scalars().all():
        if _norm(position.name) == norm_val:
            return position
    return None


async def _find_job_positions_any_department(
    session: AsyncSession,
    name_or_code: str,
) -> list[JobPosition]:
    if not name_or_code:
        return []

    raw = name_or_code.strip()
    exact = (
        await session.execute(select(JobPosition).where(JobPosition.code == raw, JobPosition.deleted_at.is_(None)))
    ).scalars().all()
    if exact:
        return list(exact)

    norm_val = _norm(raw)
    rows = await session.execute(select(JobPosition).where(JobPosition.deleted_at.is_(None)))
    return [position for position in rows.scalars().all() if _norm(position.name) == norm_val]


async def _find_employee_code_sequence(
    session: AsyncSession,
    code_or_name: str,
) -> Optional[int]:
    if not code_or_name:
        return None
    raw = code_or_name.strip()
    row = (
        await session.execute(
            select(EmployeeCodeSequence).where(
                EmployeeCodeSequence.is_active.is_(True),
                EmployeeCodeSequence.code == raw,
            )
        )
    ).scalar_one_or_none()
    if row:
        return row.id

    norm_val = _norm(raw)
    rows = await session.execute(
        select(EmployeeCodeSequence).where(EmployeeCodeSequence.is_active.is_(True))
    )
    for sequence in rows.scalars().all():
        if _norm(sequence.name) == norm_val:
            return sequence.id
    return None


def _parse_positive_int(
    raw_value: str,
    col: str,
    row: int,
    errors: list[ImportRowError],
) -> int | None:
    raw = (raw_value or "").strip()
    if not raw:
        return None
    try:
        value = int(raw)
    except ValueError:
        errors.append(
            ImportRowError(row=row, column=col, message=f"Giá trị phải là số nguyên dương, nhận được: '{raw}'")
        )
        return None
    if value <= 0:
        errors.append(
            ImportRowError(row=row, column=col, message=f"Giá trị phải lớn hơn 0, nhận được: '{raw}'")
        )
        return None
    return value


def _compute_expected_display_code(
    *,
    employee_seq: int,
    sequence: EmployeeCodeSequence,
    department: Department | None,
) -> str:
    prefix = None
    if department is not None:
        prefix = (department.display_prefix or "").strip() or department.code
    return compute_employee_display_code(
        employee_seq,
        prefix,
        min_digits=sequence.min_digits,
    )


async def _get_default_nationality_id(session: AsyncSession) -> int:
    nationality = (
        await session.execute(
            select(Nationality).where(Nationality.code == "VN")
        )
    ).scalar_one_or_none()
    if not nationality:
        raise ValueError("Không tìm thấy quốc tịch mặc định 'VN' trong danh mục quốc tịch.")
    return nationality.id


async def _find_nationality(session: AsyncSession, raw_value: str) -> Optional[int]:
    if not raw_value:
        return None
    raw = raw_value.strip()
    norm_val = _norm(raw)
    rows = await session.execute(
        select(Nationality).where(Nationality.is_active.is_(True))
    )
    for nationality in rows.scalars().all():
        if (
            nationality.code.lower() == raw.lower()
            or (nationality.iso2_code and nationality.iso2_code.lower() == raw.lower())
            or (nationality.iso3_code and nationality.iso3_code.lower() == raw.lower())
            or nationality.normalized_name == norm_val
        ):
            return nationality.id
    return None


async def _find_ethnicity(session: AsyncSession, raw_value: str) -> Optional[int]:
    if not raw_value:
        return None
    raw = raw_value.strip()
    norm_val = _norm(raw)
    rows = await session.execute(
        select(Ethnicity).where(Ethnicity.is_active.is_(True))
    )
    for ethnicity in rows.scalars().all():
        if ethnicity.code.lower() == raw.lower() or ethnicity.normalized_name == norm_val:
            return ethnicity.id
    return None


async def _find_religion(session: AsyncSession, raw_value: str) -> Optional[int]:
    if not raw_value:
        return None
    raw = raw_value.strip()
    norm_val = _norm(raw)
    rows = await session.execute(
        select(Religion).where(Religion.is_active.is_(True))
    )
    for religion in rows.scalars().all():
        if religion.code.lower() == raw.lower() or religion.normalized_name == norm_val:
            return religion.id
    return None


# ── Main import ───────────────────────────────────────────────────────────────

async def process_import(session: AsyncSession, file_bytes: bytes) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("excel_read_error", error=str(exc))
        raise ValueError("Không đọc được file Excel. Hãy dùng file .xlsx hợp lệ.") from exc

    ws = wb.worksheets[0]
    header_row, data_rows = extract_header_and_non_blank_rows(ws)
    if not header_row:
        return ImportResult(total=0, success=0, failed=0, errors=[], created_ids=[])

    # Map header → col index (0-based)
    col_idx: dict[str, int] = {h: i for i, h in enumerate(header_row)}

    def get(row_vals: tuple, col_name: str) -> str:
        idx = col_idx.get(col_name)
        if idx is None:
            return ""
        v = row_vals[idx] if idx < len(row_vals) else None
        if v is None:
            return ""
        if isinstance(v, datetime):
            return v.strftime("%d/%m/%Y")
        if isinstance(v, date):
            return v.strftime("%d/%m/%Y")
        return str(v).strip()

    if len(data_rows) > IMPORT_MAX_ROWS:
        raise ValueError(f"File có quá nhiều dòng. Tối đa {IMPORT_MAX_ROWS} dòng mỗi lần import.")

    total = 0
    success = 0
    failed = 0
    errors: list[ImportRowError] = []
    created_ids: list[int] = []
    default_nationality_id = await _get_default_nationality_id(session)

    for excel_row, row_vals in data_rows:
        total += 1
        row_errors: list[ImportRowError] = []

        # ── Required fields ───────────────────────────────────────────
        full_name  = get(row_vals, "Họ và tên")
        last_name  = get(row_vals, "Họ")
        first_name = get(row_vals, "Tên")
        id_number  = get(row_vals, "Số CCCD/CMND")
        id_issued_by = get(row_vals, "Nơi cấp CCCD")

        for field, val in [
            ("Họ và tên", full_name), ("Họ", last_name), ("Tên", first_name),
            ("Số CCCD/CMND", id_number), ("Nơi cấp CCCD", id_issued_by),
        ]:
            if not val:
                row_errors.append(ImportRowError(row=excel_row, column=field, message="Trường bắt buộc không được để trống"))

        gender_raw = get(row_vals, "Giới tính").lower()
        gender = GENDER_MAP.get(gender_raw)
        if not gender:
            row_errors.append(ImportRowError(row=excel_row, column="Giới tính", message=f"Giá trị phải là nam/nữ/khác, nhận được: '{gender_raw}'"))

        status_raw = get(row_vals, "Trạng thái").lower()
        status = STATUS_MAP.get(status_raw)
        if not status:
            row_errors.append(ImportRowError(row=excel_row, column="Trạng thái", message=f"Giá trị phải là probation/official/long_leave, nhận được: '{status_raw}'"))

        date_of_birth  = _parse_date(get(row_vals, "Ngày sinh"),       "Ngày sinh",       excel_row, row_errors)
        id_issued_on   = _parse_date(get(row_vals, "Ngày cấp CCCD"),   "Ngày cấp CCCD",   excel_row, row_errors)
        start_date     = _parse_date(get(row_vals, "Ngày vào làm"),    "Ngày vào làm",    excel_row, row_errors)
        prob_start     = _parse_date(get(row_vals, "Ngày bắt đầu thử việc"), "Ngày bắt đầu thử việc", excel_row, row_errors)
        prob_end       = _parse_date(get(row_vals, "Ngày kết thúc thử việc"), "Ngày kết thúc thử việc", excel_row, row_errors)

        if not date_of_birth and "Ngày sinh" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày sinh", message="Trường bắt buộc không được để trống"))
        if not id_issued_on and "Ngày cấp CCCD" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày cấp CCCD", message="Trường bắt buộc không được để trống"))
        if not start_date and "Ngày vào làm" not in {e.column for e in row_errors}:
            row_errors.append(ImportRowError(row=excel_row, column="Ngày vào làm", message="Trường bắt buộc không được để trống"))

        # Optional fields
        phone        = get(row_vals, "Số điện thoại") or None
        email        = get(row_vals, "Email cá nhân") or None
        tax_code     = get(row_vals, "Mã số thuế") or None
        bhxh_code    = get(row_vals, "Số BHXH") or None
        nationality_raw = get(row_vals, "Quốc tịch")
        ethnicity_raw = get(row_vals, "Dân tộc")
        religion_raw = get(row_vals, "Tôn giáo")
        dept_raw     = get(row_vals, "Phòng ban")
        jt_raw       = get(row_vals, "Chức danh")
        position_raw = get(row_vals, "Vị trí công việc")
        sequence_raw = get(row_vals, "Hệ mã nhân viên")
        employee_seq_raw = get(row_vals, "Số thứ tự mã NV")
        display_code_raw = get(row_vals, "Mã NV hiện hữu")
        permanent_new_raw = get(row_vals, "Địa chỉ thường trú (Hệ mới 2 cấp)")
        permanent_old_raw = get(row_vals, "Địa chỉ thường trú (Hệ cũ 3 cấp)")
        contact_new_raw   = get(row_vals, "Địa chỉ liên lạc (Hệ mới 2 cấp)")
        contact_old_raw   = get(row_vals, "Địa chỉ liên lạc (Hệ cũ 3 cấp)")
        employee_seq = _parse_positive_int(
            employee_seq_raw,
            "Số thứ tự mã NV",
            excel_row,
            row_errors,
        )

        if row_errors:
            errors.extend(row_errors)
            failed += 1
            continue

        # ── DB checks ─────────────────────────────────────────────────
        existing = await session.execute(
            select(Employee).where(Employee.id_number_hash == hash_sensitive(id_number))
        )
        if existing.scalar_one_or_none():
            errors.append(ImportRowError(row=excel_row, column="Số CCCD/CMND", message=f"Số CCCD/CMND '{id_number}' đã tồn tại trong hệ thống"))
            failed += 1
            continue

        dept_id      = await _find_department(session, dept_raw)
        job_title_id = await _find_job_title(session, jt_raw)
        job_position = await _find_job_position(session, position_raw, department_id=dept_id) if position_raw else None
        sequence_id  = await _find_employee_code_sequence(session, sequence_raw) if sequence_raw else None
        sequence = await session.get(EmployeeCodeSequence, sequence_id) if sequence_id else None
        nationality_id = await _find_nationality(session, nationality_raw) if nationality_raw else default_nationality_id
        ethnicity_id = await _find_ethnicity(session, ethnicity_raw) if ethnicity_raw else None
        religion_id = await _find_religion(session, religion_raw) if religion_raw else None

        if dept_raw and not dept_id:
            errors.append(ImportRowError(row=excel_row, column="Phòng ban", message=f"Không tìm thấy phòng ban '{dept_raw}' — nhân viên được tạo nhưng chưa gán phòng ban"))
        if nationality_raw and not nationality_id:
            errors.append(ImportRowError(row=excel_row, column="Quốc tịch", message=f"Không tìm thấy quốc tịch '{nationality_raw}' trong danh mục"))
            failed += 1
            continue
        if ethnicity_raw and not ethnicity_id:
            errors.append(ImportRowError(row=excel_row, column="Dân tộc", message=f"Không tìm thấy dân tộc '{ethnicity_raw}' trong danh mục"))
            failed += 1
            continue
        if religion_raw and not religion_id:
            errors.append(ImportRowError(row=excel_row, column="Tôn giáo", message=f"Không tìm thấy tôn giáo '{religion_raw}' trong danh mục"))
            failed += 1
            continue
        if position_raw and not dept_id:
            errors.append(ImportRowError(row=excel_row, column="Vị trí công việc", message="Muốn gán vị trí công việc thì phải xác định được phòng ban"))
            failed += 1
            continue
        if position_raw and not job_position:
            candidate_positions = await _find_job_positions_any_department(session, position_raw)
            if candidate_positions and dept_id:
                selected_dept = await session.get(Department, dept_id)
                if len(candidate_positions) == 1:
                    actual_dept = await session.get(Department, candidate_positions[0].department_id)
                    actual_label = (
                        f"{actual_dept.code} - {actual_dept.name}"
                        if actual_dept
                        else f"ID {candidate_positions[0].department_id}"
                    )
                    selected_label = (
                        f"{selected_dept.code} - {selected_dept.name}"
                        if selected_dept
                        else f"ID {dept_id}"
                    )
                    message = (
                        f"Vị trí công việc '{position_raw}' có tồn tại nhưng thuộc phòng ban "
                        f"'{actual_label}', không thuộc phòng ban đã chọn '{selected_label}'"
                    )
                else:
                    department_labels: list[str] = []
                    for candidate in candidate_positions[:3]:
                        actual_dept = await session.get(Department, candidate.department_id)
                        department_labels.append(
                            f"{actual_dept.code} - {actual_dept.name}"
                            if actual_dept
                            else f"ID {candidate.department_id}"
                        )
                    message = (
                        f"Vị trí công việc '{position_raw}' có tồn tại nhưng không thuộc phòng ban đã chọn. "
                        f"Tìm thấy ở: {', '.join(department_labels)}"
                    )
            else:
                message = f"Không tìm thấy vị trí công việc '{position_raw}' trong phòng ban đã chọn"
            errors.append(ImportRowError(row=excel_row, column="Vị trí công việc", message=message))
            failed += 1
            continue
        if sequence_raw and not sequence_id:
            errors.append(ImportRowError(row=excel_row, column="Hệ mã nhân viên", message=f"Không tìm thấy hệ mã nhân viên '{sequence_raw}'"))
            failed += 1
            continue
        if employee_seq is not None and not sequence_id:
            errors.append(
                ImportRowError(
                    row=excel_row,
                    column="Hệ mã nhân viên",
                    message="Muốn giữ mã nhân viên cũ thì phải nhập Hệ mã nhân viên",
                )
            )
            failed += 1
            continue
        if display_code_raw and (employee_seq is None or not sequence):
            errors.append(
                ImportRowError(
                    row=excel_row,
                    column="Mã NV hiện hữu",
                    message="Muốn đối chiếu Mã NV hiện hữu thì phải nhập cả Hệ mã nhân viên và Số thứ tự mã NV",
                )
            )
            failed += 1
            continue
        if display_code_raw and employee_seq is not None and sequence is not None:
            department = await session.get(Department, dept_id) if dept_id else None
            expected_display_code = _compute_expected_display_code(
                employee_seq=employee_seq,
                sequence=sequence,
                department=department,
            )
            if display_code_raw.strip() != expected_display_code:
                errors.append(
                    ImportRowError(
                        row=excel_row,
                        column="Mã NV hiện hữu",
                        message=(
                            f"Mã NV hiện hữu '{display_code_raw}' không khớp với mã hệ thống tính ra "
                            f"'{expected_display_code}' từ Hệ mã nhân viên + Số thứ tự mã NV + đơn vị hiện tại"
                        ),
                    )
                )
                failed += 1
                continue

        addr_errors: list[ImportRowError] = []
        permanent_payload = await _build_address_payload(
            session,
            address_type="permanent",
            new_raw=permanent_new_raw,
            old_raw=permanent_old_raw,
            new_column_name="Địa chỉ thường trú (Hệ mới 2 cấp)",
            old_column_name="Địa chỉ thường trú (Hệ cũ 3 cấp)",
            excel_row=excel_row,
            errors=addr_errors,
        )
        contact_payload = await _build_address_payload(
            session,
            address_type="contact",
            new_raw=contact_new_raw,
            old_raw=contact_old_raw,
            new_column_name="Địa chỉ liên lạc (Hệ mới 2 cấp)",
            old_column_name="Địa chỉ liên lạc (Hệ cũ 3 cấp)",
            excel_row=excel_row,
            errors=addr_errors,
        )
        if addr_errors:
            errors.extend(addr_errors)
            failed += 1
            continue

        # ── Create employee ───────────────────────────────────────────
        try:
            payload = EmployeeCreate(
                employee_seq=employee_seq,
                full_name=full_name,
                last_name=last_name,
                first_name=first_name,
                date_of_birth=date_of_birth,
                gender=gender,
                nationality_id=nationality_id,
                ethnicity_id=ethnicity_id,
                religion_id=religion_id,
                id_number=id_number,
                id_issued_on=id_issued_on,
                id_issued_by=id_issued_by,
                phone_number=phone,
                personal_email=email,
                personal_tax_code=tax_code,
                bhxh_code=bhxh_code,
                status=status,
                start_date=start_date,
                employee_code_sequence_id=sequence_id,
                initial_department_id=dept_id,
                initial_job_title_id=job_title_id,
                initial_job_position_id=job_position.id if job_position else None,
                initial_job_effective_from=start_date if dept_id else None,
                initial_probation_start_date=prob_start,
                initial_probation_end_date=prob_end,
            )
            emp = await employee_service.create_employee(session, payload)
            await employee_service.upsert_employee_address(session, emp.id, permanent_payload)
            await employee_service.upsert_employee_address(session, emp.id, contact_payload)

            await session.commit()
            created_ids.append(emp.id)
            success += 1

        except IntegrityError:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="Số CCCD/CMND", message=f"Số CCCD/CMND '{id_number}' bị trùng (race condition)"))
            failed += 1
        except HTTPException as exc:
            await session.rollback()
            column = "Vị trí công việc" if "vị trí công việc" in str(exc.detail).lower() else "—"
            errors.append(ImportRowError(row=excel_row, column=column, message=str(exc.detail)))
            failed += 1
        except Exception as exc:
            await session.rollback()
            errors.append(ImportRowError(row=excel_row, column="—", message=f"Lỗi hệ thống: {exc}"))
            failed += 1

    return ImportResult(total=total, success=success, failed=failed, errors=errors, created_ids=created_ids)
