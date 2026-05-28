from __future__ import annotations

import io
from typing import Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelStyler:
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    ALT_FILL = PatternFill("solid", fgColor="F6F8FB")
    THIN_BORDER = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    @classmethod
    def build_table_workbook(
        cls,
        title: str,
        headers: Sequence[str],
        rows: Iterable[Sequence[object]],
        sheet_name: str = "Bao cao",
    ) -> io.BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]

        if headers:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            title_cell = ws.cell(1, 1, title)
            title_cell.font = Font(bold=True, size=14, color="1E3A5F")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 22

        for column_index, header in enumerate(headers, start=1):
            cell = ws.cell(3, column_index, header)
            cell.font = cls.HEADER_FONT
            cell.fill = cls.HEADER_FILL
            cell.border = cls.THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_index, row in enumerate(rows, start=4):
            for column_index, value in enumerate(row, start=1):
                cell = ws.cell(row_index, column_index, value)
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if row_index % 2 == 0:
                    cell.fill = cls.ALT_FILL

        cls.autofit(ws)
        ws.freeze_panes = "A4"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    @staticmethod
    def autofit(ws: object) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 3, 10), 45)
